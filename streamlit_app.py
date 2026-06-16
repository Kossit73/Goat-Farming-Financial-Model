"""Interactive dashboard for the goat farming financial model."""

from __future__ import annotations

from contextlib import nullcontext
from copy import deepcopy
from importlib.util import find_spec
from io import BytesIO
import json
from pathlib import Path
import re
import sys
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

from copy import deepcopy

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pandas.api.types import (
    is_bool_dtype,
    is_datetime64_any_dtype,
    is_integer_dtype,
    is_numeric_dtype,
)
from pandas.tseries.offsets import MonthEnd, QuarterEnd
from streamlit.delta_generator import DeltaGenerator

_APP_ROOT = Path(__file__).resolve().parent
_SRC_ROOT = _APP_ROOT / "src"
if str(_SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(_SRC_ROOT))

from goat_financial_model import GoatModel, InputSchedule
from goat_financial_model.assumption_bundle import AssumptionBundle, build_assumption_bundle
from goat_financial_model.commercial_services import (
    build_pricing_validation_messages,
    sync_commercial_assumptions_to_core as sync_commercial_assumptions_to_core_service,
)
from goat_financial_model.editor_registry import ScheduleEditorSpec, render_incremental_schedule_editor
from goat_financial_model.pages.assumptions_page import (
    BiologicalEditorDefinition,
    render_biological_assumption_editor,
    render_herd_plan_editor,
    render_operating_cost_editor,
    render_pricing_manual_editor,
)
from goat_financial_model.pages.input_schedule_page import render_cogs_schedule_editor, render_schedule_summary
from goat_financial_model.pricing_engine import build_pricing_context, derive_pricing_quantities
from goat_financial_model.reporting import generate_excel_workbook, prepare_export_bundle
from goat_financial_model.scenario_runner import (
    ScenarioBuildHooks,
    ScenarioOutputSpec,
    run_scenario_suite,
)
from goat_financial_model.table_registry import ColumnSchema, TableSchema, build_default_table, ensure_table


try:
    from streamlit.runtime.scriptrunner import get_script_run_ctx
except Exception:  # pragma: no cover - fallback for older Streamlit builds
    get_script_run_ctx = None

try:  # pragma: no cover - import guard for Streamlit API variations
    from streamlit.errors import StreamlitAPIException
except Exception:  # pragma: no cover - older versions exposed the exception elsewhere
    StreamlitAPIException = Exception


_LOCAL_SESSION_STATE: Dict[str, Any] = {}
_CACHED_DEFAULT_VALUATION_INPUTS: Optional[Dict[str, float]] = None
MODEL_RESULTS_STALE_KEY = "model_results_stale"
MODEL_RUN_SIGNATURE_KEY = "model_last_run_signature"
MODEL_INPUT_VERSION_KEY = "model_input_version"
MODEL_LAST_RUN_VERSION_KEY = "model_last_run_version"
MODEL_VIEW_CACHE_KEY = "model_view_cache"
MODEL_VALIDATION_CACHE_KEY = "model_validation_cache"


def _can_rerun() -> bool:
    """Return True when the app is executing within a Streamlit runtime."""

    if get_script_run_ctx is None:
        return False
    try:
        return get_script_run_ctx() is not None
    except Exception:  # pragma: no cover - defensive guard for API changes
        return False


def _maybe_rerun() -> None:
    """Invoke Streamlit rerun when a runtime context is active."""

    rerun_fn = getattr(st, "experimental_rerun", None) or getattr(st, "rerun", None)
    if rerun_fn is None or not _can_rerun():
        return
    rerun_fn()


def _safe_session_state_get(key: str, default: Any = None) -> Any:
    """Return a session state value without raising outside a Streamlit run."""

    if not _can_rerun():
        return _LOCAL_SESSION_STATE.get(key, default)
    try:
        return st.session_state.get(key, default)
    except Exception:
        try:
            return st.session_state[key] if key in st.session_state else default
        except Exception:
            return _LOCAL_SESSION_STATE.get(key, default)


def _safe_session_state_setdefault(key: str, value: Any) -> Any:
    """Set a default session state value when a runtime context exists."""

    if not _can_rerun():
        return _LOCAL_SESSION_STATE.setdefault(key, value)
    try:
        return st.session_state.setdefault(key, value)
    except Exception:
        try:
            if key not in st.session_state:
                st.session_state[key] = value
            return st.session_state[key]
        except Exception:
            pass
        return _LOCAL_SESSION_STATE.setdefault(key, value)


def _safe_session_state_set(key: str, value: Any) -> None:
    """Assign a session state value when supported by the runtime."""

    if not _can_rerun():
        _LOCAL_SESSION_STATE[key] = value
        return
    try:
        st.session_state[key] = value
    except Exception:
        try:
            st.session_state[key] = value
            return
        except Exception:
            _LOCAL_SESSION_STATE[key] = value


def _safe_session_state_contains(key: str) -> bool:
    """Return True when the session state currently tracks the key."""

    if not _can_rerun():
        return key in _LOCAL_SESSION_STATE
    try:
        return key in st.session_state
    except Exception:
        try:
            return key in st.session_state
        except Exception:
            return key in _LOCAL_SESSION_STATE


def _safe_session_state_pop(key: str, default: Any = None) -> Any:
    """Remove a session state key without raising when unavailable."""

    if not _can_rerun():
        return _LOCAL_SESSION_STATE.pop(key, default)
    try:
        return st.session_state.pop(key, default)
    except Exception:
        try:
            if key in st.session_state:
                value = st.session_state[key]
                del st.session_state[key]
                return value
            return default
        except Exception:
            return _LOCAL_SESSION_STATE.pop(key, default)


def _frames_equal(left: Optional[pd.DataFrame], right: Optional[pd.DataFrame]) -> bool:
    """Return True when two dataframes contain the same values."""

    if left is right:
        return True
    if left is None or right is None:
        return left is None and right is None
    if not isinstance(left, pd.DataFrame) or not isinstance(right, pd.DataFrame):
        return False
    try:
        pd.testing.assert_frame_equal(
            left.reset_index(drop=True),
            right.reset_index(drop=True),
            check_dtype=False,
            check_like=False,
        )
    except AssertionError:
        return False
    return True


def _bump_model_input_version() -> int:
    """Advance the lightweight input version used for stale-state tracking."""

    current = int(_safe_session_state_get(MODEL_INPUT_VERSION_KEY, 0) or 0)
    next_value = current + 1
    _safe_session_state_set(MODEL_INPUT_VERSION_KEY, next_value)
    return next_value


def _cached_input_value(cache_name: str, builder: Callable[[], Any], *, extra: str = "") -> Any:
    """Return a value cached for the current input version."""

    version = int(_safe_session_state_get(MODEL_INPUT_VERSION_KEY, 0) or 0)
    cache = _safe_session_state_setdefault(MODEL_VALIDATION_CACHE_KEY, {})
    key = f"{version}|{cache_name}|{extra}"
    if key not in cache:
        cache[key] = builder()
        _safe_session_state_set(MODEL_VALIDATION_CACHE_KEY, cache)
    return cache[key]


def _cached_result_view(cache_name: str, scenario_name: str, builder: Callable[[], Any], *, extra: str = "") -> Any:
    """Return a value cached for the active model run and selected scenario."""

    run_version = _safe_session_state_get(MODEL_LAST_RUN_VERSION_KEY)
    cache = _safe_session_state_setdefault(MODEL_VIEW_CACHE_KEY, {})
    key = f"{run_version}|{scenario_name}|{cache_name}|{extra}"
    if key not in cache:
        cache[key] = builder()
        _safe_session_state_set(MODEL_VIEW_CACHE_KEY, cache)
    return cache[key]


def _section_selector(
    label: str,
    options: Sequence[str],
    key: str,
    *,
    help: Optional[str] = None,
) -> str:
    choices = list(options)
    if not choices:
        raise ValueError(f"{label} requires at least one option.")
    current = _safe_session_state_get(key, choices[0])
    if current not in choices:
        current = choices[0]
        _safe_session_state_set(key, current)
    return st.radio(
        label,
        options=choices,
        index=choices.index(current),
        key=key,
        horizontal=True,
        help=help,
    )

st.set_page_config(page_title="Goat Farm Financial Model", layout="wide")


AI_PROVIDER_OPTIONS = ("OpenAI", "Azure OpenAI", "Anthropic")

DEFAULT_VALUATION_INPUTS = {
    "WACC": 0.10,
    "NPV": 0.0,
    "IRR": 0.0,
    "Terminal Value": 0.0,
    "Terminal Growth Rate": 0.02,
    "Receivable Days": 12.0,
    "Inventory Days": 20.0,
    "Payable Days": 45.0,
    "Minimum Cash Reserve": 25000.0,
    "DSCR Covenant": 1.20,
    "Interest Coverage Covenant": 1.50,
}

DERIVED_VALUATION_METRICS = frozenset({"NPV", "IRR"})


def _editable_valuation_input_defaults() -> Dict[str, float]:
    return {
        metric: value
        for metric, value in DEFAULT_VALUATION_INPUTS.items()
        if metric not in DERIVED_VALUATION_METRICS
    }

ML_METHOD_LABELS = {
    "linear_regression": "Linear Regression",
    "random_forest": "Random Forest",
    "gradient_boosting": "Gradient Boosting",
}

ML_LABEL_TO_CODE = {label: code for code, label in ML_METHOD_LABELS.items()}

GEN_AI_FEATURE_LABELS = {
    "summary": "Executive Summary",
    "risk_analysis": "Risk Analysis",
    "opportunity_analysis": "Opportunity Analysis",
}

GEN_AI_LABEL_TO_CODE = {label: code for code, label in GEN_AI_FEATURE_LABELS.items()}


DEFAULT_MODEL_AUTHOR = "Goat Farmers United"
MODEL_AUTHOR_KEY = "model_author"
MODEL_AUTHOR_WIDGET_KEY = "_model_author_widget"
MODEL_AUTHOR_CACHE_KEY = "_model_author_cached"


def _inject_app_theme() -> None:
    st.markdown(
        """
        <style>
        :root {
            --goat-ink: #1f2937;
            --goat-muted: #4b5563;
            --goat-brand: #0f766e;
            --goat-brand-soft: #e8f7f3;
            --goat-panel: rgba(255, 255, 255, 0.88);
        }
        .stApp {
            background:
                radial-gradient(circle at top left, rgba(187, 247, 208, 0.28), transparent 34%),
                radial-gradient(circle at top right, rgba(191, 219, 254, 0.22), transparent 28%),
                linear-gradient(180deg, #f4fbf8 0%, #f5f7fb 58%, #eef6f4 100%);
        }
        .block-container {
            padding-top: 1.35rem;
            padding-bottom: 3rem;
            max-width: 1440px;
        }
        .designer-hero {
            margin-bottom: 1.2rem;
            padding: 1.7rem 1.8rem;
            border-radius: 28px;
            border: 1px solid rgba(15, 118, 110, 0.12);
            background:
                linear-gradient(135deg, rgba(232, 247, 243, 0.95), rgba(255, 255, 255, 0.94)),
                linear-gradient(135deg, rgba(15, 118, 110, 0.05), rgba(59, 130, 246, 0.07));
            box-shadow: 0 24px 48px rgba(15, 23, 42, 0.08);
        }
        .designer-kicker {
            margin: 0 0 0.45rem 0;
            font-size: 0.78rem;
            letter-spacing: 0.16em;
            text-transform: uppercase;
            color: var(--goat-brand);
            font-weight: 700;
        }
        .designer-title {
            margin: 0;
            font-size: clamp(2rem, 2.8vw, 3.1rem);
            line-height: 1.04;
            color: var(--goat-ink);
            font-weight: 800;
        }
        .designer-copy {
            max-width: 54rem;
            margin: 0.7rem 0 0 0;
            color: var(--goat-muted);
            font-size: 1rem;
            line-height: 1.6;
        }
        .designer-badges {
            display: flex;
            flex-wrap: wrap;
            gap: 0.55rem;
            margin-top: 1rem;
        }
        .designer-badge {
            padding: 0.42rem 0.78rem;
            border-radius: 999px;
            background: rgba(255, 255, 255, 0.92);
            border: 1px solid rgba(15, 23, 42, 0.08);
            color: var(--goat-brand);
            font-size: 0.82rem;
            font-weight: 700;
        }
        div[data-baseweb="tab-list"] {
            gap: 0.55rem;
            margin-bottom: 1rem;
        }
        div[data-baseweb="tab-list"] button {
            min-height: 3rem;
            border-radius: 999px;
            border: 1px solid rgba(15, 23, 42, 0.08);
            background: rgba(255, 255, 255, 0.72);
            color: var(--goat-muted);
            padding: 0.25rem 1rem;
        }
        div[data-baseweb="tab-list"] button[aria-selected="true"] {
            background: linear-gradient(135deg, #0f766e, #2563eb);
            color: white;
            border-color: transparent;
            box-shadow: 0 12px 24px rgba(37, 99, 235, 0.16);
        }
        div[data-testid="stMetric"],
        div[data-testid="stDataFrame"],
        div[data-testid="stExpander"] {
            border-radius: 20px;
        }
        div[data-testid="stMetric"] {
            border: 1px solid rgba(15, 23, 42, 0.08);
            background: var(--goat-panel);
            padding: 0.6rem 0.7rem;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.05);
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _render_model_hero() -> None:
    badges = "".join(
        f'<span class="designer-badge">{label}</span>'
        for label in (
            "Scenario dashboard",
            "Professional workbook",
            "Financial statements",
            "AI decision support",
        )
    )
    st.markdown(
        f"""
        <section class="designer-hero">
            <p class="designer-kicker">Livestock finance planning</p>
            <h1 class="designer-title">Goat Farm Financial Model</h1>
            <p class="designer-copy">
                Build a cleaner decision environment for herd growth, pricing, costs, financing, and
                investor reporting with a structured dashboard and polished export pack.
            </p>
            <div class="designer-badges">{badges}</div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def _style_workbook_sheet(ws, *, accent: str, accent_soft: str, is_overview: bool = False) -> None:
    ws.sheet_view.showGridLines = False
    if is_overview:
        ws.freeze_panes = "A6"
    elif ws.max_row > 1:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=accent)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.auto_filter.ref = ws.dimensions
        for row_idx in range(2, min(ws.max_row, 120) + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill("solid", fgColor=accent_soft)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 14), 34)


def _build_workbook_summary(results: Dict[str, Any], scenario_name: str) -> List[Tuple[str, object]]:
    summary: List[Tuple[str, object]] = [("Scenario", scenario_name)]
    kpis = results.get("kpis")
    if isinstance(kpis, pd.DataFrame) and not kpis.empty:
        for column in kpis.columns[:6]:
            numeric = pd.to_numeric(kpis[column], errors="coerce")
            if numeric.notna().any():
                value = float(numeric.iloc[0])
                if "margin" in str(column).lower() or "irr" in str(column).lower():
                    summary.append((str(column), f"{value:.1%}"))
                else:
                    summary.append((str(column), f"{value:,.2f}"))
    summary.append(("Prepared By", results.get("scenario_inputs", {}).get("Model author", DEFAULT_MODEL_AUTHOR)))
    return summary[:8]


def _style_professional_workbook(
    workbook_bytes: bytes,
    *,
    scenario_name: str,
    results: Dict[str, Any],
) -> bytes:
    workbook = load_workbook(BytesIO(workbook_bytes))
    accent = "0F766E"
    accent_soft = "E8F7F3"
    if "Overview" in workbook.sheetnames:
        del workbook["Overview"]
    overview = workbook.create_sheet("Overview", 0)
    overview["A1"] = "Goat Farm Financial Model"
    overview["A1"].font = Font(size=20, bold=True, color="1F2937")
    overview["A2"] = "Scenario-led workbook covering input schedules, annual KPIs, break-even, and full statements."
    overview["A2"].font = Font(size=11, color="4B5563")
    overview["A4"] = "Executive Snapshot"
    overview["A4"].font = Font(size=12, bold=True, color=accent)
    overview["A5"] = "Metric"
    overview["B5"] = "Value"
    for cell in overview[5]:
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, (label, value) in enumerate(_build_workbook_summary(results, scenario_name), start=6):
        overview.cell(row=row_idx, column=1, value=label)
        overview.cell(row=row_idx, column=2, value=value)
    overview["D4"] = "Workbook Notes"
    overview["D4"].font = Font(size=12, bold=True, color=accent)
    notes = [
        "The overview isolates the selected scenario and its KPI profile.",
        "Detailed statement tabs remain available for lender diligence and board review.",
        "Scenario inputs and supplementary tables are preserved for auditability.",
    ]
    for row_idx, note in enumerate(notes, start=5):
        overview.cell(row=row_idx, column=4, value=f"• {note}")
    overview.column_dimensions["A"].width = 28
    overview.column_dimensions["B"].width = 18
    overview.column_dimensions["D"].width = 58

    for sheet in workbook.worksheets:
        _style_workbook_sheet(
            sheet,
            accent=accent,
            accent_soft=accent_soft,
            is_overview=sheet.title == "Overview",
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _sanitize_model_author_value(value: Any) -> str:
    """Return a cleaned author string using the default when empty."""

    if not isinstance(value, str):
        value = str(value)
    return value.strip() or DEFAULT_MODEL_AUTHOR


def _handle_model_author_change() -> None:
    """Persist author edits and clear cached exports when updated."""

    if MODEL_AUTHOR_WIDGET_KEY not in st.session_state:
        return

    raw_value = st.session_state.get(MODEL_AUTHOR_WIDGET_KEY, "")
    sanitized = _sanitize_model_author_value(raw_value)
    previous = st.session_state.get(MODEL_AUTHOR_CACHE_KEY)

    # Streamlit raises when callbacks mutate keys that have not been declared yet
    # in bare execution. Ensure the storage key exists before assignment.
    if MODEL_AUTHOR_KEY not in st.session_state:
        st.session_state.setdefault(MODEL_AUTHOR_KEY, sanitized)
    else:
        st.session_state[MODEL_AUTHOR_KEY] = sanitized

    st.session_state[MODEL_AUTHOR_CACHE_KEY] = sanitized

    if sanitized != raw_value:
        st.session_state[MODEL_AUTHOR_WIDGET_KEY] = sanitized

    if previous is not None and sanitized != previous:
        st.session_state.pop("excel_bytes_map", None)


def _current_model_author() -> str:
    """Return the active model author, applying defaults when necessary."""

    current = st.session_state.get(MODEL_AUTHOR_KEY, DEFAULT_MODEL_AUTHOR)
    sanitized = _sanitize_model_author_value(current)
    if sanitized != current:
        st.session_state[MODEL_AUTHOR_KEY] = sanitized
    st.session_state.setdefault(MODEL_AUTHOR_CACHE_KEY, sanitized)
    st.session_state.setdefault(MODEL_AUTHOR_WIDGET_KEY, sanitized)
    widget_value = st.session_state.get(MODEL_AUTHOR_WIDGET_KEY)
    if widget_value != sanitized:
        st.session_state[MODEL_AUTHOR_WIDGET_KEY] = sanitized
    return sanitized


def _render_model_author_editor() -> None:
    """Display an inline editor for the model author name."""

    author_value = _current_model_author()
    st.session_state.setdefault(MODEL_AUTHOR_WIDGET_KEY, author_value)
    st.text_input(
        "Model author",
        value=st.session_state.get(MODEL_AUTHOR_WIDGET_KEY, author_value),
        key=MODEL_AUTHOR_WIDGET_KEY,
        help=(
            "Name recorded in scenario outputs and Excel downloads. "
            "Leave blank to reset to the default."
        ),
    )
    _handle_model_author_change()


def _statement_series_by_suffix(
    df: Optional[pd.DataFrame], suffixes: Sequence[str]
) -> Optional[pd.Series]:
    """Return the first numeric series whose column label ends with any suffix."""

    if df is None or df.empty:
        return None

    for suffix in suffixes:
        for column in df.columns:
            if not isinstance(column, str):
                continue
            if column.endswith(suffix):
                series = pd.to_numeric(df[column], errors="coerce")
                if series.notna().any():
                    return series
    return None


def _statement_scenario_frames(
    base_df: Optional[pd.DataFrame],
    scenario_df: Optional[pd.DataFrame],
    scenario_label: str,
) -> Dict[str, pd.DataFrame]:
    """Assemble labelled dataframes for the base case and selected scenario."""

    frames: Dict[str, pd.DataFrame] = {}

    if base_df is not None and not base_df.empty:
        frames["Base Case"] = base_df

    if scenario_df is not None and not scenario_df.empty:
        label = (scenario_label or "Selected Scenario").strip()
        if label.lower() in {"base", "base case", "base case scenario"}:
            label = "Selected Scenario"
        if label in frames:
            label = f"{label} (Selected)"
        frames[label] = scenario_df

    return frames


def _build_statement_chart_frame(
    frames: Dict[str, pd.DataFrame],
    metrics: Sequence[Tuple[str, Sequence[str]]],
) -> Optional[pd.DataFrame]:
    """Construct a combined dataframe for charting the requested metrics."""

    chart_frames: list[pd.DataFrame] = []

    for scenario_label, df in frames.items():
        if df is None or df.empty:
            continue

        scenario_columns: Dict[str, pd.Series] = {}
        for display_name, suffixes in metrics:
            series = _statement_series_by_suffix(df, suffixes)
            if series is None:
                continue
            scenario_columns[f"{scenario_label} – {display_name}"] = series

        if scenario_columns:
            chart_frames.append(pd.DataFrame(scenario_columns))

    if not chart_frames:
        return None

    combined = pd.concat(chart_frames, axis=1)
    combined = combined.loc[:, ~combined.columns.duplicated()]
    return combined.sort_index()


def _build_margin_chart_frame(
    frames: Dict[str, pd.DataFrame]
) -> Optional[pd.DataFrame]:
    """Compute gross and profit margin percentages for available scenarios."""

    margin_frames: list[pd.DataFrame] = []

    for scenario_label, df in frames.items():
        revenue = _statement_series_by_suffix(df, ("Income – Revenue",))
        if revenue is None:
            continue
        revenue = revenue.replace({0.0: np.nan})

        gross_profit = _statement_series_by_suffix(
            df, ("Cost of sales – Gross profit",)
        )
        profit_for_period = _statement_series_by_suffix(
            df, ("Profit – Profit for the period",)
        )

        margin_columns: Dict[str, pd.Series] = {}

        if gross_profit is not None:
            gross_margin = gross_profit.divide(revenue) * 100.0
            margin_columns[f"{scenario_label} – Gross margin (%)"] = gross_margin

        if profit_for_period is not None:
            profit_margin = profit_for_period.divide(revenue) * 100.0
            margin_columns[f"{scenario_label} – Profit margin (%)"] = profit_margin

        if margin_columns:
            margin_frames.append(pd.DataFrame(margin_columns))

    if not margin_frames:
        return None

    combined = pd.concat(margin_frames, axis=1)
    combined = combined.loc[:, ~combined.columns.duplicated()]
    return combined.sort_index()


def _render_financial_performance_charts(
    base_df: Optional[pd.DataFrame],
    scenario_df: Optional[pd.DataFrame],
    scenario_label: str,
) -> None:
    frames = _statement_scenario_frames(base_df, scenario_df, scenario_label)
    if not frames:
        return

    trend_metrics = [
        ("Revenue", ("Income – Revenue",)),
        ("Gross profit", ("Cost of sales – Gross profit",)),
        ("Profit for the period", ("Profit – Profit for the period",)),
    ]

    expense_metrics = [
        ("Distribution costs", ("Operating expenses – Distribution costs",)),
        ("Administrative expenses", ("Operating expenses – Administrative expenses",)),
        (
            "Depreciation and amortisation",
            ("Operating expenses – Depreciation and amortisation",),
        ),
        ("Other operating expenses", ("Operating expenses – Other operating expenses",)),
    ]

    trend_data = _build_statement_chart_frame(frames, trend_metrics)
    expense_data = _build_statement_chart_frame(frames, expense_metrics)
    margin_data = _build_margin_chart_frame(frames)

    if trend_data is not None:
        st.markdown("###### Income and profit trends")
        st.line_chart(trend_data)

    if expense_data is not None:
        st.markdown("###### Operating expense mix")
        st.bar_chart(expense_data)

    if margin_data is not None:
        st.markdown("###### Margin analysis")
        st.line_chart(margin_data)


def _render_financial_position_charts(
    base_df: Optional[pd.DataFrame],
    scenario_df: Optional[pd.DataFrame],
    scenario_label: str,
) -> None:
    frames = _statement_scenario_frames(base_df, scenario_df, scenario_label)
    if not frames:
        return

    balance_metrics = [
        ("Total assets", ("Assets – Total assets",)),
        ("Total liabilities", ("Equity and liabilities – Total liabilities",)),
        ("Equity", ("Equity and liabilities – Equity",)),
    ]

    net_metrics = [
        ("Net assets", ("Key metrics – Net assets",)),
        ("Net current assets", ("Key metrics – Net current assets",)),
    ]

    balance_data = _build_statement_chart_frame(frames, balance_metrics)
    net_data = _build_statement_chart_frame(frames, net_metrics)

    if balance_data is not None:
        st.markdown("###### Balance sheet totals")
        st.line_chart(balance_data)

    if net_data is not None:
        st.markdown("###### Net asset metrics")
        st.bar_chart(net_data)


def _render_cash_flow_charts(
    base_df: Optional[pd.DataFrame],
    scenario_df: Optional[pd.DataFrame],
    scenario_label: str,
) -> None:
    frames = _statement_scenario_frames(base_df, scenario_df, scenario_label)
    if not frames:
        return

    activity_metrics = [
        (
            "Operating activities",
            ("Operating activities – Net cash from operating activities",),
        ),
        (
            "Investing activities",
            ("Investing activities – Net cash used in investing activities",),
        ),
        (
            "Financing activities",
            ("Financing activities – Net cash from financing activities",),
        ),
    ]

    cash_balance_metrics = [
        (
            "Opening cash",
            ("Net change – Cash and cash equivalents at beginning of period",),
        ),
        (
            "Closing cash",
            ("Net change – Cash and cash equivalents at end of period",),
        ),
        (
            "Net change",
            (
                "Net change – Net increase/(decrease) in cash and cash equivalents",
            ),
        ),
    ]

    activity_data = _build_statement_chart_frame(frames, activity_metrics)
    cash_balance_data = _build_statement_chart_frame(frames, cash_balance_metrics)

    if activity_data is not None:
        st.markdown("###### Cash flow by activity")
        st.bar_chart(activity_data)

    if cash_balance_data is not None:
        st.markdown("###### Cash balance reconciliation")
        st.line_chart(cash_balance_data)


def _payload_to_ai_settings(payload: dict) -> Dict[str, Any]:
    ai_payload = payload.get("ai") or {}
    ml_methods = ai_payload.get("ml_methods") or ["linear_regression"]
    features = ai_payload.get("generative_features") or ["summary"]
    return {
        "enabled": bool(ai_payload.get("enabled", False)),
        "provider": ai_payload.get("provider", "OpenAI"),
        "model": ai_payload.get("model", "gpt-4"),
        "forecast_horizon": int(ai_payload.get("forecast_horizon", 3)),
        "ml_methods": [str(method) for method in ml_methods],
        "generative_features": [str(feature) for feature in features],
        "api_key": ai_payload.get("api_key", ""),
    }


def _ai_settings_to_payload(settings: Dict[str, Any], payload: dict) -> None:
    payload.setdefault("ai", {})
    payload["ai"].update(
        {
            "enabled": bool(settings.get("enabled", False)),
            "provider": settings.get("provider", "OpenAI"),
            "model": settings.get("model", "gpt-4"),
            "forecast_horizon": int(settings.get("forecast_horizon", 3)),
            "ml_methods": list(settings.get("ml_methods", ["linear_regression"])),
            "generative_features": list(
                settings.get("generative_features", ["summary"])
            ),
            "api_key": settings.get("api_key", ""),
        }
    )


def _render_ai_settings(payload: dict, container: Optional[DeltaGenerator] = None) -> None:
    target = container or st
    settings = st.session_state.setdefault(
        "ai_settings", _payload_to_ai_settings(payload)
    )
    st.session_state.setdefault("ai_api_key", settings.get("api_key", ""))

    if st.session_state.pop("ai_settings_saved", False):
        target.success("AI configuration updated. Rerunning the model with the new settings.")

    provider_options = list(AI_PROVIDER_OPTIONS)
    if settings.get("provider") and settings["provider"] not in provider_options:
        provider_options.append(settings["provider"])

    current_provider = settings.get("provider", provider_options[0])
    try:
        provider_index = provider_options.index(current_provider)
    except ValueError:
        provider_index = 0

    ml_defaults = [
        ML_METHOD_LABELS.get(code, code.replace("_", " ").title())
        for code in settings.get("ml_methods", ["linear_regression"])
    ]
    feature_defaults = [
        GEN_AI_FEATURE_LABELS.get(code, code.replace("_", " ").title())
        for code in settings.get("generative_features", ["summary"])
    ]

    form = target.form("ai_settings_form")
    with form:
        enabled = form.checkbox(
            "Enable AI Enhancements",
            value=bool(settings.get("enabled", False)),
            help="Toggle machine-learning forecasts and generative commentary.",
        )
        provider = form.selectbox(
            "Provider",
            provider_options,
            index=provider_index,
            help="Select the API provider powering generative insights.",
        )
        model = form.text_input(
            "Model",
            value=settings.get("model", "gpt-4"),
            help="Name of the deployed model (for example `gpt-4o-mini`).",
        )
        horizon = form.number_input(
            "Forecast Horizon (years)",
            min_value=0,
            max_value=20,
            value=int(settings.get("forecast_horizon", 3)),
            step=1,
            help="Number of additional years used for machine-learning revenue forecasts.",
        )

        ml_selection = form.multiselect(
            "Machine Learning Methods",
            list(ML_METHOD_LABELS.values()),
            default=ml_defaults,
            help="Choose algorithms applied to projected net revenue.",
        )
        feature_selection = form.multiselect(
            "Generative Features",
            list(GEN_AI_FEATURE_LABELS.values()),
            default=feature_defaults,
            help="Pick the narrative focus areas generated by the AI summary.",
        )
        api_key = form.text_input(
            "API Key",
            value=st.session_state.get("ai_api_key", ""),
            type="password",
            help="Store your provider API key securely. Keys are retained only for the current session.",
        )

        submitted = form.form_submit_button("Save AI Configuration")

    if submitted:
        ml_codes = [
            ML_LABEL_TO_CODE.get(label, label.replace(" ", "_").lower())
            for label in ml_selection
        ]
        feature_codes = [
            GEN_AI_LABEL_TO_CODE.get(label, label.replace(" ", "_").lower())
            for label in feature_selection
        ]

        settings.update(
            {
                "enabled": enabled,
                "provider": provider,
                "model": model.strip() or "gpt-4",
                "forecast_horizon": int(horizon),
                "ml_methods": ml_codes or ["linear_regression"],
                "generative_features": feature_codes or ["summary"],
                "api_key": api_key.strip(),
            }
        )
        st.session_state["ai_settings"] = settings
        st.session_state["ai_api_key"] = settings.get("api_key", "")
        _ai_settings_to_payload(settings, payload)
        st.session_state["ai_settings_saved"] = True
        _maybe_rerun()


def _analytics_override_store() -> Dict[str, Any]:
    """Return (and create) the session-backed override cache."""

    return st.session_state.setdefault("advanced_analytics_overrides", {})


def _get_analytics_override(
    scenario: str, block: str, analysis: str, table: str
) -> Optional[pd.DataFrame]:
    store = _analytics_override_store()
    return (
        store.get(scenario, {})
        .get(block, {})
        .get(analysis, {})
        .get(table)
    )


def _set_analytics_override(
    scenario: str, block: str, analysis: str, table: str, value: pd.DataFrame
) -> None:
    store = _analytics_override_store()
    scenario_store = store.setdefault(scenario, {})
    block_store = scenario_store.setdefault(block, {})
    analysis_store = block_store.setdefault(analysis, {})
    analysis_store[table] = value.copy(deep=True)
    st.session_state["advanced_analytics_overrides"] = store


def _clear_analytics_override(
    scenario: str, block: str, analysis: str, table: str
) -> None:
    store = _analytics_override_store()
    scenario_store = store.get(scenario)
    if not scenario_store:
        return
    block_store = scenario_store.get(block)
    if not block_store:
        return
    analysis_store = block_store.get(analysis)
    if not analysis_store:
        return
    analysis_store.pop(table, None)
    if not analysis_store:
        block_store.pop(analysis, None)
    if not block_store:
        scenario_store.pop(block, None)
    if not scenario_store:
        store.pop(scenario, None)
    st.session_state["advanced_analytics_overrides"] = store


def _analytics_edit_flag_key(
    scenario: str, block: str, analysis: str, table: str
) -> str:
    return f"analytics_edit::{scenario}::{block}::{analysis}::{table}"


def _analytics_editor_key(
    scenario: str, block: str, analysis: str, table: str
) -> str:
    return f"analytics_editor::{scenario}::{block}::{analysis}::{table}"


def _prepare_editor_table(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """Return an editable copy and metadata to restore the original index."""

    working = df.copy(deep=True)
    if isinstance(working.index, pd.MultiIndex):
        orig_names = list(working.index.names)
        fallback_names = [
            name if name is not None else f"Index Level {idx + 1}"
            for idx, name in enumerate(orig_names)
        ]
        working.index.set_names(fallback_names, inplace=True)
        editor_df = working.reset_index()
        return editor_df, {
            "type": "multi",
            "names": fallback_names,
            "orig_names": orig_names,
        }

    if isinstance(working.index, pd.RangeIndex):
        editor_df = working.reset_index(drop=True)
        return editor_df, {
            "type": "range",
            "start": working.index.start,
            "step": working.index.step,
        }

    index_name = working.index.name or "Index"
    working.index.name = index_name
    editor_df = working.reset_index()
    return editor_df, {
        "type": "single",
        "name": index_name,
        "orig_name": df.index.name,
    }


def _restore_editor_table(edited: pd.DataFrame, meta: Dict[str, Any]) -> pd.DataFrame:
    """Rebuild a DataFrame using the stored index metadata."""

    restored = edited.copy(deep=True)
    meta_type = meta.get("type")

    if meta_type == "multi":
        names = [name for name in meta.get("names", []) if name in restored.columns]
        if names:
            restored = restored.set_index(names)
            orig_names = meta.get("orig_names", names)
            if len(orig_names) == restored.index.nlevels:
                restored.index.names = orig_names
        else:
            restored.index = pd.RangeIndex(len(restored))
        return restored

    if meta_type == "single":
        column = meta.get("name")
        if column and column in restored.columns:
            restored = restored.set_index(column)
            restored.index.name = meta.get("orig_name")
        else:
            restored.index = pd.RangeIndex(len(restored))
            restored.index.name = meta.get("orig_name")
        return restored

    # Default to a simple RangeIndex
    restored.index = pd.RangeIndex(len(restored))
    return restored


def _format_row_label(df: pd.DataFrame, idx: int) -> str:
    """Return a compact label describing a row for the row selector."""

    if df.empty:
        return "Row"

    row = df.iloc[idx]
    parts: list[str] = []
    preferred_columns = [
        "Field",
        "Category",
        "Item",
        "Position",
        "Product",
        "Period",
        "Year",
        "Business Unit",
    ]
    ordered_columns = [
        column for column in preferred_columns if column in df.columns
    ] + [column for column in df.columns if column not in preferred_columns]
    for column in ordered_columns:
        value = row[column]
        if pd.isna(value):
            continue
        text = str(value)
        if text.strip() == "":
            continue
        parts.append(f"{column}: {text}")
        if len(parts) == 3:
            break

    if not parts:
        return f"Row {idx + 1}"
    return " | ".join(parts)


def _coerce_row_value(raw: Any, dtype: pd.Series.dtype) -> Any:
    """Coerce a raw editor input back to the column's dtype."""

    if is_bool_dtype(dtype):
        return bool(raw)

    if isinstance(raw, str):
        text = raw.strip()
        if text == "":
            if is_numeric_dtype(dtype) or is_datetime64_any_dtype(dtype):
                return np.nan
            return ""

        if is_datetime64_any_dtype(dtype):
            try:
                return pd.to_datetime(text)
            except (TypeError, ValueError):
                return text

        if is_numeric_dtype(dtype):
            coerced = pd.to_numeric([text], errors="coerce")[0]
            if pd.isna(coerced):
                return np.nan
            if is_integer_dtype(dtype):
                return int(round(coerced))
            return float(coerced)

        return text

    return raw


def _render_row_input(
    column: str, value: Any, dtype: pd.Series.dtype, widget_key: str
) -> Any:
    """Render an appropriate widget for a single row cell."""

    if is_bool_dtype(dtype):
        default = bool(value) if not pd.isna(value) else False
        return st.checkbox(column, value=default, key=widget_key)

    display_value = "" if pd.isna(value) else str(value)
    return st.text_input(column, value=display_value, key=widget_key)


def _schedule_edit_flag_key(identifier: str) -> str:
    return f"schedule_edit::{identifier}"


def _schedule_working_key(identifier: str) -> str:
    return f"schedule_editor::{identifier}::working"


def _schedule_meta_key(identifier: str) -> str:
    return f"schedule_editor::{identifier}::meta"


def _schedule_row_selector_key(identifier: str) -> str:
    return f"schedule_editor::{identifier}::row"


def _clear_schedule_editor_state(identifier: str) -> None:
    """Remove any cached working copies for a schedule editor."""

    st.session_state.pop(_schedule_working_key(identifier), None)
    st.session_state.pop(_schedule_meta_key(identifier), None)
    st.session_state.pop(_schedule_row_selector_key(identifier), None)


def _default_value_for_dtype(dtype: pd.Series.dtype) -> Any:
    if is_bool_dtype(dtype):
        return False
    if is_numeric_dtype(dtype):
        return np.nan
    return ""


def _blank_row_like(df: pd.DataFrame) -> Dict[str, Any]:
    """Return a dictionary representing a blank row for the dataframe."""

    blanks: Dict[str, Any] = {}
    for column in df.columns:
        blanks[column] = _default_value_for_dtype(df[column].dtype)
    return blanks


def _render_schedule_row_editor(
    identifier: str, table: pd.DataFrame, save_callback: Callable[[pd.DataFrame], None]
) -> None:
    """Render a row-focused editor for schedule dataframes."""

    st.dataframe(table, use_container_width=True)

    edit_flag_key = _schedule_edit_flag_key(identifier)
    editing = st.session_state.get(edit_flag_key, False)
    toggle_label = "Edit rows" if not editing else "Close row editor"
    if st.button(toggle_label, key=f"{identifier}::toggle"):
        if editing:
            _clear_schedule_editor_state(identifier)
        st.session_state[edit_flag_key] = not editing
        _maybe_rerun()
        return

    editing = st.session_state.get(edit_flag_key, False)
    if not editing:
        return

    editor_df, meta = _prepare_editor_table(table)
    working_key = _schedule_working_key(identifier)
    meta_key = _schedule_meta_key(identifier)

    if working_key not in st.session_state:
        st.session_state[working_key] = editor_df.copy(deep=True)
        st.session_state[meta_key] = meta

    working_df = st.session_state.get(working_key, editor_df.copy(deep=True))
    stored_meta = st.session_state.get(meta_key, meta)

    st.caption(
        "Update one row at a time. Use the controls below to add new rows or remove the selected row."
    )
    st.dataframe(working_df, use_container_width=True)

    template_df = working_df if not working_df.empty else editor_df

    controls = st.columns(3)
    if controls[0].button("Add row", key=f"{identifier}::add_row"):
        blank_row = _blank_row_like(template_df)
        new_row = pd.DataFrame([blank_row], columns=template_df.columns)
        updated_df = pd.concat([working_df, new_row], ignore_index=True)
        st.session_state[working_key] = updated_df
        _maybe_rerun()
        return

    selector_key = _schedule_row_selector_key(identifier)
    if working_df.empty:
        selected_row = None
        controls[1].write("No rows to delete.")
    else:
        selected_row = st.selectbox(
            "Select a row to edit",
            list(range(len(working_df))),
            format_func=lambda idx: _format_row_label(working_df, idx),
            key=selector_key,
        )
        if controls[1].button("Delete row", key=f"{identifier}::delete_row"):
            if selected_row is not None and 0 <= selected_row < len(working_df):
                updated_df = working_df.drop(index=selected_row).reset_index(drop=True)
                st.session_state[working_key] = updated_df
                _maybe_rerun()
                return

    if controls[2].button("Reset changes", key=f"{identifier}::reset"):
        st.session_state[working_key] = editor_df.copy(deep=True)
        st.session_state[meta_key] = meta
        _maybe_rerun()
        return

    working_df = st.session_state.get(working_key, editor_df.copy(deep=True))

    if working_df.empty:
        st.info("There are no rows to edit. Add a new row to begin.")
    else:
        selected_row = st.session_state.get(selector_key, 0)
        if selected_row >= len(working_df):
            selected_row = 0
            st.session_state[selector_key] = selected_row

        dtype_map = working_df.dtypes.to_dict()
        row_series = working_df.iloc[selected_row]

        with st.form(f"{identifier}::row_form"):
            updated_values: Dict[str, Any] = {}
            for column in working_df.columns:
                widget_key = f"{identifier}::{selected_row}::{column}"
                updated_values[column] = _render_row_input(
                    column, row_series[column], dtype_map[column], widget_key
                )

            submitted = st.form_submit_button("Apply row changes")
            if submitted:
                for column, raw_value in updated_values.items():
                    coerced = _coerce_row_value(raw_value, dtype_map[column])
                    working_df.iat[
                        selected_row, working_df.columns.get_loc(column)
                    ] = coerced
                st.session_state[working_key] = working_df
                _maybe_rerun()
                return

    action_cols = st.columns(3)
    if action_cols[0].button("Save changes", key=f"{identifier}::save"):
        try:
            current_df = st.session_state.get(working_key, editor_df.copy(deep=True))
            current_meta = st.session_state.get(meta_key, stored_meta)
            restored = _restore_editor_table(current_df, current_meta)
            input_version_before = int(_safe_session_state_get(MODEL_INPUT_VERSION_KEY, 0) or 0)
            save_callback(restored)
            input_version_after = int(_safe_session_state_get(MODEL_INPUT_VERSION_KEY, 0) or 0)
            if input_version_after == input_version_before:
                _reset_cached_results()
            _clear_schedule_editor_state(identifier)
            st.session_state[edit_flag_key] = False
            _maybe_rerun()
            return
        except Exception as exc:  # pragma: no cover - user feedback path
            action_cols[0].error(f"Unable to save changes: {exc}")

    if action_cols[1].button("Discard edits", key=f"{identifier}::cancel"):
        _clear_schedule_editor_state(identifier)
        st.session_state[edit_flag_key] = False
        _maybe_rerun()
        return

    if action_cols[2].button("Close editor", key=f"{identifier}::close"):
        _clear_schedule_editor_state(identifier)
        st.session_state[edit_flag_key] = False
        _maybe_rerun()


DETAIL_SCHEDULE_COLUMNS = {
    "COGS Schedule": ["COGS"],
    "Variable Expenses Schedule": ["Variable Expenses"],
    "Direct Wages Schedule": ["Direct Wages"],
    "Admin Wages Schedule": ["Admin Wages"],
    "Capex Schedule": ["Capex"],
}


BUSINESS_TYPE_OPTIONS: tuple[str, ...] = ("Breeding", "Meat", "Milk-Cheese", "Combined")
DEFAULT_BUSINESS_TYPE = "Combined"
OPERATING_MODEL_OPTIONS: tuple[str, ...] = ("Standalone", "Breeding-to-Unit")
DEFAULT_OPERATING_MODEL = "Standalone"
REPORTING_VIEW_OPTIONS: tuple[str, ...] = ("Standalone Unit", "Consolidated")
DEFAULT_REPORTING_VIEW = "Consolidated"
TRANSFER_PRICING_METHOD_OPTIONS: tuple[str, ...] = ("Cost", "Cost Plus", "Market", "Zero")
DEFAULT_TRANSFER_PRICING_METHOD = "Cost"

DAIRY_PRODUCTS: tuple[str, ...] = ("Milk", "Cheese")
LIVESTOCK_PRODUCTS: tuple[str, ...] = ("Meat", "Offal", "Pelt", "Live Herd")
BREEDING_PRODUCTS: tuple[str, ...] = ("Female Kids", "Male Kids")
ALL_MODEL_PRODUCTS: tuple[str, ...] = BREEDING_PRODUCTS + DAIRY_PRODUCTS + LIVESTOCK_PRODUCTS
ANNUAL_SLAUGHTER_RATE_COLUMN = "Annual Slaughter Rate % of Herd"
LEGACY_SLAUGHTER_RATE_COLUMN = "Slaughter Rate % of Herd per Period"

BUSINESS_TYPE_ACTIVE_PRODUCTS: Dict[str, tuple[str, ...]] = {
    "Breeding": ("Female Kids", "Male Kids", "Meat", "Offal", "Pelt", "Live Herd"),
    "Meat": ("Meat", "Offal", "Pelt", "Live Herd"),
    "Milk-Cheese": ("Milk", "Cheese", "Pelt", "Live Herd"),
    "Combined": ("Milk", "Cheese", "Meat", "Offal", "Pelt", "Live Herd"),
}

PRODUCT_FAMILY_MAP: Dict[str, str] = {
    "Female Kids": "Breeding Transfers & External Kid Sales",
    "Male Kids": "Breeding Transfers & External Kid Sales",
    "Milk": "Dairy",
    "Cheese": "Dairy",
    "Meat": "Livestock Sales & By-Products",
    "Offal": "Livestock Sales & By-Products",
    "Pelt": "Livestock Sales & By-Products",
    "Live Herd": "Livestock Sales & By-Products",
}

PRODUCT_DEFAULTS: Dict[str, dict[str, Any]] = {
    "Female Kids": {
        "unit": "Head",
        "base_price": 180.00,
        "price_growth_pct": 2.5,
        "default_active": True,
    },
    "Male Kids": {
        "unit": "Head",
        "base_price": 150.00,
        "price_growth_pct": 2.5,
        "default_active": True,
    },
    "Milk": {
        "unit": "Litre",
        "base_price": 4.00,
        "price_growth_pct": 3.0,
        "default_active": True,
    },
    "Cheese": {
        "unit": "Kg",
        "base_price": 30.00,
        "price_growth_pct": 2.5,
        "default_active": True,
    },
    "Meat": {
        "unit": "Kg",
        "base_price": 28.00,
        "price_growth_pct": 2.8,
        "default_active": True,
    },
    "Offal": {
        "unit": "Kg",
        "base_price": 12.00,
        "price_growth_pct": 2.2,
        "default_active": True,
    },
    "Pelt": {
        "unit": "Piece",
        "base_price": 20.00,
        "price_growth_pct": 2.0,
        "default_active": True,
    },
    "Live Herd": {
        "unit": "Head",
        "base_price": 220.00,
        "price_growth_pct": 2.5,
        "default_active": True,
    },
}


def _normalize_business_type(value: Any) -> str:
    cleaned = str(value or "").strip()
    for option in BUSINESS_TYPE_OPTIONS:
        if cleaned.casefold() == option.casefold():
            return option
    return DEFAULT_BUSINESS_TYPE


def _active_products_for_business_type(business_type: Any) -> list[str]:
    normalized = _normalize_business_type(business_type)
    return list(BUSINESS_TYPE_ACTIVE_PRODUCTS.get(normalized, BUSINESS_TYPE_ACTIVE_PRODUCTS[DEFAULT_BUSINESS_TYPE]))


def _normalize_operating_model(value: Any) -> str:
    cleaned = str(value or "").strip()
    for option in OPERATING_MODEL_OPTIONS:
        if cleaned.casefold() == option.casefold():
            return option
    return DEFAULT_OPERATING_MODEL


def _normalize_reporting_view(value: Any) -> str:
    cleaned = str(value or "").strip()
    for option in REPORTING_VIEW_OPTIONS:
        if cleaned.casefold() == option.casefold():
            return option
    return DEFAULT_REPORTING_VIEW


def _normalize_transfer_pricing_method(value: Any) -> str:
    cleaned = str(value or "").strip()
    for option in TRANSFER_PRICING_METHOD_OPTIONS:
        if cleaned.casefold() == option.casefold():
            return option
    return DEFAULT_TRANSFER_PRICING_METHOD


def _default_pricing_row_templates(
    products: Optional[Sequence[str]] = None,
    business_type: Any = DEFAULT_BUSINESS_TYPE,
) -> list[dict[str, object]]:
    product_list = list(products) if products is not None else list(ALL_MODEL_PRODUCTS)
    business_unit = _normalize_business_type(business_type)
    rows: list[dict[str, object]] = []
    for product in product_list:
        defaults = PRODUCT_DEFAULTS.get(str(product).strip())
        if defaults is None:
            continue
        rows.append(
            {
                "Product": str(product).strip(),
                "Business Unit": business_unit,
                "Revenue Channel": "External",
                "Unit": str(defaults.get("unit", "Unit")).strip() or "Unit",
                "Base Price": float(defaults.get("base_price", 0.0) or 0.0),
                "Price Growth %": float(defaults.get("price_growth_pct", 0.0) or 0.0),
                "Default Active": bool(defaults.get("default_active", False)),
            }
        )
    return rows


def _default_production_driver_row_templates(
    products: Optional[Sequence[str]] = None,
    business_type: Any = DEFAULT_BUSINESS_TYPE,
) -> list[dict[str, object]]:
    business_unit = _normalize_business_type(business_type)
    all_rows: Dict[str, dict[str, object]] = {
        "Female Kids": {
            "Product": "Female Kids",
            "Unit": "Head",
            "Business Unit": business_unit,
            "Quantity Source": "Breeding External Sales",
            "Quantity Mode": "Derived",
            "Lactating Herd Share %": 0.0,
            "Litres per Lactating Doe per Day": 0.0,
            "Milk Allocation to Cheese %": 0.0,
            "Cheese Yield Kg per Litre": 0.0,
            ANNUAL_SLAUGHTER_RATE_COLUMN: 0.0,
            "Live Herd Sales Share %": 0.0,
            "Meat Yield Kg per Goat": 0.0,
            "Offal Yield Kg per Goat": 0.0,
            "Pelt Units per Goat": 0.0,
            "Driver Growth %": 0.0,
        },
        "Male Kids": {
            "Product": "Male Kids",
            "Unit": "Head",
            "Business Unit": business_unit,
            "Quantity Source": "Breeding External Sales",
            "Quantity Mode": "Derived",
            "Lactating Herd Share %": 0.0,
            "Litres per Lactating Doe per Day": 0.0,
            "Milk Allocation to Cheese %": 0.0,
            "Cheese Yield Kg per Litre": 0.0,
            ANNUAL_SLAUGHTER_RATE_COLUMN: 0.0,
            "Live Herd Sales Share %": 0.0,
            "Meat Yield Kg per Goat": 0.0,
            "Offal Yield Kg per Goat": 0.0,
            "Pelt Units per Goat": 0.0,
            "Driver Growth %": 0.0,
        },
        "Milk": {
            "Product": "Milk",
            "Unit": "Litre",
            "Business Unit": business_unit,
            "Quantity Source": "Lactating Herd",
            "Quantity Mode": "Derived",
            "Lactating Herd Share %": 55.0,
            "Litres per Lactating Doe per Day": 4.0,
            "Milk Allocation to Cheese %": 0.0,
            "Cheese Yield Kg per Litre": 0.0,
            ANNUAL_SLAUGHTER_RATE_COLUMN: 0.0,
            "Live Herd Sales Share %": 0.0,
            "Meat Yield Kg per Goat": 0.0,
            "Offal Yield Kg per Goat": 0.0,
            "Pelt Units per Goat": 0.0,
            "Driver Growth %": 0.0,
        },
        "Cheese": {
            "Product": "Cheese",
            "Unit": "Kg",
            "Business Unit": business_unit,
            "Quantity Source": "Lactating Herd",
            "Quantity Mode": "Derived",
            "Lactating Herd Share %": 55.0,
            "Litres per Lactating Doe per Day": 4.0,
            "Milk Allocation to Cheese %": 25.0,
            "Cheese Yield Kg per Litre": 0.14,
            ANNUAL_SLAUGHTER_RATE_COLUMN: 0.0,
            "Live Herd Sales Share %": 0.0,
            "Meat Yield Kg per Goat": 0.0,
            "Offal Yield Kg per Goat": 0.0,
            "Pelt Units per Goat": 0.0,
            "Driver Growth %": 0.0,
        },
        "Meat": {
            "Product": "Meat",
            "Unit": "Kg",
            "Business Unit": business_unit,
            "Quantity Source": "Slaughter Output",
            "Quantity Mode": "Derived",
            "Lactating Herd Share %": 0.0,
            "Litres per Lactating Doe per Day": 0.0,
            "Milk Allocation to Cheese %": 0.0,
            "Cheese Yield Kg per Litre": 0.0,
            ANNUAL_SLAUGHTER_RATE_COLUMN: 36.0,
            "Live Herd Sales Share %": 0.0,
            "Meat Yield Kg per Goat": 22.0,
            "Offal Yield Kg per Goat": 0.0,
            "Pelt Units per Goat": 0.0,
            "Driver Growth %": 0.0,
        },
        "Offal": {
            "Product": "Offal",
            "Unit": "Kg",
            "Business Unit": business_unit,
            "Quantity Source": "Slaughter Output",
            "Quantity Mode": "Derived",
            "Lactating Herd Share %": 0.0,
            "Litres per Lactating Doe per Day": 0.0,
            "Milk Allocation to Cheese %": 0.0,
            "Cheese Yield Kg per Litre": 0.0,
            ANNUAL_SLAUGHTER_RATE_COLUMN: 36.0,
            "Live Herd Sales Share %": 0.0,
            "Meat Yield Kg per Goat": 0.0,
            "Offal Yield Kg per Goat": 4.0,
            "Pelt Units per Goat": 0.0,
            "Driver Growth %": 0.0,
        },
        "Pelt": {
            "Product": "Pelt",
            "Unit": "Piece",
            "Business Unit": business_unit,
            "Quantity Source": "Slaughter Output",
            "Quantity Mode": "Derived",
            "Lactating Herd Share %": 0.0,
            "Litres per Lactating Doe per Day": 0.0,
            "Milk Allocation to Cheese %": 0.0,
            "Cheese Yield Kg per Litre": 0.0,
            ANNUAL_SLAUGHTER_RATE_COLUMN: 36.0,
            "Live Herd Sales Share %": 0.0,
            "Meat Yield Kg per Goat": 0.0,
            "Offal Yield Kg per Goat": 0.0,
            "Pelt Units per Goat": 1.0,
            "Driver Growth %": 0.0,
        },
        "Live Herd": {
            "Product": "Live Herd",
            "Unit": "Head",
            "Business Unit": business_unit,
            "Quantity Source": "Slaughter Output",
            "Quantity Mode": "Derived",
            "Lactating Herd Share %": 0.0,
            "Litres per Lactating Doe per Day": 0.0,
            "Milk Allocation to Cheese %": 0.0,
            "Cheese Yield Kg per Litre": 0.0,
            ANNUAL_SLAUGHTER_RATE_COLUMN: 36.0,
            "Live Herd Sales Share %": 30.0,
            "Meat Yield Kg per Goat": 0.0,
            "Offal Yield Kg per Goat": 0.0,
            "Pelt Units per Goat": 0.0,
            "Driver Growth %": 0.0,
        },
    }
    product_list = list(products) if products is not None else list(ALL_MODEL_PRODUCTS)
    return [dict(all_rows[product]) for product in product_list if product in all_rows]


def _period_indexed_frame(
    rows: Sequence[dict[str, Any]],
    columns: Sequence[str],
) -> pd.DataFrame:
    frame = pd.DataFrame(rows)
    if "Period" not in frame.columns:
        frame = frame.reindex(columns=list(columns))
    return frame.set_index("Period")


def _price_change_driver(product: str) -> str:
    return f"{product} price change (%)"


def _quantity_change_driver(product: str) -> str:
    return f"{product} quantity change (%)"


BIOLOGICAL_SCENARIO_DRIVER_DEFAULTS: Dict[str, float] = {
    "Conception rate change (%)": 0.0,
    "Kid mortality change (%)": 0.0,
    "Doe mortality change (%)": 0.0,
    "Cull rate change (%)": 0.0,
    "Kids per kidding change (%)": 0.0,
    "Age at slaughter change (months)": 0.0,
    "Months to market weight change (months)": 0.0,
    "Peak yield change (%)": 0.0,
    "Lactation length change (days)": 0.0,
    "Replacement retention change (%)": 0.0,
}


DEFAULT_SCENARIO_ADJUSTMENTS: Dict[str, float] = {
    **{
        _price_change_driver(product): 0.0
        for product in ALL_MODEL_PRODUCTS
    },
    **{
        _quantity_change_driver(product): 0.0
        for product in ALL_MODEL_PRODUCTS
    },
    **BIOLOGICAL_SCENARIO_DRIVER_DEFAULTS,
    "Feed cost change (%)": 0.0,
}


SCENARIO_PRESETS: Dict[str, Dict[str, Any]] = {
    "Base Case Scenario": {
        "adjustments": DEFAULT_SCENARIO_ADJUSTMENTS,
        "description": "Baseline view using the model inputs without additional shocks.",
    },
    "Best Case Scenario": {
        "adjustments": {
            **DEFAULT_SCENARIO_ADJUSTMENTS,
            "Milk price change (%)": 12.0,
            "Cheese price change (%)": 8.0,
            "Meat price change (%)": 6.0,
            "Offal price change (%)": 5.0,
            "Pelt price change (%)": 4.0,
            "Live Herd price change (%)": 5.0,
            "Milk quantity change (%)": 5.0,
            "Cheese quantity change (%)": 4.0,
            "Meat quantity change (%)": 3.0,
            "Offal quantity change (%)": 2.5,
            "Pelt quantity change (%)": 3.0,
            "Live Herd quantity change (%)": 3.0,
            "Feed cost change (%)": -8.0,
        },
        "description": "Upside case with stronger product pricing, modest output gains, and more efficient feed spend.",
    },
    "Worst Case Scenario": {
        "adjustments": {
            **DEFAULT_SCENARIO_ADJUSTMENTS,
            "Milk price change (%)": -12.0,
            "Cheese price change (%)": -9.0,
            "Meat price change (%)": -7.0,
            "Offal price change (%)": -6.0,
            "Pelt price change (%)": -5.0,
            "Live Herd price change (%)": -6.0,
            "Milk quantity change (%)": -6.0,
            "Cheese quantity change (%)": -5.0,
            "Meat quantity change (%)": -4.0,
            "Offal quantity change (%)": -3.0,
            "Pelt quantity change (%)": -4.0,
            "Live Herd quantity change (%)": -4.0,
            "Feed cost change (%)": 10.0,
        },
        "description": "Downside case featuring weaker product pricing, lower output, and higher feed costs.",
    },
}


def _default_scenario_preset_table(name: str) -> pd.DataFrame:
    preset = SCENARIO_PRESETS.get(name, {})
    adjustments = preset.get("adjustments", {})
    if not adjustments:
        return pd.DataFrame(columns=["Driver", "Change %"])
    return pd.DataFrame(
        {
            "Driver": list(adjustments.keys()),
            "Change %": [float(value) for value in adjustments.values()],
        }
    )


def _scenario_preset_removed_store() -> Dict[str, List[str]]:
    raw_store = _safe_session_state_get("scenario_preset_removed_drivers", {})
    store = raw_store if isinstance(raw_store, dict) else {}
    normalised: Dict[str, List[str]] = {}
    updated = False

    for name in SCENARIO_PRESETS.keys():
        entries = store.get(name, [])
        cleaned = sorted(
            {
                str(entry).casefold()
                for entry in entries
                if str(entry).strip()
            }
        )
        normalised[name] = cleaned
        if entries != cleaned:
            updated = True

    if set(store.keys()) != set(normalised.keys()):
        updated = True

    if updated:
        _safe_session_state_set("scenario_preset_removed_drivers", normalised)
        store = normalised
    else:
        _safe_session_state_set("scenario_preset_removed_drivers", store)

    return store


def _unmark_removed_scenario_drivers(name: str, drivers: Iterable[str]) -> None:
    store = _scenario_preset_removed_store()
    current = set(store.get(name, []))
    lower_drivers = {str(driver).casefold() for driver in drivers if str(driver).strip()}
    new_removed = sorted(current - lower_drivers)
    if new_removed != sorted(current):
        store[name] = new_removed
        _safe_session_state_set("scenario_preset_removed_drivers", store)


def _mark_removed_scenario_driver(name: str, driver: str) -> None:
    store = _scenario_preset_removed_store()
    cleaned = str(driver).strip()
    if not cleaned:
        return
    lowered = cleaned.casefold()
    current = set(store.get(name, []))
    if lowered not in current:
        current.add(lowered)
        store[name] = sorted(current)
        _safe_session_state_set("scenario_preset_removed_drivers", store)


def _scenario_preset_tables_store() -> Dict[str, pd.DataFrame]:
    raw_store = _safe_session_state_get("scenario_preset_tables", {})
    store = raw_store if isinstance(raw_store, dict) else {}
    _safe_session_state_set("scenario_preset_tables", store)
    return store


def _ensure_scenario_preset_table(
    name: str, table: Optional[pd.DataFrame]
) -> pd.DataFrame:
    default_table = _default_scenario_preset_table(name)
    if table is None or table.empty:
        work = default_table.copy()
    else:
        work = table.copy()

    if "Driver" not in work.columns:
        work["Driver"] = ""
    work["Driver"] = _series_or_default(work, "Driver", "").astype(str).str.strip()
    work.loc[work["Driver"] == "", "Driver"] = "Driver"

    if "Change %" not in work.columns:
        work["Change %"] = np.nan
    work["Change %"] = pd.to_numeric(work.get("Change %"), errors="coerce")

    removed_store = _scenario_preset_removed_store()
    removed = set(removed_store.get(name, []))

    required_rows = []
    for _, row in default_table.iterrows():
        driver = str(row.get("Driver", "")).strip()
        if not driver:
            continue
        if driver.casefold() in removed:
            continue
        mask = work["Driver"].str.casefold() == driver.casefold()
        if not mask.any():
            required_rows.append({
                "Driver": driver,
                "Change %": float(row.get("Change %", 0.0)),
            })
        else:
            existing_index = work.index[mask][0]
            if pd.isna(work.at[existing_index, "Change %"]):
                work.at[existing_index, "Change %"] = float(row.get("Change %", 0.0))

    if required_rows:
        work = pd.concat([work, pd.DataFrame(required_rows)], ignore_index=True)

    driver_order = [
        str(val).casefold()
        for val in default_table.get("Driver", pd.Series(dtype=str)).tolist()
        if str(val).strip() and str(val).casefold() not in removed
    ]
    order_map = {driver: idx for idx, driver in enumerate(driver_order)}
    work["__order__"] = work["Driver"].str.casefold().map(order_map)
    work = work.sort_values(["__order__", "Driver"], na_position="last").drop(
        columns=["__order__"], errors="ignore"
    )

    ordered_cols = ["Driver", "Change %"]
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _set_scenario_preset_table(name: str, table: pd.DataFrame) -> None:
    store = _scenario_preset_tables_store()
    store[name] = table.copy(deep=True).reset_index(drop=True)
    drivers = store[name].get("Driver", pd.Series(dtype=str)).tolist()
    _unmark_removed_scenario_drivers(name, drivers)
    _safe_session_state_set("scenario_preset_tables", store)


def _get_scenario_preset_table(name: str) -> pd.DataFrame:
    store = _scenario_preset_tables_store()
    ensured = _ensure_scenario_preset_table(name, store.get(name))
    store[name] = ensured
    _safe_session_state_set("scenario_preset_tables", store)
    return ensured


def _scenario_preset_descriptions_store() -> Dict[str, str]:
    raw_store = _safe_session_state_get("scenario_preset_descriptions", {})
    store = raw_store if isinstance(raw_store, dict) else {}
    for name, preset in SCENARIO_PRESETS.items():
        store.setdefault(name, preset.get("description", ""))
    _safe_session_state_set("scenario_preset_descriptions", store)
    return store


def _set_scenario_preset_description(name: str, description: str) -> Optional[str]:
    store = _scenario_preset_descriptions_store()
    normalized = str(description or "").strip()
    previous = store.get(name, "")
    if normalized == previous:
        return None
    store[name] = normalized
    _safe_session_state_set("scenario_preset_descriptions", store)
    _reset_cached_results()
    return normalized


def _scenario_preset_table_to_adjustments(
    name: str, table: pd.DataFrame
) -> Dict[str, float]:
    ensured = _ensure_scenario_preset_table(name, table)
    adjustments: Dict[str, float] = {}
    for _, row in ensured.iterrows():
        driver = str(row.get("Driver", "")).strip()
        if not driver:
            continue
        value = pd.to_numeric(pd.Series([row.get("Change %")]), errors="coerce").iloc[0]
        if pd.isna(value):
            continue
        adjustments[driver] = float(value)

    removed_store = _scenario_preset_removed_store()
    removed = set(removed_store.get(name, []))
    for driver in list(adjustments.keys()):
        if driver.casefold() in removed:
            adjustments.pop(driver, None)

    return adjustments


def _add_scenario_preset_driver(name: str, driver: str, change: float) -> None:
    cleaned = str(driver).strip()
    if not cleaned:
        return

    table = _get_scenario_preset_table(name)
    mask = table["Driver"].str.casefold() == cleaned.casefold()
    if mask.any():
        table.loc[mask, "Change %"] = float(change)
    else:
        new_row = pd.DataFrame(
            [{"Driver": cleaned, "Change %": float(change)}],
            columns=table.columns,
        )
        table = pd.concat([table, new_row], ignore_index=True)

    _set_scenario_preset_table(name, table)
    _unmark_removed_scenario_drivers(name, [cleaned])
    _reset_cached_results()


def _remove_scenario_preset_driver(name: str, driver: str) -> None:
    cleaned = str(driver).strip()
    if not cleaned:
        return

    table = _get_scenario_preset_table(name)
    mask = table["Driver"].str.casefold() == cleaned.casefold()
    if not mask.any():
        return

    updated = table.loc[~mask].reset_index(drop=True)
    _set_scenario_preset_table(name, updated)
    _mark_removed_scenario_driver(name, cleaned)
    _reset_cached_results()


def _current_scenario_presets() -> Dict[str, Dict[str, Any]]:
    tables = _scenario_preset_tables_store()
    descriptions = _scenario_preset_descriptions_store()
    presets: Dict[str, Dict[str, Any]] = {}

    for name in SCENARIO_PRESETS.keys():
        table = _ensure_scenario_preset_table(name, tables.get(name))
        tables[name] = table
        adjustments = _scenario_preset_table_to_adjustments(name, table)
        description = descriptions.get(name, SCENARIO_PRESETS[name].get("description", ""))
        presets[name] = {
            "adjustments": adjustments,
            "description": description,
        }

    _safe_session_state_set("scenario_preset_tables", tables)
    _safe_session_state_set("scenario_preset_descriptions", descriptions)
    return presets


def _render_scenario_preset_editors() -> None:
    st.markdown("#### Scenario Presets")
    preset_names = list(SCENARIO_PRESETS.keys())
    tabs = st.tabs(preset_names)

    for tab, name in zip(tabs, preset_names):
        with tab:
            table = _get_scenario_preset_table(name)

            add_cols = st.columns([2, 1, 1])
            driver_key = f"scenario_preset::{_scenario_key_suffix(name)}::new_driver"
            change_key = f"scenario_preset::{_scenario_key_suffix(name)}::new_change"
            if not _safe_session_state_contains(driver_key):
                _safe_session_state_setdefault(driver_key, "")
            if not _safe_session_state_contains(change_key):
                _safe_session_state_setdefault(change_key, 0.0)

            new_driver = add_cols[0].text_input(
                "Driver name",
                key=driver_key,
            )
            new_change = add_cols[1].number_input(
                "Change (%)",
                key=change_key,
                step=0.25,
                format="%.2f",
            )
            if add_cols[2].button("Add variable", key=f"{driver_key}::add"):
                _add_scenario_preset_driver(name, new_driver, float(new_change))
                _safe_session_state_set(driver_key, "")
                _safe_session_state_set(change_key, 0.0)
                _maybe_rerun()
                return

            remove_cols = st.columns([2, 1])
            driver_options = table.get("Driver", pd.Series(dtype=str)).fillna("")
            option_labels = [value for value in driver_options.astype(str).tolist() if value.strip()]
            remove_key = f"scenario_preset::{_scenario_key_suffix(name)}::remove_choice"
            remove_selection = remove_cols[0].selectbox(
                "Select variable to remove",
                options=["-- Select --"] + option_labels,
                key=remove_key,
            )
            if remove_cols[1].button("Remove variable", key=f"{remove_key}::remove"):
                if remove_selection and remove_selection != "-- Select --":
                    _remove_scenario_preset_driver(name, remove_selection)
                    _safe_session_state_set(remove_key, "-- Select --")
                    _maybe_rerun()
                    return

            def _save(updated: pd.DataFrame, preset_name: str = name) -> None:
                ensured = _ensure_scenario_preset_table(preset_name, updated)
                _set_scenario_preset_table(preset_name, ensured)
                _reset_cached_results()

            _render_schedule_row_editor(
                f"scenario_preset::{_scenario_key_suffix(name)}",
                table,
                _save,
            )

            description_store = _scenario_preset_descriptions_store()
            default_description = SCENARIO_PRESETS[name].get("description", "")
            stored_value = description_store.get(name, default_description)
            desc_key = f"scenario_desc::{_scenario_key_suffix(name)}"
            if not _safe_session_state_contains(desc_key):
                _safe_session_state_setdefault(desc_key, stored_value)

            st.text_area("Description", key=desc_key)
            new_value = _safe_session_state_get(desc_key, "")
            updated = _set_scenario_preset_description(name, new_value)
            if updated is not None:
                _safe_session_state_set(desc_key, updated)

CAP_TABLE_COLUMNS = ["Year", "Shareholder", "Ownership %", "Investment"]


CAPEX_TABLE_COLUMNS = [
    "Year",
    "Category",
    "Spend",
    "Depreciation Rate %",
    "Depreciation",
]


ASSET_SCHEDULE_COLUMNS = [
    "Asset",
    "Year",
    "Opening NBV",
    "Additions",
    "Depreciation Rate %",
    "Depreciation",
    "Closing NBV",
]


def _normalize_period(series: pd.Series) -> pd.Series:
    values = pd.Series(series, copy=False)
    if values.empty:
        return values.astype("object")

    if is_datetime64_any_dtype(values):
        periods = pd.to_datetime(values, errors="coerce")
        return periods.dt.strftime("%Y-%m-%d").astype("object")

    if isinstance(values.dtype, pd.PeriodDtype):
        periods = values.dt.to_timestamp()
        return periods.dt.strftime("%Y-%m-%d").astype("object")

    raw_text = values.astype("string")
    stripped = raw_text.str.strip()
    normalized = stripped.astype("object")
    empty_mask = stripped.isna() | stripped.eq("")
    if empty_mask.all():
        return normalized

    canonical = stripped.str.replace("/", "-", regex=False)
    iso_mask = canonical.str.fullmatch(r"\d{4}-\d{2}-\d{2}", na=False)
    if iso_mask.any():
        iso_periods = pd.to_datetime(canonical[iso_mask], format="%Y-%m-%d", errors="coerce")
        normalized.loc[iso_mask] = iso_periods.dt.strftime("%Y-%m-%d").astype("object")

    fallback_mask = ~empty_mask & ~iso_mask
    if fallback_mask.any():
        fallback_periods = pd.to_datetime(canonical[fallback_mask], errors="coerce")
        fallback_formatted = fallback_periods.dt.strftime("%Y-%m-%d").astype("object")
        normalized.loc[fallback_mask] = fallback_formatted.where(
            ~fallback_periods.isna(),
            stripped[fallback_mask].astype("object"),
        )
    return normalized


def _format_scenario_label(primary_product: str, price_pct: int, feed_pct: int) -> str:
    product_label = str(primary_product).strip() or "Primary Product"
    if price_pct == 0 and feed_pct == 0:
        return "Base Scenario"
    return f"{product_label} {price_pct:+d}%, Feed {feed_pct:+d}%"


def _build_scenario_suite(
    custom_label: Optional[str] = None,
    custom_adjustments: Optional[Dict[str, float]] = None,
) -> Dict[str, Dict[str, Any]]:
    suite: Dict[str, Dict[str, Any]] = {}
    presets = _current_scenario_presets()
    for name, preset in presets.items():
        suite[name] = {
            "adjustments": deepcopy(preset["adjustments"]),
            "description": preset.get("description", ""),
        }

    if custom_label and custom_adjustments:
        suite[custom_label] = {
            "adjustments": {
                key: float(value)
                for key, value in custom_adjustments.items()
            },
            "description": "Custom scenario defined via scenario controls.",
        }

    return suite


def _dynamic_outputs_table(model: GoatModel, scenario_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[Dict[str, Any]] = []
    valuation = model.valuation_summary(scenario_df) if hasattr(model, "valuation_summary") else {}
    debt_capacity = (
        model.debt_capacity_schedule(scenario_df, annual=True)
        if hasattr(model, "debt_capacity_schedule")
        else pd.DataFrame()
    )
    metrics = {
        "NPV": valuation.get("npv"),
        "IRR": valuation.get("irr"),
        "Payback Period (Years)": valuation.get("payback_years"),
        "Terminal Value": valuation.get("terminal_value"),
        "Minimum DSCR": (
            pd.to_numeric(debt_capacity.get("DSCR"), errors="coerce").min()
            if not debt_capacity.empty and "DSCR" in debt_capacity.columns
            else None
        ),
        "Minimum Cash Headroom": (
            pd.to_numeric(debt_capacity.get("Cash Reserve Headroom"), errors="coerce").min()
            if not debt_capacity.empty and "Cash Reserve Headroom" in debt_capacity.columns
            else None
        ),
    }
    for metric, value in metrics.items():
        if value is None or pd.isna(value):
            continue
        rows.append({"Metric": metric, "Value": float(value)})
    return pd.DataFrame(rows)


def _dynamic_benchmark_kpis_table(kpi_df: pd.DataFrame) -> pd.DataFrame:
    if kpi_df is None or kpi_df.empty:
        return pd.DataFrame(columns=["KPI", "Benchmark"])
    last_row = kpi_df.iloc[-1]
    rows: list[Dict[str, Any]] = []
    for col in [
        "Revenue per Herd Head",
        "Feed Cost per Herd Head",
        "IRR",
        "Terminal Value",
        "Payback Period (Years)",
        "DSCR",
        "Cash Reserve Headroom",
    ]:
        value = pd.to_numeric(pd.Series([last_row.get(col)]), errors="coerce").iloc[0]
        if pd.notna(value):
            rows.append({"KPI": col, "Benchmark": float(value)})
    return pd.DataFrame(rows)


def _format_kpis_for_display(kpi_df: pd.DataFrame) -> pd.DataFrame:
    """Format KPI dataframe for UI display with selective percentage scaling."""
    if kpi_df is None or kpi_df.empty:
        return pd.DataFrame()

    formatted = kpi_df.copy()
    for column in formatted.columns:
        values = pd.to_numeric(formatted[column], errors="coerce")
        if not values.notna().any():
            continue
        is_percent_metric = "%" in str(column) or str(column).strip().upper() in {"IRR", "WACC"}
        formatted[column] = (values * 100.0) if is_percent_metric else values
    return formatted.round(2)


def _valuation_diagnostic_messages(model: GoatModel) -> List[str]:
    messages: List[str] = []
    if model.wacc() is None:
        messages.append("Valuation input missing: WACC.")
    try:
        summary = model.valuation_summary()
    except ValueError as exc:
        messages.append(f"Valuation issue: {exc}")
        return messages
    if not summary:
        messages.append("Unable to derive UFCF from the current operating schedule.")
        return messages
    ufcf_schedule = summary.get("ufcf_schedule")
    if not isinstance(ufcf_schedule, pd.DataFrame) or ufcf_schedule.empty:
        messages.append("UFCF schedule is empty after valuation assembly.")
    terminal_value = pd.to_numeric(
        pd.Series([summary.get("terminal_value")]), errors="coerce"
    ).iloc[0]
    if pd.isna(terminal_value) or terminal_value <= 0:
        messages.append("Terminal value did not compute from the current free-cash-flow profile.")
    irr_value = pd.to_numeric(pd.Series([summary.get("irr")]), errors="coerce").iloc[0]
    if pd.isna(irr_value):
        messages.append(
            "IRR is mathematically unavailable because the UFCF profile does not cross the zero-NPV threshold within the model's search range."
        )
    payback_years = pd.to_numeric(
        pd.Series([summary.get("payback_years")]), errors="coerce"
    ).iloc[0]
    if pd.isna(payback_years):
        messages.append(
            "Payback Period is mathematically unavailable because cumulative UFCF never turns positive over the modeled horizon."
        )
    return messages


def _render_model_audit(audit: Dict[str, Any]) -> None:
    if not isinstance(audit, dict):
        st.info("Model audit is not available for this scenario.")
        return

    status = str(audit.get("status", "")).strip().lower()
    headline = str(audit.get("headline", "Model audit is not available.")).strip()
    if status == "critical":
        st.error(headline)
    elif status == "warning":
        st.warning(headline)
    else:
        st.success(headline)

    summary = audit.get("summary")
    if isinstance(summary, pd.DataFrame) and not summary.empty:
        row = summary.iloc[0]
        cols = st.columns(4)
        cols[0].metric("Audit Score", f"{int(row.get('Score', 0))}/100")
        cols[1].metric("Critical", f"{int(row.get('Critical Issues', 0))}")
        cols[2].metric("Warnings", f"{int(row.get('Warnings', 0))}")
        cols[3].metric("Info", f"{int(row.get('Info', 0))}")

    reasoning = audit.get("reasoning")
    if isinstance(reasoning, list) and reasoning:
        st.markdown("**Audit reasoning**")
        for note in reasoning:
            st.write(f"- {note}")

    issues = audit.get("issues")
    if isinstance(issues, pd.DataFrame) and not issues.empty:
        display_columns = [
            col
            for col in [
                "Severity",
                "Category",
                "Metric",
                "Message",
                "Reasoning",
                "Recommendation",
                "Periods",
            ]
            if col in issues.columns
        ]
        st.markdown("**Audit findings**")
        st.dataframe(issues[display_columns], use_container_width=True)


def _scenario_viability_table(
    scenario_results: Dict[str, Dict[str, Any]]
) -> pd.DataFrame:
    rows: list[Dict[str, Any]] = []
    for scenario_name, payload in (scenario_results or {}).items():
        valuation = payload.get("valuation", {}) or {}
        debt_capacity = payload.get("debt_capacity_annual")
        if not isinstance(debt_capacity, pd.DataFrame):
            debt_capacity = pd.DataFrame()
        rows.append(
            {
                "Scenario": scenario_name,
                "NPV": valuation.get("npv"),
                "IRR": valuation.get("irr"),
                "Payback (Years)": valuation.get("payback_years"),
                "Terminal Value": valuation.get("terminal_value"),
                "Min DSCR": (
                    pd.to_numeric(debt_capacity.get("DSCR"), errors="coerce").min()
                    if not debt_capacity.empty and "DSCR" in debt_capacity.columns
                    else np.nan
                ),
                "Min DSCR Headroom": (
                    pd.to_numeric(debt_capacity.get("DSCR Headroom"), errors="coerce").min()
                    if not debt_capacity.empty and "DSCR Headroom" in debt_capacity.columns
                    else np.nan
                ),
                "Min Cash Headroom": (
                    pd.to_numeric(
                        debt_capacity.get("Cash Reserve Headroom"), errors="coerce"
                    ).min()
                    if not debt_capacity.empty
                    and "Cash Reserve Headroom" in debt_capacity.columns
                    else np.nan
                ),
                "Covenant Breach Periods": (
                    int(debt_capacity.get("Covenant Breach", pd.Series(dtype=bool)).sum())
                    if not debt_capacity.empty and "Covenant Breach" in debt_capacity.columns
                    else 0
                ),
            }
        )
    comparison = pd.DataFrame(rows)
    if comparison.empty:
        return comparison
    return comparison.set_index("Scenario")


def _scenario_output_specs() -> list[ScenarioOutputSpec]:
    return [
        ScenarioOutputSpec("Outputs", lambda ctx: _dynamic_outputs_table(ctx["model"], ctx["scenario_df"])),
        ScenarioOutputSpec("Benchmark KPIs", lambda ctx: _dynamic_benchmark_kpis_table(ctx["scenario_kpis"])),
        ScenarioOutputSpec("Debt Schedule", lambda ctx: ctx["debt_schedule_annual"]),
        ScenarioOutputSpec("Equity Schedule", lambda ctx: ctx["equity_schedule_annual"]),
        ScenarioOutputSpec("Working Capital Schedule", lambda ctx: ctx["working_capital_annual"]),
        ScenarioOutputSpec("Debt Capacity Schedule", lambda ctx: ctx["debt_capacity_annual"]),
        ScenarioOutputSpec("UFCF Schedule", lambda ctx: ctx["ufcf_annual"]),
        ScenarioOutputSpec("Model Audit Summary", lambda ctx: ctx["model_audit"].get("summary")),
        ScenarioOutputSpec("Model Audit Findings", lambda ctx: ctx["model_audit"].get("issues")),
    ]


def _execute_scenario_suite(
    schedule_df: pd.DataFrame,
    valuation_inputs: Dict[str, float],
    supplementary_tables: Dict[str, pd.DataFrame],
    scenario_suite: Dict[str, Dict[str, Any]],
) -> Tuple[GoatModel, pd.DataFrame, Dict[str, Dict[str, Any]]]:
    author_name = _current_model_author()
    hooks = ScenarioBuildHooks(
        input_schedule_cls=InputSchedule,
        biological_driver_defaults=BIOLOGICAL_SCENARIO_DRIVER_DEFAULTS,
        apply_biological_shocks=_apply_biological_scenario_shocks,
        apply_biological_assumptions_to_schedule=_apply_biological_assumptions_to_schedule,
        apply_operating_cost_assumptions_to_schedule=_apply_operating_cost_assumptions_to_schedule,
        apply_commercial_shocks_to_pricing=_apply_commercial_shocks_to_pricing,
        apply_pricing_assumptions_to_schedule=_apply_pricing_assumptions_to_schedule,
        derive_biological_schedules=_derive_biological_schedules,
        scenario_output_specs=_scenario_output_specs,
        pricing_family_summary=_pricing_family_summary,
        pricing_quantity_by_period=_pricing_quantity_by_period,
    )
    return run_scenario_suite(
        schedule_df=schedule_df,
        valuation_inputs=valuation_inputs,
        supplementary_tables=supplementary_tables,
        scenario_suite=scenario_suite,
        author_name=author_name,
        hooks=hooks,
    )


def _ensure_active_scenario_selection() -> None:
    scenario_results = st.session_state.get("all_scenario_results") or {}
    if not scenario_results:
        return

    selected = st.session_state.get("selected_scenario_name")
    if selected not in scenario_results:
        selected = next(iter(scenario_results))
        st.session_state.selected_scenario_name = selected

    current = st.session_state.get("results")
    if not isinstance(current, dict) or current.get("selected_scenario") != selected:
        st.session_state.results = scenario_results[selected]


def _render_scenario_selector(prefix: str = "main") -> None:
    scenario_results = st.session_state.get("all_scenario_results") or {}
    if not scenario_results:
        st.info("Run the scenario suite to enable comparisons.")
        return

    options = list(scenario_results.keys())
    selected_name = st.session_state.get("selected_scenario_name")
    try:
        default_index = options.index(selected_name)
    except ValueError:
        default_index = 0
        st.session_state.selected_scenario_name = options[0]
        selected_name = options[0]

    selected = st.selectbox(
        "Select scenario",
        options,
        index=default_index,
        key=f"scenario_selector::{prefix}",
    )

    if selected != selected_name:
        st.session_state.selected_scenario_name = selected
        st.session_state.results = scenario_results[selected]
        selected_name = selected

    description = scenario_results[selected_name].get("preset_description")
    adjustments = scenario_results[selected_name].get("scenario_inputs", {})
    author_name = scenario_results[selected_name].get("model_author")

    details: list[str] = []
    if description:
        details.append(description)

    if adjustments:
        numeric_adjustments = [
            f"{driver}: {value:+.1f}%"
            for driver, value in adjustments.items()
            if isinstance(value, (int, float)) and not pd.isna(value)
        ]
        if numeric_adjustments:
            details.append(f"Adjustments - {', '.join(numeric_adjustments)}")

    if isinstance(author_name, str) and author_name.strip():
        details.append(f"Prepared by - {author_name.strip()}")

    if details:
        st.caption(" \n".join(details))


def _scenario_key_suffix(label: str) -> str:
    normalized = re.sub(r"[^0-9a-zA-Z]+", "_", label).strip("_").lower()
    return normalized or "scenario"


def _sanitize_sheet_name(name: str, existing: set[str]) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in name).strip()
    cleaned = cleaned or "Sheet"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    base = cleaned
    suffix = 1
    while cleaned in existing:
        candidate = f"{base[:31 - len(str(suffix)) - 1]}_{suffix}" if len(base) >= 30 else f"{base}_{suffix}"
        cleaned = candidate[:31]
        suffix += 1
    existing.add(cleaned)
    return cleaned


def _prepare_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()

    frame = df.copy()
    if isinstance(frame, pd.Series):
        frame = frame.to_frame()

    if isinstance(frame.columns, pd.MultiIndex):
        frame.columns = [
            " - ".join(str(part) for part in tup if str(part))
            for tup in frame.columns.to_flat_index()
        ]

    index = frame.index
    if isinstance(index, pd.MultiIndex):
        frame = frame.reset_index()
    elif isinstance(index, pd.DatetimeIndex):
        frame = frame.reset_index()
        first_col = frame.columns[0]
        frame[first_col] = _normalize_period(frame[first_col])
        if first_col == "index" or not first_col:
            frame = frame.rename(columns={first_col: "Period"})
    elif not isinstance(index, pd.RangeIndex) or index.name:
        frame = frame.reset_index()
        first_col = frame.columns[0]
        if first_col == "index" and index.name:
            frame = frame.rename(columns={first_col: index.name})

    return frame


def _render_excel_download_panel(
    container: DeltaGenerator,
    results: Optional[Dict[str, Any]],
    results_stale: bool,
) -> None:
    with container:
        st.markdown("#### Excel Model Download")
        if results is None:
            st.info("Run the model to enable the Excel workbook download.")
            return

        selected_scenario = str(results.get("selected_scenario", "Scenario"))
        excel_map: Dict[str, bytes] = st.session_state.setdefault("excel_bytes_map", {})
        excel_bytes = excel_map.get(selected_scenario)
        key_suffix = _scenario_key_suffix(selected_scenario)

        if results_stale:
            st.info("Run the model again to prepare a workbook from the latest inputs.")
            if not excel_bytes:
                st.info("After rerunning, click 'Prepare Excel Model' to generate the workbook.")
            return

        if not excel_bytes:
            if st.button(
                "Prepare Excel Model",
                key=f"prepare_excel_{key_suffix}",
            ):
                with st.spinner("Preparing Excel workbook..."):
                    excel_bytes = _generate_excel_bytes(
                        results["model"],
                        results,
                        selected_scenario,
                        _current_model_author(),
                    )
                excel_map[selected_scenario] = excel_bytes
                st.session_state.excel_bytes_map = excel_map

        if excel_bytes:
            st.download_button(
                "Download Excel Model",
                data=excel_bytes,
                file_name="Goat_Farm_Financial_Model.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_excel_{key_suffix}",
            )
            if st.button(
                "Clear Prepared Excel",
                key=f"clear_excel_{key_suffix}",
            ):
                excel_map.pop(selected_scenario, None)
                st.session_state.excel_bytes_map = excel_map
        else:
            st.info("Click 'Prepare Excel Model' to generate the workbook for download.")


def _resolve_excel_writer_engine() -> str:
    if find_spec("xlsxwriter") is not None:
        return "xlsxwriter"
    if find_spec("openpyxl") is not None:
        return "openpyxl"
    raise RuntimeError(
        "Excel export requires the XlsxWriter or openpyxl package. "
        "Install one of them to enable downloads."
    )


def _generate_excel_bytes(
    model: GoatModel,
    results: Dict[str, Any],
    scenario_name: str,
    author_name: Optional[str] = None,
) -> bytes:
    bundle = prepare_export_bundle(
        model,
        scenario_name=scenario_name,
        author_name=author_name or DEFAULT_MODEL_AUTHOR,
        base_df=results.get("base"),
        scenario_df=results.get("scenario"),
        kpis_df=results.get("kpis"),
        break_even_df=results.get("break_even"),
        supplementary=results.get("supplementary", {}),
        scenario_inputs=results.get("scenario_inputs", {}),
    )
    return generate_excel_workbook(bundle)


DEFAULT_VARIABLE_ITEMS = [
    {"Item": "Feed & Supplements", "Share %": 5.0},
    {"Item": "Veterinary & Healthcare", "Share %": 4.0},
    {"Item": "Distribution & Logistics", "Share %": 3.0},
]


DEFAULT_DIRECT_WAGE_ITEMS = [
    {
        "Position": "Operations Crew",
        "Head Count": 3.0,
        "Monthly Salary per Head": 630.0,
        "Total Salary": 1890.0,
        "Share %": 67.5,
    },
    {
        "Position": "Herd Supervisor",
        "Head Count": 1.0,
        "Monthly Salary per Head": 910.0,
        "Total Salary": 910.0,
        "Share %": 32.5,
    },
]


DEFAULT_ADMIN_WAGE_ITEMS = [
    {
        "Position": "Administration",
        "Head Count": 1.0,
        "Monthly Salary per Head": 490.0,
        "Total Salary": 490.0,
        "Share %": 40.0,
    },
    {
        "Position": "Finance & Compliance",
        "Head Count": 1.0,
        "Monthly Salary per Head": 420.0,
        "Total Salary": 420.0,
        "Share %": 34.0,
    },
    {
        "Position": "Sales & Support",
        "Head Count": 1.0,
        "Monthly Salary per Head": 315.0,
        "Total Salary": 315.0,
        "Share %": 26.0,
    },
]


DEFAULT_PRICING_ROWS = _default_pricing_row_templates()


DEFAULT_PRODUCTION_DRIVER_ROWS = _default_production_driver_row_templates()

_PRODUCTION_DRIVER_REQUIRED_COLUMNS = [
    "Product",
    "Unit",
    "Business Unit",
    "Quantity Source",
    "Quantity Mode",
    "Lactating Herd Share %",
    "Litres per Lactating Doe per Day",
    "Milk Allocation to Cheese %",
    "Cheese Yield Kg per Litre",
    ANNUAL_SLAUGHTER_RATE_COLUMN,
    "Live Herd Sales Share %",
    "Meat Yield Kg per Goat",
    "Offal Yield Kg per Goat",
    "Pelt Units per Goat",
    "Driver Growth %",
]

_PRODUCTION_DRIVER_NUMERIC_COLUMNS = [
    "Lactating Herd Share %",
    "Litres per Lactating Doe per Day",
    "Milk Allocation to Cheese %",
    "Cheese Yield Kg per Litre",
    ANNUAL_SLAUGHTER_RATE_COLUMN,
    "Live Herd Sales Share %",
    "Meat Yield Kg per Goat",
    "Offal Yield Kg per Goat",
    "Pelt Units per Goat",
    "Driver Growth %",
]

_PRODUCTION_DRIVER_DAIRY_PRODUCTS = ("Milk", "Cheese")
_PRODUCTION_DRIVER_SLAUGHTER_PRODUCTS = ("Meat", "Offal", "Pelt", "Live Herd")


DEFAULT_OPERATING_COST_ROWS = [
    {
        "Year": 2024,
        "Business Unit": "General",
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "unit_cost_per_head_per_month": 6.64,
        "Inflation %": 4.0,
    },
    {
        "Year": 2025,
        "Business Unit": "General",
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "unit_cost_per_head_per_month": 6.91,
        "Inflation %": 4.0,
    },
    {
        "Year": 2024,
        "Business Unit": "General",
        "Field": "variable_healthcare_cost_per_herd",
        "Category": "Healthcare",
        "unit_cost_per_head_per_month": 1.41,
        "Inflation %": 3.5,
    },
    {
        "Year": 2025,
        "Business Unit": "General",
        "Field": "variable_healthcare_cost_per_herd",
        "Category": "Healthcare",
        "unit_cost_per_head_per_month": 1.46,
        "Inflation %": 3.5,
    },
    {
        "Year": 2024,
        "Business Unit": "General",
        "Field": "fixed_utility_cost_per_herd",
        "Category": "Utilities",
        "unit_cost_per_head_per_month": 0.94,
        "Inflation %": 2.0,
    },
    {
        "Year": 2025,
        "Business Unit": "General",
        "Field": "fixed_utility_cost_per_herd",
        "Category": "Utilities",
        "unit_cost_per_head_per_month": 0.96,
        "Inflation %": 2.0,
    },
]

OPERATING_COST_FIELD_TO_CATEGORY = {
    "variable_feed_cost_per_herd": "Feed",
    "variable_healthcare_cost_per_herd": "Healthcare",
    "fixed_utility_cost_per_herd": "Utilities",
}

BIOLOGICAL_COST_APPLIES_TO_MAP = {
    "breeding_doe": "Breeding Does",
    "breeding_buck": "Breeding Bucks",
    "replacement_doe": "Replacement Does",
    "finishing_female": "Finishing Females",
    "finishing_male": "Finishing Males",
    "lactating_doe": "Lactating Does",
    "total_herd": "Herd Size (heads)",
}


DEFAULT_INPUT_CONFIG_KEY = "default_input_templates"


def _default_input_template_config() -> Dict[str, list[dict[str, object]]]:
    return {
        "variable_items": deepcopy(DEFAULT_VARIABLE_ITEMS),
        "direct_wage_items": deepcopy(DEFAULT_DIRECT_WAGE_ITEMS),
        "admin_wage_items": deepcopy(DEFAULT_ADMIN_WAGE_ITEMS),
        "pricing_rows": deepcopy(DEFAULT_PRICING_ROWS),
        "production_driver_rows": deepcopy(DEFAULT_PRODUCTION_DRIVER_ROWS),
        "operating_rows": deepcopy(DEFAULT_OPERATING_COST_ROWS),
    }


def _ensure_default_templates() -> Dict[str, list[dict[str, object]]]:
    templates = _safe_session_state_get(DEFAULT_INPUT_CONFIG_KEY)
    if not isinstance(templates, dict):
        templates = _default_input_template_config()
        _safe_session_state_set(DEFAULT_INPUT_CONFIG_KEY, templates)
    return templates


def _template_copy(template: list[dict[str, object]]) -> list[dict[str, object]]:
    return [dict(row) for row in template]


def _get_template(name: str, fallback: list[dict[str, object]]) -> list[dict[str, object]]:
    templates = _ensure_default_templates()
    rows = templates.get(name)
    if rows:
        return deepcopy(rows)
    return deepcopy(fallback)


def _set_template(name: str, rows: list[dict[str, object]]) -> None:
    templates = _ensure_default_templates()
    templates[name] = deepcopy(rows)
    st.session_state[DEFAULT_INPUT_CONFIG_KEY] = templates


def _template_to_dataframe(
    rows: list[dict[str, object]], columns: list[str]
) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=columns)

    df = pd.DataFrame(rows)
    for column in columns:
        if column not in df.columns:
            df[column] = np.nan
    return df[columns]


def _dataframe_to_template(df: pd.DataFrame, columns: list[str]) -> list[dict[str, object]]:
    if df is None or df.empty:
        return []

    work = df.copy()
    for column in columns:
        if column not in work.columns:
            work[column] = np.nan

    records: list[dict[str, object]] = []
    for _, row in work[columns].iterrows():
        record: dict[str, object] = {}
        empty = True
        for column in columns:
            value = row[column]
            if isinstance(value, str):
                cleaned = value.strip()
                record[column] = cleaned or None
                if cleaned:
                    empty = False
            else:
                if pd.isna(value):
                    record[column] = None
                else:
                    record[column] = float(value)
                    empty = False
        if not empty:
            records.append(record)
    return records


def _default_pricing_table() -> pd.DataFrame:
    return _default_pricing_table_from_core(_default_income_schedule(periods=12))


def _default_production_driver_table() -> pd.DataFrame:
    rows = _get_template("production_driver_rows", DEFAULT_PRODUCTION_DRIVER_ROWS)
    columns = list(_PRODUCTION_DRIVER_REQUIRED_COLUMNS)
    table = _template_to_dataframe(rows, columns)
    if table.empty:
        return pd.DataFrame(columns=columns)
    return _ensure_production_driver_table(table)


def _ensure_production_driver_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = _template_to_dataframe(
            _get_template("production_driver_rows", DEFAULT_PRODUCTION_DRIVER_ROWS),
            list(_PRODUCTION_DRIVER_REQUIRED_COLUMNS),
        )
    else:
        work = table.copy()

    if (
        ANNUAL_SLAUGHTER_RATE_COLUMN not in work.columns
        and LEGACY_SLAUGHTER_RATE_COLUMN in work.columns
    ):
        work[ANNUAL_SLAUGHTER_RATE_COLUMN] = work[LEGACY_SLAUGHTER_RATE_COLUMN]
    if LEGACY_SLAUGHTER_RATE_COLUMN in work.columns:
        work = work.drop(columns=[LEGACY_SLAUGHTER_RATE_COLUMN])

    for column in _PRODUCTION_DRIVER_REQUIRED_COLUMNS:
        if column not in work.columns:
            work[column] = np.nan

    work["Product"] = _series_or_default(work, "Product", "").astype(str).str.strip()
    work.loc[work["Product"] == "", "Product"] = "Product"
    work["Unit"] = _series_or_default(work, "Unit", "").astype(str).str.strip()
    work["Business Unit"] = _series_or_default(work, "Business Unit", DEFAULT_BUSINESS_TYPE).astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = DEFAULT_BUSINESS_TYPE
    work["Quantity Source"] = _series_or_default(work, "Quantity Source", "Slaughter Output").astype(str).str.strip()
    work.loc[work["Quantity Source"] == "", "Quantity Source"] = "Slaughter Output"
    work["Quantity Mode"] = _series_or_default(work, "Quantity Mode", "Derived").astype(str).str.strip()
    work.loc[~work["Quantity Mode"].isin(["Derived", "Manual Override"]), "Quantity Mode"] = "Derived"

    for col in _PRODUCTION_DRIVER_NUMERIC_COLUMNS:
        work[col] = pd.to_numeric(work.get(col), errors="coerce").fillna(0.0)

    work = work.dropna(how="all")
    if work.empty:
        return _template_to_dataframe(
            DEFAULT_PRODUCTION_DRIVER_ROWS,
            list(_PRODUCTION_DRIVER_REQUIRED_COLUMNS),
        )

    defaults = {str(row["Product"]).strip(): row for _, row in pd.DataFrame(DEFAULT_PRODUCTION_DRIVER_ROWS).iterrows()}
    for idx in work.index:
        product = str(work.at[idx, "Product"]).strip()
        default_row = defaults.get(product)
        if default_row is not None and not str(work.at[idx, "Unit"]).strip():
            work.at[idx, "Unit"] = str(default_row.get("Unit", "")).strip()
        if default_row is not None and not str(work.at[idx, "Business Unit"]).strip():
            work.at[idx, "Business Unit"] = str(default_row.get("Business Unit", DEFAULT_BUSINESS_TYPE)).strip()
        if default_row is not None and not str(work.at[idx, "Quantity Source"]).strip():
            work.at[idx, "Quantity Source"] = str(default_row.get("Quantity Source", "Slaughter Output")).strip()

    remainder = [col for col in work.columns if col not in _PRODUCTION_DRIVER_REQUIRED_COLUMNS]
    return work[list(_PRODUCTION_DRIVER_REQUIRED_COLUMNS) + remainder].reset_index(drop=True)


def _sync_production_driver_table_to_products(
    table: Optional[pd.DataFrame],
    products: Sequence[str],
    business_type: Any = DEFAULT_BUSINESS_TYPE,
) -> pd.DataFrame:
    product_list = [str(product).strip() for product in products if str(product).strip()]
    base = _ensure_production_driver_table(
        pd.DataFrame(_default_production_driver_row_templates(product_list, business_type))
    )
    current = _ensure_production_driver_table(table)
    current = current.loc[
        current["Product"].astype(str).str.strip().isin(product_list)
    ].copy()
    if not current.empty:
        current["Product"] = current["Product"].astype(str).str.strip()
        current = current.drop_duplicates(subset=["Product"], keep="last").reset_index(drop=True)

    for column in current.columns:
        if column not in base.columns:
            base[column] = np.nan
    for column in base.columns:
        if column not in current.columns:
            current[column] = np.nan

    base = base.set_index("Product")
    current = current.set_index("Product")
    merged = base.combine_first(current)
    merged.update(current)
    merged = merged.reset_index()
    merged["__sort_order"] = merged["Product"].map({product: idx for idx, product in enumerate(product_list)}).fillna(len(product_list))
    merged = merged.sort_values("__sort_order", kind="stable").drop(columns="__sort_order")
    return _ensure_production_driver_table(merged)


def _production_driver_custom_columns(table: Optional[pd.DataFrame]) -> list[str]:
    drivers = _ensure_production_driver_table(table)
    return [col for col in drivers.columns if col not in _PRODUCTION_DRIVER_REQUIRED_COLUMNS]


def _normalise_production_driver_column_name(column_name: str) -> str:
    return re.sub(r"\s+", " ", str(column_name).strip())


def _add_production_driver_column(
    table: Optional[pd.DataFrame],
    column_name: str,
    default_value: Any = "",
) -> pd.DataFrame:
    drivers = _ensure_production_driver_table(table)
    cleaned = _normalise_production_driver_column_name(column_name)
    if not cleaned or cleaned in drivers.columns:
        return drivers
    drivers[cleaned] = default_value
    return _ensure_production_driver_table(drivers)


def _remove_production_driver_columns(
    table: Optional[pd.DataFrame],
    columns: Iterable[str],
) -> pd.DataFrame:
    drivers = _ensure_production_driver_table(table)
    removable = [
        _normalise_production_driver_column_name(column)
        for column in columns
        if _normalise_production_driver_column_name(column)
        and _normalise_production_driver_column_name(column) in drivers.columns
        and _normalise_production_driver_column_name(column) not in _PRODUCTION_DRIVER_REQUIRED_COLUMNS
    ]
    if not removable:
        return drivers
    return _ensure_production_driver_table(drivers.drop(columns=removable))


def _merge_production_driver_subset(
    table: Optional[pd.DataFrame],
    subset: pd.DataFrame,
    products: Sequence[str],
) -> pd.DataFrame:
    base = _ensure_production_driver_table(table)
    edited = subset.copy()
    edited["Product"] = _series_or_default(edited, "Product", "").astype(str).str.strip()
    product_keys = {str(product).strip().casefold() for product in products if str(product).strip()}

    for column in edited.columns:
        if column not in base.columns:
            base[column] = np.nan
    for column in base.columns:
        if column not in edited.columns:
            edited[column] = np.nan

    base_order = {
        str(product).strip().casefold(): idx
        for idx, product in enumerate(base.get("Product", pd.Series(dtype=str)).astype(str).tolist())
        if str(product).strip()
    }

    keep_mask = ~base["Product"].astype(str).str.strip().str.casefold().isin(product_keys)
    merged = pd.concat([base.loc[keep_mask].copy(), edited], ignore_index=True, sort=False)
    merged["Product"] = merged["Product"].astype(str).str.strip()
    merged = merged.drop_duplicates(subset=["Product"], keep="last").reset_index(drop=True)
    merged["_sort_order"] = (
        merged["Product"].astype(str).str.strip().str.casefold().map(base_order).fillna(len(base_order))
    )
    merged = merged.sort_values("_sort_order", kind="stable").drop(columns="_sort_order")
    return _ensure_production_driver_table(merged)


def _production_driver_column_config(
    table: Optional[pd.DataFrame],
    product_options: Optional[Sequence[str]] = None,
) -> dict[str, Any]:
    drivers = _ensure_production_driver_table(table)
    cleaned_product_options = [
        str(product).strip()
        for product in (product_options or [])
        if str(product).strip()
    ]
    config: dict[str, Any] = {
        "Product": (
            st.column_config.SelectboxColumn("Product", options=cleaned_product_options)
            if cleaned_product_options
            else st.column_config.TextColumn("Product")
        ),
        "Unit": st.column_config.TextColumn("Unit"),
        "Quantity Mode": st.column_config.SelectboxColumn(
            "Quantity Mode",
            options=["Derived", "Manual Override"],
        ),
        "Lactating Herd Share %": st.column_config.NumberColumn(
            "Lactating Herd Share (%)", format="%.2f", step=0.1
        ),
        "Litres per Lactating Doe per Day": st.column_config.NumberColumn(
            "Litres / Doe / Day", format="%.3f", step=0.1
        ),
        "Milk Allocation to Cheese %": st.column_config.NumberColumn(
            "Milk to Cheese (%)", format="%.2f", step=0.1
        ),
        "Cheese Yield Kg per Litre": st.column_config.NumberColumn(
            "Cheese Yield (Kg/Litre)", format="%.3f", step=0.01
        ),
        ANNUAL_SLAUGHTER_RATE_COLUMN: st.column_config.NumberColumn(
            "Annual Slaughter Rate (%)", format="%.2f", step=0.1
        ),
        "Live Herd Sales Share %": st.column_config.NumberColumn(
            "Live Herd Sales Share (%)", format="%.2f", step=0.1
        ),
        "Meat Yield Kg per Goat": st.column_config.NumberColumn(
            "Meat Yield (Kg/Goat)", format="%.2f", step=0.1
        ),
        "Offal Yield Kg per Goat": st.column_config.NumberColumn(
            "Offal Yield (Kg/Goat)", format="%.2f", step=0.1
        ),
        "Pelt Units per Goat": st.column_config.NumberColumn(
            "Pelt Units / Goat", format="%.2f", step=0.1
        ),
        "Driver Growth %": st.column_config.NumberColumn(
            "Annual Driver Growth (%)", format="%.2f", step=0.1
        ),
    }

    for column in drivers.columns:
        if column in config:
            continue
        series = drivers[column]
        if is_bool_dtype(series):
            config[column] = st.column_config.CheckboxColumn(column)
        elif is_numeric_dtype(series):
            config[column] = st.column_config.NumberColumn(column, format="%.2f", step=0.1)
        else:
            config[column] = st.column_config.TextColumn(column)
    return config


DEFAULT_BIOLOGICAL_SYSTEM_SETTINGS = [
    {"Setting": "Model Grain", "Value": "monthly"},
    {"Setting": "Opening Biological Start Date", "Value": "2024-01-31"},
    {"Setting": "Age Band Width (months)", "Value": "1"},
    {"Setting": "Use Herd Plan as Target", "Value": "True"},
    {"Setting": "Use Herd Plan as Hard Override", "Value": "False"},
]


DEFAULT_BREEDING_REPRODUCTION_ROWS = [
    {
        "Breeding Group": "Base Breeding Herd",
        "Age at First Kidding (months)": 18.0,
        "Gestation Months": 5.0,
        "Kiddings per Doe per Year": 1.3,
        "Kids per Kidding": 2.2,
        "Conception Rate %": 92.0,
        "Female Birth Share %": 50.0,
        "Kid Mortality %": 4.0,
        "Doe Mortality %": 2.5,
        "Buck Mortality %": 2.5,
        "Cull Rate %": 8.0,
        "Breeder Doe Cull Age (months)": 72.0,
        "Breeder Doe Cull At Parity": np.nan,
        "Breeder Doe Live Sale Share %": 0.0,
        "Breeder Buck Replacement Age (months)": 60.0,
        "Breeder Buck Live Sale Share %": 0.0,
        "Replacement Retention %": 30.0,
        "Active": True,
    }
]


DEFAULT_LACTATION_BIOLOGY_ROWS = [
    {
        "Parity Group": "1",
        "Age at First Kidding (months)": 18.0,
        "Lactation Length Days": 260.0,
        "Dry Period Days": 90.0,
        "Peak Lactation Month": 2.0,
        "Peak Yield Litres per Day": 3.08,
        "Base Yield Litres per Day": 1.90,
        "Parity Yield Factor": 0.9,
        "Curve Shape Parameter": 1.2,
        "Active": True,
    },
    {
        "Parity Group": "2",
        "Age at First Kidding (months)": 18.0,
        "Lactation Length Days": 275.0,
        "Dry Period Days": 80.0,
        "Peak Lactation Month": 2.0,
        "Peak Yield Litres per Day": 3.74,
        "Base Yield Litres per Day": 2.30,
        "Parity Yield Factor": 1.0,
        "Curve Shape Parameter": 1.15,
        "Active": True,
    },
    {
        "Parity Group": "3+",
        "Age at First Kidding (months)": 18.0,
        "Lactation Length Days": 290.0,
        "Dry Period Days": 75.0,
        "Peak Lactation Month": 2.0,
        "Peak Yield Litres per Day": 4.07,
        "Base Yield Litres per Day": 2.50,
        "Parity Yield Factor": 1.08,
        "Curve Shape Parameter": 1.1,
        "Active": True,
    },
]


DEFAULT_FINISHING_SLAUGHTER_ROWS = [
    {
        "Product Group": "Default Livestock Stream",
        "Months to Market Weight": 7.0,
        "Age at Slaughter (months)": 9.0,
        "Target Live Weight Kg": 42.0,
        "Monthly Weight Gain Kg": 4.4,
        "Breeding Retention %": 30.0,
        "Live Herd Sales Share %": 30.0,
        "Carcass Yield %": 47.0,
        "Meat Yield Kg per Goat": 22.0,
        "Offal Yield Kg per Goat": 4.0,
        "Pelt Units per Goat": 1.0,
        "Mortality %": 3.5,
        "Active": True,
    }
]


DEFAULT_OPENING_HERD_COHORTS = [
    {
        "Cohort ID": "BD-001",
        "Business Unit": "Breeding",
        "Sex": "Female",
        "Purpose": "breeding_doe",
        "Age in Months": 30.0,
        "Head Count": 150.0,
        "Parity": 2.0,
        "Pregnant": True,
        "Days in Milk": 60.0,
        "Active": True,
    },
    {
        "Cohort ID": "BB-001",
        "Business Unit": "Breeding",
        "Sex": "Male",
        "Purpose": "breeding_buck",
        "Age in Months": 36.0,
        "Head Count": 8.0,
        "Parity": 0.0,
        "Pregnant": False,
        "Days in Milk": 0.0,
        "Active": True,
    },
    {
        "Cohort ID": "RD-001",
        "Business Unit": "Breeding",
        "Sex": "Female",
        "Purpose": "replacement_doe",
        "Age in Months": 10.0,
        "Head Count": 40.0,
        "Parity": 0.0,
        "Pregnant": False,
        "Days in Milk": 0.0,
        "Active": True,
    },
    {
        "Cohort ID": "FF-001",
        "Business Unit": "Breeding",
        "Sex": "Female",
        "Purpose": "finishing_female",
        "Age in Months": 5.0,
        "Head Count": 60.0,
        "Parity": 0.0,
        "Pregnant": False,
        "Days in Milk": 0.0,
        "Active": True,
    },
    {
        "Cohort ID": "MF-001",
        "Business Unit": "Breeding",
        "Sex": "Male",
        "Purpose": "finishing_male",
        "Age in Months": 5.0,
        "Head Count": 62.0,
        "Parity": 0.0,
        "Pregnant": False,
        "Days in Milk": 0.0,
        "Active": True,
    },
]


DEFAULT_COHORT_ALLOCATION_RULES = [
    {"Rule": "Female Replacement Retention %", "Value": 35.0},
    {"Rule": "Male Live Sale %", "Value": 0.0},
    {"Rule": "Female Live Sale %", "Value": 0.0},
    {"Rule": "Finishing Entry %", "Value": 100.0},
    {"Rule": "Breeding Entry %", "Value": 100.0},
]


DEFAULT_BIOLOGICAL_COST_DRIVER_ROWS = [
    {
        "Year": 2024,
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "Applies To": "breeding_doe",
        "unit_cost_per_head_per_month": 3.96,
        "Inflation %": 4.0,
        "Active": True,
    },
    {
        "Year": 2024,
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "Applies To": "breeding_buck",
        "unit_cost_per_head_per_month": 3.60,
        "Inflation %": 4.0,
        "Active": True,
    },
    {
        "Year": 2024,
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "Applies To": "replacement_doe",
        "unit_cost_per_head_per_month": 3.06,
        "Inflation %": 4.0,
        "Active": True,
    },
    {
        "Year": 2024,
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "Applies To": "finishing_female",
        "unit_cost_per_head_per_month": 2.88,
        "Inflation %": 4.0,
        "Active": True,
    },
    {
        "Year": 2024,
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "Applies To": "finishing_male",
        "unit_cost_per_head_per_month": 2.97,
        "Inflation %": 4.0,
        "Active": True,
    },
    {
        "Year": 2024,
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "Applies To": "lactating_doe",
        "unit_cost_per_head_per_month": 1.44,
        "Inflation %": 4.0,
        "Active": True,
    },
    {
        "Year": 2024,
        "Field": "variable_healthcare_cost_per_herd",
        "Category": "Healthcare",
        "Applies To": "total_herd",
        "unit_cost_per_head_per_month": 2.25,
        "Inflation %": 3.5,
        "Active": True,
    },
    {
        "Year": 2024,
        "Field": "fixed_utility_cost_per_herd",
        "Category": "Utilities",
        "Applies To": "total_herd",
        "unit_cost_per_head_per_month": 0.75,
        "Inflation %": 2.0,
        "Active": True,
    },
]


def _merge_missing_default_rows(
    table: pd.DataFrame,
    default_rows: Sequence[dict[str, Any]],
    key_field: str,
) -> pd.DataFrame:
    work = table.copy()
    existing = {
        str(row.get(key_field, "")).strip().casefold()
        for _, row in work.iterrows()
        if str(row.get(key_field, "")).strip()
    }
    for row in default_rows:
        if str(row[key_field]).strip().casefold() not in existing:
            work = pd.concat([work, pd.DataFrame([row])], ignore_index=True)
    return work


def _series_or_default(
    frame: pd.DataFrame,
    column: str,
    default: Any,
) -> pd.Series:
    series = frame.get(column)
    if series is None:
        series = pd.Series([default] * len(frame), index=frame.index)
    return series.fillna(default)


def _normalize_biological_system_settings_table(table: pd.DataFrame) -> pd.DataFrame:
    work = table.copy()
    for column in ["Setting", "Value"]:
        if column not in work.columns:
            work[column] = ""
    work["Setting"] = _series_or_default(work, "Setting", "").astype(str).str.strip()
    work["Value"] = _series_or_default(work, "Value", "").astype(str).str.strip()
    work.loc[work["Setting"] == "", "Setting"] = "Setting"
    return _merge_missing_default_rows(work, DEFAULT_BIOLOGICAL_SYSTEM_SETTINGS, "Setting")


def _normalize_breeding_reproduction_biology_table(table: pd.DataFrame) -> pd.DataFrame:
    work = table.copy()
    work["Breeding Group"] = _series_or_default(work, "Breeding Group", "").fillna("").astype(str).str.strip()
    work.loc[work["Breeding Group"] == "", "Breeding Group"] = "Breeding Group"
    for column in [
        "Age at First Kidding (months)",
        "Gestation Months",
        "Kiddings per Doe per Year",
        "Kids per Kidding",
        "Conception Rate %",
        "Female Birth Share %",
        "Kid Mortality %",
        "Doe Mortality %",
        "Buck Mortality %",
        "Cull Rate %",
        "Breeder Doe Cull Age (months)",
        "Breeder Doe Cull At Parity",
        "Breeder Doe Live Sale Share %",
        "Breeder Buck Replacement Age (months)",
        "Breeder Buck Live Sale Share %",
        "Replacement Retention %",
    ]:
        work[column] = pd.to_numeric(work.get(column), errors="coerce")
    work["Active"] = _series_or_default(work, "Active", True).map(lambda value: _coerce_bool_value(value, True))
    return work


def _normalize_lactation_biology_table(table: pd.DataFrame) -> pd.DataFrame:
    work = table.copy()
    work["Parity Group"] = _series_or_default(work, "Parity Group", "").fillna("").astype(str).str.strip()
    work.loc[work["Parity Group"] == "", "Parity Group"] = "1"
    for column in [
        "Age at First Kidding (months)",
        "Lactation Length Days",
        "Dry Period Days",
        "Peak Lactation Month",
        "Peak Yield Litres per Day",
        "Base Yield Litres per Day",
        "Parity Yield Factor",
        "Curve Shape Parameter",
    ]:
        work[column] = pd.to_numeric(work.get(column), errors="coerce")
    work["Active"] = _series_or_default(work, "Active", True).map(lambda value: _coerce_bool_value(value, True))
    return work


def _normalize_finishing_slaughter_biology_table(table: pd.DataFrame) -> pd.DataFrame:
    work = table.copy()
    work["Product Group"] = _series_or_default(work, "Product Group", "").fillna("").astype(str).str.strip()
    work.loc[work["Product Group"] == "", "Product Group"] = "Default Livestock Stream"
    for column in [
        "Months to Market Weight",
        "Age at Slaughter (months)",
        "Target Live Weight Kg",
        "Monthly Weight Gain Kg",
        "Breeding Retention %",
        "Live Herd Sales Share %",
        "Carcass Yield %",
        "Meat Yield Kg per Goat",
        "Offal Yield Kg per Goat",
        "Pelt Units per Goat",
        "Mortality %",
    ]:
        work[column] = pd.to_numeric(work.get(column), errors="coerce")
    work["Active"] = _series_or_default(work, "Active", True).map(lambda value: _coerce_bool_value(value, True))
    return work


def _normalize_opening_herd_cohorts_table(table: pd.DataFrame) -> pd.DataFrame:
    work = table.copy()
    for column in ["Cohort ID", "Business Unit", "Sex", "Purpose"]:
        work[column] = _series_or_default(work, column, "").fillna("").astype(str).str.strip()
    work.loc[work["Cohort ID"] == "", "Cohort ID"] = "COHORT"
    work.loc[work["Business Unit"] == "", "Business Unit"] = "Breeding"
    work.loc[work["Sex"] == "", "Sex"] = "Female"
    work.loc[work["Purpose"] == "", "Purpose"] = "replacement_doe"
    for column in ["Age in Months", "Head Count", "Parity", "Days in Milk"]:
        work[column] = pd.to_numeric(work.get(column), errors="coerce")
    work["Pregnant"] = _series_or_default(work, "Pregnant", False).map(lambda value: _coerce_bool_value(value, False))
    work["Active"] = _series_or_default(work, "Active", True).map(lambda value: _coerce_bool_value(value, True))
    return work


def _normalize_cohort_allocation_rules_table(table: pd.DataFrame) -> pd.DataFrame:
    work = table.copy()
    for column in ["Rule", "Value"]:
        if column not in work.columns:
            work[column] = np.nan
    work["Rule"] = _series_or_default(work, "Rule", "").astype(str).str.strip()
    work.loc[work["Rule"] == "", "Rule"] = "Rule"
    work["Value"] = pd.to_numeric(work.get("Value"), errors="coerce")
    return _merge_missing_default_rows(work, DEFAULT_COHORT_ALLOCATION_RULES, "Rule")


def _normalize_biological_cost_drivers_table(table: pd.DataFrame) -> pd.DataFrame:
    work = table.copy()
    for column in ["Business Unit", "Field", "Category", "Applies To"]:
        work[column] = _series_or_default(work, column, "").fillna("").astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = "Breeding"
    work.loc[work["Field"] == "", "Field"] = "variable_feed_cost_per_herd"
    work.loc[work["Category"] == "", "Category"] = "Feed"
    work.loc[work["Applies To"] == "", "Applies To"] = "total_herd"
    work["Year"] = pd.to_numeric(work.get("Year"), errors="coerce").round().astype("Int64")
    work["unit_cost_per_head_per_month"] = pd.to_numeric(
        work.get("unit_cost_per_head_per_month"), errors="coerce"
    )
    work["Inflation %"] = pd.to_numeric(work.get("Inflation %"), errors="coerce")
    work["Active"] = _series_or_default(work, "Active", True).map(lambda value: _coerce_bool_value(value, True))
    return work


BIOLOGICAL_SYSTEM_SETTINGS_SCHEMA = TableSchema(
    name="Biological System Settings",
    columns=(
        ColumnSchema("Setting", ""),
        ColumnSchema("Value", ""),
    ),
    default_rows=tuple(DEFAULT_BIOLOGICAL_SYSTEM_SETTINGS),
    normalizer=_normalize_biological_system_settings_table,
)


BREEDING_REPRODUCTION_BIOLOGY_SCHEMA = TableSchema(
    name="Breeding & Reproduction Biology",
    columns=tuple(ColumnSchema(column, value) for column, value in DEFAULT_BREEDING_REPRODUCTION_ROWS[0].items()),
    default_rows=tuple(DEFAULT_BREEDING_REPRODUCTION_ROWS),
    normalizer=_normalize_breeding_reproduction_biology_table,
)


LACTATION_BIOLOGY_SCHEMA = TableSchema(
    name="Lactation Biology",
    columns=tuple(ColumnSchema(column, value) for column, value in DEFAULT_LACTATION_BIOLOGY_ROWS[0].items()),
    default_rows=tuple(DEFAULT_LACTATION_BIOLOGY_ROWS),
    normalizer=_normalize_lactation_biology_table,
)


FINISHING_SLAUGHTER_BIOLOGY_SCHEMA = TableSchema(
    name="Finishing & Slaughter Biology",
    columns=tuple(ColumnSchema(column, value) for column, value in DEFAULT_FINISHING_SLAUGHTER_ROWS[0].items()),
    default_rows=tuple(DEFAULT_FINISHING_SLAUGHTER_ROWS),
    normalizer=_normalize_finishing_slaughter_biology_table,
)


OPENING_HERD_COHORTS_SCHEMA = TableSchema(
    name="Opening Herd Cohorts",
    columns=tuple(ColumnSchema(column, value) for column, value in DEFAULT_OPENING_HERD_COHORTS[0].items()),
    default_rows=tuple(DEFAULT_OPENING_HERD_COHORTS),
    normalizer=_normalize_opening_herd_cohorts_table,
)


COHORT_ALLOCATION_RULES_SCHEMA = TableSchema(
    name="Cohort Allocation Rules",
    columns=(
        ColumnSchema("Rule", ""),
        ColumnSchema("Value", np.nan),
    ),
    default_rows=tuple(DEFAULT_COHORT_ALLOCATION_RULES),
    normalizer=_normalize_cohort_allocation_rules_table,
)


BIOLOGICAL_COST_DRIVERS_SCHEMA = TableSchema(
    name="Biological Cost Drivers",
    columns=tuple(ColumnSchema(column, value) for column, value in DEFAULT_BIOLOGICAL_COST_DRIVER_ROWS[0].items()),
    default_rows=tuple(DEFAULT_BIOLOGICAL_COST_DRIVER_ROWS),
    normalizer=_normalize_biological_cost_drivers_table,
)


def _default_biological_system_settings_table() -> pd.DataFrame:
    return _align_biological_system_settings_to_horizon(
        build_default_table(BIOLOGICAL_SYSTEM_SETTINGS_SCHEMA)
    )


def _ensure_biological_system_settings_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    return _align_biological_system_settings_to_horizon(
        ensure_table(BIOLOGICAL_SYSTEM_SETTINGS_SCHEMA, table)
    )


def _default_breeding_reproduction_biology_table() -> pd.DataFrame:
    return build_default_table(BREEDING_REPRODUCTION_BIOLOGY_SCHEMA)


def _ensure_breeding_reproduction_biology_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    return ensure_table(BREEDING_REPRODUCTION_BIOLOGY_SCHEMA, table)


def _default_lactation_biology_table() -> pd.DataFrame:
    return build_default_table(LACTATION_BIOLOGY_SCHEMA)


def _ensure_lactation_biology_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    return ensure_table(LACTATION_BIOLOGY_SCHEMA, table)


def _default_finishing_slaughter_biology_table() -> pd.DataFrame:
    return build_default_table(FINISHING_SLAUGHTER_BIOLOGY_SCHEMA)


def _ensure_finishing_slaughter_biology_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    return ensure_table(FINISHING_SLAUGHTER_BIOLOGY_SCHEMA, table)


def _default_opening_herd_cohorts_table() -> pd.DataFrame:
    return build_default_table(OPENING_HERD_COHORTS_SCHEMA)


def _ensure_opening_herd_cohorts_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    return ensure_table(OPENING_HERD_COHORTS_SCHEMA, table)


def _sync_opening_herd_cohorts_to_herd_plan(
    table: Optional[pd.DataFrame],
    herd_plan: Optional[pd.DataFrame],
) -> pd.DataFrame:
    work = _ensure_opening_herd_cohorts_table(table)
    plan = _ensure_herd_plan_table(herd_plan)

    target_heads = next(
        (
            float(value)
            for value in pd.to_numeric(plan.get("Herd Size (heads)"), errors="coerce").tolist()
            if pd.notna(value) and float(value) > 0
        ),
        np.nan,
    )
    if pd.isna(target_heads) or float(target_heads) <= 0:
        return work

    head_counts = pd.to_numeric(work.get("Head Count"), errors="coerce")
    current_total = float(head_counts.fillna(0.0).sum())
    if current_total <= 0:
        work = _default_opening_herd_cohorts_table()
        head_counts = pd.to_numeric(work.get("Head Count"), errors="coerce")
        current_total = float(head_counts.fillna(0.0).sum())
        if current_total <= 0:
            return work

    scale_factor = float(target_heads) / current_total
    work["Head Count"] = head_counts.fillna(0.0) * scale_factor
    return _ensure_opening_herd_cohorts_table(work)


def _default_cohort_allocation_rules_table() -> pd.DataFrame:
    return build_default_table(COHORT_ALLOCATION_RULES_SCHEMA)


def _ensure_cohort_allocation_rules_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    return ensure_table(COHORT_ALLOCATION_RULES_SCHEMA, table)


def _default_biological_cost_drivers_table() -> pd.DataFrame:
    return build_default_table(BIOLOGICAL_COST_DRIVERS_SCHEMA)


def _ensure_biological_cost_drivers_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    return ensure_table(BIOLOGICAL_COST_DRIVERS_SCHEMA, table)


def _string_to_bool(value: Any, default: bool = False) -> bool:
    if pd.isna(value):
        return default
    cleaned = str(value).strip().lower()
    if cleaned in {"true", "1", "yes", "y"}:
        return True
    if cleaned in {"false", "0", "no", "n"}:
        return False
    return default


def _coerce_numeric_scalar(value: Any) -> float:
    try:
        numeric = pd.to_numeric(value, errors="coerce")
    except Exception:
        return np.nan
    if isinstance(numeric, (pd.Series, pd.Index, np.ndarray, list, tuple)):
        array = np.asarray(numeric, dtype="float64").reshape(-1)
        numeric = array[0] if array.size else np.nan
    return float(numeric) if pd.notna(numeric) else np.nan


def _coerce_float_scalar(value: Any, default: float = 0.0) -> float:
    numeric = _coerce_numeric_scalar(value)
    return float(numeric) if pd.notna(numeric) else float(default)


def _coerce_int_scalar(value: Any, default: int = 0) -> int:
    numeric = _coerce_numeric_scalar(value)
    if pd.isna(numeric):
        return int(default)
    return int(round(float(numeric)))


def _assumption_bundle_ensure_map() -> dict[str, Callable[[Optional[pd.DataFrame]], pd.DataFrame]]:
    return {
        "Biological System Settings": _ensure_biological_system_settings_table,
        "Breeding & Reproduction Biology": _ensure_breeding_reproduction_biology_table,
        "Lactation Biology": _ensure_lactation_biology_table,
        "Finishing & Slaughter Biology": _ensure_finishing_slaughter_biology_table,
        "Opening Herd Cohorts": _ensure_opening_herd_cohorts_table,
        "Cohort Allocation Rules": _ensure_cohort_allocation_rules_table,
        "Biological Cost Drivers": _ensure_biological_cost_drivers_table,
        "Kid Routing Rules": _ensure_kid_routing_rules_table,
        "Internal Transfer Pricing": _ensure_internal_transfer_pricing_table,
        "Downstream Intake Rules": _ensure_downstream_intake_rules_table,
        "Transfer Elimination Rules": _ensure_transfer_elimination_rules_table,
        "Pricing": _ensure_pricing_table,
        "Production Drivers": _ensure_production_driver_table,
        "Scenario Controls": _ensure_scenario_controls_table,
        "Operating Costs": _ensure_operating_cost_table,
        "Capital & Financing": _ensure_capital_financing_table,
        "Loan Facilities": _ensure_loan_facilities_table,
        "Equity Facilities": _ensure_equity_facilities_table,
        "Valuation Inputs": _ensure_valuation_inputs_table,
        "Herd Plan": _ensure_herd_plan_table,
    }


def _build_app_assumption_bundle(assumptions: Optional[dict[str, pd.DataFrame]]) -> AssumptionBundle:
    synced_assumptions = _sync_transfer_tables_to_business_configuration(
        _sync_opening_herd_cohorts_in_assumptions(assumptions)
    )
    return build_assumption_bundle(synced_assumptions, _assumption_bundle_ensure_map())


def _setting_lookup(table: Optional[pd.DataFrame]) -> Dict[str, str]:
    settings = _ensure_biological_system_settings_table(table)
    return {
        str(row.get("Setting", "")).strip(): str(row.get("Value", "")).strip()
        for _, row in settings.iterrows()
        if str(row.get("Setting", "")).strip()
    }


def _rule_lookup(table: Optional[pd.DataFrame]) -> Dict[str, float]:
    rules = _ensure_cohort_allocation_rules_table(table)
    lookup: Dict[str, float] = {}
    for _, row in rules.iterrows():
        rule = str(row.get("Rule", "")).strip()
        value = pd.to_numeric(pd.Series([row.get("Value")]), errors="coerce").iloc[0]
        if rule and pd.notna(value):
            lookup[rule] = float(value)
    return lookup


def _apply_biological_scenario_shocks(
    assumptions: Dict[str, pd.DataFrame],
    adjustments: Dict[str, Any],
) -> Dict[str, pd.DataFrame]:
    shocked = {
        key: value.copy()
        for key, value in (assumptions or {}).items()
        if isinstance(value, pd.DataFrame)
    }
    bundle = _build_app_assumption_bundle(shocked)
    reproduction = bundle.biological.breeding_reproduction.copy()
    lactation = bundle.biological.lactation.copy()
    finishing = bundle.biological.finishing_slaughter.copy()
    allocation = bundle.biological.cohort_allocation_rules.copy()

    def _pct_adjust(table: pd.DataFrame, column: str, driver: str) -> None:
        if column not in table.columns:
            return
        shock = float(adjustments.get(driver, 0.0) or 0.0)
        if shock == 0.0:
            return
        table[column] = pd.to_numeric(table.get(column), errors="coerce") * (1.0 + shock / 100.0)

    def _abs_adjust(table: pd.DataFrame, column: str, driver: str) -> None:
        if column not in table.columns:
            return
        shock = float(adjustments.get(driver, 0.0) or 0.0)
        if shock == 0.0:
            return
        table[column] = pd.to_numeric(table.get(column), errors="coerce") + shock

    _pct_adjust(reproduction, "Conception Rate %", "Conception rate change (%)")
    _pct_adjust(reproduction, "Kid Mortality %", "Kid mortality change (%)")
    _pct_adjust(reproduction, "Doe Mortality %", "Doe mortality change (%)")
    _pct_adjust(reproduction, "Cull Rate %", "Cull rate change (%)")
    _pct_adjust(reproduction, "Kids per Kidding", "Kids per kidding change (%)")
    _pct_adjust(lactation, "Peak Yield Litres per Day", "Peak yield change (%)")
    _abs_adjust(lactation, "Lactation Length Days", "Lactation length change (days)")
    _abs_adjust(finishing, "Age at Slaughter (months)", "Age at slaughter change (months)")
    _abs_adjust(finishing, "Months to Market Weight", "Months to market weight change (months)")

    retention_shock = float(adjustments.get("Replacement retention change (%)", 0.0) or 0.0)
    if retention_shock != 0.0:
        mask = allocation["Rule"].astype(str).str.casefold() == "female replacement retention %".casefold()
        if mask.any():
            allocation.loc[mask, "Value"] = (
                pd.to_numeric(allocation.loc[mask, "Value"], errors="coerce") * (1.0 + retention_shock / 100.0)
            )

    shocked["Breeding & Reproduction Biology"] = _ensure_breeding_reproduction_biology_table(reproduction)
    shocked["Lactation Biology"] = _ensure_lactation_biology_table(lactation)
    shocked["Finishing & Slaughter Biology"] = _ensure_finishing_slaughter_biology_table(finishing)
    shocked["Cohort Allocation Rules"] = _ensure_cohort_allocation_rules_table(allocation)
    return shocked


def _active_first_row(table: pd.DataFrame, default_row: Optional[pd.Series] = None) -> pd.Series:
    if table is not None and not table.empty:
        if "Active" in table.columns:
            active = table.loc[table["Active"].fillna(True).astype(bool)]
            if not active.empty:
                return active.iloc[0]
        return table.iloc[0]
    if default_row is None:
        return pd.Series(dtype=object)
    return default_row


def _parity_key(value: Any) -> str:
    numeric = _coerce_numeric_scalar(value)
    if pd.notna(numeric):
        if numeric <= 1:
            return "1"
        if numeric <= 2:
            return "2"
    cleaned = str(value).strip()
    return cleaned if cleaned in {"1", "2", "3+"} else "3+"


def _build_lactation_lookup(table: Optional[pd.DataFrame]) -> Dict[str, dict[str, Any]]:
    lookup: Dict[str, dict[str, Any]] = {}
    for _, row in _ensure_lactation_biology_table(table).iterrows():
        if not _coerce_bool_value(row.get("Active"), True):
            continue
        lookup[_parity_key(row.get("Parity Group"))] = dict(row)
    if not lookup:
        defaults = _default_lactation_biology_table()
        for _, row in defaults.iterrows():
            lookup[_parity_key(row.get("Parity Group"))] = dict(row)
    return lookup


def _monthly_rate_from_annual_pct(percent: float, months_factor: float = 1.0) -> float:
    annual_rate = max(0.0, float(percent) / 100.0)
    return max(0.0, 1.0 - ((1.0 - annual_rate) ** max(months_factor / 12.0, 1e-9)))


def _age_bucket_index(age_months: float, max_age: int, age_band_width: float = 1.0) -> int:
    if not np.isfinite(age_months):
        return 0
    width = max(float(age_band_width), 1.0)
    max_bucket = int(np.ceil(max_age / width))
    return int(min(max(np.floor(float(age_months) / width), 0), max_bucket))


def _bucket_mid_age(bucket_index: int, age_band_width: float) -> float:
    width = max(float(age_band_width), 1.0)
    return (float(bucket_index) * width) + (width / 2.0)


def _advance_age_buckets(
    buckets: np.ndarray,
    *,
    months_factor: float,
    age_band_width: float,
    mature_age_months: float,
    weight_gain_per_month: float = 0.0,
    market_weight_kg: float = 0.0,
    months_to_market_weight: float = 0.0,
    use_weight_condition: bool = True,
) -> tuple[np.ndarray, float, float]:
    next_buckets = np.zeros_like(buckets, dtype=float)
    matured_heads = 0.0
    matured_weight_total = 0.0
    for bucket_index, head_count in enumerate(buckets.tolist()):
        if head_count <= 0:
            continue
        new_age = _bucket_mid_age(bucket_index, age_band_width) + max(months_factor, 0.0)
        estimated_weight = max(new_age * max(weight_gain_per_month, 0.0), 0.0)
        meets_weight_condition = (
            use_weight_condition
            and
            months_to_market_weight > 0
            and market_weight_kg > 0
            and weight_gain_per_month > 0
            and new_age >= months_to_market_weight
            and estimated_weight >= market_weight_kg
        )
        if new_age >= max(mature_age_months, 0.0) or meets_weight_condition:
            matured_heads += head_count
            matured_weight_total += head_count * estimated_weight
            continue
        next_index = min(
            _age_bucket_index(new_age, int(np.ceil(len(buckets) * age_band_width)), age_band_width),
            len(buckets) - 1,
        )
        next_buckets[next_index] += head_count
    return next_buckets, matured_heads, matured_weight_total


def _extract_buckets_pro_rata(
    buckets: np.ndarray,
    head_count: float,
) -> tuple[np.ndarray, np.ndarray, float]:
    source = np.array(buckets, dtype=float)
    requested = max(float(head_count), 0.0)
    total = float(source.sum())
    if total <= 0 or requested <= 0:
        return source, np.zeros_like(source, dtype=float), 0.0
    if requested >= total:
        return np.zeros_like(source, dtype=float), source.copy(), total
    ratio = requested / total
    extracted = source * ratio
    remaining = source - extracted
    return remaining, extracted, float(extracted.sum())


def _bucket_weight_total(
    buckets: np.ndarray,
    *,
    age_band_width: float,
    monthly_weight_gain: float,
    target_live_weight: float,
) -> float:
    total_weight = 0.0
    fallback_weight = max(float(target_live_weight), 0.0)
    for bucket_index, head_count in enumerate(np.array(buckets, dtype=float).tolist()):
        if head_count <= 0:
            continue
        estimated_weight = max(
            _bucket_mid_age(bucket_index, age_band_width) * max(float(monthly_weight_gain), 0.0),
            fallback_weight,
        )
        total_weight += float(head_count) * estimated_weight
    return total_weight


def _calculate_lactation_curve_yield(
    config: dict[str, Any],
    month_in_milk: float,
) -> float:
    peak_month = _coerce_float_scalar(config.get("Peak Lactation Month"), 2.0)
    base_yield = _coerce_float_scalar(config.get("Base Yield Litres per Day"), 0.0)
    peak_yield = _coerce_float_scalar(config.get("Peak Yield Litres per Day"), base_yield)
    parity_factor = _coerce_float_scalar(config.get("Parity Yield Factor"), 1.0)
    shape = _coerce_float_scalar(config.get("Curve Shape Parameter"), 1.0)
    if month_in_milk < 0:
        return 0.0
    if month_in_milk <= peak_month:
        ramp = month_in_milk / max(peak_month, 1.0)
        yield_per_day = base_yield + (peak_yield - base_yield) * min(max(ramp, 0.0), 1.0)
    else:
        decay_months = month_in_milk - peak_month
        yield_per_day = peak_yield * np.exp(-(decay_months / max(shape * 4.0, 1.0)))
        yield_per_day = max(base_yield * 0.35, yield_per_day)
    return max(0.0, yield_per_day * parity_factor)


def _derive_biological_schedules(
    schedule_df: pd.DataFrame,
    assumptions: Optional[Dict[str, pd.DataFrame]],
) -> Dict[str, pd.DataFrame]:
    if schedule_df is None or schedule_df.empty:
        return {}

    work_index = pd.to_datetime(schedule_df.index, errors="coerce")
    work_index = pd.DatetimeIndex(work_index[work_index.notna()]).sort_values()
    if work_index.empty:
        return {}

    assumption_map = assumptions or {}
    bundle = _build_app_assumption_bundle(assumption_map)
    settings_lookup = _setting_lookup(bundle.biological.system_settings)
    rules_lookup = _rule_lookup(bundle.biological.cohort_allocation_rules)
    reproduction = bundle.biological.breeding_reproduction
    lactation_lookup = _build_lactation_lookup(bundle.biological.lactation)
    finishing = bundle.biological.finishing_slaughter
    opening_cohorts = bundle.biological.opening_herd_cohorts
    herd_plan_df = bundle.get("Herd Plan", _default_herd_plan_table())

    reproduction_row = _active_first_row(reproduction, reproduction.iloc[0] if not reproduction.empty else None)
    finishing_row = _active_first_row(finishing, finishing.iloc[0] if not finishing.empty else None)
    model_grain = str(settings_lookup.get("Model Grain", "monthly")).strip().lower()
    grain_months = {"monthly": 1.0, "quarterly": 3.0, "annual": 12.0}.get(model_grain, 1.0)
    opening_start = pd.to_datetime(
        settings_lookup.get("Opening Biological Start Date", ""),
        errors="coerce",
    )
    if pd.isna(opening_start):
        opening_start = work_index.min()
    pre_roll_months = max(0.0, (work_index.min() - opening_start).days / 30.44)
    pre_roll_days = max(0.0, (work_index.min() - opening_start).days)
    age_band_width = max(_coerce_float_scalar(settings_lookup.get("Age Band Width (months)", 1.0), 1.0), 1.0)

    age_at_first_kidding = max(1, _coerce_int_scalar(reproduction_row.get("Age at First Kidding (months)"), 18))
    gestation_months = max(1, _coerce_int_scalar(reproduction_row.get("Gestation Months"), 5))
    months_to_market_weight = max(_coerce_float_scalar(finishing_row.get("Months to Market Weight"), 8.0), 1.0)
    slaughter_age = max(_coerce_float_scalar(finishing_row.get("Age at Slaughter (months)"), months_to_market_weight), 1.0)
    sale_age = max(1, int(round(max(months_to_market_weight, slaughter_age))))
    target_live_weight = max(_coerce_float_scalar(finishing_row.get("Target Live Weight Kg"), 0.0), 0.0)
    monthly_weight_gain = max(_coerce_float_scalar(finishing_row.get("Monthly Weight Gain Kg"), 0.0), 0.0)
    breeder_doe_cull_age_horizon = _coerce_numeric_scalar(reproduction_row.get("Breeder Doe Cull Age (months)"))
    breeder_buck_replacement_age_horizon = _coerce_numeric_scalar(reproduction_row.get("Breeder Buck Replacement Age (months)"))
    max_opening_age = pd.to_numeric(
        opening_cohorts.get("Age in Months"),
        errors="coerce",
    ).dropna().max()
    max_opening_age = float(max_opening_age) if pd.notna(max_opening_age) else 0.0
    max_age_horizon = max(
        age_at_first_kidding,
        sale_age,
        slaughter_age,
        months_to_market_weight,
        max_opening_age + pre_roll_months,
        float(breeder_doe_cull_age_horizon)
        if pd.notna(breeder_doe_cull_age_horizon) and float(breeder_doe_cull_age_horizon) > 0
        else 0.0,
        float(breeder_buck_replacement_age_horizon)
        if pd.notna(breeder_buck_replacement_age_horizon) and float(breeder_buck_replacement_age_horizon) > 0
        else 0.0,
    ) + (2.0 * age_band_width)
    bucket_count = max(3, int(np.ceil(max_age_horizon / age_band_width)) + 2)

    replacement_buckets = np.zeros(bucket_count, dtype=float)
    female_finishing_buckets = np.zeros(bucket_count, dtype=float)
    male_finishing_buckets = np.zeros(bucket_count, dtype=float)
    breeding_doe_buckets = {
        "1": np.zeros(bucket_count, dtype=float),
        "2": np.zeros(bucket_count, dtype=float),
        "3+": np.zeros(bucket_count, dtype=float),
    }
    breeding_buck_buckets = np.zeros(bucket_count, dtype=float)
    pregnancy_slots = max(1, int(np.ceil(gestation_months / max(grain_months, 1e-9))))
    pregnancy_queue: list[dict[str, float]] = [{"1": 0.0, "2": 0.0, "3+": 0.0} for _ in range(pregnancy_slots)]

    for _, row in opening_cohorts.loc[opening_cohorts["Active"].fillna(True).astype(bool)].iterrows():
        purpose = str(row.get("Purpose", "")).strip().lower()
        age_idx = _age_bucket_index(
            _coerce_float_scalar(row.get("Age in Months"), 0.0) + pre_roll_months,
            int(np.ceil(max_age_horizon)),
            age_band_width,
        )
        head_count = _coerce_float_scalar(row.get("Head Count"), 0.0)
        if purpose == "breeding_doe":
            breeding_doe_buckets[_parity_key(row.get("Parity"))][min(age_idx, bucket_count - 1)] += head_count
            if _coerce_bool_value(row.get("Pregnant"), False) and pregnancy_queue:
                seed_slot = min(
                    max(pregnancy_slots - 1 - int(np.floor(pre_roll_months / max(grain_months, 1e-9))), 0),
                    pregnancy_slots - 1,
                )
                pregnancy_queue[seed_slot][_parity_key(row.get("Parity"))] += head_count
        elif purpose == "breeding_buck":
            breeding_buck_buckets[min(age_idx, bucket_count - 1)] += head_count
        elif purpose == "replacement_doe":
            replacement_buckets[min(age_idx, bucket_count - 1)] += head_count
        elif purpose == "finishing_female":
            female_finishing_buckets[min(age_idx, bucket_count - 1)] += head_count
        elif purpose == "finishing_male":
            male_finishing_buckets[min(age_idx, bucket_count - 1)] += head_count

    if (
        breeding_buck_buckets.sum()
        + replacement_buckets.sum()
        + female_finishing_buckets.sum()
        + male_finishing_buckets.sum()
        + sum(bucket.sum() for bucket in breeding_doe_buckets.values())
        <= 0
    ):
        first_target = float(pd.to_numeric(herd_plan_df["Herd Size (heads)"], errors="coerce").dropna().iloc[0] or 320.0)
        breeding_doe_buckets["2"][
            min(_age_bucket_index(30.0, int(np.ceil(max_age_horizon)), age_band_width), bucket_count - 1)
        ] = first_target * 0.47
        breeding_buck_buckets[
            min(_age_bucket_index(36.0, int(np.ceil(max_age_horizon)), age_band_width), bucket_count - 1)
        ] = first_target * 0.025
        replacement_buckets[min(_age_bucket_index(age_at_first_kidding / 2.0, int(np.ceil(max_age_horizon)), age_band_width), bucket_count - 1)] = first_target * 0.125
        female_finishing_buckets[min(_age_bucket_index(sale_age / 2.0, int(np.ceil(max_age_horizon)), age_band_width), bucket_count - 1)] = first_target * 0.19
        male_finishing_buckets[min(_age_bucket_index(sale_age / 2.0, int(np.ceil(max_age_horizon)), age_band_width), bucket_count - 1)] = first_target * 0.19

    lactation_cohorts: list[dict[str, float]] = []
    if not opening_cohorts.empty:
        initial_lactating = opening_cohorts.loc[
            opening_cohorts["Purpose"].astype(str).str.strip().str.lower() == "breeding_doe"
        ]
        for _, row in initial_lactating.iterrows():
            days_in_milk = _coerce_float_scalar(row.get("Days in Milk"), 0.0) + pre_roll_days
            if days_in_milk <= 0:
                continue
            lactation_cohorts.append(
                {
                    "heads": _coerce_float_scalar(row.get("Head Count"), 0.0),
                    "parity": _parity_key(row.get("Parity")),
                    "days_in_milk": days_in_milk,
                }
            )

    period_days = _period_days_from_index(work_index)
    year_target_lookup = {
        int(row["Year"]): float(row["Herd Size (heads)"])
        for _, row in _ensure_herd_plan_table(herd_plan_df).iterrows()
        if pd.notna(row.get("Year")) and pd.notna(row.get("Herd Size (heads)"))
    }
    target_map = {
        idx.strftime("%Y-%m-%d"): float(year_target_lookup.get(int(idx.year), np.nan))
        for idx in work_index
        if pd.notna(idx)
    }
    baseline_target = next(
        (
            float(value)
            for value in pd.to_numeric(herd_plan_df.get("Herd Size (heads)"), errors="coerce").tolist()
            if pd.notna(value)
        ),
        np.nan,
    )

    use_herd_plan_target = _string_to_bool(
        settings_lookup.get("Use Herd Plan as Target", "True"),
        True,
    )
    hard_override = _string_to_bool(
        settings_lookup.get("Use Herd Plan as Hard Override", "False"),
        False,
    )

    female_share = _coerce_float_scalar(reproduction_row.get("Female Birth Share %"), 50.0) / 100.0
    conception_rate = _coerce_float_scalar(reproduction_row.get("Conception Rate %"), 85.0)
    kiddings_per_doe_year = _coerce_float_scalar(reproduction_row.get("Kiddings per Doe per Year"), 1.3)
    kids_per_kidding = _coerce_float_scalar(reproduction_row.get("Kids per Kidding"), 1.8)
    replacement_retention = _coerce_float_scalar(
        rules_lookup.get(
            "Female Replacement Retention %",
            reproduction_row.get("Replacement Retention %"),
        ),
        35.0,
    ) / 100.0
    live_sale_share = _coerce_float_scalar(finishing_row.get("Live Herd Sales Share %"), 0.0) / 100.0
    kid_mortality_pct = _coerce_float_scalar(reproduction_row.get("Kid Mortality %"), 0.0)
    doe_exit_pct = _coerce_float_scalar(reproduction_row.get("Doe Mortality %"), 0.0) + _coerce_float_scalar(
        reproduction_row.get("Cull Rate %"),
        0.0,
    )
    breeder_doe_cull_age = _coerce_numeric_scalar(reproduction_row.get("Breeder Doe Cull Age (months)"))
    breeder_doe_cull_parity = _coerce_numeric_scalar(reproduction_row.get("Breeder Doe Cull At Parity"))
    breeder_doe_live_sale_share = max(
        _coerce_float_scalar(reproduction_row.get("Breeder Doe Live Sale Share %"), 0.0),
        0.0,
    ) / 100.0
    buck_exit_pct = _coerce_float_scalar(reproduction_row.get("Buck Mortality %"), 0.0)
    breeder_buck_replacement_age = _coerce_numeric_scalar(reproduction_row.get("Breeder Buck Replacement Age (months)"))
    breeder_buck_live_sale_share = max(
        _coerce_float_scalar(reproduction_row.get("Breeder Buck Live Sale Share %"), 0.0),
        0.0,
    ) / 100.0
    finishing_mortality_pct = _coerce_float_scalar(finishing_row.get("Mortality %"), 0.0)
    meat_yield = _coerce_float_scalar(finishing_row.get("Meat Yield Kg per Goat"), 0.0)
    offal_yield = _coerce_float_scalar(finishing_row.get("Offal Yield Kg per Goat"), 0.0)
    pelt_units = _coerce_float_scalar(finishing_row.get("Pelt Units per Goat"), 0.0)

    summary_rows: list[dict[str, Any]] = []
    pregnancy_rows: list[dict[str, Any]] = []
    kidding_rows: list[dict[str, Any]] = []
    lactation_rows: list[dict[str, Any]] = []
    days_in_milk_rows: list[dict[str, Any]] = []
    finishing_rows: list[dict[str, Any]] = []
    slaughter_rows: list[dict[str, Any]] = []
    replacement_rows: list[dict[str, Any]] = []
    kid_cohort_rows: list[dict[str, Any]] = []
    breeding_rows: list[dict[str, Any]] = []

    for idx, period in enumerate(work_index):
        period_key = period.strftime("%Y-%m-%d")
        months_factor = grain_months
        if idx < len(period_days):
            actual_months = float(period_days.iloc[idx] / 30.44)
            if model_grain not in {"monthly", "quarterly", "annual"}:
                months_factor = max(actual_months, 1e-6)
        months_factor = max(months_factor, 1e-6)

        doe_exit_rate = _monthly_rate_from_annual_pct(doe_exit_pct, months_factor)
        buck_exit_rate = _monthly_rate_from_annual_pct(buck_exit_pct, months_factor)
        kid_exit_rate = _monthly_rate_from_annual_pct(kid_mortality_pct, months_factor)
        finishing_exit_rate = _monthly_rate_from_annual_pct(finishing_mortality_pct, months_factor)

        for parity in list(breeding_doe_buckets.keys()):
            breeding_doe_buckets[parity] *= max(0.0, 1.0 - doe_exit_rate)
        breeding_buck_buckets *= max(0.0, 1.0 - buck_exit_rate)
        replacement_buckets *= max(0.0, 1.0 - kid_exit_rate)
        female_finishing_buckets *= max(0.0, 1.0 - finishing_exit_rate)
        male_finishing_buckets *= max(0.0, 1.0 - finishing_exit_rate)

        breeder_doe_age_culls = 0.0
        breeder_doe_age_weight_total = 0.0
        for parity, parity_buckets in list(breeding_doe_buckets.items()):
            pre_age_total = float(np.array(parity_buckets, dtype=float).sum())
            aged_buckets, aged_culls, aged_weight_total = _advance_age_buckets(
                parity_buckets,
                months_factor=months_factor,
                age_band_width=age_band_width,
                mature_age_months=(
                    float(breeder_doe_cull_age)
                    if pd.notna(breeder_doe_cull_age) and float(breeder_doe_cull_age) > 0
                    else float("inf")
                ),
                weight_gain_per_month=monthly_weight_gain,
                market_weight_kg=target_live_weight,
                months_to_market_weight=months_to_market_weight,
                use_weight_condition=False,
            )
            breeding_doe_buckets[parity] = aged_buckets
            post_age_total = float(np.array(aged_buckets, dtype=float).sum())
            if pre_age_total > 0 and post_age_total < pre_age_total:
                retention_ratio = post_age_total / pre_age_total
                for payload in pregnancy_queue:
                    payload[parity] *= retention_ratio
            breeder_doe_age_culls += aged_culls
            breeder_doe_age_weight_total += aged_weight_total
        breeding_buck_buckets, breeder_buck_culls, breeder_buck_cull_weight_total = _advance_age_buckets(
            breeding_buck_buckets,
            months_factor=months_factor,
            age_band_width=age_band_width,
            mature_age_months=(
                float(breeder_buck_replacement_age)
                if pd.notna(breeder_buck_replacement_age) and float(breeder_buck_replacement_age) > 0
                else float("inf")
            ),
            weight_gain_per_month=monthly_weight_gain,
            market_weight_kg=target_live_weight,
            months_to_market_weight=months_to_market_weight,
            use_weight_condition=False,
        )

        replacement_buckets, matured_replacements, _ = _advance_age_buckets(
            replacement_buckets,
            months_factor=months_factor,
            age_band_width=age_band_width,
            mature_age_months=age_at_first_kidding,
        )
        female_finishing_buckets, eligible_female, female_sale_weight_total = _advance_age_buckets(
            female_finishing_buckets,
            months_factor=months_factor,
            age_band_width=age_band_width,
            mature_age_months=slaughter_age,
            weight_gain_per_month=monthly_weight_gain,
            market_weight_kg=target_live_weight,
            months_to_market_weight=months_to_market_weight,
        )
        male_finishing_buckets, eligible_male, male_sale_weight_total = _advance_age_buckets(
            male_finishing_buckets,
            months_factor=months_factor,
            age_band_width=age_band_width,
            mature_age_months=slaughter_age,
            weight_gain_per_month=monthly_weight_gain,
            market_weight_kg=target_live_weight,
            months_to_market_weight=months_to_market_weight,
        )
        breeding_doe_buckets["1"][
            min(_age_bucket_index(age_at_first_kidding, int(np.ceil(max_age_horizon)), age_band_width), bucket_count - 1)
        ] += matured_replacements

        births_payload = pregnancy_queue.pop(0) if pregnancy_queue else {"1": 0.0, "2": 0.0, "3+": 0.0}
        kidding_does = float(sum(births_payload.values()))
        kids_born = kidding_does * kids_per_kidding
        surviving_kids = kids_born * max(0.0, 1.0 - kid_exit_rate)
        female_kids = surviving_kids * female_share
        male_kids = surviving_kids - female_kids

        retained_female_kids = female_kids * max(0.0, replacement_retention)
        female_for_finishing = max(0.0, female_kids - retained_female_kids)
        replacement_buckets[_age_bucket_index(0.0, int(np.ceil(max_age_horizon)), age_band_width)] += retained_female_kids
        female_finishing_buckets[_age_bucket_index(0.0, int(np.ceil(max_age_horizon)), age_band_width)] += female_for_finishing
        male_finishing_buckets[_age_bucket_index(0.0, int(np.ceil(max_age_horizon)), age_band_width)] += male_kids

        if kidding_does > 0:
            breeding_doe_buckets["1"], promoted_from_one, _ = _extract_buckets_pro_rata(
                breeding_doe_buckets["1"],
                float(births_payload.get("1", 0.0)),
            )
            breeding_doe_buckets["2"] = breeding_doe_buckets["2"] + promoted_from_one

            breeding_doe_buckets["2"], promoted_from_two, _ = _extract_buckets_pro_rata(
                breeding_doe_buckets["2"],
                float(births_payload.get("2", 0.0)),
            )
            breeding_doe_buckets["3+"] = breeding_doe_buckets["3+"] + promoted_from_two

            for parity, heads in births_payload.items():
                if heads <= 0:
                    continue
                lactation_cohorts.append(
                    {
                        "heads": heads,
                        "parity": parity,
                        "days_in_milk": 0.0,
                    }
                )

        breeder_doe_parity_culls = 0.0
        breeder_doe_parity_cull_weight_total = 0.0
        if pd.notna(breeder_doe_cull_parity):
            parity_limit = float(breeder_doe_cull_parity)
            parity_cull_targets: list[str] = []
            if parity_limit <= 1:
                parity_cull_targets = ["2", "3+"]
            elif parity_limit <= 2:
                parity_cull_targets = ["3+"]
            for parity in parity_cull_targets:
                pre_parity_total = float(np.array(breeding_doe_buckets[parity], dtype=float).sum())
                culled_bucket = breeding_doe_buckets[parity].copy()
                breeder_doe_parity_culls += float(culled_bucket.sum())
                breeder_doe_parity_cull_weight_total += _bucket_weight_total(
                    culled_bucket,
                    age_band_width=age_band_width,
                    monthly_weight_gain=monthly_weight_gain,
                    target_live_weight=target_live_weight,
                )
                breeding_doe_buckets[parity] = np.zeros_like(culled_bucket, dtype=float)
                if pre_parity_total > 0:
                    for payload in pregnancy_queue:
                        payload[parity] = 0.0

        breeding_parity = {
            parity: float(buckets.sum())
            for parity, buckets in breeding_doe_buckets.items()
        }
        breeding_bucks = float(breeding_buck_buckets.sum())
        breeder_doe_culls = breeder_doe_age_culls + breeder_doe_parity_culls
        breeder_doe_cull_weight_total = breeder_doe_age_weight_total + breeder_doe_parity_cull_weight_total

        breeding_does = float(sum(breeding_parity.values()))
        monthly_kidding_rate = (kiddings_per_doe_year / 12.0) * (conception_rate / 100.0) * months_factor
        breeding_does_available = max(0.0, breeding_does - float(sum(sum(payload.values()) for payload in pregnancy_queue)))
        total_conceptions = breeding_does_available * max(0.0, monthly_kidding_rate)
        parity_mix = {
            parity: (breeding_parity[parity] / breeding_does) if breeding_does > 0 else 0.0
            for parity in breeding_parity
        }
        conception_payload = {
            parity: total_conceptions * share for parity, share in parity_mix.items()
        }
        pregnancy_queue.append(conception_payload)
        pregnant_does = float(sum(sum(payload.values()) for payload in pregnancy_queue))

        milk_output = 0.0
        lactating_does = 0.0
        next_lactation_cohorts: list[dict[str, float]] = []
        parity_days: Dict[str, dict[str, float]] = {}
        for cohort in lactation_cohorts:
            parity = _parity_key(cohort.get("parity"))
            config = lactation_lookup.get(parity, lactation_lookup.get("3+", {}))
            days_in_milk = float(cohort.get("days_in_milk", 0.0))
            lactation_days = float(
                pd.to_numeric(pd.Series([config.get("Lactation Length Days")]), errors="coerce").iloc[0] or 240.0
            )
            if days_in_milk > lactation_days:
                continue
            month_in_milk = max(days_in_milk / 30.44, 0.0)
            daily_yield = _calculate_lactation_curve_yield(config, month_in_milk)
            heads = float(cohort.get("heads", 0.0))
            milk_output += heads * daily_yield * float(period_days.iloc[idx])
            lactating_does += heads
            parity_days.setdefault(parity, {"heads": 0.0, "days": 0.0, "yield": 0.0})
            parity_days[parity]["heads"] += heads
            parity_days[parity]["days"] += heads * days_in_milk
            parity_days[parity]["yield"] += heads * daily_yield
            cohort["days_in_milk"] = days_in_milk + float(period_days.iloc[idx])
            next_lactation_cohorts.append(cohort)
        lactation_cohorts = next_lactation_cohorts

        finishing_saleable_goats = max(0.0, eligible_female + eligible_male)
        finishing_live_sales = finishing_saleable_goats * max(0.0, live_sale_share)
        finishing_slaughter_heads = max(0.0, finishing_saleable_goats - finishing_live_sales)
        breeder_doe_live_sales = breeder_doe_culls * breeder_doe_live_sale_share
        breeder_doe_slaughter_heads = max(0.0, breeder_doe_culls - breeder_doe_live_sales)
        breeder_buck_live_sales = breeder_buck_culls * breeder_buck_live_sale_share
        breeder_buck_slaughter_heads = max(0.0, breeder_buck_culls - breeder_buck_live_sales)
        breeder_live_sales = breeder_doe_live_sales + breeder_buck_live_sales
        breeder_slaughter_heads = breeder_doe_slaughter_heads + breeder_buck_slaughter_heads
        saleable_goats = max(0.0, finishing_saleable_goats + breeder_doe_culls + breeder_buck_culls)
        live_herd_sales = finishing_live_sales + breeder_live_sales
        slaughter_heads = max(0.0, finishing_slaughter_heads + breeder_slaughter_heads)
        breeder_slaughter_weight_total = (
            breeder_doe_cull_weight_total * (1.0 - breeder_doe_live_sale_share)
            + breeder_buck_cull_weight_total * (1.0 - breeder_buck_live_sale_share)
        )
        breeder_live_weight_total = (
            breeder_doe_cull_weight_total * breeder_doe_live_sale_share
            + breeder_buck_cull_weight_total * breeder_buck_live_sale_share
        )
        average_sale_weight = 0.0
        if saleable_goats > 0:
            average_sale_weight = (
                female_sale_weight_total
                + male_sale_weight_total
                + breeder_slaughter_weight_total
                + breeder_live_weight_total
            ) / saleable_goats

        herd_size = float(
            breeding_bucks
            + sum(breeding_parity.values())
            + replacement_buckets.sum()
            + female_finishing_buckets.sum()
            + male_finishing_buckets.sum()
        )
        target_herd = target_map.get(period_key, baseline_target if np.isfinite(baseline_target) else np.nan)
        herd_variance = herd_size - target_herd if use_herd_plan_target and pd.notna(target_herd) else np.nan

        if hard_override and pd.notna(target_herd) and herd_size > 0:
            scaling_factor = max(float(target_herd), 0.0) / herd_size
            breeding_buck_buckets *= scaling_factor
            for parity in breeding_doe_buckets:
                breeding_doe_buckets[parity] *= scaling_factor
            replacement_buckets *= scaling_factor
            female_finishing_buckets *= scaling_factor
            male_finishing_buckets *= scaling_factor
            for payload in pregnancy_queue:
                for parity in payload:
                    payload[parity] *= scaling_factor
            for cohort in lactation_cohorts:
                cohort["heads"] *= scaling_factor
            breeding_parity = {
                parity: float(buckets.sum())
                for parity, buckets in breeding_doe_buckets.items()
            }
            breeding_bucks = float(breeding_buck_buckets.sum())
            herd_size = float(target_herd)
            herd_variance = 0.0

        summary_rows.append(
            {
                "Period": period,
                "Target Herd Size (heads)": target_herd,
                "Herd Size (heads)": herd_size,
                "Herd Plan Variance (heads)": herd_variance,
                "Breeding Does": sum(breeding_parity.values()),
                "Breeding Bucks": breeding_bucks,
                "Replacement Does": replacement_buckets.sum(),
                "Finishing Females": female_finishing_buckets.sum(),
                "Finishing Males": male_finishing_buckets.sum(),
                "Finishing Goats": female_finishing_buckets.sum() + male_finishing_buckets.sum(),
                "Pregnant Does": pregnant_does,
                "Kidding Does": kidding_does,
                "Kids Born": kids_born,
                "Female Kids": female_kids,
                "Male Kids": male_kids,
                "Lactating Does": lactating_does,
                "Milk Production (L)": milk_output,
                "Breeder Doe Culls (heads)": breeder_doe_culls,
                "Breeder Buck Culls (heads)": breeder_buck_culls,
                "Breeder Live Sales (heads)": breeder_live_sales,
                "Breeder Slaughter Heads": breeder_slaughter_heads,
                "Breeder Meat Output Kg": breeder_slaughter_heads * meat_yield,
                "Breeder Offal Output Kg": breeder_slaughter_heads * offal_yield,
                "Breeder Pelt Output Units": breeder_slaughter_heads * pelt_units,
                "Saleable Goats (heads)": saleable_goats,
                "Live Herd Sales (heads)": live_herd_sales,
                "Slaughter Heads": slaughter_heads,
                "Average Sale Weight Kg": average_sale_weight,
                "Meat Output Kg": slaughter_heads * meat_yield,
                "Offal Output Kg": slaughter_heads * offal_yield,
                "Pelt Output Units": slaughter_heads * pelt_units,
            }
        )
        pregnancy_rows.append(
            {
                "Period": period,
                "Breeding Does": sum(breeding_parity.values()),
                "Conceptions": total_conceptions,
                "Pregnant Does": pregnant_does,
                "Kidding Does": kidding_does,
                "Kids Born": kids_born,
            }
        )
        kidding_rows.append(
            {
                "Period": period,
                "Kidding Does": kidding_does,
                "Kids Born": kids_born,
                "Female Kids": female_kids,
                "Male Kids": male_kids,
                "Surviving Kids": surviving_kids,
            }
        )
        lactation_rows.append(
            {
                "Period": period,
                "Lactating Does": lactating_does,
                "Milk Production (L)": milk_output,
                "Average Milk per Lactating Doe per Day": (
                    milk_output / max(lactating_does * float(period_days.iloc[idx]), 1e-9)
                    if lactating_does > 0
                    else 0.0
                ),
            }
        )
        for parity, values in parity_days.items():
            heads = values["heads"]
            days_in_milk_avg = values["days"] / heads if heads > 0 else 0.0
            yield_avg = values["yield"] / heads if heads > 0 else 0.0
            days_in_milk_rows.append(
                {
                    "Period": period,
                    "Parity Group": parity,
                    "Lactating Does": heads,
                    "Average Days in Milk": days_in_milk_avg,
                    "Average Yield per Doe per Day": yield_avg,
                }
            )
        finishing_rows.append(
            {
                "Period": period,
                "Replacement Does": replacement_buckets.sum(),
                "Finishing Females": female_finishing_buckets.sum(),
                "Finishing Males": male_finishing_buckets.sum(),
                "Finishing Saleable Goats (heads)": finishing_saleable_goats,
                "Breeder Doe Culls (heads)": breeder_doe_culls,
                "Breeder Buck Culls (heads)": breeder_buck_culls,
                "Saleable Goats (heads)": saleable_goats,
                "Average Sale Weight Kg": average_sale_weight,
            }
        )
        slaughter_rows.append(
            {
                "Period": period,
                "Saleable Goats (heads)": saleable_goats,
                "Finishing Live Herd Sales (heads)": finishing_live_sales,
                "Breeder Live Sales (heads)": breeder_live_sales,
                "Live Herd Sales (heads)": live_herd_sales,
                "Finishing Slaughter Heads": finishing_slaughter_heads,
                "Breeder Slaughter Heads": breeder_slaughter_heads,
                "Slaughter Heads": slaughter_heads,
                "Average Sale Weight Kg": average_sale_weight,
                "Meat Output Kg": slaughter_heads * meat_yield,
                "Offal Output Kg": slaughter_heads * offal_yield,
                "Pelt Output Units": slaughter_heads * pelt_units,
            }
        )
        replacement_rows.append(
            {
                "Period": period,
                "Female Kids": female_kids,
                "Replacement Retained": retained_female_kids,
                "Replacement Pool": replacement_buckets.sum(),
                "Matured Replacements": matured_replacements,
            }
        )
        breeding_rows.append(
            {
                "Period": period,
                "Parity 1": breeding_parity["1"],
                "Parity 2": breeding_parity["2"],
                "Parity 3+": breeding_parity["3+"],
                "Breeding Does": sum(breeding_parity.values()),
                "Breeding Bucks": breeding_bucks,
                "Pregnant Does": pregnant_does,
                "Breeder Doe Culls (heads)": breeder_doe_culls,
                "Breeder Buck Culls (heads)": breeder_buck_culls,
            }
        )
        for bucket_index in range(bucket_count):
            age_band_start = bucket_index * age_band_width
            age_band_end = age_band_start + age_band_width
            replacement_count = float(replacement_buckets[bucket_index])
            finishing_female_count = float(female_finishing_buckets[bucket_index])
            finishing_male_count = float(male_finishing_buckets[bucket_index])
            if replacement_count <= 0 and finishing_female_count <= 0 and finishing_male_count <= 0:
                continue
            kid_cohort_rows.append(
                {
                    "Period": period,
                    "Age Band Start (months)": age_band_start,
                    "Age Band End (months)": age_band_end,
                    "Replacement Does": replacement_count,
                    "Finishing Females": finishing_female_count,
                    "Finishing Males": finishing_male_count,
                }
            )

    summary = _period_indexed_frame(summary_rows, summary_rows[0].keys())
    pregnancy_schedule = _period_indexed_frame(pregnancy_rows, pregnancy_rows[0].keys())
    kidding_schedule = _period_indexed_frame(kidding_rows, kidding_rows[0].keys())
    lactation_schedule = _period_indexed_frame(lactation_rows, lactation_rows[0].keys())
    days_in_milk_schedule = _period_indexed_frame(
        days_in_milk_rows,
        ["Period", "Parity Group", "Lactating Does", "Average Days in Milk", "Average Yield per Doe per Day"],
    )
    finishing_schedule = _period_indexed_frame(finishing_rows, finishing_rows[0].keys())
    slaughter_schedule = _period_indexed_frame(slaughter_rows, slaughter_rows[0].keys())
    replacement_schedule = _period_indexed_frame(replacement_rows, replacement_rows[0].keys())
    kid_cohort_schedule = _period_indexed_frame(
        kid_cohort_rows,
        [
            "Period",
            "Age Band Start (months)",
            "Age Band End (months)",
            "Replacement Does",
            "Finishing Females",
            "Finishing Males",
        ],
    )
    breeding_schedule = _period_indexed_frame(breeding_rows, breeding_rows[0].keys())

    outputs = {
        "Biological Herd Summary": summary,
        "Pregnancy Schedule": pregnancy_schedule,
        "Kidding Schedule": kidding_schedule,
        "Kid Cohort Schedule": kid_cohort_schedule,
        "Lactation Schedule": lactation_schedule,
        "Days in Milk Schedule": days_in_milk_schedule,
        "Finishing Cohort Schedule": finishing_schedule,
        "Slaughter Output Schedule": slaughter_schedule,
        "Replacement Schedule": replacement_schedule,
        "Breeding Herd Schedule": breeding_schedule,
    }
    if _is_breeding_to_unit_mode(assumption_map):
        outputs.update(
            _derive_breeding_to_unit_outputs(
                schedule_df,
                assumption_map,
                outputs,
            )
        )
    return outputs


def _transfer_price_lookup(
    period: pd.Timestamp,
    destination: str,
    pricing_table: pd.DataFrame,
    transfer_class: str = "Weaned Kid",
) -> tuple[float, str]:
    pricing = _ensure_internal_transfer_pricing_table(pricing_table)
    if pricing.empty:
        return 0.0, DEFAULT_TRANSFER_PRICING_METHOD
    mask = pricing["Active"].fillna(True).astype(bool)
    mask &= pricing["Destination"].astype(str).str.strip().eq(destination)
    mask &= pricing["Transfer Class"].astype(str).str.strip().eq(transfer_class)
    period_key = period.strftime("%Y-%m-%d")
    matched = pricing.loc[
        mask
        & (
            pricing["Period"].isna()
            | pricing["Period"].astype(str).eq("")
            | pricing["Period"].astype(str).eq(period_key)
        )
    ].copy()
    if matched.empty:
        matched = pricing.loc[mask].copy()
    if matched.empty:
        return 0.0, DEFAULT_TRANSFER_PRICING_METHOD
    row = matched.iloc[-1]
    base_value = pd.to_numeric(pd.Series([row.get("Transfer Price per Head")]), errors="coerce").iloc[0]
    markup = pd.to_numeric(pd.Series([row.get("Markup %")]), errors="coerce").iloc[0]
    method = _normalize_transfer_pricing_method(row.get("Pricing Method"))
    base_price = float(base_value) if pd.notna(base_value) else 0.0
    markup_pct = float(markup) if pd.notna(markup) else 0.0
    if method == "Zero":
        return 0.0, method
    if method == "Cost Plus":
        return base_price * (1.0 + markup_pct / 100.0), method
    return base_price, method


def _external_price_lookup(
    pricing_table: pd.DataFrame,
    product: str,
    period: pd.Timestamp,
) -> float:
    pricing = _ensure_pricing_table(pricing_table)
    if pricing.empty:
        return 0.0
    period_key = period.strftime("%Y-%m-%d")
    mask = pricing["Product"].astype(str).str.strip().eq(product)
    mask &= pricing["Revenue Channel"].astype(str).str.strip().eq("External")
    period_match = pricing.loc[mask & pricing["Period"].astype(str).eq(period_key)]
    target = period_match if not period_match.empty else pricing.loc[mask]
    prices = pd.to_numeric(target.get("Base Price"), errors="coerce").dropna()
    return float(prices.iloc[-1]) if not prices.empty else 0.0


def _routing_rows_for_period(
    routing_table: pd.DataFrame,
    sex: str,
    period_key: str,
    destination: str,
    allow_external_sales: bool,
) -> pd.DataFrame:
    routing = _ensure_kid_routing_rules_table(routing_table)
    rows = routing.loc[
        routing["Active"].fillna(True).astype(bool)
        & routing["Sex"].astype(str).str.strip().str.casefold().eq(sex.casefold())
    ].copy()
    if rows.empty:
        default_destination = destination or ("External Sale" if allow_external_sales else "Meat")
        rows = pd.DataFrame(
            {
                "Period": [None],
                "Sex": [sex],
                "Transfer Class": ["Weaned Kid"],
                "Destination": [default_destination],
                "Allocation %": [100.0],
                "Priority": [1.0],
                "Min Head": [0.0],
                "Max Head": [np.nan],
                "Active": [True],
            }
        )
    period_rows = rows.loc[
        rows["Period"].isna()
        | rows["Period"].astype(str).eq("")
        | rows["Period"].astype(str).eq(period_key)
    ].copy()
    if period_rows.empty:
        period_rows = rows.copy()
    period_rows = period_rows.sort_values("Priority", kind="stable")
    return period_rows.reset_index(drop=True)


def _build_downstream_unit_schedule(
    summary: pd.DataFrame,
    intake_schedule: pd.DataFrame,
    assumptions: Dict[str, pd.DataFrame],
    destination: str,
) -> pd.DataFrame:
    if summary.empty or intake_schedule.empty:
        return pd.DataFrame()
    bundle = _build_app_assumption_bundle(assumptions)
    finishing = bundle.biological.finishing_slaughter
    lactation = bundle.biological.lactation
    production_drivers = _ensure_production_driver_table(bundle.commercial.production_drivers)
    pricing = _ensure_pricing_table(bundle.commercial.pricing)
    intake_rules = _ensure_downstream_intake_rules_table(bundle.commercial.downstream_intake_rules)
    finishing_row = _active_first_row(finishing, finishing.iloc[0] if not finishing.empty else None)
    lactation_lookup = _build_lactation_lookup(lactation)
    reproduction = bundle.biological.breeding_reproduction
    reproduction_row = _active_first_row(reproduction, reproduction.iloc[0] if not reproduction.empty else None)
    periods = pd.DatetimeIndex(pd.to_datetime(summary.index, errors="coerce")).sort_values()
    period_days = _period_days_from_index(periods)
    median_months = float(period_days.median() / 30.44) if not period_days.empty else 1.0
    grain_months = max(median_months, 1.0)

    months_to_market_weight = max(
        float(pd.to_numeric(pd.Series([finishing_row.get("Months to Market Weight")]), errors="coerce").iloc[0] or 7.0),
        1.0,
    )
    live_sale_share = max(
        float(pd.to_numeric(pd.Series([finishing_row.get("Live Herd Sales Share %")]), errors="coerce").iloc[0] or 0.0),
        0.0,
    ) / 100.0
    meat_yield = float(pd.to_numeric(pd.Series([finishing_row.get("Meat Yield Kg per Goat")]), errors="coerce").iloc[0] or 0.0)
    offal_yield = float(pd.to_numeric(pd.Series([finishing_row.get("Offal Yield Kg per Goat")]), errors="coerce").iloc[0] or 0.0)
    pelt_units = float(pd.to_numeric(pd.Series([finishing_row.get("Pelt Units per Goat")]), errors="coerce").iloc[0] or 0.0)
    age_at_first_kidding = max(
        float(pd.to_numeric(pd.Series([reproduction_row.get("Age at First Kidding (months)")]), errors="coerce").iloc[0] or 18.0),
        1.0,
    )
    gestation_months = max(
        float(pd.to_numeric(pd.Series([reproduction_row.get("Gestation Months")]), errors="coerce").iloc[0] or 5.0),
        1.0,
    )
    milk_driver = production_drivers.loc[production_drivers["Product"].astype(str).eq("Milk")]
    cheese_driver = production_drivers.loc[production_drivers["Product"].astype(str).eq("Cheese")]
    milk_per_day = float(pd.to_numeric(milk_driver.get("Litres per Lactating Doe per Day"), errors="coerce").dropna().iloc[0]) if not milk_driver.empty else 0.0
    cheese_alloc = float(pd.to_numeric(cheese_driver.get("Milk Allocation to Cheese %"), errors="coerce").dropna().iloc[0]) if not cheese_driver.empty else 0.0
    cheese_yield = float(pd.to_numeric(cheese_driver.get("Cheese Yield Kg per Litre"), errors="coerce").dropna().iloc[0]) if not cheese_driver.empty else 0.0
    lactation_days = float(pd.to_numeric(pd.Series([next(iter(lactation_lookup.values()), {}).get("Lactation Length Days")]), errors="coerce").iloc[0] or 260.0)

    meat_queue: list[float] = []
    milk_queue: list[float] = []
    rows: list[dict[str, Any]] = []
    for idx, period in enumerate(periods):
        period_key = period.strftime("%Y-%m-%d")
        period_intake = intake_schedule.loc[intake_schedule["Period"].astype(str).eq(period_key)].copy()
        female_intake = pd.to_numeric(
            period_intake.loc[period_intake["Sex"].astype(str).str.casefold().eq("female"), "Head Count"],
            errors="coerce",
        ).sum()
        male_intake = pd.to_numeric(
            period_intake.loc[period_intake["Sex"].astype(str).str.casefold().eq("male"), "Head Count"],
            errors="coerce",
        ).sum()
        finishing_heads = 0.0
        lactation_heads = 0.0
        if destination in {"Meat", "Combined"}:
            meat_rules = intake_rules.loc[
                intake_rules["Active"].fillna(True).astype(bool)
                & intake_rules["Destination"].astype(str).eq(destination)
                & intake_rules["Eligible for Finishing"].fillna(False).astype(bool)
            ]
            female_finish = female_intake if not meat_rules.empty else 0.0
            male_finish = male_intake if not meat_rules.empty else 0.0
            if destination == "Combined":
                female_finish = pd.to_numeric(
                    meat_rules.loc[meat_rules["Sex"].astype(str).str.casefold().eq("female"), "Active"], errors="coerce"
                ).count() and female_intake or 0.0
            entry_age = pd.to_numeric(meat_rules.get("Entry Age Months"), errors="coerce").dropna()
            lag_periods = max(int(np.ceil(max(months_to_market_weight - float(entry_age.iloc[0]) if not entry_age.empty else months_to_market_weight, 1.0) / grain_months)), 1)
            meat_queue.append(float(female_finish + male_finish))
            if len(meat_queue) > lag_periods:
                finishing_heads = float(meat_queue.pop(0))
        if destination in {"Milk-Cheese", "Combined"}:
            milk_rules = intake_rules.loc[
                intake_rules["Active"].fillna(True).astype(bool)
                & intake_rules["Destination"].astype(str).eq(destination)
                & intake_rules["Eligible for Lactation"].fillna(False).astype(bool)
                & intake_rules["Sex"].astype(str).str.casefold().eq("female")
            ]
            female_dairy = float(female_intake) if not milk_rules.empty else 0.0
            entry_age = pd.to_numeric(milk_rules.get("Entry Age Months"), errors="coerce").dropna()
            entry_months = float(entry_age.iloc[0]) if not entry_age.empty else 3.0
            lag_periods = max(int(np.ceil(max((age_at_first_kidding + gestation_months) - entry_months, 1.0) / grain_months)), 1)
            milk_queue.append(female_dairy)
            if len(milk_queue) > lag_periods:
                lactation_heads = float(milk_queue.pop(0))
        live_sales = finishing_heads * live_sale_share
        slaughter_heads = max(finishing_heads - live_sales, 0.0)
        milk_output = lactation_heads * milk_per_day * float(period_days.iloc[idx]) if idx < len(period_days) else 0.0
        cheese_output = milk_output * (cheese_alloc / 100.0) * cheese_yield
        direct_revenue = (
            live_sales * _external_price_lookup(pricing, "Live Herd", period)
            + (slaughter_heads * meat_yield) * _external_price_lookup(pricing, "Meat", period)
            + (slaughter_heads * offal_yield) * _external_price_lookup(pricing, "Offal", period)
            + (slaughter_heads * pelt_units) * _external_price_lookup(pricing, "Pelt", period)
            + milk_output * _external_price_lookup(pricing, "Milk", period)
            + cheese_output * _external_price_lookup(pricing, "Cheese", period)
        )
        rows.append(
            {
                "Period": period,
                "Business Unit": destination,
                "Female Transfer Intake": female_intake,
                "Male Transfer Intake": male_intake,
                "Lactating Does": lactation_heads,
                "Milk Production (L)": milk_output,
                "Cheese Output Kg": cheese_output,
                "Saleable Goats (heads)": finishing_heads,
                "Live Herd Sales (heads)": live_sales,
                "Slaughter Heads": slaughter_heads,
                "Meat Output Kg": slaughter_heads * meat_yield,
                "Offal Output Kg": slaughter_heads * offal_yield,
                "Pelt Output Units": slaughter_heads * pelt_units,
                "Revenue": direct_revenue,
            }
        )
    return pd.DataFrame(rows).set_index("Period")


def _derive_breeding_to_unit_outputs(
    schedule_df: pd.DataFrame,
    assumptions: Dict[str, pd.DataFrame],
    base_outputs: Dict[str, pd.DataFrame],
) -> Dict[str, pd.DataFrame]:
    destination = _selected_transfer_destination(assumptions)
    if not destination:
        return {}
    summary = base_outputs.get("Biological Herd Summary", pd.DataFrame())
    replacement_schedule = base_outputs.get("Replacement Schedule", pd.DataFrame())
    if summary.empty:
        return {}
    bundle = _build_app_assumption_bundle(assumptions)
    pricing = bundle.commercial.pricing
    routing_rules = bundle.commercial.kid_routing_rules
    transfer_pricing = bundle.commercial.internal_transfer_pricing
    allow_external_sales = _allow_external_kid_sales(assumptions)
    kid_rows: list[dict[str, Any]] = []
    routing_rows: list[dict[str, Any]] = []
    transfer_rows: list[dict[str, Any]] = []
    external_rows: list[dict[str, Any]] = []
    breeding_rows: list[dict[str, Any]] = []
    elimination_rows: list[dict[str, Any]] = []

    for period, row in summary.iterrows():
        timestamp = pd.Timestamp(period)
        period_key = timestamp.strftime("%Y-%m-%d")
        replacement_retained = 0.0
        if isinstance(replacement_schedule, pd.DataFrame) and not replacement_schedule.empty and period in replacement_schedule.index:
            replacement_retained = float(
                pd.to_numeric(
                    pd.Series([replacement_schedule.loc[period].get("Replacement Retained")]),
                    errors="coerce",
                ).iloc[0]
                or 0.0
            )
        female_kids = float(pd.to_numeric(pd.Series([row.get("Female Kids")]), errors="coerce").iloc[0] or 0.0)
        male_kids = float(pd.to_numeric(pd.Series([row.get("Male Kids")]), errors="coerce").iloc[0] or 0.0)
        female_available = max(female_kids - replacement_retained, 0.0)
        male_available = max(male_kids, 0.0)
        breeder_live_sales = float(pd.to_numeric(pd.Series([row.get("Breeder Live Sales (heads)")]), errors="coerce").iloc[0] or 0.0)
        breeder_slaughter_heads = float(pd.to_numeric(pd.Series([row.get("Breeder Slaughter Heads")]), errors="coerce").iloc[0] or 0.0)
        breeder_meat_output = float(pd.to_numeric(pd.Series([row.get("Breeder Meat Output Kg")]), errors="coerce").iloc[0] or 0.0)
        breeder_offal_output = float(pd.to_numeric(pd.Series([row.get("Breeder Offal Output Kg")]), errors="coerce").iloc[0] or 0.0)
        breeder_pelt_output = float(pd.to_numeric(pd.Series([row.get("Breeder Pelt Output Units")]), errors="coerce").iloc[0] or 0.0)
        kid_rows.append(
            {
                "Period": timestamp,
                "Female Kids Available": female_available,
                "Male Kids Available": male_available,
                "Replacement Retained": replacement_retained,
                "Internal Destination": destination,
            }
        )

        internal_revenue = 0.0
        kid_external_revenue = 0.0
        for sex, available in (("Female", female_available), ("Male", male_available)):
            remaining = float(available)
            period_rules = _routing_rows_for_period(
                routing_rules,
                sex,
                period_key,
                destination,
                allow_external_sales,
            )
            for _, rule in period_rules.iterrows():
                if remaining <= 0:
                    break
                alloc_pct = float(pd.to_numeric(pd.Series([rule.get("Allocation %")]), errors="coerce").iloc[0] or 0.0)
                min_head = float(pd.to_numeric(pd.Series([rule.get("Min Head")]), errors="coerce").iloc[0] or 0.0)
                max_head_raw = pd.to_numeric(pd.Series([rule.get("Max Head")]), errors="coerce").iloc[0]
                alloc_head = remaining * alloc_pct / 100.0
                alloc_head = max(alloc_head, min_head)
                if pd.notna(max_head_raw):
                    alloc_head = min(alloc_head, float(max_head_raw))
                alloc_head = min(alloc_head, remaining)
                alloc_head = max(alloc_head, 0.0)
                if alloc_head <= 0:
                    continue
                target = str(rule.get("Destination", destination)).strip() or destination
                if target == "External Sale" and not allow_external_sales:
                    target = destination
                routing_rows.append(
                    {
                        "Period": timestamp,
                        "Sex": sex,
                        "Transfer Class": "Weaned Kid",
                        "Destination": target,
                        "Head Count": alloc_head,
                        "Allocation %": alloc_pct,
                    }
                )
                if target == "External Sale":
                    product = f"{sex} Kids"
                    price = _external_price_lookup(pricing, product, timestamp)
                    revenue = alloc_head * price
                    kid_external_revenue += revenue
                    external_rows.append(
                        {
                            "Period": timestamp,
                            "Product": product,
                            "Head Count": alloc_head,
                            "Price per Head": price,
                            "Revenue": revenue,
                        }
                    )
                else:
                    price, method = _transfer_price_lookup(timestamp, target, transfer_pricing)
                    value = alloc_head * price
                    internal_revenue += value
                    transfer_rows.append(
                        {
                            "Period": timestamp,
                            "Destination": target,
                            "Sex": sex,
                            "Transfer Class": "Weaned Kid",
                            "Head Count": alloc_head,
                            "Pricing Method": method,
                            "Transfer Price per Head": price,
                            "Transfer Value": value,
                        }
                    )
                remaining -= alloc_head
            if remaining > 1e-9:
                fallback_target = "External Sale" if allow_external_sales else destination
                routing_rows.append(
                    {
                        "Period": timestamp,
                        "Sex": sex,
                        "Transfer Class": "Weaned Kid",
                        "Destination": fallback_target,
                        "Head Count": remaining,
                        "Allocation %": 100.0,
                    }
                )
        breeder_cull_revenue = (
            breeder_live_sales * _external_price_lookup(pricing, "Live Herd", timestamp)
            + breeder_meat_output * _external_price_lookup(pricing, "Meat", timestamp)
            + breeder_offal_output * _external_price_lookup(pricing, "Offal", timestamp)
            + breeder_pelt_output * _external_price_lookup(pricing, "Pelt", timestamp)
        )
        external_revenue = kid_external_revenue + breeder_cull_revenue
        breeding_rows.append(
            {
                "Period": timestamp,
                "Business Unit": "Breeding",
                "Breeding Does": row.get("Breeding Does", 0.0),
                "Breeding Bucks": row.get("Breeding Bucks", 0.0),
                "Replacement Does": row.get("Replacement Does", 0.0),
                "Female Kids Available": female_available,
                "Male Kids Available": male_available,
                "Breeder Live Sales (heads)": breeder_live_sales,
                "Breeder Slaughter Heads": breeder_slaughter_heads,
                "Breeder Meat Output Kg": breeder_meat_output,
                "Breeder Offal Output Kg": breeder_offal_output,
                "Breeder Pelt Output Units": breeder_pelt_output,
                "Breeder Cull Revenue": breeder_cull_revenue,
                "External Kid Sales Revenue": kid_external_revenue,
                "External Revenue": external_revenue,
                "Internal Transfer Revenue": internal_revenue,
                "Revenue": external_revenue + internal_revenue,
            }
        )
        elimination_rows.append(
            {
                "Period": timestamp,
                "Destination": destination,
                "Internal Transfer Revenue": internal_revenue,
                "Internal Transfer Cost": internal_revenue,
                "Revenue Elimination": -internal_revenue,
                "Cost Elimination": -internal_revenue,
            }
        )

    kid_availability = pd.DataFrame(kid_rows).set_index("Period")
    routing_schedule = pd.DataFrame(routing_rows)
    internal_transfer_schedule = pd.DataFrame(transfer_rows)
    external_sales_schedule = pd.DataFrame(external_rows)
    intake_schedule = internal_transfer_schedule.rename(columns={"Destination": "Business Unit"}).copy()
    if not intake_schedule.empty:
        intake_schedule["Intake Source"] = "Breeding"
    downstream_schedule = _build_downstream_unit_schedule(
        summary,
        intake_schedule if not intake_schedule.empty else pd.DataFrame(columns=["Period", "Head Count", "Sex"]),
        assumptions,
        destination,
    )
    breeding_unit_schedule = pd.DataFrame(breeding_rows).set_index("Period")
    unit_bridge = pd.DataFrame(index=pd.to_datetime(summary.index, errors="coerce"))
    unit_bridge.index.name = "Period"
    unit_bridge["Breeding External Revenue"] = (
        breeding_unit_schedule.reindex(unit_bridge.index)[
            "External Revenue" if "External Revenue" in breeding_unit_schedule.columns else "External Kid Sales Revenue"
        ].fillna(0.0)
        if not breeding_unit_schedule.empty
        else 0.0
    )
    unit_bridge["Breeding Internal Revenue"] = (
        breeding_unit_schedule.reindex(unit_bridge.index)["Internal Transfer Revenue"].fillna(0.0)
        if not breeding_unit_schedule.empty
        else 0.0
    )
    unit_bridge["Destination External Revenue"] = (
        downstream_schedule.reindex(unit_bridge.index)["Revenue"].fillna(0.0)
        if not downstream_schedule.empty
        else 0.0
    )
    unit_bridge["Elimination"] = -pd.to_numeric(unit_bridge["Breeding Internal Revenue"], errors="coerce").fillna(0.0)
    unit_bridge["Consolidated Revenue"] = (
        pd.to_numeric(unit_bridge["Breeding External Revenue"], errors="coerce").fillna(0.0)
        + pd.to_numeric(unit_bridge["Destination External Revenue"], errors="coerce").fillna(0.0)
    )
    outputs = {
        "Kid Availability Schedule": kid_availability,
        "Kid Routing Schedule": routing_schedule.set_index("Period") if not routing_schedule.empty else pd.DataFrame(),
        "Internal Transfer Schedule": internal_transfer_schedule.set_index("Period") if not internal_transfer_schedule.empty else pd.DataFrame(),
        "Breeding External Sales Schedule": external_sales_schedule.set_index("Period") if not external_sales_schedule.empty else pd.DataFrame(),
        "Downstream Intake Schedule": intake_schedule.set_index("Period") if not intake_schedule.empty else pd.DataFrame(),
        "Breeding Unit Schedule": breeding_unit_schedule,
        "Destination Unit Schedule": downstream_schedule,
        "Internal Transfer Elimination Schedule": pd.DataFrame(elimination_rows).set_index("Period"),
        "Unit Revenue Bridge": unit_bridge,
    }
    reporting_views = _build_reporting_unit_schedules(
        schedule_df,
        assumptions,
        outputs,
        pricing_table=pricing,
    )
    for entity, report_schedule in reporting_views.items():
        outputs[f"Reporting Schedule - {entity}"] = report_schedule
    return outputs


def _apply_biological_assumptions_to_schedule(
    schedule_df: pd.DataFrame,
    assumptions: Optional[Dict[str, pd.DataFrame]],
) -> pd.DataFrame:
    if schedule_df is None or schedule_df.empty:
        return schedule_df

    schedules = _derive_biological_schedules(schedule_df, assumptions)
    summary = schedules.get("Biological Herd Summary")
    if not isinstance(summary, pd.DataFrame) or summary.empty:
        return schedule_df

    work = schedule_df.copy()
    for column in summary.columns:
        work[column] = pd.to_numeric(summary.get(column), errors="coerce")

    herd_series = pd.to_numeric(work.get("Herd Size (heads)"), errors="coerce")
    baseline = herd_series.dropna().iloc[0] if herd_series.notna().any() else np.nan
    if pd.notna(baseline) and baseline > 0:
        work["Herd Multiplier"] = herd_series / float(baseline)
    else:
        work["Herd Multiplier"] = np.nan
    return work


def _default_pricing_table_from_core(
    core: pd.DataFrame,
    business_type: Any = DEFAULT_BUSINESS_TYPE,
) -> pd.DataFrame:
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str))).tolist()
    revenue = pd.to_numeric(core.get("Revenue"), errors="coerce").tolist()
    active_products = _active_products_for_business_type(business_type)
    template_rows = _get_template("pricing_rows", DEFAULT_PRICING_ROWS)
    row_lookup = {
        str(row.get("Product", "")).strip(): dict(row)
        for row in template_rows
        if str(row.get("Product", "")).strip()
    }
    rows = [
        row_lookup.get(product, row)
        for product in active_products
        for row in [_default_pricing_row_templates([product], business_type)[0]]
    ]
    for idx, row in enumerate(rows):
        product = str(row.get("Product", "")).strip()
        if product in row_lookup:
            rows[idx] = {**row, **row_lookup[product]}
        rows[idx]["Business Unit"] = _normalize_business_type(business_type)

    default_rows: list[dict[str, object]] = []
    for idx, period in enumerate(periods):
        period_revenue = revenue[idx] if idx < len(revenue) else np.nan
        for row in rows:
            product = str(row.get("Product", "")).strip() or "Product"
            unit = str(row.get("Unit", "")).strip() or "Unit"
            base_price = pd.to_numeric(
                pd.Series([row.get("Base Price")]), errors="coerce"
            ).iloc[0]
            is_active = bool(row.get("Default Active", False))
            quantity = (
                float(period_revenue) / float(base_price)
                if is_active and pd.notna(period_revenue) and pd.notna(base_price) and float(base_price) > 0
                else 0.0
            )
            default_rows.append(
                {
                    "Period": period,
                    "Product": product,
                    "Business Unit": str(row.get("Business Unit", business_type)).strip() or str(business_type),
                    "Revenue Channel": str(row.get("Revenue Channel", "External")).strip() or "External",
                    "Active": is_active,
                    "Allocation %": 100.0 if is_active else 0.0,
                    "Quantity Mode": "Derived",
                    "Manual Quantity Override": np.nan,
                    "Quantity per Period": quantity,
                    "Unit": unit,
                    "Base Price": base_price,
                    "Price Growth %": pd.to_numeric(
                        pd.Series([row.get("Price Growth %")]), errors="coerce"
                    ).iloc[0],
                }
            )

    if not default_rows:
        return pd.DataFrame(
            {
                "Period": ["2024-01-31"],
                "Product": ["Milk"],
                "Business Unit": [DEFAULT_BUSINESS_TYPE],
                "Revenue Channel": ["External"],
                "Active": [True],
                "Allocation %": [100.0],
                "Quantity Mode": ["Derived"],
                "Manual Quantity Override": [np.nan],
                "Quantity per Period": [0.0],
                "Unit": ["Litre"],
                "Base Price": [np.nan],
                "Price Growth %": [0.0],
                "Revenue": [0.0],
            }
        )

    return _ensure_pricing_table(pd.DataFrame(default_rows))


def _variable_default_items() -> list[tuple[str, Optional[float]]]:
    items: list[tuple[str, Optional[float]]] = []
    for row in _get_template("variable_items", DEFAULT_VARIABLE_ITEMS):
        item = str(row.get("Item", "")).strip() or "Variable Expense"
        share_value = pd.to_numeric(
            pd.Series([row.get("Share %")]), errors="coerce"
        ).iloc[0]
        share = float(share_value) / 100.0 if not pd.isna(share_value) else None
        items.append((item, share))

    if not items:
        items.append(("Variable Expense", None))

    return items


def _direct_wage_period_multiplier(source: pd.DataFrame) -> float:
    period_type = _infer_period_type_from_schedule(source)
    return 3.0 if period_type == "quarterly" else 1.0


def _coerce_direct_wage_item(
    row: Any, *, include_share: bool = True
) -> dict[str, Any]:
    position = str(row.get("Position") or row.get("Role") or "").strip() or "Direct Wage"
    headcount = pd.to_numeric(
        pd.Series([row.get("Head Count")]), errors="coerce"
    ).iloc[0]
    monthly_salary = pd.to_numeric(
        pd.Series([row.get("Monthly Salary per Head")]), errors="coerce"
    ).iloc[0]
    total_salary = pd.to_numeric(
        pd.Series([row.get("Total Salary") or row.get("Amount")]), errors="coerce"
    ).iloc[0]
    share_value = pd.to_numeric(pd.Series([row.get("Share %")]), errors="coerce").iloc[0]
    share = float(share_value) / 100.0 if not pd.isna(share_value) else None

    if pd.isna(headcount) and (
        not pd.isna(monthly_salary) or not pd.isna(total_salary) or share is not None
    ):
        headcount = 1.0

    if not pd.isna(headcount) and not pd.isna(monthly_salary):
        total_salary = float(headcount) * float(monthly_salary)
    elif not pd.isna(total_salary) and not pd.isna(headcount) and headcount > 0:
        monthly_salary = float(total_salary) / float(headcount)

    item = {
        "Position": position,
        "Head Count": None if pd.isna(headcount) else float(headcount),
        "Monthly Salary per Head": (
            None if pd.isna(monthly_salary) else float(monthly_salary)
        ),
        "Total Salary": None if pd.isna(total_salary) else float(total_salary),
    }
    if include_share:
        item["Share"] = share
    return item


def _normalize_direct_wage_template_records(
    rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for row in rows:
        normalized = _coerce_direct_wage_item(row, include_share=False)
        if any(value is not None for value in normalized.values()):
            records.append(normalized)
    return records


def _direct_wage_default_items() -> list[dict[str, Any]]:
    roles: list[dict[str, Any]] = []
    for row in _get_template("direct_wage_items", DEFAULT_DIRECT_WAGE_ITEMS):
        roles.append(_coerce_direct_wage_item(row))

    if not roles:
        roles.append(
            {
                "Position": "Direct Wage",
                "Head Count": 1.0,
                "Monthly Salary per Head": None,
                "Total Salary": None,
                "Share": None,
            }
        )

    return roles


def _coerce_admin_wage_item(
    row: Any, *, include_share: bool = True
) -> dict[str, Any]:
    position = (
        str(row.get("Position") or row.get("Function") or row.get("Role") or "").strip()
        or "Admin Wage"
    )
    headcount = pd.to_numeric(
        pd.Series([row.get("Head Count")]), errors="coerce"
    ).iloc[0]
    monthly_salary = pd.to_numeric(
        pd.Series([row.get("Monthly Salary per Head")]), errors="coerce"
    ).iloc[0]
    total_salary = pd.to_numeric(
        pd.Series([row.get("Total Salary") or row.get("Amount")]), errors="coerce"
    ).iloc[0]
    share_value = pd.to_numeric(
        pd.Series([row.get("Share %")]), errors="coerce"
    ).iloc[0]
    share = float(share_value) / 100.0 if not pd.isna(share_value) else None

    if pd.isna(headcount) and (
        not pd.isna(monthly_salary) or not pd.isna(total_salary) or share is not None
    ):
        headcount = 1.0

    if not pd.isna(headcount) and not pd.isna(monthly_salary):
        total_salary = float(headcount) * float(monthly_salary)
    elif not pd.isna(total_salary) and not pd.isna(headcount) and headcount > 0:
        monthly_salary = float(total_salary) / float(headcount)

    item = {
        "Position": position,
        "Head Count": None if pd.isna(headcount) else float(headcount),
        "Monthly Salary per Head": (
            None if pd.isna(monthly_salary) else float(monthly_salary)
        ),
        "Total Salary": None if pd.isna(total_salary) else float(total_salary),
    }
    if include_share:
        item["Share"] = share
    return item


def _normalize_admin_wage_template_records(
    rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for row in rows:
        normalized = _coerce_admin_wage_item(row, include_share=False)
        if any(value is not None for value in normalized.values()):
            records.append(normalized)
    return records


def _admin_wage_default_items() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in _get_template("admin_wage_items", DEFAULT_ADMIN_WAGE_ITEMS):
        items.append(_coerce_admin_wage_item(row))

    if not items:
        items.append(
            {
                "Position": "Admin Wage",
                "Head Count": 1.0,
                "Monthly Salary per Head": None,
                "Total Salary": None,
                "Share": None,
            }
        )

    return items


def _period_year_offsets(periods: Sequence[str]) -> list[int]:
    period_series = pd.Series(list(periods), dtype="object")
    period_dt = pd.to_datetime(period_series, errors="coerce")
    valid = period_dt.dropna()
    base_year = int(valid.iloc[0].year) if not valid.empty else pd.Timestamp.today().year
    offsets: list[int] = []
    for value in period_dt:
        if pd.isna(value):
            offsets.append(0)
        else:
            offsets.append(max(0, int(value.year) - base_year))
    return offsets


def _normalize_business_unit_name(value: Any, default: str = "General") -> str:
    unit = str(value or "").strip()
    return unit or default


def _sum_schedule_amounts_by_business_unit(
    table: pd.DataFrame,
    core: pd.DataFrame,
    ensure_fn: Callable[[Optional[pd.DataFrame], pd.DataFrame], pd.DataFrame],
    value_column: str,
) -> pd.DataFrame:
    work = ensure_fn(table, core)
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str)))
    if work.empty:
        return pd.DataFrame({"Period": periods})

    grouped = (
        work.groupby(["Period", "Business Unit"], as_index=False)[value_column].sum(min_count=1)
        if {"Period", "Business Unit", value_column}.issubset(work.columns)
        else pd.DataFrame(columns=["Period", "Business Unit", value_column])
    )
    if grouped.empty:
        return pd.DataFrame({"Period": periods})

    pivot = grouped.pivot(index="Period", columns="Business Unit", values=value_column)
    pivot = pivot.reindex(periods, fill_value=0.0)
    pivot.index.name = "Period"
    pivot = pivot.reset_index()
    ordered_cols = ["Period"] + [
        col for col in pivot.columns if col != "Period"
    ]
    return pivot[ordered_cols]


def _default_variable_expense_input_table() -> pd.DataFrame:
    base_revenue = float(_default_income_schedule(periods=1)["Revenue"].iloc[0])
    records: list[dict[str, object]] = []
    for item, share in _variable_default_items():
        amount = base_revenue * share if share is not None else np.nan
        records.append(
            {
                "Business Unit": "General",
                "Item": item,
                "Amount per Period": amount,
                "Yearly Increase %": 0.0,
            }
        )
    return pd.DataFrame(records)


def _ensure_variable_expense_input_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_variable_expense_input_table()
    else:
        work = table.copy()

    required_cols = ["Business Unit", "Item", "Amount per Period", "Yearly Increase %"]
    for column in required_cols:
        if column not in work.columns:
            work[column] = np.nan

    work["Business Unit"] = _series_or_default(work, "Business Unit", "General").astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = "General"
    work["Item"] = _series_or_default(work, "Item", "").astype(str).str.strip()
    work.loc[work["Item"] == "", "Item"] = "Variable Expense"
    work["Amount per Period"] = pd.to_numeric(
        work.get("Amount per Period"), errors="coerce"
    )
    work["Yearly Increase %"] = pd.to_numeric(
        work.get("Yearly Increase %"), errors="coerce"
    ).fillna(0.0)

    work = work.dropna(how="all")
    work = work[
        (work["Item"].notna())
        | (work["Amount per Period"].notna())
        | (work["Yearly Increase %"].notna())
    ]
    if work.empty:
        return _default_variable_expense_input_table()

    remainder = [col for col in work.columns if col not in required_cols]
    return work[required_cols + remainder].reset_index(drop=True)


def _default_direct_wage_input_table() -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for item in _direct_wage_default_items():
        records.append(
            {
                "Business Unit": "General",
                "Position": item.get("Position", "Direct Wage"),
                "Head Count": item.get("Head Count"),
                "Monthly Salary per Head": item.get("Monthly Salary per Head"),
                "Total Salary": item.get("Total Salary"),
                "Yearly Increase %": 0.0,
            }
        )
    return pd.DataFrame(records)


def _ensure_direct_wage_input_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_direct_wage_input_table()
    else:
        work = table.copy()

    required_cols = [
        "Business Unit",
        "Position",
        "Head Count",
        "Monthly Salary per Head",
        "Total Salary",
        "Yearly Increase %",
    ]
    for column in required_cols:
        if column not in work.columns:
            work[column] = np.nan

    work["Business Unit"] = _series_or_default(work, "Business Unit", "General").astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = "General"
    records: list[dict[str, object]] = []
    for _, row in work.iterrows():
        item = _coerce_direct_wage_item(row, include_share=False)
        item["Business Unit"] = str(row.get("Business Unit", "General")).strip() or "General"
        yearly_increase = pd.to_numeric(
            pd.Series([row.get("Yearly Increase %")]), errors="coerce"
        ).iloc[0]
        item["Yearly Increase %"] = 0.0 if pd.isna(yearly_increase) else float(yearly_increase)
        records.append(item)

    ensured = pd.DataFrame(records)
    if ensured.empty:
        return _default_direct_wage_input_table()

    remainder = [col for col in ensured.columns if col not in required_cols]
    return ensured[required_cols + remainder].reset_index(drop=True)


def _default_admin_wage_input_table() -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for item in _admin_wage_default_items():
        records.append(
            {
                "Business Unit": "General",
                "Position": item.get("Position", "Admin Wage"),
                "Head Count": item.get("Head Count"),
                "Monthly Salary per Head": item.get("Monthly Salary per Head"),
                "Total Salary": item.get("Total Salary"),
                "Yearly Increase %": 0.0,
            }
        )
    return pd.DataFrame(records)


def _ensure_admin_wage_input_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_admin_wage_input_table()
    else:
        work = table.copy()

    required_cols = [
        "Business Unit",
        "Position",
        "Head Count",
        "Monthly Salary per Head",
        "Total Salary",
        "Yearly Increase %",
    ]
    for column in required_cols:
        if column not in work.columns:
            work[column] = np.nan

    work["Business Unit"] = _series_or_default(work, "Business Unit", "General").astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = "General"
    records: list[dict[str, object]] = []
    for _, row in work.iterrows():
        item = _coerce_admin_wage_item(row, include_share=False)
        item["Business Unit"] = str(row.get("Business Unit", "General")).strip() or "General"
        yearly_increase = pd.to_numeric(
            pd.Series([row.get("Yearly Increase %")]), errors="coerce"
        ).iloc[0]
        item["Yearly Increase %"] = 0.0 if pd.isna(yearly_increase) else float(yearly_increase)
        records.append(item)

    ensured = pd.DataFrame(records)
    if ensured.empty:
        return _default_admin_wage_input_table()

    remainder = [col for col in ensured.columns if col not in required_cols]
    return ensured[required_cols + remainder].reset_index(drop=True)


def _apply_assumption_yearly_increase(
    table: pd.DataFrame, label_column: str, target_label: Optional[str], increment_pct: float
) -> pd.DataFrame:
    work = table.copy()
    if label_column not in work.columns or "Yearly Increase %" not in work.columns:
        return work
    target = (target_label or "").strip()
    if not target or target.startswith("All "):
        work["Yearly Increase %"] = float(increment_pct)
    else:
        labels = work[label_column].astype(str).str.strip()
        work.loc[labels == target, "Yearly Increase %"] = float(increment_pct)
    return work


def _propagate_variable_expense_inputs_to_schedule(
    input_table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    assumptions = _ensure_variable_expense_input_table(input_table)
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str))).tolist()
    period_multiplier = _direct_wage_period_multiplier(core)
    year_offsets = _period_year_offsets(periods)

    rows: list[dict[str, object]] = []
    for idx, period in enumerate(periods):
        year_offset = year_offsets[idx] if idx < len(year_offsets) else 0
        for _, row in assumptions.iterrows():
            amount = pd.to_numeric(
                pd.Series([row.get("Amount per Period")]), errors="coerce"
            ).iloc[0]
            increase = pd.to_numeric(
                pd.Series([row.get("Yearly Increase %")]), errors="coerce"
            ).iloc[0]
            if pd.notna(amount):
                factor = (1 + (float(increase) / 100.0)) ** year_offset if pd.notna(increase) else 1.0
                amount = float(amount) * period_multiplier * factor
            rows.append(
                {
                    "Period": period,
                    "Business Unit": row.get("Business Unit", "General"),
                    "Item": row.get("Item", "Variable Expense"),
                    "Amount": amount,
                }
            )
    return _ensure_variable_expense_table(pd.DataFrame(rows), core)


def _propagate_direct_wage_inputs_to_schedule(
    input_table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    assumptions = _ensure_direct_wage_input_table(input_table)
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str))).tolist()
    period_multiplier = _direct_wage_period_multiplier(core)
    year_offsets = _period_year_offsets(periods)

    rows: list[dict[str, object]] = []
    for idx, period in enumerate(periods):
        year_offset = year_offsets[idx] if idx < len(year_offsets) else 0
        for _, row in assumptions.iterrows():
            headcount = pd.to_numeric(
                pd.Series([row.get("Head Count")]), errors="coerce"
            ).iloc[0]
            monthly_salary = pd.to_numeric(
                pd.Series([row.get("Monthly Salary per Head")]), errors="coerce"
            ).iloc[0]
            increase = pd.to_numeric(
                pd.Series([row.get("Yearly Increase %")]), errors="coerce"
            ).iloc[0]
            if pd.notna(monthly_salary):
                factor = (1 + (float(increase) / 100.0)) ** year_offset if pd.notna(increase) else 1.0
                monthly_salary = float(monthly_salary) * factor
            total_salary = (
                float(headcount) * float(monthly_salary) * period_multiplier
                if pd.notna(headcount) and pd.notna(monthly_salary)
                else np.nan
            )
            rows.append(
                {
                    "Period": period,
                    "Business Unit": row.get("Business Unit", "General"),
                    "Position": row.get("Position", "Direct Wage"),
                    "Head Count": headcount,
                    "Monthly Salary per Head": monthly_salary,
                    "Total Salary": total_salary,
                }
            )
    return _ensure_direct_wage_table(pd.DataFrame(rows), core)


def _propagate_admin_wage_inputs_to_schedule(
    input_table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    assumptions = _ensure_admin_wage_input_table(input_table)
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str))).tolist()
    period_multiplier = _direct_wage_period_multiplier(core)
    year_offsets = _period_year_offsets(periods)

    rows: list[dict[str, object]] = []
    for idx, period in enumerate(periods):
        year_offset = year_offsets[idx] if idx < len(year_offsets) else 0
        for _, row in assumptions.iterrows():
            headcount = pd.to_numeric(
                pd.Series([row.get("Head Count")]), errors="coerce"
            ).iloc[0]
            monthly_salary = pd.to_numeric(
                pd.Series([row.get("Monthly Salary per Head")]), errors="coerce"
            ).iloc[0]
            increase = pd.to_numeric(
                pd.Series([row.get("Yearly Increase %")]), errors="coerce"
            ).iloc[0]
            if pd.notna(monthly_salary):
                factor = (1 + (float(increase) / 100.0)) ** year_offset if pd.notna(increase) else 1.0
                monthly_salary = float(monthly_salary) * factor
            total_salary = (
                float(headcount) * float(monthly_salary) * period_multiplier
                if pd.notna(headcount) and pd.notna(monthly_salary)
                else np.nan
            )
            rows.append(
                {
                    "Period": period,
                    "Business Unit": row.get("Business Unit", "General"),
                    "Position": row.get("Position", "Admin Wage"),
                    "Head Count": headcount,
                    "Monthly Salary per Head": monthly_salary,
                    "Total Salary": total_salary,
                }
            )
    return _ensure_admin_wage_table(pd.DataFrame(rows), core)


def _add_assumption_input_row(
    table: pd.DataFrame,
    row: dict[str, object],
    ensure_fn: Callable[[Optional[pd.DataFrame]], pd.DataFrame],
) -> pd.DataFrame:
    work = ensure_fn(table)
    return ensure_fn(pd.concat([work, pd.DataFrame([row])], ignore_index=True))


def _remove_assumption_input_row(
    table: pd.DataFrame,
    index: int,
    ensure_fn: Callable[[Optional[pd.DataFrame]], pd.DataFrame],
) -> pd.DataFrame:
    work = ensure_fn(table)
    if 0 <= index < len(work):
        work = work.drop(index=index).reset_index(drop=True)
    return ensure_fn(work)


def _assumption_input_row_labels(
    table: pd.DataFrame, label_column: str
) -> tuple[list[str], dict[str, int]]:
    labels: list[str] = []
    label_index: dict[str, int] = {}
    for idx_row, row in table.iterrows():
        label_value = str(row.get(label_column, "")).strip() or f"Row {idx_row + 1}"
        label = f"{label_value} ({idx_row + 1})"
        labels.append(label)
        label_index[label] = idx_row
    return labels, label_index


def _assumption_input_catalog_options(
    current_table: pd.DataFrame,
    default_table: pd.DataFrame,
    label_column: str,
) -> list[str]:
    options = {
        str(value).strip()
        for frame in [current_table, default_table]
        for value in frame.get(label_column, pd.Series(dtype=str)).dropna().tolist()
        if str(value).strip()
    }
    return sorted(options)


def _apply_assumption_input_schedule(
    assumption_key: str,
    schedule_name: str,
    ensure_fn: Callable[[Optional[pd.DataFrame]], pd.DataFrame],
    propagate_fn: Callable[[pd.DataFrame, pd.DataFrame], pd.DataFrame],
    editor_identifier: str,
) -> None:
    assumptions = st.session_state.get("assumptions", {})
    if not isinstance(assumptions, dict):
        return
    master_table = ensure_fn(assumptions.get(assumption_key))
    assumptions[assumption_key] = master_table
    st.session_state.assumptions = assumptions

    core_schedule = st.session_state.get("core_schedule")
    if not isinstance(core_schedule, pd.DataFrame):
        return

    propagated = propagate_fn(master_table, core_schedule)
    detail_schedules = st.session_state.get("detail_schedules", {})
    if not isinstance(detail_schedules, dict):
        detail_schedules = {}
    detail_schedules[schedule_name] = propagated
    st.session_state.detail_schedules = detail_schedules
    _clear_schedule_editor_state(editor_identifier)
    _reset_cached_results()


def _render_assumption_master_table(
    *,
    assumption_key: str,
    schedule_name: str,
    label_column: str,
    all_label: str,
    create_label: str,
    new_label_prompt: str,
    caption: str,
    propagation_note: str,
    ensure_fn: Callable[[Optional[pd.DataFrame]], pd.DataFrame],
    default_fn: Callable[[], pd.DataFrame],
    propagate_fn: Callable[[pd.DataFrame, pd.DataFrame], pd.DataFrame],
    add_row_factory: Callable[[str], dict[str, object]],
    column_config: dict[str, Any],
    editor_key: str,
    editor_identifier: str,
    add_choice_key: str,
    add_name_key: str,
    remove_choice_key: str,
    increment_target_key: str,
    increment_pct_key: str,
    disabled_columns: Optional[Sequence[str]] = None,
) -> pd.DataFrame:
    def _store_assumption_table(updated: pd.DataFrame) -> pd.DataFrame:
        ensured = ensure_fn(updated)
        if not _frames_equal(ensured, st.session_state.assumptions.get(assumption_key)):
            st.session_state.assumptions[assumption_key] = ensured
            _reset_cached_results()
        else:
            st.session_state.assumptions[assumption_key] = ensured
        return ensured

    table = ensure_fn(st.session_state.assumptions.get(assumption_key))
    st.session_state.assumptions[assumption_key] = table

    st.caption(caption)
    st.caption(propagation_note)

    add_options = [create_label] + _assumption_input_catalog_options(
        table, ensure_fn(default_fn()), label_column
    )
    st.session_state.setdefault(add_choice_key, create_label)
    st.session_state.setdefault(add_name_key, "")
    st.session_state.setdefault(remove_choice_key, "-- Select Row --")
    st.session_state.setdefault(increment_target_key, all_label)
    st.session_state.setdefault(increment_pct_key, 0.0)

    add_select_col, add_name_col, add_btn_col = st.columns([1.3, 1.7, 1])
    add_select_col.selectbox(
        "Select existing item",
        options=add_options,
        key=add_choice_key,
    )
    add_name_col.text_input(
        new_label_prompt,
        key=add_name_key,
    )
    if add_btn_col.button("Add Row", key=f"{editor_key}_add_row"):
        selected_label = str(st.session_state.get(add_choice_key, create_label)).strip()
        new_label = (
            str(st.session_state.get(add_name_key, "")).strip()
            if selected_label == create_label
            else selected_label
        )
        if new_label:
            table = _add_assumption_input_row(
                table,
                add_row_factory(new_label),
                ensure_fn,
            )
            table = _store_assumption_table(table)

    labels, label_index = _assumption_input_row_labels(table, label_column)
    remove_select_col, remove_btn_col = st.columns([3, 1])
    remove_select_col.selectbox(
        "Remove row",
        options=["-- Select Row --"] + labels,
        key=remove_choice_key,
    )
    if remove_btn_col.button("Remove", key=f"{editor_key}_remove_row"):
        choice = st.session_state.get(remove_choice_key)
        if choice in label_index:
            table = _remove_assumption_input_row(table, label_index[choice], ensure_fn)
            table = _store_assumption_table(table)
            st.session_state[remove_choice_key] = "-- Select Row --"

    increment_targets = [all_label] + sorted(
        {
            str(value).strip()
            for value in table.get(label_column, pd.Series(dtype=str)).dropna().tolist()
            if str(value).strip()
        }
    )
    inc_target_col, inc_pct_col, inc_btn_col = st.columns([2, 1, 1])
    inc_target_col.selectbox(
        "Apply yearly increase to",
        options=increment_targets,
        key=increment_target_key,
    )
    inc_pct_col.number_input(
        "Yearly increase (%)",
        min_value=-100.0,
        max_value=100.0,
        step=0.1,
        key=increment_pct_key,
    )
    if inc_btn_col.button("Set increase", key=f"{editor_key}_apply_increase"):
        table = _apply_assumption_yearly_increase(
            table,
            label_column,
            st.session_state.get(increment_target_key),
            float(st.session_state.get(increment_pct_key, 0.0)),
        )
        table = _store_assumption_table(table)

    editor = st.data_editor(
        st.session_state.assumptions[assumption_key],
        num_rows="dynamic",
        use_container_width=True,
        key=editor_key,
        column_config=column_config,
        disabled=list(disabled_columns or []),
    )
    ensured = _store_assumption_table(editor)

    if st.button("Apply to Schedule", key=f"{editor_key}_apply_schedule"):
        _apply_assumption_input_schedule(
            assumption_key,
            schedule_name,
            ensure_fn,
            propagate_fn,
            editor_identifier,
        )
        st.success(
            f"{assumption_key} propagated across the full production horizon."
        )

    return st.session_state.assumptions[assumption_key]


def _ensure_pricing_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        return _default_pricing_table()

    work = table.copy()

    if "Period" not in work.columns and "Year" in work.columns:
        year_values = pd.to_numeric(work.get("Year"), errors="coerce")
        work["Period"] = year_values.map(
            lambda value: pd.Timestamp(int(value), 12, 31).strftime("%Y-%m-%d")
            if pd.notna(value)
            else None
        )

    required_cols = [
        "Period",
        "Product",
        "Business Unit",
        "Revenue Channel",
        "Active",
        "Allocation %",
        "Quantity Mode",
        "Manual Quantity Override",
        "Quantity per Period",
        "Unit",
        "Base Price",
        "Price Growth %",
    ]
    for col in required_cols:
        if col not in work.columns:
            work[col] = np.nan

    work["Period"] = _normalize_period(work.get("Period", pd.Series(dtype=str)))
    work["Product"] = _series_or_default(work, "Product", "").astype(str).str.strip()
    work.loc[work["Product"] == "", "Product"] = "Product"
    work["Business Unit"] = _series_or_default(work, "Business Unit", DEFAULT_BUSINESS_TYPE).astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = DEFAULT_BUSINESS_TYPE
    work["Revenue Channel"] = _series_or_default(work, "Revenue Channel", "External").astype(str).str.strip()
    work.loc[~work["Revenue Channel"].isin(["External", "Internal"]), "Revenue Channel"] = "External"
    work["Active"] = _series_or_default(work, "Active", False).fillna(False).astype(bool)
    work["Allocation %"] = pd.to_numeric(work.get("Allocation %"), errors="coerce")
    work["Quantity Mode"] = _series_or_default(work, "Quantity Mode", "Derived").astype(str).str.strip()
    work.loc[~work["Quantity Mode"].isin(["Derived", "Manual Override"]), "Quantity Mode"] = "Derived"
    work["Manual Quantity Override"] = pd.to_numeric(
        work.get("Manual Quantity Override"), errors="coerce"
    )
    work["Quantity per Period"] = pd.to_numeric(
        work.get("Quantity per Period"), errors="coerce"
    )
    work["Unit"] = _series_or_default(work, "Unit", "").astype(str).str.strip()
    work["Base Price"] = pd.to_numeric(work.get("Base Price"), errors="coerce")
    work["Price Growth %"] = pd.to_numeric(
        work.get("Price Growth %"), errors="coerce"
    ).fillna(0.0)

    product_defaults = {
        str(row.get("Product", "")).strip(): row
        for row in _get_template("pricing_rows", DEFAULT_PRICING_ROWS)
    }
    for idx in work.index:
        product = str(work.at[idx, "Product"]).strip()
        default_row = product_defaults.get(product, {})
        if not str(work.at[idx, "Unit"]).strip():
            work.at[idx, "Unit"] = str(default_row.get("Unit", "Unit")).strip() or "Unit"
        if not str(work.at[idx, "Business Unit"]).strip():
            work.at[idx, "Business Unit"] = str(default_row.get("Business Unit", DEFAULT_BUSINESS_TYPE)).strip()
        if pd.isna(work.at[idx, "Base Price"]):
            default_price = pd.to_numeric(
                pd.Series([default_row.get("Base Price")]), errors="coerce"
            ).iloc[0]
            work.at[idx, "Base Price"] = default_price
        if pd.isna(work.at[idx, "Allocation %"]):
            work.at[idx, "Allocation %"] = 100.0 if bool(work.at[idx, "Active"]) else 0.0

    work = work.dropna(how="all")
    work = work.dropna(subset=["Period"], how="all")
    if work.empty:
        return _default_pricing_table()

    active_mask = work["Active"].fillna(False).astype(bool)
    work.loc[~active_mask, "Allocation %"] = work.loc[~active_mask, "Allocation %"].fillna(0.0)
    work["Revenue"] = np.where(
        active_mask,
        work["Quantity per Period"].fillna(0.0)
        * work["Base Price"].fillna(0.0)
        * (work["Allocation %"].fillna(0.0) / 100.0),
        0.0,
    )

    ordered = required_cols + ["Revenue"]
    remainder = [c for c in work.columns if c not in ordered and c != "Year"]
    work = work[ordered + remainder]
    return work.sort_values(["Period", "Product"], kind="stable").reset_index(drop=True)


def _add_pricing_row(table: pd.DataFrame) -> pd.DataFrame:
    work = _ensure_pricing_table(table)

    periods = _normalize_period(work.get("Period", pd.Series(dtype=str)))
    if not periods.empty:
        default_period = periods.iloc[-1]
    else:
        default_period = pd.Timestamp.today().strftime("%Y-%m-%d")

    new_row = {
        "Period": default_period,
        "Product": f"Product {len(work) + 1}",
        "Active": False,
        "Allocation %": 0.0,
        "Quantity Mode": "Manual Override",
        "Manual Quantity Override": np.nan,
        "Quantity per Period": 0.0,
        "Unit": "Unit",
        "Base Price": np.nan,
        "Price Growth %": 0.0,
    }

    return pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)


def _remove_pricing_row(table: pd.DataFrame, index: int) -> pd.DataFrame:
    if table is None or table.empty:
        return table

    work = table.copy()
    if 0 <= index < len(work):
        work = work.drop(index=index).reset_index(drop=True)
    return work


def _apply_pricing_yearly_increment(
    table: pd.DataFrame,
    column: str,
    increment_pct: float,
    target_product: Optional[str] = None,
) -> pd.DataFrame:
    if table is None or table.empty or increment_pct == 0:
        return table

    work = _ensure_pricing_table(table)
    if column not in work.columns:
        return work

    work["Period_dt"] = pd.to_datetime(work.get("Period"), errors="coerce")
    work["Product"] = _series_or_default(work, "Product", "").astype(str).str.strip()

    increment_factor = 1 + (increment_pct / 100.0)
    is_percent_column = column.endswith("%")

    for product, group in work.groupby("Product", dropna=False):
        product_key = product if isinstance(product, str) else ""
        if target_product and target_product != "All products" and product_key != target_product:
            continue

        group_sorted = group.sort_values("Period_dt", kind="stable")
        last_value = None
        last_year = None
        for idx, row in group_sorted.iterrows():
            current_value = pd.to_numeric(row.get(column), errors="coerce")
            period_dt = row.get("Period_dt")
            year = int(period_dt.year) if pd.notna(period_dt) else None
            if last_value is None:
                if not np.isnan(current_value):
                    last_value = current_value
                    last_year = year
                continue

            if np.isnan(last_value):
                continue

            years_elapsed = 1
            if year is not None and last_year is not None:
                years_elapsed = max(0, year - last_year)

            if is_percent_column:
                last_value = last_value + (increment_pct * years_elapsed)
            else:
                last_value = last_value * (increment_factor ** years_elapsed)

            work.at[idx, column] = last_value
            if year is not None:
                last_year = year

    work = work.drop(columns="Period_dt")
    return _ensure_pricing_table(work)


def _apply_pricing_product_plan(
    table: pd.DataFrame,
    product: str,
    *,
    active: bool,
    allocation_pct: float,
    quantity_mode: str,
    base_quantity: float,
    yearly_growth_pct: float,
    period_start: Optional[str] = None,
    period_end: Optional[str] = None,
) -> pd.DataFrame:
    work = _ensure_pricing_table(table)
    if work.empty:
        return work

    work["Period_dt"] = pd.to_datetime(work.get("Period"), errors="coerce")
    product_mask = work["Product"].astype(str).str.strip() == str(product).strip()
    if not product_mask.any():
        return work.drop(columns="Period_dt")

    start_dt = pd.to_datetime(period_start, errors="coerce")
    end_dt = pd.to_datetime(period_end, errors="coerce")
    if pd.notna(start_dt):
        product_mask &= work["Period_dt"] >= start_dt
    if pd.notna(end_dt):
        product_mask &= work["Period_dt"] <= end_dt
    if not product_mask.any():
        return work.drop(columns="Period_dt")

    product_rows = work.loc[product_mask].sort_values("Period_dt", kind="stable")
    base_year = None
    for idx, row in product_rows.iterrows():
        period_dt = row.get("Period_dt")
        if pd.isna(period_dt):
            continue
        year = int(period_dt.year)
        if base_year is None:
            base_year = year
        year_offset = max(0, year - base_year)
        quantity = float(base_quantity) * ((1 + yearly_growth_pct / 100.0) ** year_offset)
        work.at[idx, "Active"] = bool(active)
        work.at[idx, "Allocation %"] = float(allocation_pct) if active else 0.0
        work.at[idx, "Quantity Mode"] = (
            quantity_mode if quantity_mode in {"Derived", "Manual Override"} else "Derived"
        )
        if work.at[idx, "Quantity Mode"] == "Manual Override":
            work.at[idx, "Manual Quantity Override"] = quantity if active else 0.0
            work.at[idx, "Quantity per Period"] = quantity if active else 0.0
        else:
            work.at[idx, "Manual Quantity Override"] = np.nan
            work.at[idx, "Quantity per Period"] = 0.0

    work = work.drop(columns="Period_dt")
    return _ensure_pricing_table(work)


def _product_family_label(product: Any) -> str:
    product_key = str(product).strip()
    return PRODUCT_FAMILY_MAP.get(product_key, "Other")


def _pricing_family_summary(table: pd.DataFrame) -> pd.DataFrame:
    work = _ensure_pricing_table(table)
    if work.empty:
        return pd.DataFrame(columns=["Product Family", "Product", "Active Periods", "Total Quantity", "Total Revenue"])

    active = work.loc[work["Active"].fillna(False).astype(bool)].copy()
    if active.empty:
        return pd.DataFrame(columns=["Product Family", "Product", "Active Periods", "Total Quantity", "Total Revenue"])

    active["Product Family"] = active["Product"].map(_product_family_label)
    summary = (
        active.groupby(["Product Family", "Product"], as_index=False)
        .agg(
            **{
                "Active Periods": ("Period", "nunique"),
                "Total Quantity": ("Quantity per Period", "sum"),
                "Total Revenue": ("Revenue", "sum"),
            }
        )
        .sort_values(["Product Family", "Product"], kind="stable")
    )
    return summary.reset_index(drop=True)


def _pricing_quantity_by_period(table: pd.DataFrame) -> pd.DataFrame:
    work = _ensure_pricing_table(table)
    if work.empty:
        return pd.DataFrame(columns=["Period"])
    active = work.loc[work["Active"].fillna(False).astype(bool)].copy()
    if active.empty:
        return pd.DataFrame(columns=["Period"])
    summary = (
        active.pivot_table(
            index="Period",
            columns="Product",
            values="Quantity per Period",
            aggfunc="sum",
            fill_value=0.0,
        )
        .sort_index()
        .reset_index()
    )
    summary.columns.name = None
    return summary


def _pricing_validation_messages(
    pricing_table: pd.DataFrame,
    production_drivers: Optional[pd.DataFrame],
) -> list[str]:
    return build_pricing_validation_messages(
        pricing_table,
        production_drivers,
        ensure_pricing_table=_ensure_pricing_table,
        ensure_production_driver_table=_ensure_production_driver_table,
        product_family_label=_product_family_label,
        slaughter_products=_PRODUCTION_DRIVER_SLAUGHTER_PRODUCTS,
    )


def _sync_pricing_table_to_core(
    table: Optional[pd.DataFrame],
    core: pd.DataFrame,
    business_type: Any = DEFAULT_BUSINESS_TYPE,
) -> pd.DataFrame:
    active_products = _active_products_for_business_type(business_type)
    normalized_business_unit = _normalize_business_type(business_type)
    base = _default_pricing_table_from_core(core, business_type)
    if table is None or table.empty:
        return base

    current = _ensure_pricing_table(table)
    allowed_periods = set(base.get("Period", pd.Series(dtype=str)).astype(str).tolist())
    if allowed_periods:
        current = current.loc[
            current.get("Period", pd.Series(dtype=str)).astype(str).isin(allowed_periods)
        ].reset_index(drop=True)
    current = current.loc[
        current.get("Product", pd.Series(dtype=str)).astype(str).str.strip().isin(active_products)
    ].reset_index(drop=True)
    base_indexed = base.set_index(["Period", "Product"])
    current_indexed = current.set_index(["Period", "Product"])
    merged = base_indexed.combine_first(current_indexed)
    merged.update(current_indexed)
    if "Business Unit" in merged.columns:
        merged.loc[:, "Business Unit"] = normalized_business_unit
    return _ensure_pricing_table(merged.reset_index())


def _sync_commercial_assumptions_to_core(
    assumptions: Optional[Dict[str, pd.DataFrame]],
    core: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    """Align pricing periods and derived quantities to the active core schedule."""
    return sync_commercial_assumptions_to_core_service(
        assumptions,
        core,
        ensure_business_configuration_table=_ensure_business_configuration_table,
        selected_business_type=_selected_business_type,
        active_products_for_business_type=_active_products_for_business_type,
        build_app_assumption_bundle=_build_app_assumption_bundle,
        sync_production_driver_table_to_products=lambda table, products: _sync_production_driver_table_to_products(
            table,
            products,
            _selected_business_type(assumptions),
        ),
        sync_scenario_controls_to_products=_sync_scenario_controls_to_products,
        sync_pricing_table_to_core=_sync_pricing_table_to_core,
        derive_pricing_quantities_from_production=_derive_pricing_quantities_from_production,
        pricing_schedule_context=_pricing_schedule_context,
    )


def _pricing_revenue_by_period(table: pd.DataFrame) -> pd.DataFrame:
    work = _ensure_pricing_table(table)
    summary = work.groupby("Period", as_index=False)["Revenue"].sum(min_count=1)
    return summary.rename(columns={"Revenue": "Revenue from Pricing"})


def _production_driver_lookup(table: Optional[pd.DataFrame]) -> Dict[str, dict[str, Any]]:
    drivers = _ensure_production_driver_table(table)
    return {
        str(row.get("Product", "")).strip(): dict(row)
        for _, row in drivers.iterrows()
        if str(row.get("Product", "")).strip()
    }


def _period_days_from_index(index: pd.DatetimeIndex) -> pd.Series:
    if index.empty:
        return pd.Series(dtype=float)
    deltas = index.to_series().diff().dt.days.astype(float)
    valid_days = deltas.iloc[1:][np.isfinite(deltas.iloc[1:])]
    default_days = float(np.median(valid_days)) if not valid_days.empty else 30.44
    if not np.isfinite(default_days) or default_days <= 0:
        default_days = 30.44
    return deltas.fillna(default_days).clip(lower=1.0)


def _derive_pricing_quantities_from_production(
    pricing_table: pd.DataFrame,
    schedule_df: pd.DataFrame,
    production_drivers: Optional[pd.DataFrame],
) -> pd.DataFrame:
    pricing = _ensure_pricing_table(pricing_table)
    if pricing.empty or schedule_df.empty:
        return pricing

    driver_lookup = _production_driver_lookup(production_drivers)
    if not driver_lookup:
        return pricing

    work = pricing.copy()
    work["Period"] = _normalize_period(work.get("Period", pd.Series(dtype=str)))
    context = build_pricing_context(
        schedule_df,
        driver_lookup,
        tuple(_PRODUCTION_DRIVER_SLAUGHTER_PRODUCTS),
    )
    derived = derive_pricing_quantities(work, context)
    return _ensure_pricing_table(derived)


def _pricing_schedule_context(
    core_table: pd.DataFrame,
    herd_plan: Optional[pd.DataFrame],
    assumptions: Optional[Dict[str, pd.DataFrame]] = None,
) -> pd.DataFrame:
    core_clean = _clean_editor_table(core_table)
    if core_clean is None:
        return pd.DataFrame()
    prepared = _prepare_timeline_table(core_clean)
    assumption_map = dict(assumptions or {})
    if isinstance(herd_plan, pd.DataFrame) and not herd_plan.empty:
        assumption_map["Herd Plan"] = herd_plan
    context = _apply_biological_assumptions_to_schedule(prepared.copy(), assumption_map)
    if "Herd Size (heads)" not in context.columns and isinstance(herd_plan, pd.DataFrame) and not herd_plan.empty:
        context = _apply_herd_plan_to_schedule(context, herd_plan)
    return context


def _apply_pricing_assumptions_to_schedule(
    schedule_df: pd.DataFrame,
    pricing_table: Optional[pd.DataFrame],
    production_drivers: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    if schedule_df.empty or pricing_table is None or pricing_table.empty:
        return schedule_df

    work = schedule_df.copy()
    pricing = _derive_pricing_quantities_from_production(
        _ensure_pricing_table(pricing_table),
        work,
        production_drivers,
    )
    revenue_summary = _pricing_revenue_by_period(pricing)
    revenue_map = dict(
        zip(
            revenue_summary.get("Period", []),
            pd.to_numeric(revenue_summary.get("Revenue from Pricing"), errors="coerce"),
        )
    )

    period_keys = [idx.strftime("%Y-%m-%d") for idx in work.index]
    work["Revenue"] = pd.Series(period_keys, index=work.index).map(revenue_map).fillna(0.0)
    return _synchronize_financial_algorithms(work)


def _apply_commercial_shocks_to_pricing(
    pricing_table: pd.DataFrame,
    schedule_df: pd.DataFrame,
    production_drivers: Optional[pd.DataFrame],
    adjustments: Dict[str, Any],
) -> pd.DataFrame:
    pricing = _derive_pricing_quantities_from_production(
        _ensure_pricing_table(pricing_table),
        schedule_df,
        production_drivers,
    )
    if pricing.empty:
        return pricing

    work = pricing.copy()
    active_mask = work["Active"].fillna(False).astype(bool)
    products = sorted(
        {
            str(product).strip()
            for product in work.get("Product", pd.Series(dtype=str)).dropna().tolist()
            if str(product).strip()
        }
    )
    for product in products:
        product_mask = work["Product"].astype(str).str.strip() == product
        price_shock = float(adjustments.get(_price_change_driver(product), 0.0) or 0.0)
        qty_shock = float(adjustments.get(_quantity_change_driver(product), 0.0) or 0.0)
        if price_shock:
            work.loc[product_mask, "Base Price"] = (
                pd.to_numeric(work.loc[product_mask, "Base Price"], errors="coerce").fillna(0.0)
                * (1 + price_shock / 100.0)
            )
        if qty_shock:
            work.loc[product_mask & active_mask, "Quantity per Period"] = (
                pd.to_numeric(
                    work.loc[product_mask & active_mask, "Quantity per Period"], errors="coerce"
                ).fillna(0.0)
                * (1 + qty_shock / 100.0)
            )

    work["Revenue"] = np.where(
        active_mask,
        pd.to_numeric(work["Quantity per Period"], errors="coerce").fillna(0.0)
        * pd.to_numeric(work["Base Price"], errors="coerce").fillna(0.0)
        * (pd.to_numeric(work["Allocation %"], errors="coerce").fillna(0.0) / 100.0),
        0.0,
    )
    return _ensure_pricing_table(work)


def _default_operating_cost_table() -> pd.DataFrame:
    rows = _get_template("operating_rows", DEFAULT_OPERATING_COST_ROWS)
    table = _template_to_dataframe(
        rows,
        ["Year", "Business Unit", "Field", "Category", "unit_cost_per_head_per_month", "Inflation %"],
    )

    if table.empty:
        return pd.DataFrame(
            {
                "Year": [pd.Timestamp.today().year],
                "Business Unit": ["General"],
                "Field": ["variable_feed_cost_per_herd"],
                "Category": ["Operating Item"],
                "unit_cost_per_head_per_month": [np.nan],
                "Inflation %": [np.nan],
            }
        )

    return table.reset_index(drop=True)


def _ensure_operating_cost_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        return _default_operating_cost_table()

    work = table.copy()
    if "Field" not in work.columns:
        work["Field"] = np.nan
    if "Business Unit" not in work.columns:
        work["Business Unit"] = "General"
    work["Year"] = pd.to_numeric(work.get("Year"), errors="coerce")
    work["Business Unit"] = _series_or_default(work, "Business Unit", "General").astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = "General"
    work["Field"] = _series_or_default(work, "Field", "").astype(str).str.strip()
    work["Category"] = _series_or_default(work, "Category", "").astype(str).str.strip()
    work.loc[work["Category"] == "", "Category"] = np.nan
    for idx in work.index:
        category_value = work.at[idx, "Category"]
        field_value = work.at[idx, "Field"]
        category = "" if pd.isna(category_value) else str(category_value).strip()
        field = "" if pd.isna(field_value) else str(field_value).strip()
        if (not field or field.lower() == "nan") and category:
            for key, label in OPERATING_COST_FIELD_TO_CATEGORY.items():
                if label.casefold() == category.casefold():
                    field = key
                    break
            if not field or field.lower() == "nan":
                field = f"variable_{category.lower().replace(' ', '_')}_cost_per_herd"
            work.at[idx, "Field"] = field
        elif field and (not category or category.lower() == "nan"):
            work.at[idx, "Category"] = OPERATING_COST_FIELD_TO_CATEGORY.get(field, "Operating Item")
    work.loc[work["Field"] == "", "Field"] = "variable_feed_cost_per_herd"
    if "unit_cost_per_head_per_month" in work.columns:
        unit_cost = pd.to_numeric(work["unit_cost_per_head_per_month"], errors="coerce")
    else:
        unit_cost = pd.Series(np.nan, index=work.index, dtype=float)
    # Backward compatibility for historical tables still carrying "Monthly Cost".
    if "Monthly Cost" in work.columns:
        monthly_legacy = pd.to_numeric(work["Monthly Cost"], errors="coerce")
    else:
        monthly_legacy = pd.Series(np.nan, index=work.index, dtype=float)
    work["unit_cost_per_head_per_month"] = unit_cost.where(unit_cost.notna(), monthly_legacy)
    work["Inflation %"] = pd.to_numeric(work.get("Inflation %"), errors="coerce")

    work = work.dropna(how="all")

    if work.empty:
        return _default_operating_cost_table()

    if work["Category"].isna().any():
        for idx in work.index:
            if pd.isna(work.at[idx, "Category"]):
                work.at[idx, "Category"] = f"Operating Item {idx + 1}"

    if work["Year"].isna().all():
        work["Year"] = pd.Timestamp.today().year
    else:
        work["Year"] = work["Year"].ffill().fillna(work["Year"].dropna().min())
        work["Year"] = work["Year"].fillna(pd.Timestamp.today().year)
    work["Year"] = work["Year"].round().astype("Int64")

    required_cols = [
        "Year",
        "Business Unit",
        "Field",
        "Category",
        "unit_cost_per_head_per_month",
        "Inflation %",
    ]
    for col in required_cols:
        if col not in work.columns:
            work[col] = np.nan

    ordered = work[required_cols + [c for c in work.columns if c not in required_cols and c != "Monthly Cost"]]
    return ordered.sort_values(["Field", "Year"], kind="stable").reset_index(drop=True)


def _add_operating_cost_row(table: pd.DataFrame) -> pd.DataFrame:
    work = _ensure_operating_cost_table(table)
    years = pd.to_numeric(work.get("Year"), errors="coerce")
    if years.notna().any():
        default_year = int(years.dropna().max())
        default_year += 1
    else:
        default_year = pd.Timestamp.today().year
    new_row = {
        "Year": default_year,
        "Business Unit": "General",
        "Field": "variable_feed_cost_per_herd",
        "Category": "Feed",
        "unit_cost_per_head_per_month": np.nan,
        "Inflation %": np.nan,
    }
    return pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)


def _remove_operating_cost_row(table: pd.DataFrame, index: int) -> pd.DataFrame:
    if table is None or table.empty:
        return table
    work = table.copy()
    if 0 <= index < len(work):
        work = work.drop(index=index).reset_index(drop=True)
    return work


def _apply_operating_cost_increment(
    table: pd.DataFrame,
    increment_pct: float,
    target_category: Optional[str] = None,
    column: str = "unit_cost_per_head_per_month",
) -> pd.DataFrame:
    if table is None or table.empty or increment_pct == 0:
        return table

    work = _ensure_operating_cost_table(table)
    if column not in work.columns:
        return work

    work["Year"] = pd.to_numeric(work.get("Year"), errors="coerce")
    work["Category"] = _series_or_default(work, "Category", "").astype(str).str.strip()

    increment_factor = 1 + (increment_pct / 100.0)
    is_percent_column = column.endswith("%")

    for category, group in work.groupby("Category", dropna=False):
        category_key = category if isinstance(category, str) else ""
        if (
            target_category
            and target_category != "All categories"
            and category_key != target_category
        ):
            continue

        group_sorted = group.sort_values("Year", kind="stable")
        last_value = None
        for idx, row in group_sorted.iterrows():
            current_value = pd.to_numeric(row.get(column), errors="coerce")
            if last_value is None:
                if not np.isnan(current_value):
                    last_value = current_value
                continue

            if np.isnan(last_value):
                continue

            if is_percent_column:
                last_value = last_value + increment_pct
            else:
                last_value = last_value * increment_factor

            work.at[idx, column] = last_value

    work["unit_cost_per_head_per_month"] = pd.to_numeric(
        work["unit_cost_per_head_per_month"], errors="coerce"
    )

    return work.sort_values(["Field", "Year"], kind="stable").reset_index(drop=True)


def _apply_operating_cost_assumptions_to_schedule(
    schedule_df: pd.DataFrame,
    operating_table: Optional[pd.DataFrame],
    biological_cost_table: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    if schedule_df.empty:
        return schedule_df

    biological_costs = (
        _ensure_biological_cost_drivers_table(biological_cost_table)
        if isinstance(biological_cost_table, pd.DataFrame) and not biological_cost_table.empty
        else pd.DataFrame()
    )
    biological_active = (
        not biological_costs.empty
        and "Applies To" in biological_costs.columns
        and biological_costs["Active"].fillna(True).astype(bool).any()
    )

    if operating_table is None or operating_table.empty:
        assumptions = pd.DataFrame(columns=["Year", "Field", "Category", "unit_cost_per_head_per_month", "Inflation %"])
    else:
        assumptions = _ensure_operating_cost_table(operating_table)

    if assumptions.empty and not biological_active:
        return schedule_df

    work = schedule_df.copy()
    herd = pd.to_numeric(work.get("Herd Size (heads)"), errors="coerce")
    if herd.isna().all():
        return work
    herd = herd.ffill().bfill()
    period_days = work.index.to_series().diff().dt.days.astype(float)
    valid_days = period_days.iloc[1:][np.isfinite(period_days.iloc[1:])]
    default_days = float(np.median(valid_days)) if not valid_days.empty else 30.44
    if not np.isfinite(default_days) or default_days <= 0:
        default_days = 30.44
    months_factor = (period_days / 30.44).fillna(default_days / 30.44).clip(lower=1e-6)

    assumptions = assumptions.sort_values(["Field", "Year"], kind="stable")
    year_values = work.index.year

    field_period_costs: Dict[str, pd.Series] = {}
    if biological_active:
        active_rows = biological_costs.loc[biological_costs["Active"].fillna(True).astype(bool)].copy()
        for field in active_rows["Field"].dropna().astype(str).unique().tolist():
            field_rows = active_rows.loc[active_rows["Field"].astype(str) == field]
            field_cost = pd.Series(0.0, index=work.index, dtype=float)
            for _, row in field_rows.iterrows():
                applies_to = str(row.get("Applies To", "")).strip().lower()
                basis_column = BIOLOGICAL_COST_APPLIES_TO_MAP.get(applies_to)
                if not basis_column or basis_column not in work.columns:
                    continue
                basis_series = pd.to_numeric(work.get(basis_column), errors="coerce").fillna(0.0)
                row_year = pd.to_numeric(pd.Series([row.get("Year")]), errors="coerce").iloc[0]
                unit_cost = pd.to_numeric(
                    pd.Series([row.get("unit_cost_per_head_per_month")]), errors="coerce"
                ).iloc[0]
                inflation = pd.to_numeric(pd.Series([row.get("Inflation %")]), errors="coerce").iloc[0]
                if pd.isna(unit_cost):
                    continue
                base_year = int(row_year) if pd.notna(row_year) else int(year_values.min())
                inflation_rate = float(inflation) / 100.0 if pd.notna(inflation) else 0.0
                unit_series = pd.Series(
                    [
                        float(unit_cost) * ((1.0 + inflation_rate) ** max(int(year) - base_year, 0))
                        for year in year_values
                    ],
                    index=work.index,
                    dtype=float,
                )
                field_cost = field_cost.add(unit_series * basis_series * months_factor, fill_value=0.0)
            field_period_costs[field] = field_cost
    else:
        for field in assumptions["Field"].dropna().astype(str).unique().tolist():
            field_rows = assumptions[assumptions["Field"].astype(str) == field]
            unit_map = {
                int(row["Year"]): float(row["unit_cost_per_head_per_month"])
                for _, row in field_rows.iterrows()
                if pd.notna(row.get("Year")) and pd.notna(row.get("unit_cost_per_head_per_month"))
            }
            if not unit_map:
                continue
            unit_series = pd.Series(
                [unit_map.get(int(y), np.nan) for y in year_values],
                index=work.index,
                dtype=float,
            ).ffill()
            field_period_costs[field] = unit_series * herd * months_factor

    if "variable_feed_cost_per_herd" in field_period_costs:
        work["COGS"] = field_period_costs["variable_feed_cost_per_herd"]
    if "variable_healthcare_cost_per_herd" in field_period_costs:
        work["Variable Expenses"] = field_period_costs["variable_healthcare_cost_per_herd"]
    if "fixed_utility_cost_per_herd" in field_period_costs:
        utility = field_period_costs["fixed_utility_cost_per_herd"]
        if "Fixed Expenses" in work.columns:
            base_fixed = pd.to_numeric(work["Fixed Expenses"], errors="coerce").fillna(0.0)
            work["Fixed Expenses"] = base_fixed + utility
        else:
            work["Fixed Expenses"] = utility

    return _synchronize_financial_algorithms(work)


def _zero_series(index: pd.Index) -> pd.Series:
    return pd.Series(0.0, index=index, dtype=float)


def _normalized_period_series_from_index(index: pd.Index) -> pd.Series:
    period_index = pd.to_datetime(index, errors="coerce")
    return pd.Series(
        [
            value.strftime("%Y-%m-%d") if pd.notna(value) else ""
            for value in period_index
        ],
        index=index,
        dtype="object",
    )


def _series_from_period_map(
    period_map: dict[str, float],
    index: pd.Index,
) -> pd.Series:
    normalized_map: dict[str, float] = {}
    for raw_key, value in period_map.items():
        timestamp = pd.to_datetime(pd.Series([raw_key]), errors="coerce").iloc[0]
        normalized_key = (
            timestamp.strftime("%Y-%m-%d")
            if pd.notna(timestamp)
            else str(raw_key).strip()
        )
        if not normalized_key:
            continue
        normalized_map[normalized_key] = float(value)
    period_labels = _normalized_period_series_from_index(index)
    return pd.to_numeric(period_labels.map(normalized_map), errors="coerce").fillna(0.0)


def _reporting_source_series(
    schedule_df: pd.DataFrame,
    column: str,
) -> pd.Series:
    if column in schedule_df.columns:
        return pd.to_numeric(schedule_df[column], errors="coerce").fillna(0.0)
    return _zero_series(schedule_df.index)


def _sync_reporting_adjusted_columns(schedule_df: pd.DataFrame) -> pd.DataFrame:
    work = schedule_df.copy()
    adjusted_pairs = {
        "Revenue": "Revenue_adj",
        "COGS": "COGS_adj",
        "Gross Margin": "Gross Margin_adj",
        "EBITDA": "EBITDA_adj",
        "EBIT": "EBIT_adj",
        "NPAT": "NPAT_adj",
        "Interest Expense": "Interest Expense_adj",
        "CFO": "CFO_adj",
        "CFI": "CFI_adj",
        "CFF": "CFF_adj",
        "Net Cash Flow": "Net Cash Flow_adj",
        "Closing Cash Balance": "Closing Cash Balance_adj",
        "Cash and Cash Equivalents": "Cash and Cash Equivalents_adj",
        "Current Assets": "Current Assets_adj",
        "Current Liabilities": "Current Liabilities_adj",
        "Non-current Liabilities": "Non-current Liabilities_adj",
        "Equity": "Equity_adj",
        "Term Debt": "Term Debt_adj",
    }
    for base_col, adj_col in adjusted_pairs.items():
        if base_col in work.columns and adj_col in work.columns:
            work[adj_col] = pd.to_numeric(work[base_col], errors="coerce")
    return work


def _operating_cost_series_by_business_unit(
    schedule_df: pd.DataFrame,
    operating_table: Optional[pd.DataFrame],
    biological_cost_table: Optional[pd.DataFrame] = None,
) -> dict[str, dict[str, pd.Series]]:
    if schedule_df.empty:
        return {}

    work = schedule_df.copy()
    biological_costs = (
        _ensure_biological_cost_drivers_table(biological_cost_table)
        if isinstance(biological_cost_table, pd.DataFrame) and not biological_cost_table.empty
        else pd.DataFrame()
    )
    biological_active = (
        not biological_costs.empty
        and "Applies To" in biological_costs.columns
        and biological_costs["Active"].fillna(True).astype(bool).any()
    )
    assumptions = (
        _ensure_operating_cost_table(operating_table)
        if isinstance(operating_table, pd.DataFrame) and not operating_table.empty
        else pd.DataFrame()
    )
    herd_raw = work.get("Herd Size (heads)")
    herd = pd.to_numeric(herd_raw, errors="coerce")
    if not isinstance(herd, pd.Series):
        herd = pd.Series(herd, index=work.index, dtype=float)
    else:
        herd = herd.reindex(work.index)
    herd = herd.ffill().bfill() if herd.notna().any() else _zero_series(work.index)
    period_days = work.index.to_series().diff().dt.days.astype(float)
    valid_days = period_days.iloc[1:][np.isfinite(period_days.iloc[1:])]
    default_days = float(np.median(valid_days)) if not valid_days.empty else 30.44
    if not np.isfinite(default_days) or default_days <= 0:
        default_days = 30.44
    months_factor = (period_days / 30.44).fillna(default_days / 30.44).clip(lower=1e-6)
    year_values = work.index.year

    series_map: dict[str, dict[str, pd.Series]] = {}

    def _add_cost(unit: str, field: str, series: pd.Series) -> None:
        unit_key = _normalize_business_unit_name(unit, "General")
        unit_store = series_map.setdefault(unit_key, {})
        if field in unit_store:
            unit_store[field] = unit_store[field].add(series, fill_value=0.0)
        else:
            unit_store[field] = pd.to_numeric(series, errors="coerce").fillna(0.0)

    if biological_active:
        active_rows = biological_costs.loc[
            biological_costs["Active"].fillna(True).astype(bool)
        ].copy()
        for _, row in active_rows.iterrows():
            applies_to = str(row.get("Applies To", "")).strip().lower()
            basis_column = BIOLOGICAL_COST_APPLIES_TO_MAP.get(applies_to)
            if not basis_column or basis_column not in work.columns:
                continue
            basis_series = pd.to_numeric(work.get(basis_column), errors="coerce").fillna(0.0)
            field = str(row.get("Field", "")).strip()
            if not field:
                continue
            row_year = pd.to_numeric(pd.Series([row.get("Year")]), errors="coerce").iloc[0]
            unit_cost = pd.to_numeric(
                pd.Series([row.get("unit_cost_per_head_per_month")]), errors="coerce"
            ).iloc[0]
            inflation = pd.to_numeric(pd.Series([row.get("Inflation %")]), errors="coerce").iloc[0]
            if pd.isna(unit_cost):
                continue
            base_year = int(row_year) if pd.notna(row_year) else int(year_values.min())
            inflation_rate = float(inflation) / 100.0 if pd.notna(inflation) else 0.0
            unit_series = pd.Series(
                [
                    float(unit_cost) * ((1.0 + inflation_rate) ** max(int(year) - base_year, 0))
                    for year in year_values
                ],
                index=work.index,
                dtype=float,
            )
            _add_cost(
                row.get("Business Unit", "Breeding"),
                field,
                unit_series * basis_series * months_factor,
            )
        return series_map

    if assumptions.empty:
        return series_map

    assumptions = assumptions.sort_values(["Business Unit", "Field", "Year"], kind="stable")
    for (unit, field), group in assumptions.groupby(["Business Unit", "Field"], dropna=False):
        unit_map = {
            int(row["Year"]): float(row["unit_cost_per_head_per_month"])
            for _, row in group.iterrows()
            if pd.notna(row.get("Year")) and pd.notna(row.get("unit_cost_per_head_per_month"))
        }
        if not unit_map:
            continue
        unit_series = pd.Series(
            [unit_map.get(int(year), np.nan) for year in year_values],
            index=work.index,
            dtype=float,
        ).ffill()
        _add_cost(unit, str(field), unit_series * herd * months_factor)
    return series_map


def _pricing_revenue_by_business_unit(
    pricing_table: Optional[pd.DataFrame],
    schedule_index: pd.Index,
) -> dict[str, pd.Series]:
    pricing = (
        _ensure_pricing_table(pricing_table)
        if isinstance(pricing_table, pd.DataFrame) and not pricing_table.empty
        else pd.DataFrame()
    )
    if pricing.empty:
        return {}

    work = pricing.copy()
    work["Revenue"] = pd.to_numeric(work.get("Revenue"), errors="coerce").fillna(0.0)
    work["Active"] = work["Active"].fillna(False).astype(bool)
    work["Business Unit"] = work.get("Business Unit", pd.Series(dtype=str)).apply(
        _normalize_business_unit_name
    )
    work["Revenue Channel"] = (
        _series_or_default(work, "Revenue Channel", "External").astype(str).str.strip().str.casefold()
    )
    work = work.loc[work["Active"] & work["Revenue Channel"].ne("internal")].copy()
    if work.empty:
        return {}

    grouped = work.groupby(["Period", "Business Unit"], as_index=False)["Revenue"].sum(min_count=1)
    result: dict[str, pd.Series] = {}
    for unit, group in grouped.groupby("Business Unit", dropna=False):
        period_map = {
            str(period): float(value)
            for period, value in zip(group["Period"], group["Revenue"])
            if str(period).strip() and pd.notna(value)
        }
        result[_normalize_business_unit_name(unit)] = _series_from_period_map(period_map, schedule_index)
    return result


def _supplementary_unit_weight_frame(
    schedule_index: pd.Index,
    units: Sequence[str],
    revenue_series_map: dict[str, pd.Series],
    supplementary: Optional[dict[str, pd.DataFrame]] = None,
) -> pd.DataFrame:
    weight_frame = pd.DataFrame(index=schedule_index)
    for unit in units:
        weight_frame[unit] = pd.to_numeric(
            revenue_series_map.get(unit, _zero_series(schedule_index)),
            errors="coerce",
        ).fillna(0.0)

    row_totals = weight_frame.sum(axis=1)
    zero_mask = row_totals <= 0
    if zero_mask.any():
        breeding_schedule = (supplementary or {}).get("Breeding Unit Schedule")
        destination_schedule = (supplementary or {}).get("Destination Unit Schedule")
        fallback = pd.DataFrame(index=schedule_index)
        for unit in units:
            fallback[unit] = 0.0
        if isinstance(breeding_schedule, pd.DataFrame) and not breeding_schedule.empty and units:
            breeding_cols = [
                col
                for col in [
                    "Breeding Does",
                    "Breeding Bucks",
                    "Replacement Does",
                    "Female Kids Available",
                    "Male Kids Available",
                ]
                if col in breeding_schedule.columns
            ]
            if breeding_cols:
                fallback[units[0]] = pd.to_numeric(
                    breeding_schedule.reindex(schedule_index)[breeding_cols].sum(axis=1, min_count=1),
                    errors="coerce",
                ).fillna(0.0)
        if isinstance(destination_schedule, pd.DataFrame) and not destination_schedule.empty and len(units) > 1:
            destination_cols = [
                col
                for col in [
                    "Female Transfer Intake",
                    "Male Transfer Intake",
                    "Lactating Does",
                    "Saleable Goats (heads)",
                ]
                if col in destination_schedule.columns
            ]
            if destination_cols:
                fallback[units[1]] = pd.to_numeric(
                    destination_schedule.reindex(schedule_index)[destination_cols].sum(axis=1, min_count=1),
                    errors="coerce",
                ).fillna(0.0)
        fallback_totals = fallback.sum(axis=1)
        zero_both = fallback_totals <= 0
        for unit in units:
            weight_frame.loc[zero_mask, unit] = fallback.loc[zero_mask, unit]
        row_totals = weight_frame.sum(axis=1)
        if zero_both.any():
            for unit in units:
                weight_frame.loc[zero_both, unit] = 1.0 / max(len(units), 1)
            row_totals = weight_frame.sum(axis=1)

    row_totals = row_totals.replace(0.0, np.nan)
    for unit in units:
        weight_frame[unit] = (weight_frame[unit] / row_totals).fillna(1.0 / max(len(units), 1))
    return weight_frame


def _detail_table_for_reporting(
    detail_tables: Optional[dict[str, pd.DataFrame]],
    assumptions: dict[str, pd.DataFrame],
    schedule_name: str,
    schedule_index: pd.Index,
) -> pd.DataFrame:
    detail_map = detail_tables or {}
    if isinstance(detail_map.get(schedule_name), pd.DataFrame) and not detail_map[schedule_name].empty:
        return detail_map[schedule_name].copy()

    core_stub = pd.DataFrame({"Period": _normalized_period_series_from_index(schedule_index).tolist()})
    if schedule_name == "Variable Expenses Schedule":
        return _propagate_variable_expense_inputs_to_schedule(
            assumptions.get("Variable Expenses", pd.DataFrame()),
            core_stub,
        )
    if schedule_name == "Direct Wages Schedule":
        return _propagate_direct_wage_inputs_to_schedule(
            assumptions.get("Direct Wages", pd.DataFrame()),
            core_stub,
        )
    if schedule_name == "Admin Wages Schedule":
        return _propagate_admin_wage_inputs_to_schedule(
            assumptions.get("Admin Wages", pd.DataFrame()),
            core_stub,
        )
    return pd.DataFrame()


def _unit_series_lookup(
    aggregated_table: pd.DataFrame,
    unit: str,
    schedule_index: pd.Index,
) -> pd.Series:
    if aggregated_table.empty or unit not in aggregated_table.columns:
        return _zero_series(schedule_index)
    period_map = {
        str(period): float(value)
        for period, value in zip(aggregated_table["Period"], aggregated_table[unit])
        if str(period).strip() and pd.notna(value)
    }
    return _series_from_period_map(period_map, schedule_index)


def _allocate_reporting_column(
    total_series: pd.Series,
    units: Sequence[str],
    weights: pd.DataFrame,
    explicit_map: Optional[dict[str, pd.Series]] = None,
    extra_map: Optional[dict[str, pd.Series]] = None,
) -> dict[str, pd.Series]:
    explicit = explicit_map or {}
    extra = extra_map or {}
    explicit_total = _zero_series(total_series.index)
    for unit in units:
        explicit_total = explicit_total.add(
            pd.to_numeric(explicit.get(unit, _zero_series(total_series.index)), errors="coerce").fillna(0.0),
            fill_value=0.0,
        )
    residual = pd.to_numeric(total_series, errors="coerce").fillna(0.0) - explicit_total
    allocated: dict[str, pd.Series] = {}
    for unit in units:
        base_series = pd.to_numeric(explicit.get(unit, _zero_series(total_series.index)), errors="coerce").fillna(0.0)
        extra_series = pd.to_numeric(extra.get(unit, _zero_series(total_series.index)), errors="coerce").fillna(0.0)
        weight_series = pd.to_numeric(weights.get(unit), errors="coerce").fillna(0.0)
        allocated[unit] = base_series + (residual * weight_series) + extra_series
    return allocated


def _build_reporting_unit_schedules(
    schedule_df: pd.DataFrame,
    assumptions: Optional[dict[str, pd.DataFrame]],
    supplementary: Optional[dict[str, pd.DataFrame]] = None,
    detail_tables: Optional[dict[str, pd.DataFrame]] = None,
    pricing_table: Optional[pd.DataFrame] = None,
) -> dict[str, pd.DataFrame]:
    if schedule_df is None or schedule_df.empty:
        return {}

    assumption_map = dict(assumptions or {})
    working = schedule_df.copy()
    working.index = pd.to_datetime(working.index, errors="coerce")
    working = working.loc[working.index.notna()].sort_index()
    if working.empty:
        return {}

    if not _is_breeding_to_unit_mode(assumption_map):
        return {"Consolidated": _sync_reporting_adjusted_columns(_synchronize_financial_algorithms(working))}

    destination = _selected_transfer_destination(assumption_map)
    if not destination:
        return {"Consolidated": _sync_reporting_adjusted_columns(_synchronize_financial_algorithms(working))}

    units = ["Breeding", destination]
    pricing_source = pricing_table if isinstance(pricing_table, pd.DataFrame) and not pricing_table.empty else assumption_map.get("Pricing")
    external_revenue = _pricing_revenue_by_business_unit(pricing_source, working.index)
    weights = _supplementary_unit_weight_frame(working.index, units, external_revenue, supplementary)
    numeric_working = working.apply(pd.to_numeric, errors="coerce")

    unit_frames = {
        unit: _sync_reporting_adjusted_columns(
            numeric_working.copy().mul(weights[unit], axis=0)
        )
        for unit in units
    }

    operating_costs = _operating_cost_series_by_business_unit(
        working,
        assumption_map.get("Operating Costs"),
        assumption_map.get("Biological Cost Drivers"),
    )
    variable_detail = _aggregate_variable_expenses_by_business_unit(
        _detail_table_for_reporting(detail_tables, assumption_map, "Variable Expenses Schedule", working.index),
        pd.DataFrame({"Period": _normalized_period_series_from_index(working.index).tolist()}),
    )
    direct_detail = _aggregate_direct_wages_by_business_unit(
        _detail_table_for_reporting(detail_tables, assumption_map, "Direct Wages Schedule", working.index),
        pd.DataFrame({"Period": _normalized_period_series_from_index(working.index).tolist()}),
    )
    admin_detail = _aggregate_admin_wages_by_business_unit(
        _detail_table_for_reporting(detail_tables, assumption_map, "Admin Wages Schedule", working.index),
        pd.DataFrame({"Period": _normalized_period_series_from_index(working.index).tolist()}),
    )

    transfer_schedule = (supplementary or {}).get("Internal Transfer Schedule", pd.DataFrame())
    internal_revenue_map = {unit: _zero_series(working.index) for unit in units}
    internal_cost_map = {unit: _zero_series(working.index) for unit in units}
    if isinstance(transfer_schedule, pd.DataFrame) and not transfer_schedule.empty:
        transfer_work = transfer_schedule.copy()
        if "Period" not in transfer_work.columns:
            transfer_work = transfer_work.reset_index()
        transfer_work["Transfer Value"] = pd.to_numeric(transfer_work.get("Transfer Value"), errors="coerce").fillna(0.0)
        transfer_work["Destination"] = transfer_work.get("Destination", pd.Series(dtype=str)).apply(
            lambda value: _normalize_business_unit_name(value, destination)
        )
        grouped_transfer = transfer_work.groupby(["Period", "Destination"], as_index=False)["Transfer Value"].sum(min_count=1)
        total_internal = _series_from_period_map(
            {
                str(period): float(value)
                for period, value in zip(grouped_transfer["Period"], grouped_transfer["Transfer Value"])
                if str(period).strip() and pd.notna(value)
            },
            working.index,
        )
        internal_revenue_map["Breeding"] = total_internal
        destination_transfer = grouped_transfer.loc[grouped_transfer["Destination"].eq(destination)].copy()
        internal_cost_map[destination] = _series_from_period_map(
            {
                str(period): float(value)
                for period, value in zip(destination_transfer["Period"], destination_transfer["Transfer Value"])
                if str(period).strip() and pd.notna(value)
            },
            working.index,
        )

    explicit_variable = {
        unit: pd.to_numeric(
            operating_costs.get(unit, {}).get("variable_healthcare_cost_per_herd", _zero_series(working.index)),
            errors="coerce",
        ).fillna(0.0).add(_unit_series_lookup(variable_detail, unit, working.index), fill_value=0.0)
        for unit in units
    }
    explicit_direct = {
        unit: _unit_series_lookup(direct_detail, unit, working.index)
        for unit in units
    }
    explicit_admin = {
        unit: _unit_series_lookup(admin_detail, unit, working.index)
        for unit in units
    }
    explicit_fixed = {
        unit: pd.to_numeric(
            operating_costs.get(unit, {}).get("fixed_utility_cost_per_herd", _zero_series(working.index)),
            errors="coerce",
        ).fillna(0.0)
        for unit in units
    }
    explicit_cogs = {
        unit: pd.to_numeric(
            operating_costs.get(unit, {}).get("variable_feed_cost_per_herd", _zero_series(working.index)),
            errors="coerce",
        ).fillna(0.0)
        for unit in units
    }

    revenue_allocated = _allocate_reporting_column(
        _reporting_source_series(working, "Revenue_adj" if "Revenue_adj" in working.columns else "Revenue"),
        units,
        weights,
        explicit_map={unit: external_revenue.get(unit, _zero_series(working.index)) for unit in units},
        extra_map=internal_revenue_map,
    )
    cogs_allocated = _allocate_reporting_column(
        _reporting_source_series(working, "COGS_adj" if "COGS_adj" in working.columns else "COGS"),
        units,
        weights,
        explicit_map=explicit_cogs,
        extra_map=internal_cost_map,
    )
    variable_allocated = _allocate_reporting_column(
        _reporting_source_series(working, "Variable Expenses"),
        units,
        weights,
        explicit_map=explicit_variable,
    )
    fixed_allocated = _allocate_reporting_column(
        _reporting_source_series(working, "Fixed Expenses"),
        units,
        weights,
        explicit_map=explicit_fixed,
    )
    direct_allocated = _allocate_reporting_column(
        _reporting_source_series(working, "Direct Wages"),
        units,
        weights,
        explicit_map=explicit_direct,
    )
    admin_allocated = _allocate_reporting_column(
        _reporting_source_series(working, "Admin Wages"),
        units,
        weights,
        explicit_map=explicit_admin,
    )

    for unit in units:
        frame = unit_frames[unit].copy()
        for column, allocated in [
            ("Revenue", revenue_allocated),
            ("COGS", cogs_allocated),
            ("Variable Expenses", variable_allocated),
            ("Fixed Expenses", fixed_allocated),
            ("Direct Wages", direct_allocated),
            ("Admin Wages", admin_allocated),
        ]:
            frame[column] = pd.to_numeric(allocated.get(unit), errors="coerce").fillna(0.0)
        frame = _sync_reporting_adjusted_columns(_synchronize_financial_algorithms(frame))
        unit_frames[unit] = frame

    consolidated = unit_frames[units[0]].copy()
    for unit in units[1:]:
        consolidated = consolidated.add(unit_frames[unit], fill_value=0.0)

    elimination_table = (supplementary or {}).get("Internal Transfer Elimination Schedule", pd.DataFrame())
    if isinstance(elimination_table, pd.DataFrame) and not elimination_table.empty:
        elimination_work = elimination_table.copy()
        if "Period" not in elimination_work.columns:
            elimination_work = elimination_work.reset_index()
        revenue_elimination = _series_from_period_map(
            {
                str(period): float(value)
                for period, value in zip(
                    elimination_work.get("Period", pd.Series(dtype=str)),
                    pd.to_numeric(elimination_work.get("Revenue Elimination"), errors="coerce").fillna(0.0),
                )
                if str(period).strip() and pd.notna(value)
            },
            working.index,
        )
        cost_elimination = _series_from_period_map(
            {
                str(period): float(value)
                for period, value in zip(
                    elimination_work.get("Period", pd.Series(dtype=str)),
                    pd.to_numeric(elimination_work.get("Cost Elimination"), errors="coerce").fillna(0.0),
                )
                if str(period).strip() and pd.notna(value)
            },
            working.index,
        )
        if "Revenue" in consolidated.columns:
            consolidated["Revenue"] = pd.to_numeric(consolidated["Revenue"], errors="coerce").fillna(0.0) + revenue_elimination
        if "COGS" in consolidated.columns:
            consolidated["COGS"] = pd.to_numeric(consolidated["COGS"], errors="coerce").fillna(0.0) + cost_elimination

    consolidated = _sync_reporting_adjusted_columns(_synchronize_financial_algorithms(consolidated))
    return {
        "Breeding": unit_frames["Breeding"],
        destination: unit_frames[destination],
        "Consolidated": consolidated,
    }


def _reporting_entity_options(assumptions: Optional[dict[str, pd.DataFrame]]) -> list[str]:
    assumption_map = dict(assumptions or {})
    if not _is_breeding_to_unit_mode(assumption_map):
        return ["Consolidated"]
    destination = _selected_transfer_destination(assumption_map)
    options = ["Consolidated", "Breeding"]
    if destination:
        options.append(destination)
    return options


def _default_reporting_entity(assumptions: Optional[dict[str, pd.DataFrame]]) -> str:
    assumption_map = dict(assumptions or {})
    if _selected_reporting_view(assumption_map) == "Standalone Unit":
        return "Breeding"
    return "Consolidated"


def _default_herd_plan_table() -> pd.DataFrame:
    current_year = pd.Timestamp.today().year
    return pd.DataFrame(
        {
            "Year": [current_year, current_year + 1, current_year + 2],
            "Herd Size (heads)": [320.0, 336.0, 353.0],
            "Herd Growth %": [np.nan, 5.0, 5.0],
        }
    )


def _ensure_herd_plan_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_herd_plan_table()
    else:
        work = table.copy()

    for column in ["Year", "Herd Size (heads)", "Herd Growth %", "yearly_increment_percent"]:
        if column not in work.columns:
            work[column] = np.nan
        work[column] = pd.to_numeric(work.get(column), errors="coerce")

    work = work.dropna(how="all").dropna(subset=["Year"]).sort_values("Year").reset_index(drop=True)
    if work.empty:
        return _default_herd_plan_table()

    previous_size: Optional[float] = None
    for idx in work.index:
        size = work.at[idx, "Herd Size (heads)"]
        growth = work.at[idx, "Herd Growth %"]
        if pd.isna(growth):
            growth = work.at[idx, "yearly_increment_percent"]
        if pd.isna(size):
            if previous_size is not None and pd.notna(growth):
                size = previous_size * (1.0 + float(growth) / 100.0)
            elif previous_size is not None:
                size = previous_size
            else:
                size = 100.0
            work.at[idx, "Herd Size (heads)"] = float(size)

        if previous_size is not None and previous_size > 0 and pd.isna(growth):
            work.at[idx, "Herd Growth %"] = (float(work.at[idx, "Herd Size (heads)"]) / previous_size - 1.0) * 100.0
        previous_size = float(work.at[idx, "Herd Size (heads)"])

    work["yearly_increment_percent"] = pd.to_numeric(work["Herd Growth %"], errors="coerce")
    ordered = ["Year", "Herd Size (heads)", "Herd Growth %", "yearly_increment_percent"]
    remainder = [col for col in work.columns if col not in ordered]
    return work[ordered + remainder].reset_index(drop=True)


def _apply_herd_yearly_increment(table: pd.DataFrame, yearly_increment_percent: float) -> pd.DataFrame:
    work = _ensure_herd_plan_table(table)
    if work.empty:
        return work

    increment = float(yearly_increment_percent)
    base_size = pd.to_numeric(pd.Series([work.iloc[0].get("Herd Size (heads)")]), errors="coerce").iloc[0]
    if pd.isna(base_size) or base_size <= 0:
        base_size = 100.0
    work.at[0, "Herd Size (heads)"] = float(base_size)
    work.at[0, "Herd Growth %"] = np.nan

    for idx in work.index[1:]:
        prev_size = float(work.at[idx - 1, "Herd Size (heads)"])
        next_size = prev_size * (1.0 + increment / 100.0)
        work.at[idx, "Herd Size (heads)"] = float(next_size)
        work.at[idx, "Herd Growth %"] = increment

    work["yearly_increment_percent"] = pd.to_numeric(work["Herd Growth %"], errors="coerce")
    return _ensure_herd_plan_table(work)


def _apply_herd_plan_to_schedule(schedule_df: pd.DataFrame, herd_plan: Optional[pd.DataFrame]) -> pd.DataFrame:
    if schedule_df.empty or herd_plan is None or herd_plan.empty:
        return schedule_df

    plan = _ensure_herd_plan_table(herd_plan)
    size_map = {
        int(row["Year"]): float(row["Herd Size (heads)"])
        for _, row in plan.iterrows()
        if pd.notna(row.get("Year")) and pd.notna(row.get("Herd Size (heads)"))
    }
    if not size_map:
        return schedule_df

    baseline_size = next((v for _, v in sorted(size_map.items()) if v > 0), None)
    if baseline_size is None:
        return schedule_df

    work = schedule_df.copy()
    herd_sizes = work.index.year.map(lambda year: size_map.get(int(year), float(baseline_size))).astype(float)
    multipliers = herd_sizes / float(baseline_size)
    work["Herd Size (heads)"] = herd_sizes
    work["Herd Multiplier"] = multipliers

    for col in ["Revenue", "COGS", "Variable Expenses", "Direct Wages"]:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce") * multipliers

    return _synchronize_financial_algorithms(work)


def _revenue_map(core: pd.DataFrame) -> Dict[str, float]:
    if "Period" not in core.columns or "Revenue" not in core.columns:
        return {}
    periods = _normalize_period(core["Period"])
    revenue = pd.to_numeric(core["Revenue"], errors="coerce")
    return {
        period: float(value)
        for period, value in zip(periods, revenue)
        if period and not np.isnan(value)
    }


def _sync_cogs_table(
    table: pd.DataFrame, core: pd.DataFrame, default_pct: float = 45.0
) -> pd.DataFrame:
    if table is None or table.empty:
        return table

    work = table.copy()
    work["Period"] = _normalize_period(work.get("Period", pd.Series(dtype=str)))

    if "COGS" not in work.columns:
        work["COGS"] = np.nan
    if "COGS %" not in work.columns:
        work["COGS %"] = np.nan

    revenue_lookup = _revenue_map(core)
    revenue_series = work["Period"].map(revenue_lookup)
    amount = pd.to_numeric(work["COGS"], errors="coerce")
    percent = pd.to_numeric(work["COGS %"], errors="coerce")

    # Treat fractions (0.45) as percentages (45%)
    fraction_mask = percent.notna() & (percent <= 1.0)
    percent.loc[fraction_mask] = percent.loc[fraction_mask] * 100.0

    percent_provided = percent.notna()
    amount_provided = amount.notna()
    revenue_available = revenue_series.notna() & (revenue_series != 0)

    # Where amount is provided alongside revenue, derive the percentage
    mask = revenue_available & amount_provided
    percent.loc[mask] = (amount.loc[mask] / revenue_series.loc[mask]) * 100.0

    # Where only the percentage is provided, back into the amount
    mask = revenue_available & percent_provided & ~amount_provided
    amount.loc[mask] = revenue_series.loc[mask] * (percent.loc[mask] / 100.0)

    # Where neither is supplied but revenue exists, fall back to default
    mask = revenue_available & ~amount_provided & ~percent_provided
    amount.loc[mask] = revenue_series.loc[mask] * (default_pct / 100.0)
    percent.loc[mask] = default_pct

    work["COGS"] = amount
    work["COGS %"] = percent

    ordered_cols = ["Period", "COGS %", "COGS"]
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _ensure_cogs_schedule(
    table: Optional[pd.DataFrame], core: pd.DataFrame, default_pct: float = 45.0
) -> pd.DataFrame:
    if table is None or table.empty:
        base_periods = _normalize_period(core.get("Period", pd.Series(dtype=str)))
        table = pd.DataFrame({"Period": base_periods, "COGS": np.nan})

    return _sync_cogs_table(table, core, default_pct=default_pct)


def _sync_cogs_from_operating_assumptions(
    cogs_table: pd.DataFrame,
    core_schedule: pd.DataFrame,
    assumptions: Optional[Dict[str, pd.DataFrame]] = None,
    default_pct: float = 45.0,
) -> pd.DataFrame:
    """Project COGS from Herd Plan + Operating Costs and sync COGS/% rows by period."""
    if cogs_table is None or cogs_table.empty or core_schedule is None or core_schedule.empty:
        return cogs_table

    assumptions_map = assumptions or {}
    herd_plan = assumptions_map.get("Herd Plan")
    operating_costs = assumptions_map.get("Operating Costs")
    if not isinstance(operating_costs, pd.DataFrame) or operating_costs.empty:
        return _sync_cogs_table(cogs_table, core_schedule, default_pct=default_pct)

    projected = core_schedule.copy()
    if "Period" in projected.columns:
        period_values = pd.to_datetime(projected["Period"], errors="coerce")
        projected = projected.set_index(period_values)
    projected.index = pd.to_datetime(projected.index, errors="coerce")
    projected = projected.loc[projected.index.notna()].sort_index()
    if projected.empty:
        return _sync_cogs_table(cogs_table, core_schedule, default_pct=default_pct)

    if isinstance(herd_plan, pd.DataFrame) and not herd_plan.empty:
        projected = _apply_herd_plan_to_schedule(projected, herd_plan)
    projected = _apply_biological_assumptions_to_schedule(projected, assumptions_map)
    projected = _apply_operating_cost_assumptions_to_schedule(
        projected,
        operating_costs,
        assumptions_map.get("Biological Cost Drivers"),
    )

    mapped = cogs_table.copy()
    mapped["Period"] = _normalize_period(mapped.get("Period", pd.Series(dtype=str)))
    projected_map = {
        idx.strftime("%Y-%m-%d"): float(value)
        for idx, value in zip(
            projected.index,
            pd.to_numeric(projected.get("COGS"), errors="coerce"),
        )
        if pd.notna(idx) and pd.notna(value)
    }
    mapped["COGS"] = mapped["Period"].map(projected_map).combine_first(
        pd.to_numeric(mapped.get("COGS"), errors="coerce")
    )
    return _sync_cogs_table(mapped, core_schedule, default_pct=default_pct)


# ---------- Direct wages helpers ----------


def _default_direct_wage_table(core: pd.DataFrame) -> pd.DataFrame:
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str))).tolist()
    totals_raw = core.get("Direct Wages", pd.Series(dtype=float))
    totals = pd.to_numeric(totals_raw, errors="coerce")
    period_multiplier = _direct_wage_period_multiplier(core)

    if isinstance(totals, pd.Series):
        total_series = totals.reset_index(drop=True)
    else:
        # When a scalar or unsupported type is returned, broadcast across periods
        total_series = pd.Series([totals])

    total_series = total_series.reindex(range(len(periods)), fill_value=np.nan)
    total_values = total_series.to_list()

    rows: list[dict[str, object]] = []
    if periods:
        for idx, period in enumerate(periods):
            total = total_values[idx] if idx < len(total_values) else np.nan
            for item in _direct_wage_default_items():
                headcount = pd.to_numeric(
                    pd.Series([item.get("Head Count")]), errors="coerce"
                ).iloc[0]
                monthly_salary = pd.to_numeric(
                    pd.Series([item.get("Monthly Salary per Head")]), errors="coerce"
                ).iloc[0]
                total_salary = pd.to_numeric(
                    pd.Series([item.get("Total Salary")]), errors="coerce"
                ).iloc[0]
                share = item.get("Share")

                if pd.isna(total_salary) and not pd.isna(monthly_salary) and not pd.isna(headcount):
                    total_salary = float(headcount) * float(monthly_salary) * period_multiplier
                elif pd.isna(monthly_salary) and not pd.isna(total_salary):
                    divisor = float(headcount) if not pd.isna(headcount) and headcount > 0 else 1.0
                    monthly_salary = float(total_salary) / (divisor * period_multiplier)

                if share is not None and total is not None and not np.isnan(total):
                    total_salary = float(total) * float(share)
                    divisor = float(headcount) if not pd.isna(headcount) and headcount > 0 else 1.0
                    monthly_salary = total_salary / (divisor * period_multiplier)

                rows.append(
                    {
                        "Period": period,
                        "Business Unit": "General",
                        "Position": item.get("Position", "Direct Wage"),
                        "Head Count": headcount,
                        "Monthly Salary per Head": monthly_salary,
                        "Total Salary": total_salary,
                    }
                )

    if not rows:
        period_type = _infer_period_type_from_schedule(core)
        today = _next_period_from_last(None, period_type).strftime("%Y-%m-%d")
        rows.append(
            {
                "Period": today,
                "Business Unit": "General",
                "Position": "Direct Wage",
                "Head Count": 1.0,
                "Monthly Salary per Head": np.nan,
                "Total Salary": np.nan,
            }
        )

    return pd.DataFrame(rows)


def _ensure_direct_wage_table(
    table: Optional[pd.DataFrame], core: pd.DataFrame
) -> pd.DataFrame:
    if table is None or table.empty:
        return _default_direct_wage_table(core)

    work = table.copy()
    work["Period"] = _normalize_period(work.get("Period", pd.Series(dtype=str)))
    if "Business Unit" not in work.columns:
        work["Business Unit"] = "General"
    work["Business Unit"] = _series_or_default(work, "Business Unit", "General").astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = "General"
    if "Position" not in work.columns:
        work["Position"] = work.get("Role", "")
    work["Position"] = _series_or_default(work, "Position", "").astype(str).str.strip()
    work.loc[work["Position"] == "", "Position"] = "Direct Wage"

    if "Head Count" not in work.columns:
        work["Head Count"] = 1.0
    work["Head Count"] = pd.to_numeric(work.get("Head Count"), errors="coerce")

    if "Monthly Salary per Head" not in work.columns:
        work["Monthly Salary per Head"] = np.nan
    work["Monthly Salary per Head"] = pd.to_numeric(
        work.get("Monthly Salary per Head"), errors="coerce"
    )

    if "Total Salary" not in work.columns:
        work["Total Salary"] = work.get("Amount", np.nan)
    work["Total Salary"] = pd.to_numeric(work.get("Total Salary"), errors="coerce")

    period_multiplier = _direct_wage_period_multiplier(work)
    divisor = work["Head Count"].where(work["Head Count"] > 0, 1.0)
    missing_monthly = work["Monthly Salary per Head"].isna() & work["Total Salary"].notna()
    work.loc[missing_monthly, "Monthly Salary per Head"] = (
        work.loc[missing_monthly, "Total Salary"] / (divisor.loc[missing_monthly] * period_multiplier)
    )
    computable_total = work["Head Count"].notna() & work["Monthly Salary per Head"].notna()
    work.loc[computable_total, "Total Salary"] = (
        work.loc[computable_total, "Head Count"]
        * work.loc[computable_total, "Monthly Salary per Head"]
        * period_multiplier
    )

    work = work.dropna(how="all")
    work = work[
        (work["Position"].notna())
        | (work["Head Count"].notna())
        | (work["Monthly Salary per Head"].notna())
        | (work["Total Salary"].notna())
    ]
    work = work.dropna(subset=["Period"], how="all")

    if work.empty:
        return _default_direct_wage_table(core)

    ordered_cols = [
        "Period",
        "Business Unit",
        "Position",
        "Head Count",
        "Monthly Salary per Head",
        "Total Salary",
    ]
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _add_direct_wage_row(table: pd.DataFrame, core: pd.DataFrame) -> pd.DataFrame:
    work = _ensure_direct_wage_table(table, core)
    periods = work.get("Period", pd.Series(dtype=str))
    default_period = None
    if not periods.empty:
        default_period = periods.iloc[-1]
    else:
        core_periods = _normalize_period(core.get("Period", pd.Series(dtype=str)))
        if not core_periods.empty:
            default_period = core_periods.iloc[-1]
    if default_period is None or pd.isna(default_period):
        period_type = _infer_period_type_from_schedule(core)
        default_period = _next_period_from_last(None, period_type).strftime("%Y-%m-%d")

    new_row = {
        "Period": default_period,
        "Business Unit": "General",
        "Position": f"Direct Wage Position {len(work) + 1}",
        "Head Count": 1.0,
        "Monthly Salary per Head": np.nan,
        "Total Salary": np.nan,
    }
    return pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)


def _remove_direct_wage_row(table: pd.DataFrame, index: int) -> pd.DataFrame:
    if table is None or table.empty:
        return table
    work = table.copy()
    if 0 <= index < len(work):
        work = work.drop(index=index).reset_index(drop=True)
    return work


def _apply_direct_wage_increment(
    table: pd.DataFrame, increment_pct: float, target_position: Optional[str] = None
) -> pd.DataFrame:
    if table is None or table.empty or increment_pct == 0:
        return table

    work = _ensure_direct_wage_table(table, pd.DataFrame({"Period": table.get("Period", pd.Series(dtype=str))}))
    work["Period_dt"] = pd.to_datetime(work.get("Period"), errors="coerce")
    work["Monthly Salary per Head"] = pd.to_numeric(
        work.get("Monthly Salary per Head"), errors="coerce"
    )
    work["Position"] = _series_or_default(work, "Position", "").astype(str).str.strip()

    increment_factor = 1 + (increment_pct / 100.0)

    def _should_update(position: str) -> bool:
        if not target_position or target_position == "All positions":
            return True
        return position == target_position

    for position, group in work.groupby("Position", dropna=False):
        position_key = position if isinstance(position, str) else ""
        if not _should_update(position_key):
            continue
        group = group.sort_values("Period_dt", kind="stable")
        prev_amount = None
        prev_year = None
        for idx, row in group.iterrows():
            period_dt = row["Period_dt"]
            amount = row["Monthly Salary per Head"]
            if pd.isna(period_dt):
                continue
            year = int(period_dt.year)
            if prev_amount is None and not pd.isna(amount):
                prev_amount = amount
                prev_year = year
                continue
            if prev_amount is None or prev_year is None:
                continue
            year_gap = year - prev_year
            if year_gap <= 0:
                if not pd.isna(amount):
                    prev_amount = amount
                    prev_year = year
                continue
            new_amount = prev_amount * (increment_factor ** year_gap)
            work.at[idx, "Monthly Salary per Head"] = new_amount
            prev_amount = new_amount
            prev_year = year

    work = work.drop(columns="Period_dt")
    return _ensure_direct_wage_table(work, pd.DataFrame({"Period": work.get("Period", pd.Series(dtype=str))}))


def _aggregate_direct_wages(
    table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    work = _ensure_direct_wage_table(table, core)
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str)))
    summary = (
        work.groupby("Period", as_index=False)["Total Salary"].sum(min_count=1)
        if not work.empty
        else pd.DataFrame(columns=["Period", "Total Salary"])
    )
    result = pd.DataFrame({"Period": periods})
    summary_map = dict(zip(summary.get("Period", []), summary.get("Total Salary", [])))
    result["Direct Wages"] = result["Period"].map(summary_map)
    if result["Direct Wages"].notna().any():
        result["Direct Wages"] = result["Direct Wages"].astype(float)
    else:
        result["Direct Wages"] = 0.0
    return result


def _aggregate_direct_wages_by_business_unit(
    table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    return _sum_schedule_amounts_by_business_unit(
        table,
        core,
        _ensure_direct_wage_table,
        "Total Salary",
    )


# ---------- Admin wages helpers ----------


def _default_admin_wage_table(core: pd.DataFrame) -> pd.DataFrame:
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str))).tolist()
    totals_raw = core.get("Admin Wages", pd.Series(dtype=float))
    totals = pd.to_numeric(totals_raw, errors="coerce")
    period_multiplier = _direct_wage_period_multiplier(core)

    if isinstance(totals, pd.Series):
        total_series = totals.reset_index(drop=True)
    else:
        total_series = pd.Series([totals])

    total_series = total_series.reindex(range(len(periods)), fill_value=np.nan)
    total_values = total_series.to_list()

    rows: list[dict[str, object]] = []
    if periods:
        for idx, period in enumerate(periods):
            total = total_values[idx] if idx < len(total_values) else np.nan
            for item in _admin_wage_default_items():
                headcount = pd.to_numeric(
                    pd.Series([item.get("Head Count")]), errors="coerce"
                ).iloc[0]
                monthly_salary = pd.to_numeric(
                    pd.Series([item.get("Monthly Salary per Head")]), errors="coerce"
                ).iloc[0]
                total_salary = pd.to_numeric(
                    pd.Series([item.get("Total Salary")]), errors="coerce"
                ).iloc[0]
                share = item.get("Share")

                if pd.isna(total_salary) and not pd.isna(monthly_salary) and not pd.isna(headcount):
                    total_salary = float(headcount) * float(monthly_salary) * period_multiplier
                elif pd.isna(monthly_salary) and not pd.isna(total_salary):
                    divisor = float(headcount) if not pd.isna(headcount) and headcount > 0 else 1.0
                    monthly_salary = float(total_salary) / (divisor * period_multiplier)

                if share is not None and total is not None and not np.isnan(total):
                    total_salary = float(total) * float(share)
                    divisor = float(headcount) if not pd.isna(headcount) and headcount > 0 else 1.0
                    monthly_salary = total_salary / (divisor * period_multiplier)

                rows.append(
                    {
                        "Period": period,
                        "Business Unit": "General",
                        "Position": item.get("Position", "Admin Wage"),
                        "Head Count": headcount,
                        "Monthly Salary per Head": monthly_salary,
                        "Total Salary": total_salary,
                    }
                )

    if not rows:
        period_type = _infer_period_type_from_schedule(core)
        today = _next_period_from_last(None, period_type).strftime("%Y-%m-%d")
        rows.append(
            {
                "Period": today,
                "Business Unit": "General",
                "Position": "Admin Wage",
                "Head Count": 1.0,
                "Monthly Salary per Head": np.nan,
                "Total Salary": np.nan,
            }
        )

    return pd.DataFrame(rows)


def _ensure_admin_wage_table(
    table: Optional[pd.DataFrame], core: pd.DataFrame
) -> pd.DataFrame:
    if table is None or table.empty:
        return _default_admin_wage_table(core)

    work = table.copy()
    if "Business Unit" not in work.columns:
        work["Business Unit"] = "General"
    work["Business Unit"] = _series_or_default(work, "Business Unit", "General").astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = "General"
    if "Admin Wages" in work.columns and "Amount" not in work.columns and "Total Salary" not in work.columns:
        periods = _normalize_period(work.get("Period", pd.Series(dtype=str)))
        totals = pd.to_numeric(work.get("Admin Wages"), errors="coerce")
        if isinstance(totals, pd.Series):
            total_values = totals.tolist()
        elif totals is None:
            total_values = []
        else:
            try:
                total_values = [float(totals)] * len(periods)
            except (TypeError, ValueError):
                total_values = []
        reconstructed: list[dict[str, object]] = []
        for idx, period in enumerate(periods):
            total = total_values[idx] if idx < len(total_values) else np.nan
            for item in _admin_wage_default_items():
                headcount = pd.to_numeric(
                    pd.Series([item.get("Head Count")]), errors="coerce"
                ).iloc[0]
                monthly_salary = pd.to_numeric(
                    pd.Series([item.get("Monthly Salary per Head")]), errors="coerce"
                ).iloc[0]
                total_salary = pd.to_numeric(
                    pd.Series([item.get("Total Salary")]), errors="coerce"
                ).iloc[0]
                share = item.get("Share")

                if share is not None and total is not None and not np.isnan(total):
                    total_salary = float(total) * float(share)
                    divisor = float(headcount) if not pd.isna(headcount) and headcount > 0 else 1.0
                    monthly_salary = total_salary / (
                        divisor * _direct_wage_period_multiplier(pd.DataFrame({"Period": periods}))
                    )
                reconstructed.append(
                    {
                        "Period": period,
                        "Position": item.get("Position", "Admin Wage"),
                        "Head Count": headcount,
                        "Monthly Salary per Head": monthly_salary,
                        "Total Salary": total_salary,
                    }
                )
        work = pd.DataFrame(reconstructed)

    work["Period"] = _normalize_period(work.get("Period", pd.Series(dtype=str)))
    if "Position" not in work.columns:
        if "Function" in work.columns:
            work["Position"] = work["Function"]
        elif "Role" in work.columns:
            work["Position"] = work["Role"]
        else:
            work["Position"] = ""
    work["Position"] = _series_or_default(work, "Position", "").astype(str).str.strip()
    work.loc[work["Position"] == "", "Position"] = "Admin Wage"

    if "Head Count" not in work.columns:
        work["Head Count"] = 1.0
    work["Head Count"] = pd.to_numeric(work.get("Head Count"), errors="coerce")

    if "Monthly Salary per Head" not in work.columns:
        work["Monthly Salary per Head"] = np.nan
    work["Monthly Salary per Head"] = pd.to_numeric(
        work.get("Monthly Salary per Head"), errors="coerce"
    )

    if "Total Salary" not in work.columns:
        work["Total Salary"] = work.get("Amount", np.nan)
    work["Total Salary"] = pd.to_numeric(work.get("Total Salary"), errors="coerce")

    period_multiplier = _direct_wage_period_multiplier(work)
    divisor = work["Head Count"].where(work["Head Count"] > 0, 1.0)
    missing_monthly = work["Monthly Salary per Head"].isna() & work["Total Salary"].notna()
    work.loc[missing_monthly, "Monthly Salary per Head"] = (
        work.loc[missing_monthly, "Total Salary"] / (divisor.loc[missing_monthly] * period_multiplier)
    )
    computable_total = work["Head Count"].notna() & work["Monthly Salary per Head"].notna()
    work.loc[computable_total, "Total Salary"] = (
        work.loc[computable_total, "Head Count"]
        * work.loc[computable_total, "Monthly Salary per Head"]
        * period_multiplier
    )

    work = work.dropna(how="all")
    work = work[
        (work["Position"].notna())
        | (work["Head Count"].notna())
        | (work["Monthly Salary per Head"].notna())
        | (work["Total Salary"].notna())
    ]
    work = work.dropna(subset=["Period"], how="all")

    if work.empty:
        return _default_admin_wage_table(core)

    ordered_cols = [
        "Period",
        "Business Unit",
        "Position",
        "Head Count",
        "Monthly Salary per Head",
        "Total Salary",
    ]
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _add_admin_wage_row(table: pd.DataFrame, core: pd.DataFrame) -> pd.DataFrame:
    work = _ensure_admin_wage_table(table, core)
    periods = work.get("Period", pd.Series(dtype=str))
    default_period = None
    if not periods.empty:
        default_period = periods.iloc[-1]
    else:
        core_periods = _normalize_period(core.get("Period", pd.Series(dtype=str)))
        if not core_periods.empty:
            default_period = core_periods.iloc[-1]
    if default_period is None or pd.isna(default_period):
        period_type = _infer_period_type_from_schedule(core)
        default_period = _next_period_from_last(None, period_type).strftime("%Y-%m-%d")

    new_row = {
        "Period": default_period,
        "Business Unit": "General",
        "Position": f"Admin Wage Position {len(work) + 1}",
        "Head Count": 1.0,
        "Monthly Salary per Head": np.nan,
        "Total Salary": np.nan,
    }
    return pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)


def _remove_admin_wage_row(table: pd.DataFrame, index: int) -> pd.DataFrame:
    if table is None or table.empty:
        return table
    work = table.copy()
    if 0 <= index < len(work):
        work = work.drop(index=index).reset_index(drop=True)
    return work


def _apply_admin_wage_increment(
    table: pd.DataFrame, increment_pct: float, target_position: Optional[str] = None
) -> pd.DataFrame:
    if table is None or table.empty or increment_pct == 0:
        return table

    work = _ensure_admin_wage_table(table, pd.DataFrame({"Period": table.get("Period", pd.Series(dtype=str))}))
    work["Period_dt"] = pd.to_datetime(work.get("Period"), errors="coerce")
    work["Monthly Salary per Head"] = pd.to_numeric(
        work.get("Monthly Salary per Head"), errors="coerce"
    )
    work["Position"] = _series_or_default(work, "Position", "").astype(str).str.strip()

    increment_factor = 1 + (increment_pct / 100.0)

    def _should_update(position: str) -> bool:
        if not target_position or target_position == "All positions":
            return True
        return position == target_position

    for position, group in work.groupby("Position", dropna=False):
        position_key = position if isinstance(position, str) else ""
        if not _should_update(position_key):
            continue
        group = group.sort_values("Period_dt", kind="stable")
        prev_amount = None
        prev_year = None
        for idx, row in group.iterrows():
            period_dt = row["Period_dt"]
            amount = row["Monthly Salary per Head"]
            if pd.isna(period_dt):
                continue
            year = int(period_dt.year)
            if prev_amount is None and not pd.isna(amount):
                prev_amount = amount
                prev_year = year
                continue
            if prev_amount is None or prev_year is None:
                continue
            year_gap = year - prev_year
            if year_gap <= 0:
                if not pd.isna(amount):
                    prev_amount = amount
                    prev_year = year
                continue
            new_amount = prev_amount * (increment_factor ** year_gap)
            work.at[idx, "Monthly Salary per Head"] = new_amount
            prev_amount = new_amount
            prev_year = year

    work = work.drop(columns="Period_dt")
    return _ensure_admin_wage_table(work, pd.DataFrame({"Period": work.get("Period", pd.Series(dtype=str))}))


def _aggregate_admin_wages(
    table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    work = _ensure_admin_wage_table(table, core)
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str)))
    summary = (
        work.groupby("Period", as_index=False)["Total Salary"].sum(min_count=1)
        if not work.empty
        else pd.DataFrame(columns=["Period", "Total Salary"])
    )
    result = pd.DataFrame({"Period": periods})
    summary_map = dict(zip(summary.get("Period", []), summary.get("Total Salary", [])))
    result["Admin Wages"] = result["Period"].map(summary_map)
    if result["Admin Wages"].notna().any():
        result["Admin Wages"] = result["Admin Wages"].astype(float)
    else:
        result["Admin Wages"] = 0.0
    return result


def _aggregate_admin_wages_by_business_unit(
    table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    return _sum_schedule_amounts_by_business_unit(
        table,
        core,
        _ensure_admin_wage_table,
        "Total Salary",
    )


# ---------- Variable expenses helpers ----------


def _default_variable_expense_table(core: pd.DataFrame) -> pd.DataFrame:
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str))).tolist()
    revenue = pd.to_numeric(core.get("Revenue"), errors="coerce")
    revenue_values = revenue.tolist() if revenue is not None else []

    rows: list[dict[str, object]] = []
    if periods:
        for idx, period in enumerate(periods):
            rev = revenue_values[idx] if idx < len(revenue_values) else np.nan
            for item, share in _variable_default_items():
                amount = (
                    rev * share if share is not None and rev is not None and not np.isnan(rev)
                    else np.nan
                )
                rows.append({
                    "Period": period,
                    "Business Unit": "General",
                    "Item": item,
                    "Amount": amount,
                })

    if not rows:
        period_type = _infer_period_type_from_schedule(core)
        today = _next_period_from_last(None, period_type).strftime("%Y-%m-%d")
        rows.append({"Period": today, "Business Unit": "General", "Item": "Variable Expense", "Amount": np.nan})

    return pd.DataFrame(rows)


def _ensure_variable_expense_table(
    table: Optional[pd.DataFrame], core: pd.DataFrame
) -> pd.DataFrame:
    if table is None or table.empty:
        return _default_variable_expense_table(core)

    work = table.copy()
    work["Period"] = _normalize_period(work.get("Period", pd.Series(dtype=str)))
    if "Business Unit" not in work.columns:
        work["Business Unit"] = "General"
    work["Business Unit"] = _series_or_default(work, "Business Unit", "General").astype(str).str.strip()
    work.loc[work["Business Unit"] == "", "Business Unit"] = "General"
    work["Item"] = _series_or_default(work, "Item", "").astype(str).str.strip()
    work.loc[work["Item"] == "", "Item"] = "Variable Expense"
    work["Amount"] = pd.to_numeric(work.get("Amount"), errors="coerce")

    work = work.dropna(how="all")
    work = work[(work["Item"].notna()) | (work["Amount"].notna())]
    work = work.dropna(subset=["Period"], how="all")

    if work.empty:
        return _default_variable_expense_table(core)

    return work.reset_index(drop=True)


def _add_variable_expense_row(table: pd.DataFrame, core: pd.DataFrame) -> pd.DataFrame:
    work = _ensure_variable_expense_table(table, core)
    periods = work.get("Period", pd.Series(dtype=str))
    default_period = None
    if not periods.empty:
        default_period = periods.iloc[-1]
    else:
        core_periods = _normalize_period(core.get("Period", pd.Series(dtype=str)))
        if not core_periods.empty:
            default_period = core_periods.iloc[-1]
    if default_period is None or pd.isna(default_period):
        period_type = _infer_period_type_from_schedule(core)
        default_period = _next_period_from_last(None, period_type).strftime("%Y-%m-%d")

    new_row = {
        "Period": default_period,
        "Business Unit": "General",
        "Item": f"Variable Item {len(work) + 1}",
        "Amount": np.nan,
    }
    return pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)


def _remove_variable_expense_row(table: pd.DataFrame, index: int) -> pd.DataFrame:
    if table is None or table.empty:
        return table
    work = table.copy()
    if 0 <= index < len(work):
        work = work.drop(index=index).reset_index(drop=True)
    return work


def _apply_variable_expense_increment(
    table: pd.DataFrame, increment_pct: float, target_item: Optional[str] = None
) -> pd.DataFrame:
    if table is None or table.empty or increment_pct == 0:
        return table

    work = table.copy()
    work["Period_dt"] = pd.to_datetime(work.get("Period"), errors="coerce")
    work["Amount"] = pd.to_numeric(work.get("Amount"), errors="coerce")
    work["Item"] = _series_or_default(work, "Item", "").astype(str).str.strip()

    increment_factor = 1 + (increment_pct / 100.0)

    def _should_update(item: str) -> bool:
        if not target_item or target_item == "All items":
            return True
        return item == target_item

    for item, group in work.groupby("Item", dropna=False):
        if not _should_update(item if isinstance(item, str) else ""):
            continue
        group = group.sort_values("Period_dt", kind="stable")
        prev_amount = None
        prev_year = None
        for idx, row in group.iterrows():
            period_dt = row["Period_dt"]
            amount = row["Amount"]
            if pd.isna(period_dt):
                continue
            year = int(period_dt.year)
            if prev_amount is None and not pd.isna(amount):
                prev_amount = amount
                prev_year = year
                continue
            if prev_amount is None or prev_year is None:
                continue
            year_gap = year - prev_year
            if year_gap <= 0:
                if not pd.isna(amount):
                    prev_amount = amount
                    prev_year = year
                continue
            new_amount = prev_amount * (increment_factor ** year_gap)
            work.at[idx, "Amount"] = new_amount
            prev_amount = new_amount
            prev_year = year

    return work.drop(columns="Period_dt")


def _aggregate_variable_expenses(
    table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    work = _ensure_variable_expense_table(table, core)
    periods = _normalize_period(core.get("Period", pd.Series(dtype=str)))
    summary = (
        work.groupby("Period", as_index=False)["Amount"].sum(min_count=1)
        if not work.empty
        else pd.DataFrame(columns=["Period", "Amount"])
    )
    result = pd.DataFrame({"Period": periods})
    summary_map = dict(zip(summary.get("Period", []), summary.get("Amount", [])))
    result["Variable Expenses"] = result["Period"].map(summary_map)
    if result["Variable Expenses"].notna().any():
        result["Variable Expenses"] = result["Variable Expenses"].astype(float)
    else:
        result["Variable Expenses"] = 0.0
    return result


def _aggregate_variable_expenses_by_business_unit(
    table: pd.DataFrame, core: pd.DataFrame
) -> pd.DataFrame:
    return _sum_schedule_amounts_by_business_unit(
        table,
        core,
        _ensure_variable_expense_table,
        "Amount",
    )


def _schedule_editor_save_table(schedule_name: str, table: pd.DataFrame) -> None:
    current = st.session_state.detail_schedules.get(schedule_name)
    if _frames_equal(current, table):
        return
    st.session_state.detail_schedules[schedule_name] = table
    _reset_cached_results()


def _schedule_editor_target_options(
    table: pd.DataFrame,
    label_column: str,
    all_label: str,
) -> list[str]:
    return [all_label] + sorted(
        {
            str(value)
            for value in table.get(label_column, pd.Series(dtype=str)).dropna().unique().tolist()
            if str(value).strip()
        }
    )


def _variable_expense_schedule_editor_spec() -> ScheduleEditorSpec:
    return ScheduleEditorSpec(
        schedule_name="Variable Expenses Schedule",
        editor_key="detail::variable_expenses",
        remove_choice_key="var_exp_remove_choice",
        increment_target_key="var_exp_increment_target",
        increment_pct_key="var_exp_increment_pct",
        add_button_key="var_exp_add_row",
        remove_button_key="var_exp_remove_row",
        increment_button_key="var_exp_apply_increment",
        ensure_fn=_ensure_variable_expense_table,
        add_row_fn=_add_variable_expense_row,
        remove_row_fn=_remove_variable_expense_row,
        apply_increment_fn=_apply_variable_expense_increment,
        aggregate_fn=_aggregate_variable_expenses,
        label_builder=lambda row: f"{row.get('Period') or 'Unknown Period'} - {row.get('Item') or 'Item'}",
        target_options_builder=lambda table: _schedule_editor_target_options(table, "Item", "All items"),
        all_items_label="All items",
    )


def _direct_wage_schedule_editor_spec() -> ScheduleEditorSpec:
    return ScheduleEditorSpec(
        schedule_name="Direct Wages Schedule",
        editor_key="detail::direct_wages",
        remove_choice_key="direct_wage_remove_choice",
        increment_target_key="direct_wage_increment_target",
        increment_pct_key="direct_wage_increment_pct",
        add_button_key="direct_wage_add_row",
        remove_button_key="direct_wage_remove_row",
        increment_button_key="direct_wage_apply_increment",
        ensure_fn=_ensure_direct_wage_table,
        add_row_fn=_add_direct_wage_row,
        remove_row_fn=_remove_direct_wage_row,
        apply_increment_fn=_apply_direct_wage_increment,
        aggregate_fn=_aggregate_direct_wages,
        label_builder=lambda row: f"{row.get('Period') or 'Unknown Period'} - {row.get('Position') or row.get('Role') or 'Position'}",
        target_options_builder=lambda table: _schedule_editor_target_options(table, "Position", "All positions"),
        all_items_label="All positions",
    )


def _admin_wage_schedule_editor_spec() -> ScheduleEditorSpec:
    return ScheduleEditorSpec(
        schedule_name="Admin Wages Schedule",
        editor_key="detail::admin_wages",
        remove_choice_key="admin_wage_remove_choice",
        increment_target_key="admin_wage_increment_target",
        increment_pct_key="admin_wage_increment_pct",
        add_button_key="admin_wage_add_row",
        remove_button_key="admin_wage_remove_row",
        increment_button_key="admin_wage_apply_increment",
        ensure_fn=_ensure_admin_wage_table,
        add_row_fn=_add_admin_wage_row,
        remove_row_fn=_remove_admin_wage_row,
        apply_increment_fn=_apply_admin_wage_increment,
        aggregate_fn=_aggregate_admin_wages,
        label_builder=lambda row: f"{row.get('Period') or 'Unknown Period'} - {row.get('Position') or row.get('Function') or 'Position'}",
        target_options_builder=lambda table: _schedule_editor_target_options(table, "Position", "All positions"),
        all_items_label="All positions",
    )


def _apply_cogs_percentage(
    table: pd.DataFrame, core: pd.DataFrame, percent: float
) -> pd.DataFrame:
    work = table.copy()
    work["COGS %"] = percent
    return _sync_cogs_table(work, core, default_pct=percent)


def _apply_yearly_increment(
    table: pd.DataFrame, core: pd.DataFrame, increment_pct: float, default_pct: float
) -> pd.DataFrame:
    if table is None or table.empty or increment_pct == 0:
        return table

    work = _sync_cogs_table(table, core, default_pct=default_pct)
    temp = work.copy()
    temp["__period_dt"] = pd.to_datetime(temp["Period"], errors="coerce")
    temp = temp.sort_values("__period_dt", kind="stable")

    increment_factor = 1 + (increment_pct / 100.0)
    current_pct = None
    current_year = None

    for idx, row in temp.iterrows():
        period_dt = row["__period_dt"]
        pct_value = row["COGS %"]
        if pd.isna(period_dt):
            continue
        if current_pct is None:
            current_pct = default_pct if pd.isna(pct_value) else float(pct_value)
            current_year = period_dt.year
            temp.at[idx, "COGS %"] = current_pct
            continue

        year_gap = period_dt.year - current_year if current_year is not None else 0
        if year_gap > 0:
            current_pct = current_pct * (increment_factor ** year_gap)
            current_year = period_dt.year
        temp.at[idx, "COGS %"] = current_pct

    temp = temp.drop(columns="__period_dt")
    work.loc[temp.index, "COGS %"] = temp["COGS %"]
    return _sync_cogs_table(work, core, default_pct=default_pct)


def _add_cogs_row(
    table: pd.DataFrame, core: pd.DataFrame, default_pct: float
) -> pd.DataFrame:
    work = _sync_cogs_table(table, core, default_pct=default_pct)
    periods = pd.to_datetime(work["Period"], errors="coerce")
    period_type = _infer_period_type_from_schedule(core)

    if periods.notna().any():
        last_period = periods.max()
    else:
        core_periods = pd.to_datetime(core.get("Period", pd.Series(dtype=str)), errors="coerce")
        if core_periods.notna().any():
            last_period = core_periods.max()
        else:
            last_period = _next_period_from_last(None, period_type)

    next_period = _next_period_from_last(last_period, period_type)
    existing_periods = set(work["Period"].astype(str))
    while next_period.strftime("%Y-%m-%d") in existing_periods:
        next_period = _next_period_from_last(next_period, period_type)

    next_period_str = next_period.strftime("%Y-%m-%d")

    revenue_lookup = _revenue_map(core)
    revenue = revenue_lookup.get(next_period_str)

    last_pct_series = pd.to_numeric(work.get("COGS %"), errors="coerce").dropna()
    pct_value = float(last_pct_series.iloc[-1]) if not last_pct_series.empty else default_pct
    amount = revenue * (pct_value / 100.0) if revenue is not None else np.nan

    new_row = {"Period": next_period_str, "COGS %": pct_value, "COGS": amount}
    return _sync_cogs_table(pd.concat([work, pd.DataFrame([new_row])], ignore_index=True), core, default_pct=pct_value)


def _remove_cogs_row(table: pd.DataFrame, period: str) -> pd.DataFrame:
    if table is None or table.empty:
        return table
    mask = table["Period"].astype(str) != str(period)
    return table.loc[mask].reset_index(drop=True)


def _ensure_asset_schedule(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = pd.DataFrame(columns=ASSET_SCHEDULE_COLUMNS)
    else:
        work = table.copy()

    for column in ASSET_SCHEDULE_COLUMNS:
        if column not in work.columns:
            work[column] = np.nan

    work["Asset"] = work["Asset"].fillna("").astype(str)
    work["Year"] = pd.to_numeric(work["Year"], errors="coerce")

    numeric_cols = [
        "Opening NBV",
        "Additions",
        "Depreciation Rate %",
        "Depreciation",
        "Closing NBV",
    ]
    for column in numeric_cols:
        work[column] = pd.to_numeric(work[column], errors="coerce")

    ordered = [col for col in ASSET_SCHEDULE_COLUMNS if col in work.columns]
    remaining = [col for col in work.columns if col not in ordered]

    work = work[ordered + remaining].reset_index(drop=True)
    return _recalculate_asset_schedule(work)


def _recalculate_asset_schedule(table: pd.DataFrame) -> pd.DataFrame:
    if table is None or table.empty:
        return pd.DataFrame(columns=ASSET_SCHEDULE_COLUMNS)

    work = table.copy()

    opening = pd.to_numeric(work.get("Opening NBV"), errors="coerce").fillna(0.0)
    additions = pd.to_numeric(work.get("Additions"), errors="coerce").fillna(0.0)
    rate = pd.to_numeric(work.get("Depreciation Rate %"), errors="coerce")
    depreciation = pd.to_numeric(work.get("Depreciation"), errors="coerce")

    total_basis = opening + additions

    if not rate.isna().all():
        calc_dep = total_basis * (rate.fillna(0.0) / 100.0)
        dep_mask = depreciation.isna() & rate.notna()
        depreciation.loc[dep_mask] = calc_dep.loc[dep_mask]

    with np.errstate(divide="ignore", invalid="ignore"):
        rate_mask = rate.isna() & depreciation.notna() & (total_basis != 0)
        rate.loc[rate_mask] = (depreciation.loc[rate_mask] / total_basis.loc[rate_mask]) * 100.0

    closing = opening + additions - depreciation.fillna(0.0)

    work["Depreciation Rate %"] = rate
    work["Depreciation"] = depreciation
    work["Closing NBV"] = closing

    return work


def _derive_asset_schedule_from_capex(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    capex = _ensure_capex_schedule(table)
    if capex.empty:
        return pd.DataFrame(columns=ASSET_SCHEDULE_COLUMNS)

    work = capex.copy()
    work["Category"] = work["Category"].fillna("").astype(str).str.strip()
    work["Year"] = pd.to_numeric(work["Year"], errors="coerce")
    work["Spend"] = pd.to_numeric(work["Spend"], errors="coerce").fillna(0.0)
    work["Depreciation"] = pd.to_numeric(work["Depreciation"], errors="coerce").fillna(0.0)

    work = work.loc[
        work["Category"].ne("") & work["Year"].notna()
    ].copy()
    if work.empty:
        return pd.DataFrame(columns=ASSET_SCHEDULE_COLUMNS)

    grouped = (
        work.groupby(["Category", "Year"], as_index=False, sort=True)
        .agg(
            **{
                "Additions": ("Spend", "sum"),
                "Depreciation": ("Depreciation", "sum"),
            }
        )
        .sort_values(["Category", "Year"], kind="stable")
        .reset_index(drop=True)
    )
    grouped["Year"] = grouped["Year"].round().astype("Int64")

    asset_rows: list[dict[str, Any]] = []
    for asset, asset_group in grouped.groupby("Category", sort=False):
        opening_nbv = 0.0
        for _, row in asset_group.sort_values("Year", kind="stable").iterrows():
            additions = float(row.get("Additions", 0.0) or 0.0)
            depreciation = float(row.get("Depreciation", 0.0) or 0.0)
            year = int(row["Year"]) if pd.notna(row.get("Year")) else pd.NA
            rate = (depreciation / additions * 100.0) if additions else np.nan
            closing_nbv = opening_nbv + additions - depreciation
            asset_rows.append(
                {
                    "Asset": asset,
                    "Year": year,
                    "Opening NBV": opening_nbv,
                    "Additions": additions,
                    "Depreciation Rate %": rate,
                    "Depreciation": depreciation,
                    "Closing NBV": closing_nbv,
                }
            )
            opening_nbv = closing_nbv

    return _ensure_asset_schedule(pd.DataFrame(asset_rows, columns=ASSET_SCHEDULE_COLUMNS))


def _capex_asset_reconciliation_issues(
    capex_table: Optional[pd.DataFrame],
    asset_table: Optional[pd.DataFrame],
) -> list[str]:
    issues: list[str] = []
    capex = _ensure_capex_schedule(capex_table)
    assets = _ensure_asset_schedule(asset_table)

    if capex.empty and assets.empty:
        return issues

    capex_work = capex.copy()
    capex_work["Category"] = capex_work["Category"].fillna("").astype(str).str.strip()
    capex_work["Year"] = pd.to_numeric(capex_work["Year"], errors="coerce").round().astype("Int64")
    capex_work["Spend"] = pd.to_numeric(capex_work["Spend"], errors="coerce").fillna(0.0)
    capex_work["Depreciation"] = pd.to_numeric(capex_work["Depreciation"], errors="coerce").fillna(0.0)
    capex_work = capex_work.loc[capex_work["Category"].ne("") & capex_work["Year"].notna()].copy()
    capex_summary = (
        capex_work.groupby(["Category", "Year"], as_index=False)
        .agg(
            **{
                "Capex Spend": ("Spend", "sum"),
                "Capex Depreciation": ("Depreciation", "sum"),
            }
        )
        if not capex_work.empty
        else pd.DataFrame(columns=["Category", "Year", "Capex Spend", "Capex Depreciation"])
    )

    asset_work = assets.copy()
    asset_work["Asset"] = asset_work["Asset"].fillna("").astype(str).str.strip()
    asset_work["Year"] = pd.to_numeric(asset_work["Year"], errors="coerce").round().astype("Int64")
    asset_work["Additions"] = pd.to_numeric(asset_work["Additions"], errors="coerce").fillna(0.0)
    asset_work["Depreciation"] = pd.to_numeric(asset_work["Depreciation"], errors="coerce").fillna(0.0)
    asset_work = asset_work.loc[asset_work["Asset"].ne("") & asset_work["Year"].notna()].copy()
    asset_summary = (
        asset_work.groupby(["Asset", "Year"], as_index=False)
        .agg(
            **{
                "Asset Additions": ("Additions", "sum"),
                "Asset Depreciation": ("Depreciation", "sum"),
            }
        )
        if not asset_work.empty
        else pd.DataFrame(columns=["Asset", "Year", "Asset Additions", "Asset Depreciation"])
    )
    if not asset_summary.empty:
        asset_summary = asset_summary.rename(columns={"Asset": "Category"})

    merged = capex_summary.merge(
        asset_summary,
        how="outer",
        on=["Category", "Year"],
    ).fillna(0.0)

    spend_mismatch = merged.loc[
        ~np.isclose(
            pd.to_numeric(merged["Capex Spend"], errors="coerce").fillna(0.0),
            pd.to_numeric(merged["Asset Additions"], errors="coerce").fillna(0.0),
            atol=1e-6,
            rtol=0.0,
        )
    ]
    if not spend_mismatch.empty:
        issues.append("Asset Schedule additions do not reconcile to Capex Schedule spend by year/category.")

    dep_mismatch = merged.loc[
        ~np.isclose(
            pd.to_numeric(merged["Capex Depreciation"], errors="coerce").fillna(0.0),
            pd.to_numeric(merged["Asset Depreciation"], errors="coerce").fillna(0.0),
            atol=1e-6,
            rtol=0.0,
        )
    ]
    if not dep_mismatch.empty:
        issues.append("Asset Schedule depreciation does not reconcile to Capex Schedule depreciation by year/category.")

    for asset, asset_group in asset_work.groupby("Asset", sort=False):
        ordered = asset_group.sort_values("Year", kind="stable").reset_index(drop=True)
        previous_closing = None
        for _, row in ordered.iterrows():
            opening = float(pd.to_numeric(pd.Series([row.get("Opening NBV")]), errors="coerce").iloc[0] or 0.0)
            closing = float(pd.to_numeric(pd.Series([row.get("Closing NBV")]), errors="coerce").iloc[0] or 0.0)
            additions = float(pd.to_numeric(pd.Series([row.get("Additions")]), errors="coerce").iloc[0] or 0.0)
            depreciation = float(pd.to_numeric(pd.Series([row.get("Depreciation")]), errors="coerce").iloc[0] or 0.0)
            expected_closing = opening + additions - depreciation
            if not np.isclose(closing, expected_closing, atol=1e-6, rtol=0.0):
                issues.append(f"Asset Schedule row for {asset} does not satisfy Opening NBV + Additions - Depreciation = Closing NBV.")
                break
            if previous_closing is not None and not np.isclose(opening, previous_closing, atol=1e-6, rtol=0.0):
                issues.append(f"Asset Schedule opening NBV for {asset} does not roll forward from the prior closing NBV.")
                break
            previous_closing = closing

    return issues


def _apply_asset_rate(table: pd.DataFrame, rate: float) -> pd.DataFrame:
    work = _ensure_asset_schedule(table)
    work["Depreciation Rate %"] = rate
    return _recalculate_asset_schedule(work)


def _apply_asset_additions_pattern(
    table: pd.DataFrame, base_amount: float, increment_pct: float
) -> pd.DataFrame:
    work = _ensure_asset_schedule(table)
    if work.empty:
        return work

    temp = work.copy()
    temp["__year"] = pd.to_numeric(temp["Year"], errors="coerce")
    temp = temp.sort_values("__year", kind="stable")

    factor = 1 + (increment_pct / 100.0)
    position = 0
    for idx, row in temp.iterrows():
        year = row["__year"]
        if pd.isna(year):
            continue
        temp.at[idx, "Additions"] = base_amount * (factor ** position)
        position += 1

    temp = temp.drop(columns="__year")
    work.loc[temp.index, "Additions"] = temp["Additions"]
    return _recalculate_asset_schedule(work)


def _apply_asset_yearly_increment(
    table: pd.DataFrame, column: str, increment_pct: float
) -> pd.DataFrame:
    work = _ensure_asset_schedule(table)
    if work.empty or column not in work.columns or increment_pct == 0:
        return work

    temp = work.copy()
    temp["__year"] = pd.to_numeric(temp["Year"], errors="coerce")
    temp = temp.sort_values("__year", kind="stable")

    factor = 1 + (increment_pct / 100.0)
    base_value = None
    last_year = None

    for idx, row in temp.iterrows():
        year = row["__year"]
        if pd.isna(year):
            continue

        current = pd.to_numeric(row[column], errors="coerce")
        if base_value is None:
            base_value = 0.0 if pd.isna(current) else float(current)
            last_year = int(year)
            temp.at[idx, column] = base_value
            continue

        year_gap = int(year) - int(last_year)
        if year_gap > 0:
            base_value = base_value * (factor ** year_gap)
            last_year = int(year)

        temp.at[idx, column] = base_value

    temp = temp.drop(columns="__year")
    work.loc[temp.index, column] = temp[column]
    return _recalculate_asset_schedule(work)


def _add_asset_row(
    table: pd.DataFrame, default_rate: float = 5.0, default_additions: float = 0.0
) -> pd.DataFrame:
    work = _ensure_asset_schedule(table)

    if work.empty:
        next_year = pd.Timestamp.today().year
    else:
        years = pd.to_numeric(work["Year"], errors="coerce")
        next_year = int(years.dropna().max()) + 1 if years.notna().any() else pd.Timestamp.today().year

    new_row = {
        "Asset": "New Asset",
        "Year": next_year,
        "Opening NBV": 0.0,
        "Additions": default_additions,
        "Depreciation Rate %": default_rate,
        "Depreciation": np.nan,
        "Closing NBV": np.nan,
    }

    work = pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)
    return _recalculate_asset_schedule(work)


def _remove_asset_row(table: pd.DataFrame, index: int) -> pd.DataFrame:
    work = _ensure_asset_schedule(table)
    if index not in work.index:
        return work
    reduced = work.drop(index=index).reset_index(drop=True)
    return _recalculate_asset_schedule(reduced)


def _ensure_capitalisation_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        return pd.DataFrame({col: [] for col in CAP_TABLE_COLUMNS})

    work = table.copy()
    for column in CAP_TABLE_COLUMNS:
        if column not in work.columns:
            work[column] = np.nan

    work["Year"] = pd.to_numeric(work["Year"], errors="coerce").astype("Int64")
    shareholders = work.get("Shareholder")
    if shareholders is None:
        shareholders = pd.Series(["" for _ in range(len(work))])
    else:
        shareholders = shareholders.astype("string")
    work["Shareholder"] = shareholders.fillna("").replace({pd.NA: "", "<NA>": "", "nan": ""})
    work["Ownership %"] = pd.to_numeric(work["Ownership %"], errors="coerce")
    work["Investment"] = pd.to_numeric(work["Investment"], errors="coerce")

    ordered = CAP_TABLE_COLUMNS + [col for col in work.columns if col not in CAP_TABLE_COLUMNS]
    return work[ordered].reset_index(drop=True)


def _add_capitalisation_row(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    work = _ensure_capitalisation_table(table)

    years = pd.to_numeric(work.get("Year"), errors="coerce")
    if years.notna().any():
        next_year = int(years.max()) + 1
    else:
        next_year = pd.Timestamp.today().year

    new_row = {
        "Year": next_year,
        "Shareholder": "New Investor",
        "Ownership %": np.nan,
        "Investment": np.nan,
    }

    combined = pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)
    return _ensure_capitalisation_table(combined)


def _remove_capitalisation_row(table: Optional[pd.DataFrame], index: int) -> pd.DataFrame:
    work = _ensure_capitalisation_table(table)
    if work.empty or index not in work.index:
        return work
    reduced = work.drop(index=index)
    return _ensure_capitalisation_table(reduced)


def _apply_capitalisation_increment(
    table: Optional[pd.DataFrame], column: str, increment_pct: float
) -> pd.DataFrame:
    if column not in {"Ownership %", "Investment"}:
        return _ensure_capitalisation_table(table)

    work = _ensure_capitalisation_table(table)
    if work.empty or increment_pct == 0:
        return work

    temp = work.copy()
    temp["__year"] = pd.to_numeric(temp["Year"], errors="coerce")
    temp = temp.sort_values(["Shareholder", "__year"], kind="stable")

    increment_factor = 1 + (increment_pct / 100.0)

    last_values: Dict[str, tuple[Optional[float], Optional[float]]] = {}
    for idx, row in temp.iterrows():
        shareholder = row["Shareholder"] or "Unnamed"
        year = row["__year"]
        if pd.isna(year):
            continue

        previous_value, previous_year = last_values.get(shareholder, (None, None))
        current_value = row[column]

        if previous_year is None or pd.isna(previous_year):
            if pd.notna(current_value):
                last_values[shareholder] = (float(current_value), float(year))
            else:
                last_values[shareholder] = (None, float(year))
            continue

        years_elapsed = int(float(year) - previous_year)
        if years_elapsed <= 0:
            if pd.notna(current_value):
                last_values[shareholder] = (float(current_value), float(year))
            continue

        base_value = previous_value
        if base_value is None:
            if pd.isna(current_value):
                continue
            base_value = float(current_value)

        new_value = float(base_value) * (increment_factor ** years_elapsed)
        temp.at[idx, column] = new_value
        last_values[shareholder] = (new_value, float(year))

    temp = temp.drop(columns="__year")
    work.loc[temp.index, column] = temp[column]
    return _ensure_capitalisation_table(work)


def _ensure_capex_schedule(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = pd.DataFrame(columns=CAPEX_TABLE_COLUMNS)
    else:
        work = table.copy()

    for column in CAPEX_TABLE_COLUMNS:
        if column not in work.columns:
            work[column] = np.nan

    work["Category"] = work["Category"].fillna("").astype(str)
    work["Year"] = pd.to_numeric(work["Year"], errors="coerce").astype("Int64")

    work["Spend"] = pd.to_numeric(work["Spend"], errors="coerce")
    rate = pd.to_numeric(work["Depreciation Rate %"], errors="coerce")

    fraction_mask = rate.notna() & (rate.abs() <= 1.0)
    rate.loc[fraction_mask] = rate.loc[fraction_mask] * 100.0

    work["Depreciation Rate %"] = rate
    work["Depreciation"] = pd.to_numeric(work["Depreciation"], errors="coerce")

    ordered = [col for col in CAPEX_TABLE_COLUMNS if col in work.columns]
    remaining = [col for col in work.columns if col not in ordered]

    return _recalculate_capex_schedule(work[ordered + remaining].reset_index(drop=True))


def _recalculate_capex_schedule(table: pd.DataFrame) -> pd.DataFrame:
    if table is None or table.empty:
        return pd.DataFrame(columns=CAPEX_TABLE_COLUMNS)

    work = table.copy()

    spend = pd.to_numeric(work.get("Spend"), errors="coerce")
    rate = pd.to_numeric(work.get("Depreciation Rate %"), errors="coerce").fillna(0.0)

    depreciation = spend * (rate / 100.0)
    override = pd.to_numeric(work.get("Depreciation"), errors="coerce")
    depreciation = depreciation.where(override.isna(), override)

    work["Spend"] = spend
    work["Depreciation Rate %"] = rate
    work["Depreciation"] = depreciation

    return work


def _apply_capex_rate(table: Optional[pd.DataFrame], rate: float) -> pd.DataFrame:
    work = _ensure_capex_schedule(table)
    if work.empty:
        return work
    work["Depreciation Rate %"] = rate
    return _recalculate_capex_schedule(work)


def _apply_capex_yearly_increment(
    table: Optional[pd.DataFrame], column: str, increment_pct: float
) -> pd.DataFrame:
    work = _ensure_capex_schedule(table)
    if work.empty or column not in work.columns or increment_pct == 0:
        return work

    temp = work.copy()
    temp["__year"] = pd.to_numeric(temp["Year"], errors="coerce")
    temp = temp.sort_values(["__year"], kind="stable")

    factor = 1 + (increment_pct / 100.0)
    base_value = None
    last_year = None

    for idx, row in temp.iterrows():
        year = row["__year"]
        if pd.isna(year):
            continue

        current_value = pd.to_numeric(row[column], errors="coerce")
        if base_value is None:
            base_value = 0.0 if pd.isna(current_value) else float(current_value)
            last_year = int(year)
            temp.at[idx, column] = base_value
            continue

        year_gap = int(year) - int(last_year)
        if year_gap > 0:
            base_value = base_value * (factor ** year_gap)
            last_year = int(year)

        temp.at[idx, column] = base_value

    temp = temp.drop(columns="__year")
    work.loc[temp.index, column] = temp[column]
    return _recalculate_capex_schedule(work)


def _add_capex_row(
    table: Optional[pd.DataFrame], default_rate: float = 10.0, default_spend: float = 0.0
) -> pd.DataFrame:
    work = _ensure_capex_schedule(table)

    years = pd.to_numeric(work.get("Year"), errors="coerce")
    if years.notna().any():
        next_year = int(years.max()) + 1
    else:
        next_year = pd.Timestamp.today().year

    new_row = {
        "Year": next_year,
        "Category": "New Capex",
        "Spend": default_spend,
        "Depreciation Rate %": default_rate,
        "Depreciation": np.nan,
    }

    combined = pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)
    return _recalculate_capex_schedule(combined)


def _remove_capex_row(table: Optional[pd.DataFrame], index: int) -> pd.DataFrame:
    work = _ensure_capex_schedule(table)
    if work.empty or index not in work.index:
        return work
    reduced = work.drop(index=index).reset_index(drop=True)
    return _recalculate_capex_schedule(reduced)


def _normalize_period_type(value: Optional[str]) -> str:
    candidate = str(value or "").strip().lower()
    return "quarterly" if candidate == "quarterly" else "monthly"


def _period_end_offset(period_type: str) -> Union[MonthEnd, QuarterEnd]:
    return QuarterEnd(1) if _normalize_period_type(period_type) == "quarterly" else MonthEnd(1)


def _period_label(period_type: str) -> str:
    return "Quarterly" if _normalize_period_type(period_type) == "quarterly" else "Monthly"


def _period_index_for_horizon(start_year: int, end_year: int, period_type: str) -> pd.DatetimeIndex:
    normalized = _normalize_period_type(period_type)
    if normalized == "quarterly":
        start_date = pd.Timestamp(start_year, 3, 31)
        end_date = pd.Timestamp(end_year, 12, 31)
        return pd.date_range(start=start_date, end=end_date, freq=QuarterEnd())

    start_date = pd.Timestamp(start_year, 1, 1) + MonthEnd(0)
    end_date = pd.Timestamp(end_year, 12, 1) + MonthEnd(0)
    return pd.date_range(start=start_date, end=end_date, freq=MonthEnd())


def _infer_period_type_from_schedule(core_schedule: Optional[pd.DataFrame]) -> str:
    if not isinstance(core_schedule, pd.DataFrame) or core_schedule.empty:
        return "monthly"
    raw_periods = core_schedule.get("Period", pd.Series(dtype=str))
    periods = pd.to_datetime(raw_periods, errors="coerce")
    if not isinstance(periods, pd.Series):
        periods = pd.Series(periods)
    periods = periods.dropna()
    if len(periods) < 2:
        return "monthly"
    month_diffs = periods.sort_values().diff().dt.days.dropna()
    if month_diffs.empty:
        return "monthly"
    median_days = float(month_diffs.median())
    return "quarterly" if median_days >= 75 else "monthly"


def _next_period_from_last(last_period: Optional[pd.Timestamp], period_type: str) -> pd.Timestamp:
    normalized = _normalize_period_type(period_type)
    if last_period is None or pd.isna(last_period):
        return pd.Timestamp.today() + (QuarterEnd(0) if normalized == "quarterly" else MonthEnd(0))
    return last_period + _period_end_offset(normalized)


def _default_income_schedule(
    periods: int = 12, start: str = "2024-01-31", period_type: str = "monthly"
) -> pd.DataFrame:
    dates = pd.date_range(start, periods=periods, freq=_period_end_offset(period_type))
    revenue = np.linspace(45000, 70000, periods)
    cogs = revenue * 0.45
    variable = revenue * 0.12
    fixed = np.full(periods, 9000.0)
    direct = revenue * 0.08
    admin = np.full(periods, 3500.0)
    gross_margin = revenue - cogs
    ebitda = gross_margin - variable - fixed - direct - admin
    depreciation = np.full(periods, 2000.0)
    ebit = ebitda - depreciation
    interest_expense = np.full(periods, 1500.0)
    npbt = ebit - interest_expense
    tax_expense = npbt * 0.28
    npat = npbt - tax_expense
    cfo = ebitda - 2500.0
    cfi = np.full(periods, -3000.0)
    cff = np.full(periods, 1500.0)
    capex = np.full(periods, 2500.0)
    net_cf = cfo + cfi + cff
    opening_cash = np.empty(periods)
    opening_cash[0] = 25000.0
    for i in range(1, periods):
        opening_cash[i] = opening_cash[i - 1] + net_cf[i - 1]
    closing_cash = opening_cash + net_cf
    current_assets = np.linspace(95000, 120000, periods)
    non_current_assets = np.linspace(210000, 245000, periods)
    current_liabilities = np.linspace(40000, 45000, periods)
    non_current_liabilities = np.linspace(85000, 90000, periods)
    equity = current_assets + non_current_assets - current_liabilities - non_current_liabilities

    df = pd.DataFrame(
        {
            "Period": dates.strftime("%Y-%m-%d"),
            "Revenue": revenue,
            "COGS": cogs,
            "Variable Expenses": variable,
            "Fixed Expenses": fixed,
            "Direct Wages": direct,
            "Admin Wages": admin,
            "Gross Margin": gross_margin,
            "EBITDA": ebitda,
            "Depreciation & Amortization": depreciation,
            "EBIT": ebit,
            "Interest Expense": interest_expense,
            "NPBT": npbt,
            "Tax Expense": tax_expense,
            "NPAT": npat,
            "CFO": cfo,
            "CFI": cfi,
            "CFF": cff,
            "Capex": capex,
            "Net Cash Flow": net_cf,
            "Opening Cash Balance": opening_cash,
            "Closing Cash Balance": closing_cash,
            "Cash and Cash Equivalents": closing_cash,
            "Current Assets": current_assets,
            "Non-current Assets": non_current_assets,
            "Current Liabilities": current_liabilities,
            "Non-current Liabilities": non_current_liabilities,
            "Equity": equity,
        }
    )
    return df


def _derive_horizon_years(
    production_horizon: Optional[pd.DataFrame],
) -> tuple[int, int]:
    """Return start and end years inferred from the production horizon table."""

    default_start = 2024
    default_end = 2024

    if production_horizon is None or production_horizon.empty:
        return default_start, default_end

    start_years = pd.to_numeric(
        production_horizon.get("Start Year"), errors="coerce"
    ).dropna()
    end_years = pd.to_numeric(
        production_horizon.get("End Year"), errors="coerce"
    ).dropna()

    start_year = int(start_years.iloc[0]) if not start_years.empty else default_start
    end_year = int(end_years.iloc[0]) if not end_years.empty else default_end

    if end_year < start_year:
        end_year = start_year

    return start_year, end_year


def _default_schedule_components(
    periods: Optional[int] = None,
    start: Optional[str] = None,
    production_horizon: Optional[pd.DataFrame] = None,
    period_type: str = "monthly",
    assumptions: Optional[Dict[str, pd.DataFrame]] = None,
) -> tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    if production_horizon is None:
        production_horizon = _default_production_horizon_table()

    start_year, end_year = _derive_horizon_years(production_horizon)
    normalized_period_type = _normalize_period_type(period_type)

    if periods is None:
        per_year = 4 if normalized_period_type == "quarterly" else 12
        periods = max(1, (end_year - start_year + 1) * per_year)

    if start is None:
        if normalized_period_type == "quarterly":
            start_date = pd.Timestamp(start_year, 3, 31)
        else:
            start_date = pd.Timestamp(start_year, 1, 1) + MonthEnd(0)
        start = start_date.strftime("%Y-%m-%d")

    base = _default_income_schedule(
        periods=periods, start=start, period_type=normalized_period_type
    )

    core_columns = [
        "Period",
        "Revenue",
        "Variable Expenses",
        "Direct Wages",
        "Admin Wages",
        "Fixed Expenses",
        "Depreciation & Amortization",
        "EBIT",
        "Interest Expense",
        "NPBT",
        "Tax Expense",
        "NPAT",
        "CFO",
        "CFI",
        "CFF",
        "Net Cash Flow",
        "Opening Cash Balance",
        "Closing Cash Balance",
        "Cash and Cash Equivalents",
        "Current Assets",
        "Non-current Assets",
        "Current Liabilities",
        "Non-current Liabilities",
        "Equity",
    ]
    core_columns = [col for col in core_columns if col in base.columns]
    core = base[core_columns].copy()

    assumption_map = assumptions or {}
    variable_inputs = _ensure_variable_expense_input_table(
        assumption_map.get("Variable Expenses")
    )
    direct_inputs = _ensure_direct_wage_input_table(
        assumption_map.get("Direct Wages")
    )
    admin_inputs = _ensure_admin_wage_input_table(
        assumption_map.get("Admin Wages")
    )

    detail_tables: Dict[str, pd.DataFrame] = {}
    for name, cols in DETAIL_SCHEDULE_COLUMNS.items():
        if name == "Variable Expenses Schedule":
            detail_tables[name] = _propagate_variable_expense_inputs_to_schedule(
                variable_inputs, base
            )
            continue
        if name == "Direct Wages Schedule":
            detail_tables[name] = _propagate_direct_wage_inputs_to_schedule(
                direct_inputs, base
            )
            continue
        if name == "Admin Wages Schedule":
            detail_tables[name] = _propagate_admin_wage_inputs_to_schedule(
                admin_inputs, base
            )
            continue
        detail_cols = ["Period"] + [col for col in cols if col in base.columns]
        detail_tables[name] = base[detail_cols].copy()

    return core, detail_tables


def _default_supplementary_tables() -> Dict[str, pd.DataFrame]:
    capex_schedule = pd.DataFrame(
        {
            "Year": [2024, 2025],
            "Category": ["Production Equipment", "Housing Upgrades"],
            "Spend": [500000.0, 175000.0],
            "Depreciation Rate %": [8.0, 6.5],
            "Depreciation": [40000.0, 11375.0],
        }
    )
    return {
        "Loan Facilities": _default_loan_facilities_table(),
        "Equity Facilities": _default_equity_facilities_table(),
        "Capitalisation Table": pd.DataFrame(
            {
                "Shareholder": ["Founder", "Investor"],
                "Ownership %": [60.0, 40.0],
                "Investment": [0.0, 250000.0],
            }
        ),
        "Capex Schedule": capex_schedule,
        "Asset Schedules": _derive_asset_schedule_from_capex(capex_schedule),
        "Outputs": pd.DataFrame(
            {
                "Metric": ["IRR", "Payback Period (Years)"],
                "Value": [0.17, 4.2],
            }
        ),
        "Benchmark KPIs": pd.DataFrame(
            {
                "KPI": ["Revenue per Herd Head", "Feed Cost per Herd Head"],
                "Benchmark": [1500.0, 320.0],
            }
        ),
    }


def _sync_asset_schedule_from_capex_in_supplementary(
    supplementary: Optional[Dict[str, pd.DataFrame]],
) -> Dict[str, pd.DataFrame]:
    synced = dict(supplementary or {})
    capex = _ensure_capex_schedule(synced.get("Capex Schedule"))
    synced["Capex Schedule"] = capex
    synced["Asset Schedules"] = _derive_asset_schedule_from_capex(capex)
    return synced


def _default_scenario_controls_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Driver": list(DEFAULT_SCENARIO_ADJUSTMENTS.keys()),
            "Change %": list(DEFAULT_SCENARIO_ADJUSTMENTS.values()),
        }
    )


def _default_business_configuration_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Business Type": [DEFAULT_BUSINESS_TYPE],
            "Operating Model": [DEFAULT_OPERATING_MODEL],
            "Transfer Destination": [""],
            "Reporting View": [DEFAULT_REPORTING_VIEW],
            "Transfer Pricing Method": [DEFAULT_TRANSFER_PRICING_METHOD],
            "Allow External Kid Sales": [True],
        }
    )


def _ensure_business_configuration_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_business_configuration_table()
    else:
        work = table.copy()

    if "Business Type" not in work.columns:
        work["Business Type"] = DEFAULT_BUSINESS_TYPE
    if "Operating Model" not in work.columns:
        work["Operating Model"] = DEFAULT_OPERATING_MODEL
    if "Transfer Destination" not in work.columns:
        work["Transfer Destination"] = ""
    if "Reporting View" not in work.columns:
        work["Reporting View"] = DEFAULT_REPORTING_VIEW
    if "Transfer Pricing Method" not in work.columns:
        work["Transfer Pricing Method"] = DEFAULT_TRANSFER_PRICING_METHOD
    if "Allow External Kid Sales" not in work.columns:
        work["Allow External Kid Sales"] = True
    work["Business Type"] = _series_or_default(work, "Business Type", DEFAULT_BUSINESS_TYPE).apply(_normalize_business_type)
    work["Operating Model"] = _series_or_default(work, "Operating Model", DEFAULT_OPERATING_MODEL).apply(
        _normalize_operating_model
    )
    work["Transfer Destination"] = _series_or_default(work, "Transfer Destination", "").astype(str).str.strip()
    work["Transfer Destination"] = work["Transfer Destination"].where(
        work["Transfer Destination"].isin(["", "Meat", "Milk-Cheese", "Combined"]),
        "",
    )
    work["Reporting View"] = _series_or_default(work, "Reporting View", DEFAULT_REPORTING_VIEW).apply(
        _normalize_reporting_view
    )
    work["Transfer Pricing Method"] = _series_or_default(
        work, "Transfer Pricing Method", DEFAULT_TRANSFER_PRICING_METHOD
    ).apply(_normalize_transfer_pricing_method)
    work["Allow External Kid Sales"] = _series_or_default(work, "Allow External Kid Sales", True).map(
        lambda value: _coerce_bool_value(value, True)
    )
    ordered_cols = [
        "Business Type",
        "Operating Model",
        "Transfer Destination",
        "Reporting View",
        "Transfer Pricing Method",
        "Allow External Kid Sales",
    ]
    return work[ordered_cols].head(1).reset_index(drop=True)


def _selected_business_type(table_or_assumptions: Any) -> str:
    if isinstance(table_or_assumptions, dict):
        table = _ensure_business_configuration_table(table_or_assumptions.get("Business Configuration"))
    else:
        table = _ensure_business_configuration_table(table_or_assumptions)
    value = table.iloc[0].get("Business Type") if not table.empty else DEFAULT_BUSINESS_TYPE
    return _normalize_business_type(value)


def _selected_operating_model(table_or_assumptions: Any) -> str:
    if isinstance(table_or_assumptions, dict):
        table = _ensure_business_configuration_table(table_or_assumptions.get("Business Configuration"))
    else:
        table = _ensure_business_configuration_table(table_or_assumptions)
    value = table.iloc[0].get("Operating Model") if not table.empty else DEFAULT_OPERATING_MODEL
    return _normalize_operating_model(value)


def _selected_transfer_destination(table_or_assumptions: Any) -> str:
    if isinstance(table_or_assumptions, dict):
        table = _ensure_business_configuration_table(table_or_assumptions.get("Business Configuration"))
    else:
        table = _ensure_business_configuration_table(table_or_assumptions)
    value = str(table.iloc[0].get("Transfer Destination", "")).strip() if not table.empty else ""
    return value if value in {"", "Meat", "Milk-Cheese", "Combined"} else ""


def _selected_reporting_view(table_or_assumptions: Any) -> str:
    if isinstance(table_or_assumptions, dict):
        table = _ensure_business_configuration_table(table_or_assumptions.get("Business Configuration"))
    else:
        table = _ensure_business_configuration_table(table_or_assumptions)
    value = table.iloc[0].get("Reporting View") if not table.empty else DEFAULT_REPORTING_VIEW
    return _normalize_reporting_view(value)


def _selected_transfer_pricing_method(table_or_assumptions: Any) -> str:
    if isinstance(table_or_assumptions, dict):
        table = _ensure_business_configuration_table(table_or_assumptions.get("Business Configuration"))
    else:
        table = _ensure_business_configuration_table(table_or_assumptions)
    value = (
        table.iloc[0].get("Transfer Pricing Method")
        if not table.empty
        else DEFAULT_TRANSFER_PRICING_METHOD
    )
    return _normalize_transfer_pricing_method(value)


def _allow_external_kid_sales(table_or_assumptions: Any) -> bool:
    if isinstance(table_or_assumptions, dict):
        table = _ensure_business_configuration_table(table_or_assumptions.get("Business Configuration"))
    else:
        table = _ensure_business_configuration_table(table_or_assumptions)
    value = table.iloc[0].get("Allow External Kid Sales", True) if not table.empty else True
    return _coerce_bool_value(value, True)


def _is_breeding_to_unit_mode(table_or_assumptions: Any) -> bool:
    return _selected_operating_model(table_or_assumptions) == "Breeding-to-Unit"


def _default_kid_routing_rules_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Period": [None, None],
            "Sex": ["Female", "Male"],
            "Transfer Class": ["Weaned Kid", "Weaned Kid"],
            "Destination": ["External Sale", "External Sale"],
            "Allocation %": [100.0, 100.0],
            "Priority": [1, 1],
            "Min Head": [0.0, 0.0],
            "Max Head": [np.nan, np.nan],
            "Active": [True, True],
        }
    )


def _ensure_kid_routing_rules_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_kid_routing_rules_table()
    else:
        work = table.copy()
    required_cols = [
        "Period",
        "Sex",
        "Transfer Class",
        "Destination",
        "Allocation %",
        "Priority",
        "Min Head",
        "Max Head",
        "Active",
    ]
    for col in required_cols:
        if col not in work.columns:
            work[col] = np.nan
    work["Period"] = _normalize_period(work.get("Period", pd.Series(dtype=str)))
    work["Sex"] = _series_or_default(work, "Sex", "Female").astype(str).str.strip().replace({"": "Female"})
    work["Transfer Class"] = _series_or_default(work, "Transfer Class", "Weaned Kid").astype(str).str.strip()
    work["Transfer Class"] = work["Transfer Class"].replace({"": "Weaned Kid"})
    work["Destination"] = _series_or_default(work, "Destination", "External Sale").astype(str).str.strip()
    work["Destination"] = work["Destination"].replace({"": "External Sale"})
    for col in ["Allocation %", "Priority", "Min Head", "Max Head"]:
        work[col] = pd.to_numeric(work.get(col), errors="coerce")
    work["Allocation %"] = work["Allocation %"].fillna(0.0)
    work["Priority"] = work["Priority"].fillna(1.0)
    work["Min Head"] = work["Min Head"].fillna(0.0)
    work["Active"] = _series_or_default(work, "Active", True).map(lambda value: _coerce_bool_value(value, True))
    work = work.dropna(how="all")
    if work.empty:
        return _default_kid_routing_rules_table()
    remainder = [col for col in work.columns if col not in required_cols]
    return work[required_cols + remainder].reset_index(drop=True)


def _default_internal_transfer_pricing_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Period": [None],
            "Transfer Class": ["Weaned Kid"],
            "Destination": ["Meat"],
            "Pricing Method": [DEFAULT_TRANSFER_PRICING_METHOD],
            "Transfer Price per Head": [150.0],
            "Markup %": [0.0],
            "Active": [True],
        }
    )


def _ensure_internal_transfer_pricing_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_internal_transfer_pricing_table()
    else:
        work = table.copy()
    required_cols = [
        "Period",
        "Transfer Class",
        "Destination",
        "Pricing Method",
        "Transfer Price per Head",
        "Markup %",
        "Active",
    ]
    for col in required_cols:
        if col not in work.columns:
            work[col] = np.nan
    work["Period"] = _normalize_period(work.get("Period", pd.Series(dtype=str)))
    work["Transfer Class"] = _series_or_default(work, "Transfer Class", "Weaned Kid").astype(str).str.strip()
    work["Transfer Class"] = work["Transfer Class"].replace({"": "Weaned Kid"})
    work["Destination"] = _series_or_default(work, "Destination", "Meat").astype(str).str.strip()
    work["Pricing Method"] = _series_or_default(
        work, "Pricing Method", DEFAULT_TRANSFER_PRICING_METHOD
    ).apply(_normalize_transfer_pricing_method)
    work["Transfer Price per Head"] = pd.to_numeric(work.get("Transfer Price per Head"), errors="coerce")
    work["Markup %"] = pd.to_numeric(work.get("Markup %"), errors="coerce").fillna(0.0)
    work["Active"] = _series_or_default(work, "Active", True).map(lambda value: _coerce_bool_value(value, True))
    work = work.dropna(how="all")
    if work.empty:
        return _default_internal_transfer_pricing_table()
    remainder = [col for col in work.columns if col not in required_cols]
    return work[required_cols + remainder].reset_index(drop=True)


def _default_downstream_intake_rules_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Destination": ["Meat", "Milk-Cheese", "Combined", "Combined"],
            "Sex": ["Female", "Female", "Female", "Male"],
            "Entry Age Months": [3.0, 3.0, 3.0, 3.0],
            "Entry Weight Kg": [12.0, 12.0, 12.0, 12.0],
            "Mortality %": [3.5, 2.5, 2.5, 3.5],
            "Eligible for Lactation": [False, True, True, False],
            "Eligible for Finishing": [True, False, True, True],
            "Active": [True, True, True, True],
        }
    )


def _ensure_downstream_intake_rules_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_downstream_intake_rules_table()
    else:
        work = table.copy()
    required_cols = [
        "Destination",
        "Sex",
        "Entry Age Months",
        "Entry Weight Kg",
        "Mortality %",
        "Eligible for Lactation",
        "Eligible for Finishing",
        "Active",
    ]
    for col in required_cols:
        if col not in work.columns:
            work[col] = np.nan
    work["Destination"] = _series_or_default(work, "Destination", "Meat").astype(str).str.strip()
    work["Sex"] = _series_or_default(work, "Sex", "Female").astype(str).str.strip()
    for col in ["Entry Age Months", "Entry Weight Kg", "Mortality %"]:
        work[col] = pd.to_numeric(work.get(col), errors="coerce")
    for col in ["Eligible for Lactation", "Eligible for Finishing", "Active"]:
        work[col] = _series_or_default(work, col, True).map(lambda value: _coerce_bool_value(value, True))
    work = work.dropna(how="all")
    if work.empty:
        return _default_downstream_intake_rules_table()
    remainder = [col for col in work.columns if col not in required_cols]
    return work[required_cols + remainder].reset_index(drop=True)


def _default_transfer_elimination_rules_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Destination": ["Meat", "Milk-Cheese", "Combined"],
            "Eliminate Internal Revenue": [True, True, True],
            "Eliminate Internal Cost": [True, True, True],
            "Active": [True, True, True],
        }
    )


def _ensure_transfer_elimination_rules_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_transfer_elimination_rules_table()
    else:
        work = table.copy()
    required_cols = [
        "Destination",
        "Eliminate Internal Revenue",
        "Eliminate Internal Cost",
        "Active",
    ]
    for col in required_cols:
        if col not in work.columns:
            work[col] = np.nan
    work["Destination"] = _series_or_default(work, "Destination", "Meat").astype(str).str.strip()
    for col in ["Eliminate Internal Revenue", "Eliminate Internal Cost", "Active"]:
        work[col] = _series_or_default(work, col, True).map(lambda value: _coerce_bool_value(value, True))
    work = work.dropna(how="all")
    if work.empty:
        return _default_transfer_elimination_rules_table()
    remainder = [col for col in work.columns if col not in required_cols]
    return work[required_cols + remainder].reset_index(drop=True)


def _sync_transfer_tables_to_business_configuration(
    assumptions: Optional[Dict[str, pd.DataFrame]],
) -> Dict[str, pd.DataFrame]:
    synced = dict(assumptions or {})
    business_config = _ensure_business_configuration_table(synced.get("Business Configuration"))
    destination = _selected_transfer_destination(business_config)
    pricing_method = _selected_transfer_pricing_method(business_config)
    allow_external = _allow_external_kid_sales(business_config)

    routing = _ensure_kid_routing_rules_table(synced.get("Kid Routing Rules"))
    if destination:
        default_destination = destination
        if routing.empty or not routing["Active"].fillna(True).astype(bool).any():
            routing = _default_kid_routing_rules_table()
        if "Destination" in routing.columns:
            no_internal_destinations = ~routing["Destination"].isin(["Meat", "Milk-Cheese", "Combined"])
            if no_internal_destinations.all():
                routing = routing.copy()
                female_mask = routing["Sex"].astype(str).str.strip().str.casefold() == "female"
                male_mask = routing["Sex"].astype(str).str.strip().str.casefold() == "male"
                if female_mask.any():
                    routing.loc[female_mask, "Destination"] = default_destination
                    routing.loc[female_mask, "Allocation %"] = 100.0
                if male_mask.any():
                    routing.loc[male_mask, "Destination"] = (
                        "External Sale" if allow_external else default_destination
                    )
                    routing.loc[male_mask, "Allocation %"] = 100.0
    synced["Kid Routing Rules"] = _ensure_kid_routing_rules_table(routing)

    transfer_pricing = _ensure_internal_transfer_pricing_table(synced.get("Internal Transfer Pricing"))
    if destination:
        transfer_pricing = transfer_pricing.copy()
        transfer_pricing["Destination"] = destination
        transfer_pricing["Pricing Method"] = pricing_method
    synced["Internal Transfer Pricing"] = _ensure_internal_transfer_pricing_table(transfer_pricing)
    synced["Downstream Intake Rules"] = _ensure_downstream_intake_rules_table(
        synced.get("Downstream Intake Rules")
    )
    synced["Transfer Elimination Rules"] = _ensure_transfer_elimination_rules_table(
        synced.get("Transfer Elimination Rules")
    )
    synced["Business Configuration"] = business_config
    return synced


def _ensure_scenario_controls_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_scenario_controls_table()
    else:
        work = table.copy()

    if "Driver" not in work.columns:
        work["Driver"] = ""
    work["Driver"] = _series_or_default(work, "Driver", "").astype(str).str.strip()
    work.loc[work["Driver"] == "", "Driver"] = "Driver"

    if "Change %" not in work.columns:
        work["Change %"] = np.nan
    work["Change %"] = pd.to_numeric(work.get("Change %"), errors="coerce")

    ordered_cols = ["Driver", "Change %"]
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _sync_scenario_controls_to_products(
    table: Optional[pd.DataFrame],
    products: Sequence[str],
) -> pd.DataFrame:
    work = _ensure_scenario_controls_table(table)
    active_products = [str(product).strip() for product in products if str(product).strip()]
    active_drivers = {
        _price_change_driver(product) for product in active_products
    } | {
        _quantity_change_driver(product) for product in active_products
    }
    all_product_drivers = {
        _price_change_driver(product) for product in ALL_MODEL_PRODUCTS
    } | {
        _quantity_change_driver(product) for product in ALL_MODEL_PRODUCTS
    }

    keep_mask = work["Driver"].astype(str).apply(
        lambda value: value not in all_product_drivers or value in active_drivers
    )
    filtered = work.loc[keep_mask].copy()
    filtered_drivers = set(filtered["Driver"].astype(str))
    biological_drivers = list(BIOLOGICAL_SCENARIO_DRIVER_DEFAULTS.keys())
    for driver in list(active_drivers) + biological_drivers + ["Feed cost change (%)"]:
        if driver not in filtered_drivers:
            default_value = float(DEFAULT_SCENARIO_ADJUSTMENTS.get(driver, 0.0))
            filtered = pd.concat(
                [
                    filtered,
                    pd.DataFrame({"Driver": [driver], "Change %": [default_value]}),
                ],
                ignore_index=True,
            )

    driver_order = [
        *[_price_change_driver(product) for product in active_products],
        *[_quantity_change_driver(product) for product in active_products],
        *biological_drivers,
        "Feed cost change (%)",
    ]
    order_map = {driver: idx for idx, driver in enumerate(driver_order)}
    filtered["__sort_order"] = filtered["Driver"].map(order_map).fillna(len(order_map))
    filtered = filtered.sort_values(["__sort_order", "Driver"], kind="stable").drop(columns="__sort_order")
    return _ensure_scenario_controls_table(filtered)


def _scenario_controls_value_map(table: pd.DataFrame) -> Dict[str, float]:
    work = _ensure_scenario_controls_table(table)
    values: Dict[str, float] = {}
    for _, row in work.iterrows():
        driver = str(row.get("Driver", "")).strip()
        change = pd.to_numeric(pd.Series([row.get("Change %")]), errors="coerce").iloc[0]
        if driver:
            values[driver] = float(change) if not pd.isna(change) else 0.0
    return values


def _update_scenario_control_value(
    table: pd.DataFrame, driver: str, value: float
) -> pd.DataFrame:
    work = _ensure_scenario_controls_table(table)
    driver_key = str(driver).strip()
    if not driver_key:
        return work

    mask = work["Driver"].str.casefold() == driver_key.casefold()
    if mask.any():
        work.loc[mask, "Change %"] = float(value)
    else:
        work = pd.concat(
            [
                work,
                pd.DataFrame({"Driver": [driver_key], "Change %": [float(value)]}),
            ],
            ignore_index=True,
        )
    return _ensure_scenario_controls_table(work)


def _default_production_horizon_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Start Year": [2024],
            "End Year": [2030],
        }
    )


def _ensure_production_horizon_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_production_horizon_table()
    else:
        work = table.copy()

    for column in ["Start Year", "End Year"]:
        if column not in work.columns:
            work[column] = np.nan
        work[column] = pd.to_numeric(work.get(column), errors="coerce")

    work = work.dropna(how="all")
    if work.empty:
        return _default_production_horizon_table()

    defaults = _default_production_horizon_table().iloc[0]
    work["Start Year"] = work["Start Year"].fillna(defaults["Start Year"])
    work["End Year"] = work["End Year"].fillna(defaults["End Year"])

    for column in ["Start Year", "End Year"]:
        work[column] = work[column].round().astype("Int64")

    ordered_cols = ["Start Year", "End Year"]
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _opening_biological_start_date_for_horizon(
    production_horizon: Optional[pd.DataFrame],
    period_type: str = "monthly",
) -> str:
    start_year, end_year = _derive_horizon_years(
        _ensure_production_horizon_table(production_horizon)
    )
    period_index = _period_index_for_horizon(start_year, end_year, period_type)
    if period_index.empty:
        aligned_start = pd.Timestamp(start_year, 1, 1) + MonthEnd(0)
    else:
        aligned_start = pd.Timestamp(period_index.min())
    return aligned_start.strftime("%Y-%m-%d")


def _align_biological_system_settings_to_horizon(
    table: Optional[pd.DataFrame],
    production_horizon: Optional[pd.DataFrame] = None,
    period_type: str = "monthly",
) -> pd.DataFrame:
    work = ensure_table(BIOLOGICAL_SYSTEM_SETTINGS_SCHEMA, table)
    aligned_start = _opening_biological_start_date_for_horizon(
        production_horizon,
        period_type=period_type,
    )
    setting_key = "Opening Biological Start Date"
    mask = work["Setting"].astype(str).str.strip().eq(setting_key)
    if mask.any():
        work.loc[mask, "Value"] = aligned_start
    else:
        work = pd.concat(
            [
                work,
                pd.DataFrame([{"Setting": setting_key, "Value": aligned_start}]),
            ],
            ignore_index=True,
        )
    return _normalize_biological_system_settings_table(work)


def _sync_biological_start_date_in_assumptions(
    assumptions: Optional[Dict[str, pd.DataFrame]],
    period_type: str = "monthly",
) -> Dict[str, pd.DataFrame]:
    synced = dict(assumptions or {})
    production_horizon = _ensure_production_horizon_table(
        synced.get("Production Horizon")
    )
    synced["Production Horizon"] = production_horizon
    synced["Biological System Settings"] = _align_biological_system_settings_to_horizon(
        synced.get("Biological System Settings"),
        production_horizon=production_horizon,
        period_type=period_type,
    )
    return synced


def _sync_opening_herd_cohorts_in_assumptions(
    assumptions: Optional[Dict[str, pd.DataFrame]],
) -> Dict[str, pd.DataFrame]:
    synced = dict(assumptions or {})
    herd_plan = _ensure_herd_plan_table(synced.get("Herd Plan"))
    synced["Herd Plan"] = herd_plan
    synced["Opening Herd Cohorts"] = _sync_opening_herd_cohorts_to_herd_plan(
        synced.get("Opening Herd Cohorts"),
        herd_plan,
    )
    return synced


def _save_herd_plan_and_sync_opening_cohorts(table: pd.DataFrame) -> None:
    assumptions = dict(st.session_state.get("assumptions", {}))
    assumptions["Herd Plan"] = _ensure_herd_plan_table(table)
    assumptions = _sync_opening_herd_cohorts_in_assumptions(assumptions)
    if (
        _frames_equal(st.session_state.get("assumptions", {}).get("Herd Plan"), assumptions.get("Herd Plan"))
        and _frames_equal(
            st.session_state.get("assumptions", {}).get("Opening Herd Cohorts"),
            assumptions.get("Opening Herd Cohorts"),
        )
    ):
        return
    st.session_state.assumptions = assumptions
    _reset_cached_results()


def _production_year_options(start_year: int, end_year: int) -> list[int]:
    """Return a flexible list of year options covering the supplied range."""

    current_year = pd.Timestamp.today().year
    minimum = min(start_year, end_year, current_year - 10, 2000)
    maximum = max(start_year, end_year, current_year + 20)
    return list(range(minimum, maximum + 1))


def _safe_timeline(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Convert a schedule dataframe into a timeline index, ignoring errors."""

    if isinstance(table, pd.DataFrame) and not table.empty:
        try:
            return _prepare_timeline_table(table)
        except ValueError:
            return pd.DataFrame(index=pd.DatetimeIndex([], name="Period"))
    return pd.DataFrame(index=pd.DatetimeIndex([], name="Period"))


def _timeline_to_schedule_frame(
    timeline: pd.DataFrame, column_order: Optional[Sequence[str]] = None
) -> pd.DataFrame:
    """Convert a datetime-indexed timeline back into a schedule dataframe."""

    if timeline.empty:
        base_columns: list[str] = ["Period"]
        if column_order:
            base_columns.extend(
                [col for col in column_order if col not in {"Period"}]
            )
        return pd.DataFrame(columns=base_columns)

    ordered = timeline.sort_index()
    frame = ordered.reset_index()
    first_column = frame.columns[0]
    if first_column != "Period":
        frame = frame.rename(columns={first_column: "Period"})

    frame["Period"] = _normalize_period(frame.get("Period", pd.Series(dtype=str)))

    if column_order:
        normalized_order = [
            col for col in column_order if col not in {"Period"} and col in frame.columns
        ]
    else:
        normalized_order = []

    remainder = [
        col
        for col in frame.columns
        if col not in {"Period", *normalized_order}
    ]
    return frame[["Period", *normalized_order, *remainder]]


def _merge_schedule_table(
    existing: Optional[pd.DataFrame],
    defaults: Optional[pd.DataFrame],
    period_index: pd.DatetimeIndex,
) -> pd.DataFrame:
    """Merge an existing schedule with defaults to match the target horizon."""

    column_order: list[str] = []
    if isinstance(existing, pd.DataFrame) and not existing.empty:
        column_order = list(existing.columns)
    elif isinstance(defaults, pd.DataFrame) and not defaults.empty:
        column_order = list(defaults.columns)

    existing_timeline = _safe_timeline(existing)
    default_timeline = _safe_timeline(defaults)

    aligned_defaults = default_timeline.reindex(period_index)
    aligned_existing = existing_timeline.reindex(period_index)

    merged = aligned_defaults.copy()
    if not aligned_existing.empty:
        merged = merged.combine_first(aligned_existing)
        merged.update(aligned_existing)

    merged = merged.reindex(period_index)
    return _timeline_to_schedule_frame(merged, column_order)


def _rebase_schedule_to_horizon(
    core: Optional[pd.DataFrame],
    detail_tables: Optional[Dict[str, pd.DataFrame]],
    start_year: int,
    end_year: int,
    period_type: str = "monthly",
    assumptions: Optional[Dict[str, pd.DataFrame]] = None,
) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """Return schedule tables that span the requested production horizon."""

    if start_year > end_year:
        start_year, end_year = end_year, start_year

    horizon_table = pd.DataFrame({"Start Year": [start_year], "End Year": [end_year]})
    default_core, default_details = _default_schedule_components(
        production_horizon=horizon_table,
        period_type=period_type,
        assumptions=assumptions,
    )

    period_index = _period_index_for_horizon(start_year, end_year, period_type)

    merged_core = _merge_schedule_table(core, default_core, period_index)

    existing_details = detail_tables or {}
    detail_names = set(default_details.keys()) | set(existing_details.keys())

    merged_details: Dict[str, pd.DataFrame] = {}
    for name in detail_names:
        merged_details[name] = _merge_schedule_table(
            existing_details.get(name),
            default_details.get(name),
            period_index,
        )

    return merged_core, merged_details


def _reset_cached_results() -> None:
    """Mark outputs stale after input changes without auto-running the model again."""

    _bump_model_input_version()
    _safe_session_state_set(MODEL_RESULTS_STALE_KEY, True)
    _safe_session_state_set("excel_bytes_map", {})
    _safe_session_state_set(MODEL_VIEW_CACHE_KEY, {})
    _safe_session_state_set(MODEL_VALIDATION_CACHE_KEY, {})


def _sync_production_horizon(start_year: int, end_year: int) -> None:
    """Ensure schedules and cached results reflect the selected horizon."""

    st.session_state.assumptions = _sync_biological_start_date_in_assumptions(
        st.session_state.get("assumptions"),
        period_type=_normalize_period_type(st.session_state.get("schedule_period_type")),
    )
    st.session_state.assumptions = _sync_opening_herd_cohorts_in_assumptions(
        st.session_state.assumptions
    )

    core_table = st.session_state.get("core_schedule")
    detail_tables = st.session_state.get("detail_schedules")
    period_type = _normalize_period_type(st.session_state.get("schedule_period_type"))

    merged_core, merged_details = _rebase_schedule_to_horizon(
        core_table,
        detail_tables,
        start_year,
        end_year,
        period_type=period_type,
        assumptions=st.session_state.get("assumptions"),
    )

    st.session_state.core_schedule = merged_core
    st.session_state.detail_schedules = merged_details

    _clear_schedule_editor_state("core_schedule")
    for name in merged_details:
        identifier = f"detail::{_scenario_key_suffix(name)}"
        _clear_schedule_editor_state(identifier)

    _sync_horizon_dependent_state(start_year, end_year)
    _reset_cached_results()


def _sync_schedule_period_type(period_type: str) -> None:
    """Rebuild period-based schedules to monthly/quarterly while preserving data."""
    production_table = _ensure_production_horizon_table(
        st.session_state.assumptions.get("Production Horizon")
    )
    start_year, end_year = _derive_horizon_years(production_table)
    st.session_state["schedule_period_type"] = _normalize_period_type(period_type)
    st.session_state.assumptions = _sync_biological_start_date_in_assumptions(
        st.session_state.get("assumptions"),
        period_type=st.session_state["schedule_period_type"],
    )
    st.session_state.assumptions = _sync_opening_herd_cohorts_in_assumptions(
        st.session_state.assumptions
    )

    core_table = st.session_state.get("core_schedule")
    detail_tables = st.session_state.get("detail_schedules")
    merged_core, merged_details = _rebase_schedule_to_horizon(
        core_table,
        detail_tables,
        start_year,
        end_year,
        period_type=st.session_state["schedule_period_type"],
        assumptions=st.session_state.get("assumptions"),
    )

    st.session_state.core_schedule = merged_core
    st.session_state.detail_schedules = merged_details
    _clear_schedule_editor_state("core_schedule")
    for name in merged_details:
        _clear_schedule_editor_state(f"detail::{_scenario_key_suffix(name)}")
    _reset_cached_results()


def _dedupe_horizon_assumption_rows(table_name: str, table: pd.DataFrame) -> pd.DataFrame:
    if table.empty:
        return table

    work = table.copy()
    if table_name == "Operating Costs":
        if "Field" in work.columns:
            work["Field"] = work["Field"].astype(str).str.strip()
            return work.drop_duplicates(subset=["Year", "Field"], keep="last").reset_index(drop=True)
        if "Category" in work.columns:
            work["Category"] = work["Category"].astype(str).str.strip()
            return work.drop_duplicates(subset=["Year", "Category"], keep="last").reset_index(drop=True)
    if table_name == "Pricing" and {"Period", "Product"}.issubset(work.columns):
        work["Period"] = _normalize_period(work["Period"])
        work["Product"] = work["Product"].astype(str).str.strip()
        return work.drop_duplicates(subset=["Period", "Product"], keep="last").reset_index(drop=True)
    if table_name == "Herd Plan":
        return work.drop_duplicates(subset=["Year"], keep="last").reset_index(drop=True)
    return work


def _sync_horizon_dependent_state(start_year: int, end_year: int) -> None:
    """Propagate horizon edits to dependent pages/tables."""

    if start_year > end_year:
        start_year, end_year = end_year, start_year

    # Keep year-based assumption tables aligned to the active horizon.
    assumptions = st.session_state.get("assumptions", {})
    if isinstance(assumptions, dict):
        core_schedule = st.session_state.get("core_schedule", pd.DataFrame())
        assumptions = _sync_commercial_assumptions_to_core(assumptions, core_schedule)
        assumptions = _sync_opening_herd_cohorts_in_assumptions(assumptions)

        for table_name in ["Operating Costs", "Herd Plan"]:
            table = assumptions.get(table_name)
            if not isinstance(table, pd.DataFrame) or table.empty or "Year" not in table:
                continue
            work = table.copy()
            year_col = pd.to_numeric(work.get("Year"), errors="coerce")
            work["Year"] = year_col
            mask = year_col.between(start_year, end_year, inclusive="both")
            if mask.any():
                work = work.loc[mask].reset_index(drop=True)
            elif not work.empty:
                work.loc[:, "Year"] = start_year
            work = _dedupe_horizon_assumption_rows(table_name, work)
            assumptions[table_name] = work
        st.session_state.assumptions = assumptions

    # Keep supplementary schedules aligned where year columns exist.
    supplementary = st.session_state.get("supplementary", {})
    if isinstance(supplementary, dict):
        for name, table in supplementary.items():
            if not isinstance(table, pd.DataFrame) or table.empty or "Year" not in table:
                continue
            work = table.copy()
            years = pd.to_numeric(work.get("Year"), errors="coerce")
            mask = years.between(start_year, end_year, inclusive="both")
            if mask.any():
                work = work.loc[mask].reset_index(drop=True)
            elif not work.empty:
                work.loc[:, "Year"] = start_year
            supplementary[name] = work
        st.session_state.supplementary = supplementary

    # Sync analytics-framework period labels to the rebased core schedule.
    framework = st.session_state.get("analytics_framework", {})
    core_periods = _normalize_period(
        st.session_state.get("core_schedule", pd.DataFrame()).get("Period", pd.Series(dtype=str))
    ).tolist()
    if isinstance(framework, dict) and core_periods:
        for tool_key, config in framework.items():
            if not isinstance(config, dict):
                continue
            data_table = config.get("data")
            if not isinstance(data_table, pd.DataFrame) or data_table.empty:
                continue
            if "Period" not in data_table.columns:
                continue
            work = data_table.copy()
            count = len(work)
            if count > 0:
                work["Period"] = core_periods[:count] if len(core_periods) >= count else (
                    core_periods + [core_periods[-1]] * (count - len(core_periods))
                )
            config["data"] = work
            framework[tool_key] = config
        st.session_state.analytics_framework = framework

    _sync_shared_model_context(st.session_state.get("results"))

def _default_capital_financing_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Source": ["Bank Loan", "Equity"],
            "Amount": [250000.0, 150000.0],
            "Interest/Return %": [6.5, 0.0],
            "Term (years)": [7, None],
        }
    )


def _ensure_capital_financing_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_capital_financing_table()
    else:
        work = table.copy()

    if "Source" not in work.columns:
        work["Source"] = ""
    work["Source"] = _series_or_default(work, "Source", "").astype(str).str.strip()
    work.loc[work["Source"] == "", "Source"] = "Source"

    for column in ["Amount", "Interest/Return %", "Term (years)"]:
        if column not in work.columns:
            work[column] = np.nan
        work[column] = pd.to_numeric(work.get(column), errors="coerce")

    ordered_cols = ["Source", "Amount", "Interest/Return %", "Term (years)"]
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _default_loan_facilities_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Loan Name": ["Bank Loan"],
            "Lender": ["Commercial Bank"],
            "Start Period": [pd.Timestamp("2024-01-31")],
            "Drawdown Amount": [250000.0],
            "Interest Rate %": [6.5],
            "Term (years)": [7.0],
            "Repayment Type": ["straight_line"],
            "Grace Periods": [0],
            "Balloon Amount": [0.0],
            "Fees": [0.0],
            "Active": [True],
        }
    )


def _coerce_bool_value(value: Any, default: bool = True) -> bool:
    if pd.isna(value):
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return default


def _ensure_loan_facilities_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_loan_facilities_table()
    else:
        work = table.copy()

    defaults = _default_loan_facilities_table()
    for column in defaults.columns:
        if column not in work.columns:
            work[column] = defaults.iloc[0][column]

    for column in ["Loan Name", "Lender", "Repayment Type"]:
        work[column] = work[column].fillna("").astype(str).str.strip()

    work.loc[work["Loan Name"] == "", "Loan Name"] = "Loan Facility"
    work["Lender"] = work["Lender"].replace({"nan": "", "<NA>": ""})
    work["Repayment Type"] = work["Repayment Type"].replace({"": "straight_line"})
    work["Start Period"] = pd.to_datetime(work.get("Start Period"), errors="coerce")

    for column in [
        "Drawdown Amount",
        "Interest Rate %",
        "Term (years)",
        "Grace Periods",
        "Balloon Amount",
        "Fees",
    ]:
        work[column] = pd.to_numeric(work.get(column), errors="coerce")

    work["Active"] = _series_or_default(work, "Active", True).map(
        lambda value: _coerce_bool_value(value, True)
    )

    ordered_cols = list(defaults.columns)
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _default_equity_facilities_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Investor Name": ["Founder Equity", "Growth Investor"],
            "Start Period": [pd.Timestamp("2024-01-31"), pd.Timestamp("2025-01-31")],
            "Contribution Amount": [150000.0, 250000.0],
            "Ownership %": [60.0, 40.0],
            "Share Class": ["Ordinary", "Ordinary"],
            "Issue Costs": [0.0, 0.0],
            "Active": [True, True],
        }
    )


def _ensure_equity_facilities_table(table: Optional[pd.DataFrame]) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_equity_facilities_table()
    else:
        work = table.copy()

    defaults = _default_equity_facilities_table()
    for column in defaults.columns:
        if column not in work.columns:
            work[column] = defaults.iloc[0][column]

    for column in ["Investor Name", "Share Class"]:
        work[column] = work[column].fillna("").astype(str).str.strip()

    work.loc[work["Investor Name"] == "", "Investor Name"] = "Investor"
    work["Share Class"] = work["Share Class"].replace({"": "Ordinary", "nan": "Ordinary"})
    work["Start Period"] = pd.to_datetime(work.get("Start Period"), errors="coerce")

    for column in ["Contribution Amount", "Ownership %", "Issue Costs"]:
        work[column] = pd.to_numeric(work.get(column), errors="coerce")

    work["Active"] = _series_or_default(work, "Active", True).map(
        lambda value: _coerce_bool_value(value, True)
    )

    ordered_cols = list(defaults.columns)
    remainder = [col for col in work.columns if col not in ordered_cols]
    return work[ordered_cols + remainder].reset_index(drop=True)


def _estimate_default_wacc_from_capital_table() -> float:
    cap_table = _default_capital_financing_table()
    amounts = pd.to_numeric(cap_table.get("Amount"), errors="coerce")
    returns = pd.to_numeric(cap_table.get("Interest/Return %"), errors="coerce") / 100.0
    valid = amounts.notna() & (amounts > 0)
    if not valid.any():
        return float(DEFAULT_VALUATION_INPUTS.get("WACC", 0.12))
    total = float(amounts[valid].sum())
    if total <= 0:
        return float(DEFAULT_VALUATION_INPUTS.get("WACC", 0.12))
    returns = returns.where(returns > 0, float(DEFAULT_VALUATION_INPUTS.get("WACC", 0.12)))
    weighted = float((amounts[valid] * returns[valid]).sum() / total)
    return weighted if np.isfinite(weighted) and weighted > 0 else float(
        DEFAULT_VALUATION_INPUTS.get("WACC", 0.12)
    )


def _prepare_detail_tables_for_schedule(
    core_table: pd.DataFrame,
    detail_tables: Dict[str, pd.DataFrame],
) -> Dict[str, pd.DataFrame]:
    prepared_details: Dict[str, pd.DataFrame] = {}
    for name, table in (detail_tables or {}).items():
        cleaned = _clean_editor_table(table)
        if cleaned is None:
            continue

        prepared_source = cleaned.copy()
        if name == "Variable Expenses Schedule":
            prepared_source = _aggregate_variable_expenses(
                _ensure_variable_expense_table(cleaned, core_table),
                core_table,
            )
        elif name == "Direct Wages Schedule":
            prepared_source = _aggregate_direct_wages(
                _ensure_direct_wage_table(cleaned, core_table),
                core_table,
            )
        elif name == "Admin Wages Schedule":
            prepared_source = _aggregate_admin_wages(
                _ensure_admin_wage_table(cleaned, core_table),
                core_table,
            )

        prepared = _prepare_timeline_table(prepared_source)
        expected_cols = DETAIL_SCHEDULE_COLUMNS.get(name)
        if expected_cols:
            missing = [col for col in expected_cols if col not in prepared.columns]
            if missing:
                raise ValueError(
                    f"{name} is missing required column(s): {', '.join(missing)}"
                )
            prepared = prepared[expected_cols]
        prepared_details[name] = prepared
    return prepared_details


def _build_schedule_dataframe(
    core_table: pd.DataFrame,
    detail_tables: Dict[str, pd.DataFrame],
    assumptions: Optional[Dict[str, pd.DataFrame]] = None,
) -> pd.DataFrame:
    core_clean = _clean_editor_table(core_table)
    if core_clean is None:
        raise ValueError("Provide at least one period in the core schedule.")

    core_prepared = _prepare_timeline_table(core_clean)
    prepared_details = _prepare_detail_tables_for_schedule(core_clean, detail_tables)
    schedule_df = _assemble_schedule(core_prepared, prepared_details)

    if isinstance(assumptions, dict):
        synced_assumptions = _sync_commercial_assumptions_to_core(assumptions, core_clean)
        schedule_df = _apply_biological_assumptions_to_schedule(
            schedule_df,
            synced_assumptions,
        )
        if "Herd Size (heads)" not in schedule_df.columns:
            herd_plan = synced_assumptions.get("Herd Plan")
            if isinstance(herd_plan, pd.DataFrame) and not herd_plan.empty:
                schedule_df = _apply_herd_plan_to_schedule(schedule_df, herd_plan)
        operating_costs = synced_assumptions.get("Operating Costs")
        if isinstance(operating_costs, pd.DataFrame) and not operating_costs.empty:
            schedule_df = _apply_operating_cost_assumptions_to_schedule(
                schedule_df,
                operating_costs,
                synced_assumptions.get("Biological Cost Drivers"),
            )
        pricing = synced_assumptions.get("Pricing")
        if isinstance(pricing, pd.DataFrame) and not pricing.empty:
            production_drivers = synced_assumptions.get("Production Drivers")
            schedule_df = _apply_pricing_assumptions_to_schedule(
                schedule_df,
                pricing,
                production_drivers if isinstance(production_drivers, pd.DataFrame) else None,
            )

    return schedule_df


def _results_assumption_tables(results: Optional[dict[str, Any]]) -> dict[str, pd.DataFrame]:
    if isinstance(results, dict):
        assumption_tables = results.get("assumption_tables")
        if isinstance(assumption_tables, dict):
            return {
                name: table.copy()
                for name, table in assumption_tables.items()
                if isinstance(table, pd.DataFrame)
            }
    session_assumptions = st.session_state.get("assumptions", {})
    return {
        name: table.copy()
        for name, table in (session_assumptions or {}).items()
        if isinstance(table, pd.DataFrame)
    }


def _results_detail_tables(results: Optional[dict[str, Any]]) -> dict[str, pd.DataFrame]:
    if isinstance(results, dict):
        detail_tables = results.get("detail_tables")
        if isinstance(detail_tables, dict):
            return {
                name: table.copy()
                for name, table in detail_tables.items()
                if isinstance(table, pd.DataFrame)
            }
    session_details = st.session_state.get("detail_schedules", {})
    return {
        name: table.copy()
        for name, table in (session_details or {}).items()
        if isinstance(table, pd.DataFrame)
    }


def _reporting_views_for_result(
    result_payload: dict[str, Any],
) -> dict[str, Any]:
    assumptions = _results_assumption_tables(result_payload)
    detail_tables = _results_detail_tables(result_payload)
    scenario_supplementary = result_payload.get("supplementary", {})
    if not isinstance(scenario_supplementary, dict):
        scenario_supplementary = {}

    scenario_pricing = result_payload.get("pricing_assumptions")
    if not isinstance(scenario_pricing, pd.DataFrame):
        scenario_pricing = assumptions.get("Pricing", pd.DataFrame())

    base_schedule = result_payload.get("base")
    scenario_schedule = result_payload.get("scenario")
    if not isinstance(base_schedule, pd.DataFrame) or not isinstance(scenario_schedule, pd.DataFrame):
        return {"assumptions": assumptions, "entity_options": ["Consolidated"], "default_entity": "Consolidated", "base_schedules": {}, "scenario_schedules": {}}

    base_supplementary = _derive_biological_schedules(base_schedule, assumptions)
    scenario_views = _build_reporting_unit_schedules(
        scenario_schedule,
        assumptions,
        scenario_supplementary,
        detail_tables,
        scenario_pricing,
    )
    base_views = _build_reporting_unit_schedules(
        base_schedule,
        assumptions,
        base_supplementary,
        detail_tables,
        assumptions.get("Pricing"),
    )
    entity_options = _reporting_entity_options(assumptions)
    default_entity = _default_reporting_entity(assumptions)
    return {
        "assumptions": assumptions,
        "entity_options": entity_options,
        "default_entity": default_entity if default_entity in entity_options else entity_options[0],
        "base_schedules": base_views,
        "scenario_schedules": scenario_views,
    }


def _reporting_schedule_for_entity(
    result_payload: dict[str, Any],
    entity: str,
) -> pd.DataFrame:
    views = _reporting_views_for_result(result_payload)
    scenario_schedules = views.get("scenario_schedules", {})
    if entity in scenario_schedules:
        return scenario_schedules[entity]
    return result_payload.get("scenario", pd.DataFrame())


def _reporting_scenario_viability_table(
    results_map: dict[str, dict[str, Any]],
    entity: str,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    use_cached_consolidated = str(entity).strip().casefold() == "consolidated"
    for scenario_name, payload in (results_map or {}).items():
        if not isinstance(payload, dict):
            continue
        model = payload.get("model")
        if model is None:
            continue
        schedule_view = _reporting_schedule_for_entity(payload, entity)
        if not isinstance(schedule_view, pd.DataFrame) or schedule_view.empty:
            continue
        if use_cached_consolidated:
            valuation = payload.get("valuation")
            if not isinstance(valuation, dict) or not valuation:
                valuation = model.valuation_summary(schedule_view)
            debt_capacity = payload.get("debt_capacity_annual")
            if not isinstance(debt_capacity, pd.DataFrame):
                debt_capacity = model.debt_capacity_schedule(schedule_view, annual=True)
        else:
            valuation = model.valuation_summary(schedule_view, annual=True)
            debt_capacity = model.debt_capacity_schedule(schedule_view, annual=True)
        rows.append(
            {
                "Scenario": scenario_name,
                "NPV": valuation.get("npv"),
                "IRR": valuation.get("irr"),
                "Payback (Years)": valuation.get("payback_years"),
                "Terminal Value": valuation.get("terminal_value"),
                "Min DSCR": (
                    pd.to_numeric(debt_capacity.get("DSCR"), errors="coerce").min()
                    if not debt_capacity.empty and "DSCR" in debt_capacity.columns
                    else np.nan
                ),
                "Min DSCR Headroom": (
                    pd.to_numeric(debt_capacity.get("DSCR Headroom"), errors="coerce").min()
                    if not debt_capacity.empty and "DSCR Headroom" in debt_capacity.columns
                    else np.nan
                ),
                "Min Cash Headroom": (
                    pd.to_numeric(
                        debt_capacity.get("Cash Reserve Headroom"), errors="coerce"
                    ).min()
                    if not debt_capacity.empty
                    and "Cash Reserve Headroom" in debt_capacity.columns
                    else np.nan
                ),
                "Covenant Breach Periods": (
                    int(debt_capacity.get("Covenant Breach", pd.Series(dtype=bool)).sum())
                    if not debt_capacity.empty and "Covenant Breach" in debt_capacity.columns
                    else 0
                ),
            }
        )
    comparison = pd.DataFrame(rows)
    if comparison.empty:
        return comparison
    return comparison.set_index("Scenario")


def _dashboard_outputs_for_entity(
    result_payload: dict[str, Any],
    entity: str,
) -> dict[str, Any]:
    scenario = _reporting_schedule_for_entity(result_payload, entity)
    model = result_payload["model"]
    use_cached_consolidated = str(entity).strip().casefold() == "consolidated"

    if use_cached_consolidated:
        valuation_summary = result_payload.get("valuation_annual")
        if not isinstance(valuation_summary, dict) or not valuation_summary:
            valuation_summary = result_payload.get("valuation", {})
        model_audit = result_payload.get("model_audit")
        if not isinstance(model_audit, dict):
            model_audit = model.model_audit(scenario, annual=True)
        working_capital_annual = result_payload.get("working_capital_annual")
        if not isinstance(working_capital_annual, pd.DataFrame):
            working_capital_annual = model.working_capital_schedule(scenario, annual=True)
        debt_capacity_annual = result_payload.get("debt_capacity_annual")
        if not isinstance(debt_capacity_annual, pd.DataFrame):
            debt_capacity_annual = model.debt_capacity_schedule(scenario, annual=True)
        ufcf_schedule_annual = result_payload.get("ufcf_schedule_annual")
        if not isinstance(ufcf_schedule_annual, pd.DataFrame):
            ufcf_schedule_annual = model.ufcf_schedule(scenario, annual=True)
        kpis = result_payload.get("kpis")
        if not isinstance(kpis, pd.DataFrame):
            kpis = model.kpis(scenario, annual=True)
        break_even = result_payload.get("break_even")
        if not isinstance(break_even, pd.DataFrame):
            break_even = model.break_even(scenario, annual=True)
    else:
        valuation_summary = model.valuation_summary(scenario, annual=True)
        model_audit = model.model_audit(scenario, annual=True)
        working_capital_annual = model.working_capital_schedule(scenario, annual=True)
        debt_capacity_annual = model.debt_capacity_schedule(scenario, annual=True)
        ufcf_schedule_annual = model.ufcf_schedule(scenario, annual=True)
        kpis = model.kpis(scenario, annual=True)
        break_even = model.break_even(scenario, annual=True)

    supplementary = result_payload.get("supplementary", {})
    if not isinstance(supplementary, dict):
        supplementary = {}
    pricing_assumptions = result_payload.get("pricing_assumptions")
    if not isinstance(pricing_assumptions, pd.DataFrame):
        pricing_assumptions = pd.DataFrame()
    product_revenue_summary = supplementary.get("Commercial Revenue by Product")
    if not isinstance(product_revenue_summary, pd.DataFrame):
        product_revenue_summary = (
            _pricing_family_summary(pricing_assumptions)
            if not pricing_assumptions.empty
            else pd.DataFrame()
        )
    product_qty_summary = supplementary.get("Commercial Quantity by Period")
    if not isinstance(product_qty_summary, pd.DataFrame):
        product_qty_summary = (
            _pricing_quantity_by_period(pricing_assumptions)
            if not pricing_assumptions.empty
            else pd.DataFrame()
        )

    return {
        "scenario": scenario,
        "valuation_summary": valuation_summary,
        "model_audit": model_audit,
        "working_capital_annual": working_capital_annual,
        "debt_capacity_annual": debt_capacity_annual,
        "ufcf_schedule_annual": ufcf_schedule_annual,
        "kpis": kpis,
        "break_even": break_even,
        "pricing_assumptions": pricing_assumptions,
        "product_revenue_summary": product_revenue_summary,
        "product_qty_summary": product_qty_summary,
        "supplementary": supplementary,
    }


def _computed_default_valuation_inputs() -> Dict[str, float]:
    global _CACHED_DEFAULT_VALUATION_INPUTS
    if _CACHED_DEFAULT_VALUATION_INPUTS is not None:
        return dict(_CACHED_DEFAULT_VALUATION_INPUTS)

    fallback = {k: float(v) for k, v in DEFAULT_VALUATION_INPUTS.items() if pd.notna(v)}
    computed = dict(fallback)
    try:
        assumptions = {
            "Business Configuration": _default_business_configuration_table(),
            "Production Horizon": _default_production_horizon_table(),
            "Herd Plan": _default_herd_plan_table(),
            "Biological System Settings": _default_biological_system_settings_table(),
            "Breeding & Reproduction Biology": _default_breeding_reproduction_biology_table(),
            "Lactation Biology": _default_lactation_biology_table(),
            "Finishing & Slaughter Biology": _default_finishing_slaughter_biology_table(),
            "Opening Herd Cohorts": _default_opening_herd_cohorts_table(),
            "Cohort Allocation Rules": _default_cohort_allocation_rules_table(),
            "Biological Cost Drivers": _default_biological_cost_drivers_table(),
            "Kid Routing Rules": _default_kid_routing_rules_table(),
            "Internal Transfer Pricing": _default_internal_transfer_pricing_table(),
            "Downstream Intake Rules": _default_downstream_intake_rules_table(),
            "Transfer Elimination Rules": _default_transfer_elimination_rules_table(),
            "Pricing": _default_pricing_table(),
            "Production Drivers": _default_production_driver_table(),
            "Operating Costs": _default_operating_cost_table(),
            "Variable Expenses": _default_variable_expense_input_table(),
            "Direct Wages": _default_direct_wage_input_table(),
            "Admin Wages": _default_admin_wage_input_table(),
            "Capital & Financing": _default_capital_financing_table(),
            "Valuation Inputs": pd.DataFrame(
                {
                    "Metric": list(_editable_valuation_input_defaults().keys()),
                    "Value": list(_editable_valuation_input_defaults().values()),
                }
            ),
        }
        core, detail_tables = _default_schedule_components(
            production_horizon=assumptions.get("Production Horizon"),
            assumptions=assumptions,
        )
        schedule_df = _build_schedule_dataframe(core, detail_tables, assumptions)

        valuation_inputs = _valuation_table_to_inputs(assumptions["Valuation Inputs"])
        valuation_inputs["WACC"] = float(
            valuation_inputs.get("WACC", _estimate_default_wacc_from_capital_table())
        )

        supplementary_tables = _default_supplementary_tables()
        supplementary_tables["Capital & Financing"] = _ensure_capital_financing_table(
            assumptions.get("Capital & Financing")
        )
        model = InputSchedule(
            data=schedule_df,
            valuation_inputs=valuation_inputs,
            supplementary_tables=supplementary_tables,
        ).to_model()
        summary = model.valuation_summary()

        computed.update(valuation_inputs)
        if summary:
            npv_value = summary.get("npv")
            irr_value = summary.get("irr")
            computed["WACC"] = float(summary.get("discount_rate", computed["WACC"]))
            computed["NPV"] = (
                float(npv_value) if npv_value is not None else float(computed.get("NPV", 0.0))
            )
            computed["IRR"] = (
                float(irr_value) if irr_value is not None else float(computed.get("IRR", 0.0))
            )
            computed["Terminal Value"] = float(
                summary.get("terminal_value", computed.get("Terminal Value", 0.0))
            )
        _CACHED_DEFAULT_VALUATION_INPUTS = dict(computed)
        return dict(computed)
    except Exception:
        _CACHED_DEFAULT_VALUATION_INPUTS = dict(computed)
        return dict(computed)


def _default_valuation_inputs_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Metric": list(_editable_valuation_input_defaults().keys()),
            "Value": list(_editable_valuation_input_defaults().values()),
        }
    )


def _ensure_valuation_inputs_table(
    table: Optional[pd.DataFrame],
) -> pd.DataFrame:
    if table is None or table.empty:
        work = _default_valuation_inputs_table()
    else:
        work = table.copy()

    if "Metric" not in work.columns:
        work["Metric"] = ""
    work["Metric"] = _series_or_default(work, "Metric", "").astype(str).str.strip()
    work.loc[work["Metric"] == "", "Metric"] = "Metric"
    derived_metric_keys = {metric.casefold() for metric in DERIVED_VALUATION_METRICS}
    work = work.loc[
        ~work["Metric"].astype(str).str.casefold().isin(derived_metric_keys)
    ].reset_index(drop=True)

    if "Value" not in work.columns:
        work["Value"] = np.nan
    work["Value"] = pd.to_numeric(work.get("Value"), errors="coerce")

    existing_metrics = set(work["Metric"].astype(str))
    missing_rows = [
        {"Metric": metric, "Value": value}
        for metric, value in _editable_valuation_input_defaults().items()
        if metric not in existing_metrics
    ]
    if missing_rows:
        work = pd.concat([work, pd.DataFrame(missing_rows)], ignore_index=True)

    ordered_cols = ["Metric", "Value"]
    remainder = [col for col in work.columns if col not in ordered_cols]
    ordered = work[ordered_cols + remainder].reset_index(drop=True)
    metric_order = {
        metric: idx for idx, metric in enumerate(_editable_valuation_input_defaults())
    }
    ordered = ordered.assign(
        __metric_order=ordered["Metric"].map(metric_order).fillna(len(metric_order))
    ).sort_values(["__metric_order", "Metric"], kind="stable")
    return ordered.drop(columns="__metric_order").reset_index(drop=True)


def _valuation_table_to_inputs(table: pd.DataFrame) -> Dict[str, float]:
    work = _ensure_valuation_inputs_table(table)
    allowed_metrics = set(_editable_valuation_input_defaults().keys())
    inputs: Dict[str, float] = {}
    for _, row in work.iterrows():
        metric = str(row.get("Metric", "")).strip()
        value = pd.to_numeric(pd.Series([row.get("Value")]), errors="coerce").iloc[0]
        if metric in allowed_metrics and not pd.isna(value):
            inputs[metric] = float(value)
    return inputs


def _default_assumption_tables() -> Dict[str, pd.DataFrame]:
    return _sync_transfer_tables_to_business_configuration({
        "Business Configuration": _default_business_configuration_table(),
        "Scenario Controls": _default_scenario_controls_table(),
        "Production Horizon": _default_production_horizon_table(),
        "Herd Plan": _default_herd_plan_table(),
        "Biological System Settings": _default_biological_system_settings_table(),
        "Breeding & Reproduction Biology": _default_breeding_reproduction_biology_table(),
        "Lactation Biology": _default_lactation_biology_table(),
        "Finishing & Slaughter Biology": _default_finishing_slaughter_biology_table(),
        "Opening Herd Cohorts": _default_opening_herd_cohorts_table(),
        "Cohort Allocation Rules": _default_cohort_allocation_rules_table(),
        "Biological Cost Drivers": _default_biological_cost_drivers_table(),
        "Kid Routing Rules": _default_kid_routing_rules_table(),
        "Internal Transfer Pricing": _default_internal_transfer_pricing_table(),
        "Downstream Intake Rules": _default_downstream_intake_rules_table(),
        "Transfer Elimination Rules": _default_transfer_elimination_rules_table(),
        "Pricing": _default_pricing_table(),
        "Production Drivers": _default_production_driver_table(),
        "Operating Costs": _default_operating_cost_table(),
        "Variable Expenses": _default_variable_expense_input_table(),
        "Direct Wages": _default_direct_wage_input_table(),
        "Admin Wages": _default_admin_wage_input_table(),
        "Capital & Financing": _default_capital_financing_table(),
        "Valuation Inputs": _default_valuation_inputs_table(),
    })


def _ensure_default_results_loaded() -> None:
    """Initialise run-state flags without auto-executing the model."""

    _safe_session_state_setdefault(MODEL_RESULTS_STALE_KEY, True)
    _safe_session_state_setdefault(MODEL_RUN_SIGNATURE_KEY, None)
    _safe_session_state_setdefault(MODEL_INPUT_VERSION_KEY, 0)
    _safe_session_state_setdefault(MODEL_LAST_RUN_VERSION_KEY, None)
    _safe_session_state_setdefault(MODEL_VIEW_CACHE_KEY, {})
    _safe_session_state_setdefault(MODEL_VALIDATION_CACHE_KEY, {})


def _serialize_run_state(value: Any) -> Any:
    """Convert run-driving state into a stable JSON-serialisable structure."""

    if isinstance(value, pd.DataFrame):
        work = value.copy()
        work.columns = [str(column) for column in work.columns]
        rows: list[list[Any]] = []
        for row in work.itertuples(index=False, name=None):
            rows.append([_serialize_run_state(item) for item in row])
        return {
            "columns": list(work.columns),
            "rows": rows,
        }

    if isinstance(value, dict):
        return {
            str(key): _serialize_run_state(item)
            for key, item in sorted(value.items(), key=lambda pair: str(pair[0]))
        }

    if isinstance(value, (list, tuple)):
        return [_serialize_run_state(item) for item in value]

    if isinstance(value, pd.Timestamp):
        return value.isoformat()

    if isinstance(value, np.generic):
        value = value.item()

    if value is pd.NA or pd.isna(value):
        return None

    if isinstance(value, (str, int, float, bool)) or value is None:
        return value

    return str(value)


def _current_model_run_signature() -> str:
    """Return a stable signature for all inputs that affect model execution."""

    payload = {
        "core_schedule": _safe_session_state_get("core_schedule"),
        "detail_schedules": _safe_session_state_get("detail_schedules"),
        "assumptions": _safe_session_state_get("assumptions"),
        "supplementary": _safe_session_state_get("supplementary"),
        "schedule_period_type": _safe_session_state_get("schedule_period_type"),
        "scenario_preset_tables": _safe_session_state_get("scenario_preset_tables", {}),
        "scenario_preset_descriptions": _safe_session_state_get(
            "scenario_preset_descriptions", {}
        ),
        "scenario_preset_removed_drivers": _safe_session_state_get(
            "scenario_preset_removed_drivers", {}
        ),
    }
    return json.dumps(_serialize_run_state(payload), sort_keys=True, separators=(",", ":"))


def _refresh_results_stale_state() -> None:
    """Mark cached outputs stale whenever model-driving inputs change."""

    current_version = int(_safe_session_state_get(MODEL_INPUT_VERSION_KEY, 0) or 0)
    last_run_version = _safe_session_state_get(MODEL_LAST_RUN_VERSION_KEY)
    has_results = isinstance(_safe_session_state_get("results"), dict)
    is_stale = (not has_results) or (last_run_version is None) or (current_version != last_run_version)
    _safe_session_state_set(MODEL_RESULTS_STALE_KEY, is_stale)


def _prepare_timeline_table(df: pd.DataFrame) -> pd.DataFrame:
    if "Period" not in df.columns:
        raise ValueError("The schedule must include a 'Period' column with dates.")

    work = df.copy()
    work = work[work["Period"].astype(str).str.strip() != ""]
    periods = pd.to_datetime(work["Period"], errors="coerce")
    if periods.isna().any():
        raise ValueError("One or more period values could not be parsed as dates.")

    values = work.drop(columns=["Period"]).apply(pd.to_numeric, errors="coerce")
    values.index = pd.DatetimeIndex(periods)
    values.index.name = "Period"

    mask = values.notna().any(axis=1)
    values = values.loc[mask]
    if values.empty:
        raise ValueError("No numeric data supplied in the schedule.")
    if values.index.has_duplicates:
        raise ValueError("Each period in the schedule must be unique.")
    return values.sort_index()


def _synchronize_financial_algorithms(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    numeric_cols = [
        "Revenue",
        "COGS",
        "Gross Margin",
        "Variable Expenses",
        "Fixed Expenses",
        "Direct Wages",
        "Admin Wages",
        "EBITDA",
        "Depreciation & Amortization",
        "EBIT",
        "Interest Expense",
        "NPBT",
        "Tax Expense",
        "NPAT",
        "CFO",
        "CFI",
        "CFF",
        "Net Cash Flow",
        "Opening Cash Balance",
        "Closing Cash Balance",
        "Cash and Cash Equivalents",
    ]
    for col in numeric_cols:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce")

    if {"Revenue", "COGS"}.issubset(work.columns):
        work["Gross Margin"] = work["Revenue"] - work["COGS"]
    if {"Gross Margin", "Variable Expenses", "Fixed Expenses", "Direct Wages", "Admin Wages"}.issubset(work.columns):
        work["EBITDA"] = (
            work["Gross Margin"]
            - work["Variable Expenses"].fillna(0)
            - work["Fixed Expenses"].fillna(0)
            - work["Direct Wages"].fillna(0)
            - work["Admin Wages"].fillna(0)
        )
    if {"EBITDA", "Depreciation & Amortization"}.issubset(work.columns):
        work["EBIT"] = work["EBITDA"] - work["Depreciation & Amortization"].fillna(0)
    if {"EBIT", "Interest Expense"}.issubset(work.columns):
        work["NPBT"] = work["EBIT"] - work["Interest Expense"].fillna(0)
    if {"NPBT", "Tax Expense"}.issubset(work.columns):
        tax_ratio = pd.Series(0.28, index=work.index)
        if work["NPBT"].notna().any():
            implied = work["Tax Expense"] / work["NPBT"].replace(0, np.nan)
            implied = implied.replace([np.inf, -np.inf], np.nan).dropna()
            if not implied.empty:
                tax_ratio[:] = float(np.clip(implied.median(), 0.0, 0.6))
        work["Tax Expense"] = np.maximum(work["NPBT"], 0.0) * tax_ratio
        work["NPAT"] = work["NPBT"] - work["Tax Expense"]
    if {"CFO", "CFI", "CFF"}.issubset(work.columns):
        work["Net Cash Flow"] = work[["CFO", "CFI", "CFF"]].sum(axis=1, min_count=1)
    if {"Opening Cash Balance", "Net Cash Flow"}.issubset(work.columns):
        opening = work["Opening Cash Balance"].ffill()
        work["Closing Cash Balance"] = opening + work["Net Cash Flow"].fillna(0.0)
        if "Cash and Cash Equivalents" in work.columns:
            work["Cash and Cash Equivalents"] = work["Closing Cash Balance"]
    return work


def _assemble_schedule(
    core: pd.DataFrame, detail_tables: Dict[str, pd.DataFrame]
) -> pd.DataFrame:
    combined = core.copy()
    for table in detail_tables.values():
        if table is None or table.empty:
            continue
        detail = table.copy()
        union_index = combined.index.union(detail.index)
        combined = combined.reindex(union_index)
        detail = detail.reindex(union_index)

        overlap = [col for col in detail.columns if col in combined.columns]
        new_cols = [col for col in detail.columns if col not in combined.columns]

        for col in overlap:
            combined[col] = pd.to_numeric(detail[col], errors="coerce").combine_first(
                pd.to_numeric(combined[col], errors="coerce")
            )
        for col in new_cols:
            combined[col] = detail[col]

    required_columns = [
        "Revenue",
        "COGS",
        "Gross Margin",
        "Variable Expenses",
        "Fixed Expenses",
        "Direct Wages",
        "Admin Wages",
        "EBITDA",
        "Depreciation & Amortization",
        "EBIT",
        "Interest Expense",
        "NPBT",
        "Tax Expense",
        "NPAT",
        "CFO",
        "CFI",
        "CFF",
        "Capex",
        "Net Cash Flow",
        "Opening Cash Balance",
        "Closing Cash Balance",
        "Cash and Cash Equivalents",
        "Current Assets",
        "Non-current Assets",
        "Current Liabilities",
        "Non-current Liabilities",
        "Equity",
    ]

    for col in required_columns:
        if col not in combined.columns:
            combined[col] = np.nan

    combined = _synchronize_financial_algorithms(combined)

    ordered_columns = [
        "Revenue",
        "COGS",
        "Gross Margin",
        "Variable Expenses",
        "Fixed Expenses",
        "Direct Wages",
        "Admin Wages",
        "EBITDA",
        "Depreciation & Amortization",
        "EBIT",
        "Interest Expense",
        "NPBT",
        "Tax Expense",
        "NPAT",
        "CFO",
        "CFI",
        "CFF",
        "Capex",
        "Net Cash Flow",
        "Opening Cash Balance",
        "Closing Cash Balance",
        "Cash and Cash Equivalents",
        "Current Assets",
        "Non-current Assets",
        "Current Liabilities",
        "Non-current Liabilities",
        "Equity",
    ]

    ordered = [col for col in ordered_columns if col in combined.columns]
    remaining = [col for col in combined.columns if col not in ordered]

    combined = combined[ordered + remaining]
    return combined.sort_index()


def _clean_editor_table(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    if df is None or df.empty:
        return None
    work = df.dropna(how="all").dropna(axis=1, how="all")
    if work.empty:
        return None
    return work.reset_index(drop=True)


def _render_table(title: str, table: Optional[pd.DataFrame]) -> None:
    if table is None:
        st.info(f"No **{title}** data was provided.")
        return
    st.subheader(title)
    st.dataframe(table)


def _render_workflow_status_strip() -> None:
    cols = st.columns(4)
    assumptions_ready = isinstance(st.session_state.get("assumptions"), dict)
    schedule_ready = isinstance(st.session_state.get("core_schedule"), pd.DataFrame)
    last_run = st.session_state.get("model_last_run_at")
    results = st.session_state.get("results")
    results_stale = bool(st.session_state.get(MODEL_RESULTS_STALE_KEY, True))
    if results is None:
        results_status = "Not run"
    else:
        results_status = "Stale" if results_stale else "Current"
    cols[0].metric("Assumptions", "Ready" if assumptions_ready else "Pending")
    cols[1].metric("Schedule", "Ready" if schedule_ready else "Pending")
    cols[2].metric("Results", results_status)
    cols[3].metric("Last Recalculated", str(last_run) if last_run else "Not run")


def _render_stale_results_notice() -> None:
    """Explain that displayed outputs are from the previous model run."""

    if not st.session_state.get(MODEL_RESULTS_STALE_KEY):
        return
    if st.session_state.get("results") is None:
        return
    st.warning(
        "Inputs have changed since the last model run. The displayed outputs are from the previous run. "
        "Click `Run model` to refresh the financials, dashboard, analytics, and downloads."
    )


def _assumption_validation_issues(assumptions: Dict[str, pd.DataFrame]) -> list[str]:
    issues: list[str] = []
    if _is_breeding_to_unit_mode(assumptions):
        destination = _selected_transfer_destination(assumptions)
        if not destination:
            issues.append("Business Configuration requires a Transfer Destination in Breeding-to-Unit mode.")
        routing = _ensure_kid_routing_rules_table(assumptions.get("Kid Routing Rules"))
        if not routing.empty:
            internal_destinations = routing.loc[
                routing["Active"].fillna(True).astype(bool)
                & routing["Destination"].isin(["Meat", "Milk-Cheese", "Combined"]),
                "Destination",
            ].dropna().astype(str).unique().tolist()
            if len(internal_destinations) > 1:
                issues.append("Kid Routing Rules supports only one internal destination in the current implementation.")
            grouped = routing.loc[routing["Active"].fillna(True).astype(bool)].groupby("Sex")["Allocation %"].sum()
            for sex, total in grouped.items():
                if pd.notna(total) and float(total) > 100.001:
                    issues.append(f"Kid Routing Rules allocation exceeds 100% for {sex} kids.")
        transfer_pricing = _ensure_internal_transfer_pricing_table(assumptions.get("Internal Transfer Pricing"))
        if destination and transfer_pricing.loc[
            transfer_pricing["Active"].fillna(True).astype(bool)
            & transfer_pricing["Destination"].astype(str).eq(destination)
        ].empty:
            issues.append("Internal Transfer Pricing is missing an active row for the selected transfer destination.")
        intake_rules = _ensure_downstream_intake_rules_table(assumptions.get("Downstream Intake Rules"))
        if destination and intake_rules.loc[
            intake_rules["Active"].fillna(True).astype(bool)
            & intake_rules["Destination"].astype(str).eq(destination)
        ].empty:
            issues.append("Downstream Intake Rules is missing active rows for the selected transfer destination.")
    operating = assumptions.get("Operating Costs", pd.DataFrame())
    if isinstance(operating, pd.DataFrame) and not operating.empty and {"Year", "Field"}.issubset(operating.columns):
        duplicate_mask = operating.duplicated(subset=["Year", "Field"], keep=False)
        if duplicate_mask.any():
            issues.append("Operating Costs has duplicate Year+Field rows.")
    herd = assumptions.get("Herd Plan", pd.DataFrame())
    if isinstance(herd, pd.DataFrame) and not herd.empty and "Year" in herd.columns:
        duplicate_years = herd.duplicated(subset=["Year"], keep=False)
        if duplicate_years.any():
            issues.append("Herd Plan has duplicate years.")
    core_for_validation = st.session_state.get("core_schedule", pd.DataFrame())
    synced_assumptions = (
        _sync_commercial_assumptions_to_core(assumptions, core_for_validation)
        if isinstance(assumptions, dict)
        else {}
    )
    pricing = synced_assumptions.get("Pricing", pd.DataFrame())
    if isinstance(pricing, pd.DataFrame) and not pricing.empty and {"Period", "Product"}.issubset(pricing.columns):
        dup = pricing.duplicated(subset=["Period", "Product"], keep=False)
        if dup.any():
            issues.append("Pricing has duplicate Period+Product rows.")
        issues.extend(
            _pricing_validation_messages(
                pricing,
                synced_assumptions.get("Production Drivers"),
            )
        )
    supplementary = _sync_asset_schedule_from_capex_in_supplementary(
        st.session_state.get("supplementary", {})
    )
    capex_recon_issues = _capex_asset_reconciliation_issues(
        supplementary.get("Capex Schedule"),
        supplementary.get("Asset Schedules"),
    )
    issues.extend(capex_recon_issues)
    return issues


def _render_assumption_validation_summary(assumptions: Dict[str, pd.DataFrame]) -> None:
    issues = _cached_input_value(
        "assumption_validation_summary",
        lambda: _assumption_validation_issues(assumptions),
    )

    if issues:
        st.warning("Validation summary: " + " ".join(f"- {msg}" for msg in issues))
    else:
        st.success("Validation summary: commercial, schedule, and capex-to-asset keys are consistent.")


ANALYTICS_FRAMEWORK_TOOLS: List[Dict[str, str]] = [
    {
        "key": "sensitivity_analysis",
        "title": "Sensitivity Analysis",
        "methodology": "One-at-a-time and multi-driver elasticity tests against profitability and cash flow outputs.",
        "visualization": "Spider plot + elasticity heatmap",
    },
    {
        "key": "scenario_stress_testing",
        "title": "Scenario & Stress Testing",
        "methodology": "Apply severe but plausible shocks and compare resilience across operating, liquidity, and valuation KPIs.",
        "visualization": "Scenario bridge chart + downside table",
    },
    {
        "key": "trend_seasonality",
        "title": "Trend & Seasonality Decomposition",
        "methodology": "Decompose historical production/revenue signals into trend, seasonal, and residual components.",
        "visualization": "Trend-season-residual line charts",
    },
    {
        "key": "customer_product_segmentation",
        "title": "Customer & Product Segmentation",
        "methodology": "Segment products/channels by margin, growth, and volatility to prioritize strategic focus areas.",
        "visualization": "Segment matrix + contribution waterfall",
    },
    {
        "key": "monte_carlo_simulation",
        "title": "Monte Carlo Simulation",
        "methodology": "Run probabilistic simulations for revenue, costs, NPV, and IRR based on configured distributions.",
        "visualization": "Distribution histogram + percentile fan chart",
    },
    {
        "key": "what_if_analysis",
        "title": "What-If Analysis",
        "methodology": "Interactive assumption perturbation with immediate recalc of key KPI outputs.",
        "visualization": "Delta KPI cards + before/after bars",
    },
    {
        "key": "goal_seek",
        "title": "Goal Seek Routines",
        "methodology": "Solve for required input values that satisfy target profitability, liquidity, and return constraints.",
        "visualization": "Target vs solved-input table",
    },
    {
        "key": "tornado_spider",
        "title": "Tornado Charts & Spider Diagrams",
        "methodology": "Rank assumptions by marginal impact on NPV/IRR and display response curves.",
        "visualization": "Tornado bar chart + spider line chart",
    },
    {
        "key": "regression_modeling",
        "title": "Regression Modeling",
        "methodology": "Estimate explanatory relationships between historical drivers and financial outcomes.",
        "visualization": "Coefficient chart + fitted vs actual scatter",
    },
    {
        "key": "time_series_models",
        "title": "Time Series (ARIMA/Prophet/LSTM)",
        "methodology": "Forecast cyclical and seasonal patterns in revenues, prices, and expenses using multiple model classes.",
        "visualization": "Forecast bands + model comparison table",
    },
    {
        "key": "classification_models",
        "title": "Classification Models",
        "methodology": "Classify credit risk/churn/segment outcomes from labeled features and operational indicators.",
        "visualization": "Confusion matrix + lift chart",
    },
    {
        "key": "linear_nonlinear_optimization",
        "title": "Linear/Nonlinear Optimization",
        "methodology": "Optimize objective functions under operational and capital constraints.",
        "visualization": "Optimal allocation table + constraint slack chart",
    },
    {
        "key": "portfolio_optimization",
        "title": "Portfolio Optimization",
        "methodology": "Balance risk and return across herds/product lines using mean-variance and robust alternatives.",
        "visualization": "Efficient frontier + allocation pie",
    },
    {
        "key": "real_options_analysis",
        "title": "Real Options Analysis",
        "methodology": "Value defer/expand/abandon flexibility embedded in strategic initiatives.",
        "visualization": "Option decision tree + value uplift table",
    },
    {
        "key": "var_cvar",
        "title": "VaR & Conditional VaR",
        "methodology": "Estimate downside tail risk at configurable confidence levels.",
        "visualization": "Loss distribution + tail expectation chart",
    },
    {
        "key": "copula_models",
        "title": "Copula Models",
        "methodology": "Model joint tail dependencies across multiple risk factors.",
        "visualization": "Dependence heatmap + tail copula diagnostics",
    },
    {
        "key": "macroeconomic_linking",
        "title": "Macroeconomic Linking",
        "methodology": "Link inflation/GDP/rates/FX assumptions into model drivers for macro-consistent projections.",
        "visualization": "Macro-to-driver linkage table",
    },
    {
        "key": "esg_sustainability",
        "title": "ESG & Sustainability Metrics",
        "methodology": "Quantify financial impact of emissions, carbon pricing, and renewable adoption pathways.",
        "visualization": "ESG scorecard + cost-benefit trend",
    },
    {
        "key": "market_intelligence",
        "title": "Market Intelligence Integration",
        "methodology": "Blend external sentiment and industry outlook data into dynamic demand forecasts.",
        "visualization": "Sentiment index + demand response chart",
    },
    {
        "key": "probabilistic_valuation",
        "title": "Probabilistic Valuation",
        "methodology": "Produce valuation ranges and confidence intervals instead of single-point estimates.",
        "visualization": "Valuation percentile chart",
    },
    {
        "key": "comparative_valuation_clustering",
        "title": "Comparative Valuation with Clustering",
        "methodology": "Benchmark against statistically similar peers using clustering and relative multiples.",
        "visualization": "Peer cluster map + valuation spread",
    },
    {
        "key": "ml_based_valuation",
        "title": "Machine Learning–Based Valuation",
        "methodology": "Predict fair value/multiples from historical market and operational feature sets.",
        "visualization": "Feature importance + predicted range chart",
    },
]


def _default_framework_table(
    columns: Sequence[str], rows: Sequence[Sequence[Any]]
) -> pd.DataFrame:
    return pd.DataFrame(list(rows), columns=list(columns))


def _analytics_framework_store() -> Dict[str, Dict[str, Any]]:
    store = st.session_state.setdefault("analytics_framework", {})
    for tool in ANALYTICS_FRAMEWORK_TOOLS:
        key = tool["key"]
        if key in store:
            continue
        store[key] = {
            "enabled": False,
            "data_sources": ["Scenario Output"],
            "model_mode": "Balanced",
            "tool_shock_override": 0.0,
            "data": _default_framework_table(
                ["Period", "Revenue", "Cost", "Volume"],
                [
                    ("P1", 150000.0, 105000.0, 1200.0),
                    ("P2", 158000.0, 109000.0, 1240.0),
                    ("P3", 166000.0, 112500.0, 1280.0),
                    ("P4", 172000.0, 115000.0, 1310.0),
                ],
            ),
            "inputs": _default_framework_table(
                ["Input", "Source", "Transform", "Active"],
                [
                    ("Primary Product Price Series", "Scenario Output", "none", True),
                    ("Feed Cost Series", "Scenario Output", "none", True),
                    ("Herd Productivity", "Input Schedule", "rolling_mean", True),
                ],
            ),
            "assumptions": _default_framework_table(
                ["Assumption", "Value", "Units"],
                [("Confidence Level", 95.0, "%"), ("Lookback Window", 12.0, "months")],
            ),
            "drivers": _default_framework_table(
                ["Driver", "Base", "Low", "High"],
                [
                    ("Primary Product Price", 1.0, -20.0, 20.0),
                    ("Feed Cost", 1.0, -20.0, 20.0),
                    ("Herd Productivity", 1.0, -15.0, 15.0),
                ],
            ),
            "scenarios": _default_framework_table(
                ["Scenario", "Shock %", "Probability %", "Active"],
                [
                    ("Base", 0.0, 60.0, True),
                    ("Downside", -15.0, 25.0, True),
                    ("Severe Stress", -35.0, 15.0, False),
                ],
            ),
        }
    st.session_state["analytics_framework"] = store
    return store


def _analytics_framework_control_store() -> Dict[str, Any]:
    default_controls = {
        "scenario": "Base",
        "custom_shock_pct": 0.0,
        "focus_metric": "Profit",
        "period_filter": "All",
    }
    return st.session_state.setdefault("analytics_framework_controls", default_controls)


def _numeric_column_mean(df: pd.DataFrame, column: str) -> float:
    if column not in df.columns:
        return 0.0
    numeric = pd.to_numeric(df[column], errors="coerce")
    if numeric.dropna().empty:
        return 0.0
    return float(numeric.mean())


def _scenario_shock_value(controls: Dict[str, Any]) -> float:
    scenario = controls.get("scenario", "Base")
    if scenario == "Upside":
        return 8.0
    if scenario == "Downside":
        return -10.0
    if scenario == "Stress":
        return -25.0
    if scenario == "Custom":
        return float(controls.get("custom_shock_pct", 0.0))
    return 0.0


def _analytics_framework_output(
    tool_config: Dict[str, Any],
    results: Optional[Dict[str, Any]],
    controls: Dict[str, Any],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    assumptions = tool_config.get("assumptions", pd.DataFrame())
    drivers = tool_config.get("drivers", pd.DataFrame())
    scenarios = tool_config.get("scenarios", pd.DataFrame())
    data_table = tool_config.get("data", pd.DataFrame())

    assumption_level = _numeric_column_mean(assumptions, "Value")
    driver_range = (
        abs(_numeric_column_mean(drivers, "High"))
        + abs(_numeric_column_mean(drivers, "Low"))
    ) / 2.0
    tool_shock = _numeric_column_mean(scenarios, "Shock %")
    scenario_shock = _scenario_shock_value(controls)
    override_shock = float(tool_config.get("tool_shock_override", 0.0))
    total_shock = tool_shock + scenario_shock + override_shock

    base_npv = np.nan
    base_irr = np.nan
    if results is not None and isinstance(results.get("kpis"), pd.DataFrame):
        kpi_df = results["kpis"]
        if "NPV" in kpi_df.columns:
            base_npv = float(pd.to_numeric(kpi_df["NPV"], errors="coerce").iloc[0])
        if "IRR" in kpi_df.columns:
            base_irr = float(pd.to_numeric(kpi_df["IRR"], errors="coerce").iloc[0])

    modeled = data_table.copy(deep=True) if isinstance(data_table, pd.DataFrame) else pd.DataFrame()
    if modeled.empty:
        modeled = _default_framework_table(
            ["Period", "Revenue", "Cost", "Volume"],
            [("P1", 150000.0, 105000.0, 1200.0)],
        )
    for col in ["Revenue", "Cost", "Volume"]:
        if col not in modeled.columns:
            modeled[col] = 0.0
        modeled[col] = pd.to_numeric(modeled[col], errors="coerce").fillna(0.0)

    modeled["Revenue_adj"] = modeled["Revenue"] * (1 + (total_shock / 100.0))
    modeled["Cost_adj"] = modeled["Cost"] * (1 - (total_shock * 0.35 / 100.0))
    modeled["Profit_adj"] = modeled["Revenue_adj"] - modeled["Cost_adj"]
    modeled["Margin_adj"] = np.where(
        modeled["Revenue_adj"] != 0,
        modeled["Profit_adj"] / modeled["Revenue_adj"],
        0.0,
    )

    base_profit = float(modeled["Profit_adj"].mean())
    base_margin = float(modeled["Margin_adj"].mean())
    impact_score = (driver_range * 0.35) + (assumption_level * 0.2) + (total_shock * -0.25)
    resilience_score = max(0.0, 100.0 - abs(total_shock) - max(0.0, -base_margin * 40))
    npv_proxy = base_profit * max(1.0, assumption_level / 12.0)

    summary = pd.DataFrame(
        {
            "Metric": [
                "Configured Data Sources",
                "Average Assumption Level",
                "Average Driver Stress Range",
                "Applied Shock (%)",
                "Modeled Profit",
                "Modeled Margin",
                "Resilience Score",
                "Indicative Impact Score",
                "Modeled NPV Proxy",
                "Reference NPV",
                "Reference IRR",
            ],
            "Value": [
                len(tool_config.get("data_sources", [])),
                round(assumption_level, 2),
                round(driver_range, 2),
                round(total_shock, 2),
                round(base_profit, 2),
                round(base_margin, 4),
                round(resilience_score, 2),
                round(impact_score, 2),
                round(npv_proxy, 2),
                round(base_npv, 2) if pd.notna(base_npv) else np.nan,
                round(base_irr, 4) if pd.notna(base_irr) else np.nan,
            ],
        }
    )

    sensitivity = pd.DataFrame({"Shock %": [-20, -10, 0, 10, 20]})
    sensitivity["Profit"] = sensitivity["Shock %"].apply(
        lambda shock: base_profit * (1 + (shock / 100.0))
    )

    scenario_compare = scenarios.copy(deep=True) if isinstance(scenarios, pd.DataFrame) else pd.DataFrame()
    if scenario_compare.empty:
        scenario_compare = pd.DataFrame(
            {"Scenario": ["Base"], "Shock %": [0.0], "Probability %": [100.0]}
        )
    scenario_compare["Shock %"] = pd.to_numeric(
        scenario_compare.get("Shock %", pd.Series(dtype=float)), errors="coerce"
    ).fillna(0.0)
    scenario_compare["Modeled Profit"] = scenario_compare["Shock %"].apply(
        lambda shock: base_profit * (1 + (shock / 100.0))
    )

    return summary, modeled, sensitivity, scenario_compare


def _render_analytics_framework(results: Optional[Dict[str, Any]]) -> None:
    st.markdown("### Editable Analytics Schedule Framework")
    st.caption(
        "Each analytical capability has editable inputs, assumptions, model drivers, and "
        "scenario settings. Outputs refresh automatically on every change."
    )
    framework = _analytics_framework_store()
    controls = _analytics_framework_control_store()
    shared_context = _sync_shared_model_context(results)

    st.markdown("#### Global Analytics Controls")
    g1, g2, g3, g4 = st.columns(4)
    controls["scenario"] = g1.selectbox(
        "Scenario View",
        options=["Base", "Upside", "Downside", "Stress", "Custom"],
        index=["Base", "Upside", "Downside", "Stress", "Custom"].index(
            controls.get("scenario", "Base")
        ),
        key="analytics_global_scenario",
    )
    custom_shock_options = [round(x * 0.5, 1) for x in range(-100, 101)]
    custom_shock_default = float(controls.get("custom_shock_pct", 0.0))
    if custom_shock_default not in custom_shock_options:
        custom_shock_default = 0.0
    controls["custom_shock_pct"] = g2.selectbox(
        "Custom Shock (%)",
        options=custom_shock_options,
        index=custom_shock_options.index(custom_shock_default),
        key="analytics_custom_shock",
    )
    controls["focus_metric"] = g3.selectbox(
        "Focus Metric",
        options=["Profit", "Margin", "NPV Proxy", "Resilience Score"],
        index=["Profit", "Margin", "NPV Proxy", "Resilience Score"].index(
            controls.get("focus_metric", "Profit")
        ),
        key="analytics_focus_metric",
    )
    controls["period_filter"] = g4.selectbox(
        "Period Filter",
        options=["All", "P1", "P2", "P3", "P4"],
        index=["All", "P1", "P2", "P3", "P4"].index(
            controls.get("period_filter", "All")
        ),
        key="analytics_period_filter",
    )
    st.session_state["analytics_framework_controls"] = controls
    shared_context = _sync_shared_model_context(results)
    current_scenario_label = shared_context.get("active_result_scenario") or shared_context.get(
        "selected_scenario_name", "Scenario"
    )
    st.caption(f"Synced model context scenario: **{current_scenario_label}**")

    module_options = [tool["title"] for tool in ANALYTICS_FRAMEWORK_TOOLS]
    active_modules = st.multiselect(
        "Module Filter",
        options=module_options,
        default=module_options,
        key="analytics_module_filter",
        help="Show only selected modules below.",
    )
    if active_modules:
        active_module_title = _section_selector(
            "Analytics Module",
            active_modules,
            key="analytics_active_module_selector",
            help="Only the selected analytics module is rendered.",
        )
    else:
        active_module_title = ""

    linked_sources = [
        "Input Schedule",
        "Assumptions",
        "Scenario Output",
        "Financial Statements",
        "Dashboard KPIs",
        "Supplementary Schedules",
        "External Market Data",
        "ESG Data",
        "Peer Benchmark Data",
    ]

    module_summary_rows: list[Dict[str, Any]] = []

    for tool in ANALYTICS_FRAMEWORK_TOOLS:
        if tool["title"] not in active_modules:
            continue
        if tool["title"] != active_module_title:
            continue
        tool_key = tool["key"]
        config = framework[tool_key]
        with st.expander(tool["title"], expanded=False):
            left, right = st.columns([1, 2])
            config["enabled"] = left.checkbox(
                "Enable tool",
                value=bool(config.get("enabled", False)),
                key=f"framework_enabled::{tool_key}",
            )
            selected_sources = right.multiselect(
                "Linked data inputs",
                options=linked_sources,
                default=config.get("data_sources", ["Scenario Output"]),
                key=f"framework_sources::{tool_key}",
                help="Choose datasets this tool should consume.",
            )
            config["data_sources"] = selected_sources
            config["model_mode"] = left.selectbox(
                "Model Mode",
                options=["Conservative", "Balanced", "Aggressive"],
                index=["Conservative", "Balanced", "Aggressive"].index(
                    config.get("model_mode", "Balanced")
                ),
                key=f"framework_mode::{tool_key}",
            )
            tool_shock_options = [round(x * 0.5, 1) for x in range(-60, 61)]
            tool_shock_default = float(config.get("tool_shock_override", 0.0))
            if tool_shock_default not in tool_shock_options:
                tool_shock_default = 0.0
            config["tool_shock_override"] = right.selectbox(
                "Tool-level shock override (%)",
                options=tool_shock_options,
                index=tool_shock_options.index(tool_shock_default),
                key=f"framework_shock::{tool_key}",
            )

            st.markdown("**Methodology**")
            st.write(tool["methodology"])

            st.markdown("**Underlying Data Schedule (Editable)**")
            config["data"] = st.data_editor(
                config.get("data", pd.DataFrame()),
                num_rows="dynamic",
                use_container_width=True,
                key=f"framework_data::{tool_key}",
            )

            st.markdown("**Configurable Input Mapping**")
            config["inputs"] = st.data_editor(
                config.get("inputs", pd.DataFrame()),
                num_rows="dynamic",
                use_container_width=True,
                key=f"framework_inputs::{tool_key}",
            )

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Adjustable Assumptions**")
                config["assumptions"] = st.data_editor(
                    config.get("assumptions", pd.DataFrame()),
                    num_rows="dynamic",
                    use_container_width=True,
                    key=f"framework_assumptions::{tool_key}",
                )
            with c2:
                st.markdown("**Model Drivers & Ranges**")
                config["drivers"] = st.data_editor(
                    config.get("drivers", pd.DataFrame()),
                    num_rows="dynamic",
                    use_container_width=True,
                    key=f"framework_drivers::{tool_key}",
                )

            st.markdown("**Scenario Settings**")
            config["scenarios"] = st.data_editor(
                config.get("scenarios", pd.DataFrame()),
                num_rows="dynamic",
                use_container_width=True,
                key=f"framework_scenarios::{tool_key}",
            )

            output_df, modeled_df, sensitivity_df, scenario_compare_df = _analytics_framework_output(
                config, results, controls
            )
            st.markdown("**Dynamic Outputs**")
            st.dataframe(output_df, use_container_width=True)
            module_summary_rows.append(
                {
                    "Module": tool["title"],
                    "Modeled Profit": float(
                        output_df.loc[output_df["Metric"] == "Modeled Profit", "Value"].iloc[0]
                    ),
                    "Applied Shock (%)": float(
                        output_df.loc[output_df["Metric"] == "Applied Shock (%)", "Value"].iloc[0]
                    ),
                    "Resilience Score": float(
                        output_df.loc[output_df["Metric"] == "Resilience Score", "Value"].iloc[0]
                    ),
                }
            )

            if controls.get("period_filter") != "All" and "Period" in modeled_df.columns:
                modeled_view = modeled_df.loc[
                    modeled_df["Period"].astype(str) == str(controls["period_filter"])
                ]
                if modeled_view.empty:
                    modeled_view = modeled_df
            else:
                modeled_view = modeled_df

            st.markdown("**Modeled Time-Series Output**")
            st.dataframe(modeled_view, use_container_width=True)
            if {"Revenue_adj", "Cost_adj", "Profit_adj"}.issubset(modeled_view.columns):
                st.line_chart(
                    modeled_view.set_index("Period")[["Revenue_adj", "Cost_adj", "Profit_adj"]]
                )

            c_left, c_right = st.columns(2)
            with c_left:
                st.markdown("**What-if & Sensitivity Grid**")
                st.dataframe(sensitivity_df, use_container_width=True)
                st.area_chart(sensitivity_df.set_index("Shock %")[["Profit"]])
            with c_right:
                st.markdown("**Scenario Comparison**")
                st.dataframe(scenario_compare_df, use_container_width=True)
                if {"Scenario", "Modeled Profit"}.issubset(scenario_compare_df.columns):
                    compare_chart = scenario_compare_df.set_index("Scenario")[
                        ["Modeled Profit"]
                    ]
                    st.bar_chart(compare_chart)

            st.caption(f"Suggested visualisation: {tool['visualization']}")

            framework[tool_key] = config

    if module_summary_rows:
        st.markdown("#### Cross-Module Scenario Scorecard")
        module_summary_df = pd.DataFrame(module_summary_rows)
        st.dataframe(module_summary_df, use_container_width=True)
        st.bar_chart(module_summary_df.set_index("Module")[["Modeled Profit"]])

    st.session_state["analytics_framework"] = framework


def _sync_shared_model_context(results: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    assumptions = st.session_state.get("assumptions", {})
    production = assumptions.get("Production Horizon", pd.DataFrame()) if isinstance(assumptions, dict) else pd.DataFrame()
    controls_table = assumptions.get("Scenario Controls", pd.DataFrame()) if isinstance(assumptions, dict) else pd.DataFrame()
    valuation_table = assumptions.get("Valuation Inputs", pd.DataFrame()) if isinstance(assumptions, dict) else pd.DataFrame()

    production_context: Dict[str, Any] = {}
    if isinstance(production, pd.DataFrame) and not production.empty:
        production_context = {
            "start_year": int(pd.to_numeric(production.iloc[0].get("Start Year"), errors="coerce") or 0),
            "end_year": int(pd.to_numeric(production.iloc[0].get("End Year"), errors="coerce") or 0),
        }

    scenario_controls = (
        _scenario_controls_value_map(controls_table)
        if isinstance(controls_table, pd.DataFrame)
        else {}
    )
    valuation_inputs = (
        _valuation_table_to_inputs(valuation_table)
        if isinstance(valuation_table, pd.DataFrame)
        else {}
    )
    business_type = _selected_business_type(assumptions) if isinstance(assumptions, dict) else DEFAULT_BUSINESS_TYPE

    context = {
        "selected_scenario_name": st.session_state.get("selected_scenario_name"),
        "active_result_scenario": (results or {}).get("selected_scenario"),
        "business_type": business_type,
        "active_products": _active_products_for_business_type(business_type),
        "production_horizon": production_context,
        "scenario_controls": scenario_controls,
        "valuation_inputs": valuation_inputs,
        "analytics_controls": st.session_state.get("analytics_framework_controls", {}).copy(),
        "ai_orchestration_config": st.session_state.get("ai_orchestration_config", {}).copy(),
    }
    st.session_state["shared_model_context"] = context
    return context


def _rag_store() -> Dict[str, Any]:
    store = st.session_state.setdefault(
        "rag_framework",
        {
            "documents": [],
            "index": pd.DataFrame(),
            "version": 0,
            "last_reindexed_at": None,
        },
    )
    return store


def _chunk_text(text: str, chunk_size: int = 450) -> List[str]:
    cleaned = (text or "").strip()
    if not cleaned:
        return []
    return [cleaned[i : i + chunk_size] for i in range(0, len(cleaned), chunk_size)]


def _ingest_rag_document(title: str, content: str, source: str) -> bool:
    title_clean = (title or "").strip()
    content_clean = (content or "").strip()
    if not title_clean or not content_clean:
        return False

    store = _rag_store()
    docs = list(store.get("documents", []))
    doc_id = f"doc_{len(docs) + 1}"
    docs.append(
        {
            "id": doc_id,
            "title": title_clean,
            "source": source,
            "content": content_clean,
            "updated_at": pd.Timestamp.utcnow().isoformat(),
        }
    )
    store["documents"] = docs
    st.session_state["rag_framework"] = store
    return True


def _reindex_rag(snapshot: Dict[str, Any]) -> pd.DataFrame:
    store = _rag_store()
    records: List[Dict[str, Any]] = []

    snapshot_index = _snapshot_index(snapshot)
    if not snapshot_index.empty:
        for _, row in snapshot_index.iterrows():
            records.append(
                {
                    "Chunk ID": f"snap::{row.get('Section')}::{row.get('Key')}",
                    "Source": "model_snapshot",
                    "Section": row.get("Section"),
                    "Text": str(row.get("Text", "")),
                }
            )

    for doc in store.get("documents", []):
        chunks = _chunk_text(str(doc.get("content", "")))
        for idx, chunk in enumerate(chunks, start=1):
            records.append(
                {
                    "Chunk ID": f"{doc.get('id')}::chunk_{idx}",
                    "Source": doc.get("source", "manual"),
                    "Section": doc.get("title", "document"),
                    "Text": chunk,
                }
            )

    index_df = pd.DataFrame(records)
    store["index"] = index_df
    store["version"] = int(store.get("version", 0)) + 1
    store["last_reindexed_at"] = pd.Timestamp.utcnow().isoformat()
    st.session_state["rag_framework"] = store
    return index_df


def _retrieve_rag_context(query: str, top_n: int = 8) -> pd.DataFrame:
    store = _rag_store()
    index_df = store.get("index")
    if not isinstance(index_df, pd.DataFrame) or index_df.empty:
        return pd.DataFrame(columns=["Chunk ID", "Source", "Section", "Text", "score"])

    tokens = {token for token in re.findall(r"[a-zA-Z0-9_]+", (query or "").lower()) if token}
    scored = index_df.copy()
    if not tokens:
        scored["score"] = 0
        return scored.head(top_n)

    scored["score"] = scored["Text"].str.lower().apply(
        lambda text: sum(1 for token in tokens if token in text)
    )
    scored = scored.sort_values(["score", "Source"], ascending=[False, True])
    filtered = scored[scored["score"] > 0]
    return (filtered if not filtered.empty else scored).head(top_n)


def _render_rag_admin(
    snapshot: Dict[str, Any],
    show_header: bool = True,
    use_expander: bool = True,
) -> pd.DataFrame:
    store = _rag_store()
    if show_header:
        st.markdown("### Retrieval-Augmented Generation (RAG) Hub")
        st.caption(
            "Ingest documents/data, re-index knowledge, and retrieve grounded context for the "
            "orchestration engine."
        )

    rag_container = (
        st.expander("RAG Ingestion & Indexing", expanded=False)
        if use_expander
        else nullcontext()
    )
    with rag_container:
        title = st.text_input("Document title", key="rag_doc_title")
        content = st.text_area(
            "Document or data content",
            key="rag_doc_content",
            placeholder="Paste policies, research notes, investor memos, or model assumptions...",
        )
        add_col, reindex_col = st.columns(2)
        if add_col.button("Ingest Document", key="rag_ingest_btn"):
            if _ingest_rag_document(title, content, source="manual_text"):
                st.success("Document ingested.")
            else:
                st.warning("Provide both title and content before ingestion.")

        uploaded_files = st.file_uploader(
            "Upload file(s) for ingestion (up to 200MB total)",
            key="rag_uploader",
            accept_multiple_files=True,
        )
        if uploaded_files and st.button("Ingest Uploaded File(s)", key="rag_ingest_file"):
            total_size = sum(len(file.getvalue()) for file in uploaded_files)
            max_size_bytes = 200 * 1024 * 1024
            if total_size > max_size_bytes:
                st.error(
                    "Uploaded files exceed the 200MB combined limit. "
                    "Please reduce the selection and try again."
                )
            else:
                ingested_count = 0
                for uploaded in uploaded_files:
                    raw_bytes = uploaded.getvalue()
                    try:
                        raw_text = raw_bytes.decode("utf-8")
                    except UnicodeDecodeError:
                        raw_text = (
                            f"Binary document uploaded: {uploaded.name}\n"
                            f"Content-Type: {uploaded.type}\n"
                            f"Size (bytes): {len(raw_bytes)}\n"
                            "Note: This binary file requires specialized parsing for full text extraction."
                        )
                    if _ingest_rag_document(
                        uploaded.name, raw_text, source="uploaded_file"
                    ):
                        ingested_count += 1
                if ingested_count:
                    st.success(f"Ingested {ingested_count} uploaded file(s).")

        if reindex_col.button("Re-index Knowledge", key="rag_reindex_btn"):
            _reindex_rag(snapshot)
            st.success("RAG index refreshed.")

        docs_df = pd.DataFrame(store.get("documents", []))
        if not docs_df.empty:
            st.markdown("**Ingested Documents**")
            st.dataframe(
                docs_df[[col for col in ["id", "title", "source", "updated_at"] if col in docs_df.columns]],
                use_container_width=True,
            )

        index_df = store.get("index", pd.DataFrame())
        st.markdown("**Index Status**")
        st.write(
            f"Version: {store.get('version', 0)} | "
            f"Chunks: {len(index_df) if isinstance(index_df, pd.DataFrame) else 0} | "
            f"Last Reindex: {store.get('last_reindexed_at')}"
        )

    if not isinstance(store.get("index"), pd.DataFrame) or store.get("index").empty:
        return _reindex_rag(snapshot)
    return store["index"]


def _orchestration_default_config() -> Dict[str, Any]:
    return {
        "investor_profile": "Growth + resilience",
        "planning_horizon_years": 5,
        "target_irr": 0.18,
        "target_ebitda_margin": 0.25,
        "min_governance_score": 80.0,
        "response_style": "Strategic and concise",
        "proactive_mode": True,
    }


def _ai_orchestration_store() -> Dict[str, Any]:
    config = st.session_state.setdefault(
        "ai_orchestration_config", _orchestration_default_config()
    )
    st.session_state.setdefault("ai_orchestration_chat_history", [])
    st.session_state.setdefault("ai_orchestration_last_query", "")
    return config


def _flatten_numeric_summary(df: Optional[pd.DataFrame], label: str) -> Dict[str, float]:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return {}
    output: Dict[str, float] = {}
    for col in df.columns:
        series = pd.to_numeric(df[col], errors="coerce")
        valid = series.dropna()
        if valid.empty:
            continue
        mean_value = valid.mean()
        if pd.isna(mean_value):
            continue
        try:
            output[f"{label}:{col}"] = float(mean_value)
        except (TypeError, ValueError):
            continue
    return output


def _build_orchestration_snapshot(results: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    assumptions = st.session_state.get("assumptions", {})
    supplementary = st.session_state.get("supplementary", {})
    analytics_framework = st.session_state.get("analytics_framework", {})

    assumption_summaries: Dict[str, float] = {}
    if isinstance(assumptions, dict):
        for name, table in assumptions.items():
            assumption_summaries.update(_flatten_numeric_summary(table, f"assumption.{name}"))

    supplementary_summaries: Dict[str, float] = {}
    if isinstance(supplementary, dict):
        for name, table in supplementary.items():
            supplementary_summaries.update(_flatten_numeric_summary(table, f"supplementary.{name}"))

    kpi_summary: Dict[str, float] = {}
    if results and isinstance(results.get("kpis"), pd.DataFrame):
        kpis = results["kpis"]
        kpi_summary = {
            col: float(pd.to_numeric(kpis[col], errors="coerce").iloc[0])
            for col in kpis.columns
            if pd.to_numeric(kpis[col], errors="coerce").notna().any()
        }

    framework_enabled = [
        key
        for key, value in analytics_framework.items()
        if isinstance(value, dict) and value.get("enabled")
    ]

    return {
        "selected_scenario": (results or {}).get("selected_scenario", "Scenario"),
        "kpis": kpi_summary,
        "assumptions": assumption_summaries,
        "supplementary": supplementary_summaries,
        "framework_enabled": framework_enabled,
    }


def _snapshot_index(snapshot: Dict[str, Any]) -> pd.DataFrame:
    records: list[Dict[str, Any]] = []
    for section in ["kpis", "assumptions", "supplementary"]:
        values = snapshot.get(section, {})
        if isinstance(values, dict):
            for key, value in values.items():
                records.append(
                    {
                        "Section": section,
                        "Key": key,
                        "Value": value,
                        "Text": f"{section}::{key}={value}",
                    }
                )

    enabled = snapshot.get("framework_enabled", [])
    for item in enabled:
        records.append(
            {
                "Section": "framework",
                "Key": item,
                "Value": 1,
                "Text": f"framework::{item}=enabled",
            }
        )

    return pd.DataFrame(records)


def _retrieve_snapshot_context(index_df: pd.DataFrame, query: str, top_n: int = 6) -> pd.DataFrame:
    if index_df.empty:
        return index_df
    tokens = {token for token in re.findall(r"[a-zA-Z0-9_]+", query.lower()) if token}
    if not tokens:
        return index_df.head(top_n)

    scored = index_df.copy()
    scored["score"] = scored["Text"].str.lower().apply(
        lambda text: sum(1 for token in tokens if token in text)
    )
    scored = scored.sort_values(["score", "Section"], ascending=[False, True])
    filtered = scored[scored["score"] > 0]
    if filtered.empty:
        return scored.head(top_n)
    return filtered.head(top_n)


def _infer_orchestration_intent(query: str) -> str:
    text = (query or "").lower()
    intent_map = {
        "risk": ["risk", "downside", "stress", "resilience", "volatility", "uncertainty"],
        "valuation": ["valuation", "irr", "npv", "return", "discount", "wacc", "enterprise value"],
        "assumptions": ["assumption", "assumptions", "driver", "input", "price", "yield", "cost"],
        "operations": ["operations", "capacity", "production", "feed", "capex", "opex", "efficiency"],
        "governance": ["governance", "control", "compliance", "board", "audit"],
        "planning": ["plan", "roadmap", "milestone", "timeline", "execution", "90-day"],
    }
    for intent, keywords in intent_map.items():
        if any(keyword in text for keyword in keywords):
            return intent
    return "strategy"


def _response_depth_from_query(query: str) -> str:
    q = (query or "").strip()
    lower = q.lower()
    if len(q) > 140 or any(
        token in lower for token in ["compare", "trade-off", "why", "how", "sensitivity"]
    ):
        return "deep"
    if any(token in lower for token in ["quick", "brief", "summary", "tldr"]):
        return "brief"
    return "standard"


def _classify_orchestration_task(query: str, intent: str) -> str:
    text = (query or "").lower()
    if any(token in text for token in ["compare", "versus", "vs ", "difference"]):
        return "comparative_assessment"
    if any(token in text for token in ["run", "execute", "perform", "do ", "action"]):
        return "action_plan"
    if intent == "valuation":
        return "valuation_diagnostic"
    if intent in {"risk", "governance"}:
        return "risk_governance_check"
    if intent in {"assumptions", "operations"}:
        return "driver_diagnostic"
    return "strategic_summary"


def _confidence_from_evidence(
    *,
    available_kpis: int,
    retrieved_hits: int,
    missing_requirements: int,
) -> float:
    score = 0.35
    score += min(0.35, available_kpis * 0.04)
    score += min(0.2, retrieved_hits * 0.03)
    score -= min(0.3, missing_requirements * 0.08)
    return float(max(0.05, min(0.95, score)))


def _run_orchestration_engine(
    query: str,
    config: Dict[str, Any],
    snapshot: Dict[str, Any],
    retrieved: pd.DataFrame,
    history: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    kpis = snapshot.get("kpis", {})
    irr = float(kpis.get("IRR", np.nan))
    ebitda_margin = float(kpis.get("EBITDA Margin", np.nan))

    target_irr = float(config.get("target_irr", 0.18))
    target_ebitda = float(config.get("target_ebitda_margin", 0.25))

    irr_gap = irr - target_irr if pd.notna(irr) else np.nan
    margin_gap = ebitda_margin - target_ebitda if pd.notna(ebitda_margin) else np.nan

    investor_score = 100.0
    if pd.notna(irr_gap):
        investor_score -= max(0.0, (target_irr - irr) * 220.0)
    if pd.notna(margin_gap):
        investor_score -= max(0.0, (target_ebitda - ebitda_margin) * 180.0)
    investor_score = float(max(0.0, min(100.0, investor_score)))

    governance_score = 100.0
    if not snapshot.get("framework_enabled"):
        governance_score -= 10.0
    if retrieved.empty:
        governance_score -= 15.0

    benchmark_table = pd.DataFrame(
        {
            "Metric": ["IRR", "EBITDA Margin", "Investor Readiness", "Governance"],
            "Actual": [irr, ebitda_margin, investor_score, governance_score],
            "Target": [target_irr, target_ebitda, 85.0, config.get("min_governance_score", 80.0)],
            "Gap": [
                irr_gap if pd.notna(irr_gap) else np.nan,
                margin_gap if pd.notna(margin_gap) else np.nan,
                investor_score - 85.0,
                governance_score - float(config.get("min_governance_score", 80.0)),
            ],
        }
    )

    recommendations: list[str] = []
    if pd.notna(irr_gap) and irr_gap < 0:
        recommendations.append(
            "Improve return profile by prioritising higher-margin products and phasing capex."
        )
    if pd.notna(margin_gap) and margin_gap < 0:
        recommendations.append(
            "Tighten feed and variable-cost controls; run sensitivity scenarios on feed inflation."
        )
    if governance_score < float(config.get("min_governance_score", 80.0)):
        recommendations.append(
            "Strengthen governance with monthly KPI packs, data-quality checks, and approval workflows."
        )
    if not recommendations:
        recommendations.append(
            "Current plan is investor-ready; focus on scaling strategy and downside protection."
        )

    context_strings = retrieved.get("Text", pd.Series(dtype=str)).astype(str).tolist()
    intent = _infer_orchestration_intent(query)
    task_type = _classify_orchestration_task(query, intent)
    depth = _response_depth_from_query(query)
    response_style = str(config.get("response_style", "Strategic and concise"))
    scenario_name = snapshot.get("selected_scenario", "Scenario")
    recent_history = history[-3:] if history else []
    prior_focus = recent_history[-1]["Query"] if recent_history else ""

    irr_text = f"{irr:.2%}" if pd.notna(irr) else "not available"
    margin_text = f"{ebitda_margin:.2%}" if pd.notna(ebitda_margin) else "not available"
    context_text = " | ".join(context_strings[:3]) if context_strings else "No indexed context found."
    lead_by_style = {
        "Strategic and concise": "Strategic view:",
        "Detailed and technical": "Technical assessment:",
        "Investor memo": "Investor memo:",
    }.get(response_style, "Strategic view:")
    intent_guidance = {
        "risk": "Downside resilience is sensitive to margin compression and governance coverage.",
        "valuation": "Valuation leans on return quality versus target hurdle rates.",
        "assumptions": "Key assumptions should be validated for pricing, yield, and cost inflation.",
        "operations": "Operational efficiency and cost control are the primary execution levers.",
        "governance": "Governance sufficiency depends on controls, cadence, and quality signals.",
        "planning": "Execution quality depends on sequencing, ownership, and measurable milestones.",
        "strategy": "Overall strategy should balance growth, resilience, and investor confidence.",
    }
    task_guidance = {
        "comparative_assessment": "Workflow selected: comparative assessment using KPI gaps and available indexed records.",
        "action_plan": "Workflow selected: actionable execution plan based on model constraints and risk posture.",
        "valuation_diagnostic": "Workflow selected: valuation diagnostic with hurdle-rate and margin validation.",
        "risk_governance_check": "Workflow selected: risk-governance check with control-threshold validation.",
        "driver_diagnostic": "Workflow selected: driver diagnostic focused on assumptions and operating levers.",
        "strategic_summary": "Workflow selected: strategic summary from current model state.",
    }
    limitations: list[str] = []
    if pd.isna(irr):
        limitations.append("IRR is unavailable in current model outputs.")
    if pd.isna(ebitda_margin):
        limitations.append("EBITDA margin is unavailable in current model outputs.")
    if retrieved.empty:
        limitations.append("No indexed retrieval evidence matched the query.")
    if not snapshot.get("framework_enabled"):
        limitations.append("No analytics framework modules are enabled for cross-checking.")

    requirements_by_task = {
        "comparative_assessment": ["IRR", "EBITDA Margin"],
        "valuation_diagnostic": ["IRR"],
        "risk_governance_check": ["EBITDA Margin"],
        "driver_diagnostic": ["EBITDA Margin"],
        "action_plan": [],
        "strategic_summary": [],
    }
    missing_requirements = [
        req
        for req in requirements_by_task.get(task_type, [])
        if (req == "IRR" and pd.isna(irr)) or (req == "EBITDA Margin" and pd.isna(ebitda_margin))
    ]
    if missing_requirements:
        limitations.append(
            f"Requested workflow has missing required metrics: {', '.join(missing_requirements)}."
        )

    assumptions_used = [
        f"Target IRR={target_irr:.2%}",
        f"Target EBITDA margin={target_ebitda:.2%}",
        f"Minimum governance score={float(config.get('min_governance_score', 80.0)):.1f}",
    ]
    confidence = _confidence_from_evidence(
        available_kpis=len(kpis),
        retrieved_hits=len(context_strings),
        missing_requirements=len(missing_requirements),
    )
    continuity_line = (
        f"Building on your prior question ('{prior_focus}'), "
        if prior_focus and prior_focus.strip().lower() != (query or "").strip().lower()
        else ""
    )
    depth_line = {
        "brief": "Net: maintain focus on the top one to two value levers this quarter.",
        "standard": "Recommended focus: protect margin while improving return quality and governance consistency.",
        "deep": "Priority sequence: 1) stabilize unit economics, 2) run downside stress pack, 3) tighten governance reporting against investor thresholds.",
    }[depth]
    if missing_requirements and task_type in {"comparative_assessment", "valuation_diagnostic"}:
        grounded_answer = (
            f"{lead_by_style} {continuity_line}I cannot fully execute the requested {task_type.replace('_', ' ')} "
            f"because required evidence is missing ({', '.join(missing_requirements)}). "
            f"Available evidence for scenario '{scenario_name}': IRR={irr_text}, EBITDA margin={margin_text}. "
            f"Indexed context: {context_text} "
            "Please provide the missing data or run the scenario again."
        )
    else:
        grounded_answer = (
            f"{lead_by_style} {continuity_line}For scenario '{scenario_name}', IRR is {irr_text} and EBITDA margin is {margin_text}. "
            f"{task_guidance.get(task_type, task_guidance['strategic_summary'])} "
            f"{intent_guidance.get(intent, intent_guidance['strategy'])} "
            f"Indexed context: {context_text} "
            f"{depth_line}"
        )

    business_plan = pd.DataFrame(
        {
            "Workstream": ["Profitability", "Resilience", "Investor Readiness", "Execution"],
            "Priority": ["High", "High", "Medium", "Medium"],
            "90-Day Action": [
                "Re-price weak-margin products and optimize feed contracts.",
                "Run downside stress pack (drought + price collapse + rates up).",
                "Produce benchmark scorecard and governance dashboard.",
                "Publish quarterly milestones with owners and KPIs.",
            ],
        }
    )

    explainability = pd.DataFrame(
        {
            "Driver": ["IRR Gap", "EBITDA Margin Gap", "Governance Buffer"],
            "Contribution": [
                float(irr_gap) if pd.notna(irr_gap) else 0.0,
                float(margin_gap) if pd.notna(margin_gap) else 0.0,
                float(governance_score - config.get("min_governance_score", 80.0)),
            ],
            "Interpretation": [
                "Positive means IRR is above target.",
                "Positive means operating margin is above target.",
                "Positive means governance controls exceed threshold.",
            ],
        }
    )

    return {
        "investor_readiness_score": investor_score,
        "governance_score": governance_score,
        "benchmark_table": benchmark_table,
        "recommendations": recommendations,
        "grounded_answer": grounded_answer,
        "task_type": task_type,
        "confidence": confidence,
        "limitations": limitations,
        "assumptions_used": assumptions_used,
        "business_plan": business_plan,
        "explainability": explainability,
        "retrieved_context": retrieved,
    }


def _render_ai_orchestration_layer(results: Optional[Dict[str, Any]]) -> None:
    st.subheader("AI Decision Making — Unified Orchestration Layer")
    st.caption(
        "A single context-aware intelligence engine that unifies benchmarking, governance, "
        "RAG-style retrieval, Q&A, scenario reasoning, and planning."
    )

    config = _ai_orchestration_store()
    shared_context = _sync_shared_model_context(results)
    scenario_label = shared_context.get("active_result_scenario") or shared_context.get(
        "selected_scenario_name", "Scenario"
    )
    st.caption(f"Using shared model context for scenario: **{scenario_label}**")
    snapshot = _build_orchestration_snapshot(results)
    with st.expander("Unified Orchestration (Config + Runtime + RAG)", expanded=True):
        st.markdown("#### Unified Configuration Model")
        c1, c2, c3 = st.columns(3)
        config["investor_profile"] = c1.text_input(
            "Investor Profile",
            value=str(config.get("investor_profile", "Growth + resilience")),
        )
        config["planning_horizon_years"] = int(
            c1.number_input(
                "Planning Horizon (Years)",
                min_value=1,
                max_value=15,
                value=int(config.get("planning_horizon_years", 5)),
                step=1,
            )
        )
        config["target_irr"] = float(
            c2.number_input(
                "Target IRR",
                min_value=0.0,
                max_value=1.0,
                value=float(config.get("target_irr", 0.18)),
                step=0.01,
                format="%.2f",
            )
        )
        config["target_ebitda_margin"] = float(
            c2.number_input(
                "Target EBITDA Margin",
                min_value=0.0,
                max_value=1.0,
                value=float(config.get("target_ebitda_margin", 0.25)),
                step=0.01,
                format="%.2f",
            )
        )
        config["min_governance_score"] = float(
            c3.number_input(
                "Minimum Governance Score",
                min_value=0.0,
                max_value=100.0,
                value=float(config.get("min_governance_score", 80.0)),
                step=1.0,
            )
        )
        config["response_style"] = c3.selectbox(
            "Response Style",
            options=["Strategic and concise", "Detailed and technical", "Investor memo"],
            index=[
                "Strategic and concise",
                "Detailed and technical",
                "Investor memo",
            ].index(config.get("response_style", "Strategic and concise")),
        )
        config["proactive_mode"] = st.checkbox(
            "Proactive Recommendation Mode",
            value=bool(config.get("proactive_mode", True)),
        )
        st.session_state["ai_orchestration_config"] = config

        st.markdown("#### LLM & ML Runtime Settings")
        _render_ai_settings(st.session_state.setdefault("ai_payload", {}))

        st.markdown("#### Retrieval-Augmented Generation (RAG)")
        index_df = _render_rag_admin(snapshot, show_header=False, use_expander=False)

    messages = st.session_state.setdefault("ai_orchestration_chat_messages", [])
    st.markdown("### Intelligent Orchestration Chat")
    for msg in messages[-12:]:
        with st.chat_message(msg.get("role", "assistant")):
            st.markdown(msg.get("content", ""))

    with st.form("ai_orchestration_chat_form", clear_on_submit=True):
        prompt = st.text_area(
            "Ask a strategic question",
            key="ai_orchestration_prompt",
            placeholder="Compare downside resilience and investor readiness.",
            height=100,
        ).strip()
        submit_prompt = st.form_submit_button("Generate response", type="primary")

    history = st.session_state.setdefault("ai_orchestration_chat_history", [])
    current_query = ""
    should_record_history = False
    cache_signature = {
        "scenario": snapshot.get("selected_scenario", "Scenario"),
        "run_version": _safe_session_state_get(MODEL_LAST_RUN_VERSION_KEY),
    }
    cached_payload = st.session_state.get("ai_orchestration_cached_output")
    output = None
    if submit_prompt and prompt:
        st.session_state["ai_orchestration_last_query"] = prompt
        current_query = prompt
        retrieved = _retrieve_rag_context(prompt)
        output = _run_orchestration_engine(prompt, config, snapshot, retrieved, history)
        assistant_reply = (
            f"{output['grounded_answer']}\n\n"
            f"Top recommendations:\n- " + "\n- ".join(output["recommendations"])
        )
        messages.append({"role": "user", "content": prompt})
        messages.append({"role": "assistant", "content": assistant_reply})
        st.session_state["ai_orchestration_chat_messages"] = messages
        st.session_state["ai_orchestration_cached_output"] = {
            "signature": cache_signature,
            "query": prompt,
            "output": output,
            "retrieved": retrieved,
        }
        should_record_history = True
    elif submit_prompt:
        st.warning("Enter a question before generating a response.")
    else:
        if (
            isinstance(cached_payload, dict)
            and cached_payload.get("signature") == cache_signature
            and isinstance(cached_payload.get("output"), dict)
        ):
            current_query = str(cached_payload.get("query", "")).strip()
            output = cached_payload["output"]
        else:
            st.info(
                "Generate a response to analyze the current scenario. The orchestration engine now stays idle until requested."
            )
            return

    if should_record_history and current_query.strip():
        history.append(
            {
                "Query": current_query.strip(),
                "Response": output["grounded_answer"],
                "Scenario": snapshot.get("selected_scenario", "Scenario"),
            }
        )
        st.session_state["ai_orchestration_chat_history"] = history[-20:]
    if history:
        st.markdown("**Context Retention (Recent Q&A)**")
        st.dataframe(pd.DataFrame(history[-10:]), use_container_width=True)

    score_cols = st.columns(3)
    score_cols[0].metric(
        "Investor Readiness",
        f"{output['investor_readiness_score']:.1f}/100",
    )
    score_cols[1].metric("Governance", f"{output['governance_score']:.1f}/100")
    score_cols[2].metric(
        "Indexed Knowledge Records",
        len(index_df) if isinstance(index_df, pd.DataFrame) else 0,
    )

    st.markdown("### Unified Decision Support Output")
    st.markdown("**Grounded Strategic Response**")
    st.write(output["grounded_answer"])
    st.caption(
        f"Workflow: **{output.get('task_type', 'strategic_summary').replace('_', ' ')}** | "
        f"Confidence: **{output.get('confidence', 0.0):.0%}**"
    )
    limitations = output.get("limitations", [])
    if limitations:
        st.warning("Limitations: " + " ".join(f"- {item}" for item in limitations))
    assumptions_used = output.get("assumptions_used", [])
    if assumptions_used:
        st.info("Assumptions used: " + "; ".join(assumptions_used))

    st.markdown("**Investor Benchmarking & Scorecard**")
    st.dataframe(output["benchmark_table"], use_container_width=True)

    st.markdown("**Governance + Explainability**")
    g1, g2 = st.columns(2)
    with g1:
        st.dataframe(output["explainability"], use_container_width=True)
    with g2:
        st.markdown("**Proactive Recommendations**")
        for rec in output["recommendations"]:
            st.write(f"- {rec}")

    st.markdown("**Business Planning Actions (Scenario-Aware)**")
    st.dataframe(output["business_plan"], use_container_width=True)

    st.markdown("**RAG Retrieval Context (Indexed Knowledge)**")
    st.dataframe(output["retrieved_context"], use_container_width=True)


def main() -> None:
    _inject_app_theme()
    _render_model_hero()

    if "schedule" in st.session_state:
        st.session_state.pop("schedule")

    if DEFAULT_INPUT_CONFIG_KEY not in st.session_state:
        st.session_state[DEFAULT_INPUT_CONFIG_KEY] = _default_input_template_config()

    if "assumptions" not in st.session_state:
        st.session_state.assumptions = _default_assumption_tables()
    else:
        defaults = _default_assumption_tables()
        for name, table in defaults.items():
            st.session_state.assumptions.setdefault(name, table.copy())

    production_horizon_defaults = st.session_state.assumptions.get("Production Horizon")
    st.session_state.setdefault("schedule_period_type", "quarterly")
    st.session_state["schedule_period_type"] = _normalize_period_type(
        st.session_state.get("schedule_period_type")
    )
    st.session_state.assumptions = _sync_biological_start_date_in_assumptions(
        st.session_state.assumptions,
        period_type=st.session_state["schedule_period_type"],
    )
    st.session_state.assumptions = _sync_opening_herd_cohorts_in_assumptions(
        st.session_state.assumptions
    )
    production_horizon_defaults = st.session_state.assumptions.get("Production Horizon")

    if "core_schedule" not in st.session_state or "detail_schedules" not in st.session_state:
        core_default, detail_defaults = _default_schedule_components(
            production_horizon=production_horizon_defaults,
            period_type=st.session_state["schedule_period_type"],
            assumptions=st.session_state.assumptions,
        )
        if "core_schedule" not in st.session_state:
            st.session_state.core_schedule = core_default
        if "detail_schedules" not in st.session_state:
            st.session_state.detail_schedules = detail_defaults
    else:
        st.session_state["schedule_period_type"] = _infer_period_type_from_schedule(
            st.session_state.get("core_schedule")
        )

    st.session_state.assumptions = _sync_commercial_assumptions_to_core(
        st.session_state.assumptions,
        st.session_state.core_schedule,
    )
    if "supplementary" not in st.session_state:
        st.session_state.supplementary = _default_supplementary_tables()
    st.session_state.supplementary = _sync_asset_schedule_from_capex_in_supplementary(
        st.session_state.get("supplementary")
    )
    if "all_scenario_results" not in st.session_state:
        st.session_state.all_scenario_results = {}
    if "selected_scenario_name" not in st.session_state:
        st.session_state.selected_scenario_name = next(iter(SCENARIO_PRESETS))
    if "results" not in st.session_state:
        st.session_state.results = None

    _ensure_default_results_loaded()
    _refresh_results_stale_state()

    _ensure_active_scenario_selection()
    _sync_shared_model_context(st.session_state.get("results"))

    excel_download_container = st.container()

    ai_payload = st.session_state.setdefault("ai_payload", {})

    milk_price = 0
    feed_cost = 0
    valuation_inputs: Dict[str, float] = {}
    include_valuation = False
    run_clicked = False
    run_feedback_slot = None

    supplementary_tables: Dict[str, pd.DataFrame] = {}
    detail_tables_for_run: Dict[str, pd.DataFrame] = {}
    core_editor: Optional[pd.DataFrame] = None

    main_sections = [
        "Assumptions",
        "Input Schedule",
        "Financials",
        "Dashboard",
        "Advanced Analytics",
        "AI Decision Making",
    ]
    pending_main_section = st.session_state.pop("_pending_main_section_selector", None)
    if pending_main_section in main_sections:
        _safe_session_state_set("main_section_selector", pending_main_section)

    active_main_section = _section_selector(
        "Workspace Section",
        main_sections,
        key="main_section_selector",
        help="Only the selected section is rendered to keep the app responsive.",
    )

    if active_main_section == "Input Schedule":
        st.subheader("Input Schedule")
        _render_workflow_status_strip()
        st.caption(
            "Translate the planning, commercial, operating, labour, and funding assumptions into period-by-period "
            "schedules. Build the operating model first, then complete the capital and asset support schedules."
        )
        current_period_type = _normalize_period_type(
            st.session_state.get("schedule_period_type")
        )
        period_choice = st.selectbox(
            "Schedule period type",
            options=["Monthly", "Quarterly"],
            index=0 if current_period_type == "monthly" else 1,
            key="schedule_period_type_selector",
            help=(
                "Applies to Core, COGS, Variable Expenses, Direct Wages, and Admin Wages schedules. "
                "Quarterly is the recommended planning default; monthly remains available for operating-detail use. "
                "Capex remains year-based and aligned to the active production horizon."
            ),
        )
        selected_period_type = _normalize_period_type(period_choice)
        if selected_period_type != current_period_type:
            _sync_schedule_period_type(selected_period_type)
            st.success(f"Updated schedule period type to {_period_label(selected_period_type)}.")

        st.markdown("### Scenario Explorer")
        _render_model_author_editor()
        _render_scenario_selector()
        _render_scenario_preset_editors()

        st.markdown("### Operating Model Schedules")
        st.caption(
            "These schedules should follow the Assumptions page flow: planning and herd assumptions shape the core "
            "schedule, commercial assumptions support revenue and COGS, and labour assumptions support the wage schedules."
        )

        schedule_tab_names = [
            "Core Schedule",
            "COGS Schedule",
            "Variable Expenses Schedule",
            "Direct Wages Schedule",
            "Admin Wages Schedule",
        ]
        current_assumption_snapshot = {
            name: table.copy()
            for name, table in (st.session_state.get("assumptions") or {}).items()
            if isinstance(table, pd.DataFrame)
        }
        generated_schedule_tab_names: list[str] = []
        if _is_breeding_to_unit_mode(current_assumption_snapshot):
            generated_schedule_tab_names = [
                "Kid Availability",
                "Kid Routing",
                "Internal Transfers",
                "Downstream Intake",
                "Unit Herd Schedules",
            ]
        schedule_section_labels = [
            "1. Core Schedule",
            "2. COGS Schedule",
            "3. Variable Expenses",
            "4. Direct Wages",
            "5. Admin Wages",
            *[
                f"{idx}. {name}"
                for idx, name in enumerate(generated_schedule_tab_names, start=6)
            ],
        ]
        active_schedule_section = _section_selector(
            "Schedule Section",
            schedule_section_labels,
            key="input_schedule_section_selector",
            help="Only the selected schedule editor is rendered.",
        )

        if active_schedule_section == schedule_section_labels[0]:
            core_table = st.session_state.get("core_schedule")
            if not isinstance(core_table, pd.DataFrame):
                core_table = pd.DataFrame()
                st.session_state.core_schedule = core_table

            _render_schedule_row_editor(
                "core_schedule",
                st.session_state.core_schedule,
                lambda updated: st.session_state.__setitem__(
                    "core_schedule", updated
                ),
            )
            core_editor = st.session_state.core_schedule
            st.markdown("### Capital & Asset Support Schedules")
            st.caption(
                "Complete these support schedules after the operating model tabs. They should stay aligned with the "
                "Capital & Financing assumptions and provide the asset, capex, and ownership detail that supports the model."
            )
            for name in list(st.session_state.supplementary.keys()):
                if name == "Loan Facilities":
                    st.markdown("#### Loan Facilities")
                    loan_table = _ensure_loan_facilities_table(
                        st.session_state.supplementary.get(name)
                    )
                    st.session_state.supplementary[name] = loan_table

                    def _save_loans(updated: pd.DataFrame) -> None:
                        st.session_state.supplementary[name] = _ensure_loan_facilities_table(
                            updated
                        )

                    _render_schedule_row_editor(
                        "supp::loan_facilities",
                        loan_table,
                        _save_loans,
                    )

                    cleaned_loans = _clean_editor_table(
                        st.session_state.supplementary[name]
                    )
                    if cleaned_loans is not None:
                        supplementary_tables[name] = _ensure_loan_facilities_table(
                            cleaned_loans
                        )
                    continue

                if name == "Equity Facilities":
                    st.markdown("#### Equity Facilities")
                    equity_table = _ensure_equity_facilities_table(
                        st.session_state.supplementary.get(name)
                    )
                    st.session_state.supplementary[name] = equity_table

                    def _save_equity(updated: pd.DataFrame) -> None:
                        st.session_state.supplementary[name] = _ensure_equity_facilities_table(
                            updated
                        )

                    _render_schedule_row_editor(
                        "supp::equity_facilities",
                        equity_table,
                        _save_equity,
                    )

                    cleaned_equity = _clean_editor_table(
                        st.session_state.supplementary[name]
                    )
                    if cleaned_equity is not None:
                        supplementary_tables[name] = _ensure_equity_facilities_table(
                            cleaned_equity
                        )
                    continue

                if name == "Capitalisation Table":
                    st.markdown("#### Capitalisation Table Schedule")
                    cap_table = _ensure_capitalisation_table(
                        st.session_state.supplementary.get(name)
                    )
                    st.session_state.supplementary[name] = cap_table

                    add_col, remove_select_col, remove_btn_col, inc_col_col, inc_pct_col = st.columns(
                        [1, 2, 1, 2, 2]
                    )

                    with add_col:
                        if st.button("Add Row", key="cap_table_add_row"):
                            cap_table = _add_capitalisation_row(cap_table)
                            st.session_state.supplementary[name] = cap_table
                            _reset_cached_results()
                            _clear_schedule_editor_state("supp::capitalisation_table")

                    option_labels = []
                    option_map: Dict[str, int] = {}
                    for idx, row in cap_table.iterrows():
                        year = row.get("Year")
                        shareholder = row.get("Shareholder") or "Unnamed"
                        if pd.isna(year):
                            label = f"{shareholder}"
                        else:
                            label = f"{int(year)} – {shareholder}"
                        option_labels.append(label)
                        option_map[label] = idx

                    remove_choice = None
                    with remove_select_col:
                        if option_labels:
                            remove_choice = st.selectbox(
                                "Select row",
                                options=["-- Select Row --"] + option_labels,
                                key="cap_table_remove_choice",
                            )
                        else:
                            st.write("No rows available to remove.")

                    with remove_btn_col:
                        if (
                            st.button("Remove Row", key="cap_table_remove_button")
                            and remove_choice
                            and remove_choice in option_map
                        ):
                            cap_table = _remove_capitalisation_row(
                                cap_table, option_map[remove_choice]
                            )
                            st.session_state.supplementary[name] = cap_table
                            _reset_cached_results()
                            _clear_schedule_editor_state("supp::capitalisation_table")

                    with inc_col_col:
                        increment_column = st.selectbox(
                            "Increment column",
                            options=["Ownership %", "Investment"],
                            key="cap_table_increment_column",
                        )

                    with inc_pct_col:
                        increment_pct = st.number_input(
                            "Yearly increment (%)",
                            value=0.0,
                            step=0.5,
                            key="cap_table_increment_pct",
                        )
                        if st.button("Apply", key="cap_table_increment_button"):
                            cap_table = _apply_capitalisation_increment(
                                cap_table, increment_column, increment_pct
                            )

                            st.session_state.supplementary[name] = cap_table
                            _reset_cached_results()
                            _clear_schedule_editor_state("supp::capitalisation_table")

                    cap_table = st.session_state.supplementary[name]

                    def _save_capitalisation(updated: pd.DataFrame) -> None:
                        ensured = _ensure_capitalisation_table(updated)
                        st.session_state.supplementary[name] = ensured

                    _render_schedule_row_editor(
                        "supp::capitalisation_table", cap_table, _save_capitalisation
                    )

                    cleaned_cap = _clean_editor_table(
                        st.session_state.supplementary[name]
                    )
                    if cleaned_cap is not None:
                        supplementary_tables[name] = _ensure_capitalisation_table(cleaned_cap)
                    continue
                if name == "Capex Schedule":
                    st.markdown("#### Capex Schedule")
                    capex_table = _ensure_capex_schedule(
                        st.session_state.supplementary.get(name)
                    )
                    st.session_state.supplementary[name] = capex_table

                    rate_series = pd.to_numeric(
                        capex_table.get("Depreciation Rate %"), errors="coerce"
                    ).dropna()
                    spend_series = pd.to_numeric(
                        capex_table.get("Spend"), errors="coerce"
                    ).dropna()

                    default_rate = float(rate_series.iloc[0]) if not rate_series.empty else 10.0
                    default_spend = float(spend_series.iloc[0]) if not spend_series.empty else 0.0

                    st.session_state.setdefault("capex_rate_default", round(default_rate, 2))
                    st.session_state.setdefault("capex_default_spend", default_spend)
                    st.session_state.setdefault("capex_remove_choice", "-- Select Row --")
                    st.session_state.setdefault("capex_increment_column", "Spend")
                    st.session_state.setdefault("capex_increment_pct", 0.0)

                    rate_col, rate_btn_col, spend_col = st.columns([1, 1, 1])

                    rate_value = rate_col.number_input(
                        "Default depreciation rate (%)",
                        min_value=0.0,
                        max_value=100.0,
                        step=0.1,
                        key="capex_rate_default",
                    )
                    if rate_btn_col.button("Apply rate to all", key="capex_apply_rate"):
                        capex_table = _apply_capex_rate(capex_table, rate_value)
                        st.session_state.supplementary[name] = capex_table
                        st.session_state.supplementary = _sync_asset_schedule_from_capex_in_supplementary(
                            st.session_state.supplementary
                        )
                        _reset_cached_results()
                        _clear_schedule_editor_state("supp::capex_schedule")

                    spend_default = spend_col.number_input(
                        "Default spend for new row",
                        min_value=0.0,
                        step=100.0,
                        key="capex_default_spend",
                    )

                    add_col, remove_select_col, remove_btn_col = st.columns([1, 2, 1])

                    if add_col.button("Add Row", key="capex_add_row"):
                        capex_table = _add_capex_row(
                            capex_table,
                            default_rate=rate_value,
                            default_spend=spend_default,
                        )
                        st.session_state.supplementary[name] = capex_table
                        st.session_state.supplementary = _sync_asset_schedule_from_capex_in_supplementary(
                            st.session_state.supplementary
                        )
                        _reset_cached_results()
                        _clear_schedule_editor_state("supp::capex_schedule")

                    option_labels: list[str] = []
                    option_map: Dict[str, int] = {}
                    for idx, row in capex_table.iterrows():
                        label_year = row.get("Year")
                        label_category = row.get("Category") or "Unnamed"
                        if pd.notna(label_year):
                            label = f"{int(label_year)} – {label_category}"
                        else:
                            label = str(label_category)
                        option_labels.append(label)
                        option_map[label] = idx

                    remove_choice = remove_select_col.selectbox(
                        "Select row",
                        options=["-- Select Row --"] + option_labels,
                        key="capex_remove_choice",
                    )
                    if remove_btn_col.button("Remove Row", key="capex_remove_row"):
                        if remove_choice in option_map:
                            capex_table = _remove_capex_row(
                                capex_table, option_map[remove_choice]
                            )
                            st.session_state.capex_remove_choice = "-- Select Row --"
                            st.session_state.supplementary[name] = capex_table
                            st.session_state.supplementary = _sync_asset_schedule_from_capex_in_supplementary(
                                st.session_state.supplementary
                            )
                            _reset_cached_results()
                            _clear_schedule_editor_state("supp::capex_schedule")

                    inc_col, inc_pct_col, inc_btn_col = st.columns([1.5, 1, 1])

                    increment_column = inc_col.selectbox(
                        "Increment column",
                        options=["Spend", "Depreciation Rate %"],
                        key="capex_increment_column",
                    )
                    increment_pct = inc_pct_col.number_input(
                        "Yearly increment (%)",
                        min_value=-100.0,
                        max_value=100.0,
                        step=0.1,
                        key="capex_increment_pct",
                    )
                    if inc_btn_col.button("Apply increment", key="capex_apply_increment"):
                        capex_table = _apply_capex_yearly_increment(
                            capex_table, increment_column, increment_pct
                        )
                        st.session_state.supplementary[name] = capex_table
                        st.session_state.supplementary = _sync_asset_schedule_from_capex_in_supplementary(
                            st.session_state.supplementary
                        )
                        _reset_cached_results()
                        _clear_schedule_editor_state("supp::capex_schedule")

                    capex_table = st.session_state.supplementary[name]

                    def _save_capex(updated: pd.DataFrame) -> None:
                        ensured = _ensure_capex_schedule(updated)
                        st.session_state.supplementary[name] = ensured
                        st.session_state.supplementary = _sync_asset_schedule_from_capex_in_supplementary(
                            st.session_state.supplementary
                        )

                    _render_schedule_row_editor(
                        "supp::capex_schedule", capex_table, _save_capex
                    )

                    cleaned_capex = _clean_editor_table(
                        st.session_state.supplementary[name]
                    )
                    if cleaned_capex is not None:
                        supplementary_tables[name] = _ensure_capex_schedule(cleaned_capex)
                    continue
                if name == "Asset Schedules":
                    st.markdown("#### Asset Schedule")
                    st.caption(
                        "Derived automatically from the Capex Schedule. Edit capex rows above to change asset additions, depreciation, and NBV roll-forwards."
                    )
                    st.session_state.supplementary = _sync_asset_schedule_from_capex_in_supplementary(
                        st.session_state.supplementary
                    )
                    asset_table = _ensure_asset_schedule(
                        st.session_state.supplementary.get(name)
                    )
                    recon_issues = _capex_asset_reconciliation_issues(
                        st.session_state.supplementary.get("Capex Schedule"),
                        asset_table,
                    )
                    if recon_issues:
                        st.warning(" ".join(f"- {issue}" for issue in recon_issues))
                    else:
                        st.success("Asset Schedule reconciles to Capex Schedule by year/category.")
                    st.dataframe(asset_table, use_container_width=True)
                    supplementary_tables[name] = asset_table.copy()
                    continue
                if False and name == "Asset Schedules":
                    st.markdown("#### Asset Schedule")
                    asset_table = _ensure_asset_schedule(
                        st.session_state.supplementary.get(name)
                    )
                    st.session_state.supplementary[name] = asset_table

                    rate_series = pd.to_numeric(
                        asset_table.get("Depreciation Rate %"), errors="coerce"
                    ).dropna()
                    addition_series = pd.to_numeric(
                        asset_table.get("Additions"), errors="coerce"
                    ).dropna()

                    default_rate = float(rate_series.iloc[0]) if not rate_series.empty else 5.0
                    default_addition = (
                        float(addition_series.iloc[0]) if not addition_series.empty else 0.0
                    )

                    st.session_state.setdefault("asset_rate_default", round(default_rate, 2))
                    st.session_state.setdefault("asset_add_base", default_addition)
                    st.session_state.setdefault("asset_add_increment", 0.0)
                    st.session_state.setdefault("asset_increment_column", "Depreciation Rate %")
                    st.session_state.setdefault("asset_increment_pct", 0.0)
                    st.session_state.setdefault("asset_remove_choice", "-- Select Row --")

                    rate_col, rate_btn_col, add_base_col, add_inc_col, add_btn_col = st.columns(
                        [1, 1, 1, 1, 1]
                    )

                    rate_value = rate_col.number_input(
                        "Default depreciation rate (%)",
                        min_value=0.0,
                        max_value=100.0,
                        step=0.1,
                        key="asset_rate_default",
                    )
                    if rate_btn_col.button("Apply rate to all", key="asset_apply_rate"):
                        asset_table = _apply_asset_rate(asset_table, rate_value)
                        st.session_state.supplementary[name] = asset_table
                        _clear_schedule_editor_state("supp::asset_schedule")

                    base_add_value = add_base_col.number_input(
                        "Base additions",
                        min_value=0.0,
                        step=100.0,
                        key="asset_add_base",
                    )
                    add_increment = add_inc_col.number_input(
                        "Additions yearly increment (%)",
                        min_value=-100.0,
                        max_value=100.0,
                        step=0.1,
                        key="asset_add_increment",
                    )
                    if add_btn_col.button(
                        "Apply additions pattern", key="asset_apply_additions"
                    ):
                        asset_table = _apply_asset_additions_pattern(
                            asset_table, base_add_value, add_increment
                        )
                        st.session_state.supplementary[name] = asset_table
                        _clear_schedule_editor_state("supp::asset_schedule")

                    (
                        add_row_col,
                        remove_select_col,
                        remove_btn_col,
                        inc_column_col,
                        inc_pct_col,
                        inc_btn_col,
                    ) = st.columns([1, 2, 1, 1.5, 1, 1])

                    if add_row_col.button("Add Asset", key="asset_add_row"):
                        asset_table = _add_asset_row(
                            asset_table,
                            default_rate=rate_value,
                            default_additions=base_add_value,
                        )
                        st.session_state.supplementary[name] = asset_table
                        _clear_schedule_editor_state("supp::asset_schedule")

                    option_labels: list[str] = []
                    option_map: Dict[str, int] = {}
                    for idx, row in asset_table.iterrows():
                        label_year = row.get("Year")
                        label_asset = row.get("Asset") or "Unnamed"
                        if pd.notna(label_year):
                            label = f"{int(label_year)} – {label_asset}"
                        else:
                            label = str(label_asset)
                        option_labels.append(label)
                        option_map[label] = idx

                    remove_choice = remove_select_col.selectbox(
                        "Select asset",
                        options=["-- Select Row --"] + option_labels,
                        key="asset_remove_choice",
                    )
                    if remove_btn_col.button("Remove Asset", key="asset_remove_row"):
                        if remove_choice in option_map:
                            asset_table = _remove_asset_row(
                                asset_table, option_map[remove_choice]
                            )
                            st.session_state.asset_remove_choice = "-- Select Row --"
                            st.session_state.supplementary[name] = asset_table
                            _clear_schedule_editor_state("supp::asset_schedule")

                    increment_column = inc_column_col.selectbox(
                        "Increment column",
                        options=[
                            "Depreciation Rate %",
                            "Additions",
                            "Opening NBV",
                            "Depreciation",
                        ],
                        key="asset_increment_column",
                    )
                    increment_pct = inc_pct_col.number_input(
                        "Yearly increment (%)",
                        min_value=-100.0,
                        max_value=100.0,
                        step=0.1,
                        key="asset_increment_pct",
                    )
                    if inc_btn_col.button("Apply increment", key="asset_apply_increment"):
                        asset_table = _apply_asset_yearly_increment(
                            asset_table, increment_column, increment_pct
                        )
                        st.session_state.supplementary[name] = asset_table
                        _clear_schedule_editor_state("supp::asset_schedule")

                    asset_table = st.session_state.supplementary[name]

                    def _save_asset(updated: pd.DataFrame) -> None:
                        ensured = _ensure_asset_schedule(updated)
                        st.session_state.supplementary[name] = ensured

                    _render_schedule_row_editor(
                        "supp::asset_schedule", asset_table, _save_asset
                    )

                    cleaned_asset = _clean_editor_table(
                        st.session_state.supplementary[name]
                    )
                    if cleaned_asset is not None:
                        supplementary_tables[name] = _ensure_asset_schedule(cleaned_asset)
                    continue

                with st.expander(name, expanded=False):
                    table = st.session_state.supplementary.get(name, pd.DataFrame())
                    if not isinstance(table, pd.DataFrame):
                        table = pd.DataFrame()
                        st.session_state.supplementary[name] = table

                    def _save_generic(updated: pd.DataFrame, table_name: str = name) -> None:
                        st.session_state.supplementary[table_name] = updated

                    _render_schedule_row_editor(
                        f"supp::{_scenario_key_suffix(name)}",
                        table,
                        _save_generic,
                    )

                cleaned = _clean_editor_table(
                    st.session_state.supplementary.get(name, pd.DataFrame())
                )
                if cleaned is not None:
                    supplementary_tables[name] = cleaned
                else:
                    st.session_state.supplementary[name] = (
                        st.session_state.supplementary.get(name, pd.DataFrame())
                    )

        for idx, name in enumerate(schedule_tab_names[1:], start=1):
            if active_schedule_section != schedule_section_labels[idx]:
                continue
            if name == "COGS Schedule":
                st.markdown("#### Cost of Goods Sold Schedule")
                st.caption(
                    "Adjust COGS as a percentage of revenue, add or remove periods, and apply "
                    "automatic yearly increments. Amounts update automatically when revenue is available."
                )

                cogs_table = _ensure_cogs_schedule(
                    st.session_state.detail_schedules.get(name, pd.DataFrame()),
                    st.session_state.core_schedule,
                )
                cogs_table = _sync_cogs_from_operating_assumptions(
                    cogs_table,
                    st.session_state.core_schedule,
                    st.session_state.get("assumptions"),
                )
                cogs_table = render_cogs_schedule_editor(
                    cogs_table=cogs_table,
                    core_schedule=st.session_state.core_schedule,
                    save_table=lambda updated, schedule_name=name: _schedule_editor_save_table(
                        schedule_name,
                        updated,
                    ),
                    render_row_editor=_render_schedule_row_editor,
                    clear_editor_state=_clear_schedule_editor_state,
                    apply_pct_fn=_apply_cogs_percentage,
                    apply_increment_fn=_apply_yearly_increment,
                    add_row_fn=_add_cogs_row,
                    remove_row_fn=_remove_cogs_row,
                    sync_fn=_sync_cogs_table,
                    ensure_fn=_ensure_cogs_schedule,
                )
                detail_tables_for_run[name] = cogs_table
            elif name == "Variable Expenses Schedule":
                st.markdown("#### Variable Expenses Schedule")
                st.caption(
                    "Manage individual variable cost items, add or remove rows, and apply yearly increments "
                    "to quickly escalate recurring expenses. Amounts roll up into the income statement automatically."
                )

                variable_spec = _variable_expense_schedule_editor_spec()
                _, aggregated_variable = render_incremental_schedule_editor(
                    spec=variable_spec,
                    table=st.session_state.detail_schedules.get(name, pd.DataFrame()),
                    core_schedule=st.session_state.core_schedule,
                    save_table=_schedule_editor_save_table,
                    render_row_editor=_render_schedule_row_editor,
                    clear_editor_state=_clear_schedule_editor_state,
                )

                st.info(
                    "Master variable-expense inputs now live on the Assumptions page. Use `Apply to Schedule` "
                    "there to regenerate this schedule across the full production horizon, then refine period "
                    "rows here only when needed."
                )

                detail_tables_for_run[name] = aggregated_variable

                render_schedule_summary("Variable Expenses Summary", aggregated_variable)
            elif name == "Direct Wages Schedule":
                st.markdown("#### Direct Wages Schedule")
                st.caption(
                    "Capture direct labour by position, headcount, and monthly salary per head. Position totals "
                    "roll up automatically into the model's EBITDA calculations."
                )

                direct_spec = _direct_wage_schedule_editor_spec()
                _, aggregated_direct = render_incremental_schedule_editor(
                    spec=direct_spec,
                    table=st.session_state.detail_schedules.get(name, pd.DataFrame()),
                    core_schedule=st.session_state.core_schedule,
                    save_table=_schedule_editor_save_table,
                    render_row_editor=_render_schedule_row_editor,
                    clear_editor_state=_clear_schedule_editor_state,
                )

                st.info(
                    "Master direct-wage inputs now live on the Assumptions page. Use `Apply to Schedule` there "
                    "to rebuild labour rows across the full horizon, then use this schedule for downstream "
                    "refinements only."
                )

                detail_tables_for_run[name] = aggregated_direct

                render_schedule_summary("Direct Wages Summary", aggregated_direct)
            elif name == "Admin Wages Schedule":
                st.markdown("#### Admin Wages Schedule")
                st.caption(
                    "Capture administrative labour by position, headcount, and monthly salary per head. Position totals "
                    "roll up automatically into the income statement."
                )

                admin_spec = _admin_wage_schedule_editor_spec()
                _, aggregated_admin = render_incremental_schedule_editor(
                    spec=admin_spec,
                    table=st.session_state.detail_schedules.get(name, pd.DataFrame()),
                    core_schedule=st.session_state.core_schedule,
                    save_table=_schedule_editor_save_table,
                    render_row_editor=_render_schedule_row_editor,
                    clear_editor_state=_clear_schedule_editor_state,
                )

                st.info(
                    "Master admin-wage inputs now live on the Assumptions page. Use `Apply to Schedule` there "
                    "to regenerate this schedule across the active production horizon, then fine-tune rows here "
                    "if the live plan needs adjustments."
                )

                detail_tables_for_run[name] = aggregated_admin

                render_schedule_summary("Admin Wages Summary", aggregated_admin)
            else:
                table = st.session_state.detail_schedules.get(name, pd.DataFrame())
                if not isinstance(table, pd.DataFrame):
                    table = pd.DataFrame()
                    st.session_state.detail_schedules[name] = table

                def _save_other(updated: pd.DataFrame, schedule_name: str = name) -> None:
                    st.session_state.detail_schedules[schedule_name] = updated

                _render_schedule_row_editor(
                    f"detail::{_scenario_key_suffix(name)}",
                    st.session_state.detail_schedules.get(name, table),
                    _save_other,
                )

        preview_generated_schedule = (
            active_schedule_section in schedule_section_labels[len(schedule_tab_names):]
        )
        if generated_schedule_tab_names and preview_generated_schedule:
            preview_outputs: dict[str, pd.DataFrame] = {}
            preview_error: Optional[str] = None
            try:
                preview_schedule = _build_schedule_dataframe(
                    st.session_state.core_schedule,
                    st.session_state.detail_schedules,
                    current_assumption_snapshot,
                )
                preview_outputs = _derive_biological_schedules(
                    preview_schedule,
                    current_assumption_snapshot,
                )
            except ValueError as exc:
                preview_error = str(exc)

            generated_tab_lookup = {
                "Kid Availability": "Kid Availability Schedule",
                "Kid Routing": "Kid Routing Schedule",
                "Internal Transfers": "Internal Transfer Schedule",
                "Downstream Intake": "Downstream Intake Schedule",
            }
            for offset, tab_name in enumerate(generated_schedule_tab_names, start=len(schedule_tab_names)):
                if active_schedule_section != schedule_section_labels[offset]:
                    continue
                if preview_error:
                    st.info(preview_error)
                    continue
                if tab_name == "Unit Herd Schedules":
                    st.markdown("#### Breeding Unit Schedule")
                    _render_table(
                        "Breeding Unit Schedule",
                        preview_outputs.get("Breeding Unit Schedule"),
                    )
                    st.markdown("#### Destination Unit Schedule")
                    _render_table(
                        "Destination Unit Schedule",
                        preview_outputs.get("Destination Unit Schedule"),
                    )
                    st.markdown("#### Consolidated Reporting Schedule")
                    _render_table(
                        "Reporting Schedule - Consolidated",
                        preview_outputs.get("Reporting Schedule - Consolidated"),
                    )
                    continue
                schedule_name = generated_tab_lookup.get(tab_name, "")
                _render_table(
                    schedule_name,
                    preview_outputs.get(schedule_name),
                )
                detail_tables_for_run[name] = st.session_state.detail_schedules.get(
                    name, table
                )

    if core_editor is None:
        core_editor = st.session_state.core_schedule

    for name, table in st.session_state.detail_schedules.items():
        detail_tables_for_run.setdefault(name, table)

    assumption_tables: Dict[str, pd.DataFrame] = {}

    if active_main_section == "Assumptions":
        st.subheader("Assumptions")
        _render_workflow_status_strip()
        _render_assumption_validation_summary(st.session_state.assumptions)
        st.caption(
            "Define the planning, commercial, operating, labour, and funding rules here first. Then use the Input "
            "Schedule page to translate the same logic into period-by-period operating and capital schedules."
        )
        st.info(
            "Recommended flow: 1. Planning & Scenario Design -> 2. Commercial Drivers -> 3. Operating Drivers -> "
            "4. Capital, Funding & Valuation -> 5. Input Schedule."
        )

        st.markdown("### 1. Planning & Scenario Design")
        st.caption(
            "These assumptions set the planning horizon, scenario stress, and herd-growth logic that feeds the Core "
            "Schedule and the Scenario Explorer on the Input Schedule page."
        )

        st.markdown("#### Business Structure")
        business_config = _ensure_business_configuration_table(
            st.session_state.assumptions.get("Business Configuration")
        )
        st.session_state.assumptions["Business Configuration"] = business_config
        current_business_type = _selected_business_type(business_config)
        current_operating_model = _selected_operating_model(business_config)
        current_destination = _selected_transfer_destination(business_config)
        current_reporting_view = _selected_reporting_view(business_config)
        current_transfer_pricing_method = _selected_transfer_pricing_method(business_config)
        current_allow_external_sales = _allow_external_kid_sales(business_config)

        structure_cols = st.columns(3)
        selected_business_type = structure_cols[0].selectbox(
            "Business type",
            options=list(BUSINESS_TYPE_OPTIONS),
            index=list(BUSINESS_TYPE_OPTIONS).index(current_business_type),
            key="assump::business_type",
        )
        selected_operating_model = structure_cols[1].selectbox(
            "Operating model",
            options=list(OPERATING_MODEL_OPTIONS),
            index=list(OPERATING_MODEL_OPTIONS).index(current_operating_model),
            key="assump::operating_model",
        )
        destination_options = ["", "Meat", "Milk-Cheese", "Combined"]
        selected_destination = structure_cols[2].selectbox(
            "Transfer destination",
            options=destination_options,
            index=destination_options.index(current_destination),
            key="assump::transfer_destination",
            disabled=selected_operating_model != "Breeding-to-Unit",
        )
        structure_cols_2 = st.columns(3)
        selected_reporting_view = structure_cols_2[0].selectbox(
            "Reporting view",
            options=list(REPORTING_VIEW_OPTIONS),
            index=list(REPORTING_VIEW_OPTIONS).index(current_reporting_view),
            key="assump::reporting_view",
        )
        selected_transfer_pricing_method = structure_cols_2[1].selectbox(
            "Transfer pricing method",
            options=list(TRANSFER_PRICING_METHOD_OPTIONS),
            index=list(TRANSFER_PRICING_METHOD_OPTIONS).index(current_transfer_pricing_method),
            key="assump::transfer_pricing_method",
            disabled=selected_operating_model != "Breeding-to-Unit",
        )
        selected_allow_external_sales = structure_cols_2[2].checkbox(
            "Allow external kid sales",
            value=current_allow_external_sales,
            key="assump::allow_external_kid_sales",
            disabled=selected_operating_model != "Breeding-to-Unit",
        )
        if (
            selected_business_type != current_business_type
            or selected_operating_model != current_operating_model
            or selected_destination != current_destination
            or selected_reporting_view != current_reporting_view
            or selected_transfer_pricing_method != current_transfer_pricing_method
            or bool(selected_allow_external_sales) != bool(current_allow_external_sales)
        ):
            st.session_state.assumptions["Business Configuration"] = _ensure_business_configuration_table(
                pd.DataFrame(
                    {
                        "Business Type": [selected_business_type],
                        "Operating Model": [selected_operating_model],
                        "Transfer Destination": [selected_destination],
                        "Reporting View": [selected_reporting_view],
                        "Transfer Pricing Method": [selected_transfer_pricing_method],
                        "Allow External Kid Sales": [selected_allow_external_sales],
                    }
                )
            )
            st.session_state.assumptions = _sync_transfer_tables_to_business_configuration(
                st.session_state.assumptions
            )
            st.session_state.assumptions = _sync_commercial_assumptions_to_core(
                st.session_state.assumptions,
                st.session_state.core_schedule,
            )
            _clear_schedule_editor_state("assump::scenario_controls")
            _clear_schedule_editor_state("assump::production_drivers")
            _clear_schedule_editor_state("assump::pricing")
            _reset_cached_results()
            _maybe_rerun()
        current_business_type = _selected_business_type(st.session_state.assumptions)
        active_products = _active_products_for_business_type(current_business_type)
        assumption_tables["Business Configuration"] = st.session_state.assumptions[
            "Business Configuration"
        ]
        st.caption(
            f"Active product universe for {current_business_type}: {', '.join(active_products)}."
        )

        st.markdown("#### Scenario Controls")
        st.caption(
            "Use the table to configure product-specific price and quantity shocks. The quick selectors below follow "
            "the active business type and keep the main product-price and feed-cost stresses close at hand."
        )
        scenario_table = _sync_scenario_controls_to_products(
            st.session_state.assumptions.get("Scenario Controls")
            if isinstance(st.session_state.assumptions.get("Scenario Controls"), pd.DataFrame)
            else pd.DataFrame(),
            active_products,
        )
        st.session_state.assumptions["Scenario Controls"] = scenario_table

        def _save_scenario_controls(updated: pd.DataFrame) -> None:
            ensured = _sync_scenario_controls_to_products(updated, active_products)
            if _frames_equal(
                st.session_state.assumptions.get("Scenario Controls"),
                ensured,
            ):
                return
            st.session_state.assumptions["Scenario Controls"] = ensured
            _reset_cached_results()

        _render_schedule_row_editor(
            "assump::scenario_controls",
            st.session_state.assumptions["Scenario Controls"],
            _save_scenario_controls,
        )

        control_values = _scenario_controls_value_map(
            st.session_state.assumptions["Scenario Controls"]
        )
        feed_default = control_values.get("Feed cost change (%)", 0.0)
        shock_options = list(range(-50, 51))
        quick_price_products = active_products[: min(len(active_products), 3)]
        if quick_price_products:
            price_cols = st.columns(len(quick_price_products))
            for idx_product, product in enumerate(quick_price_products):
                driver_name = _price_change_driver(product)
                current_value = control_values.get(driver_name, 0.0)
                selected_default = int(round(current_value))
                if selected_default not in shock_options:
                    selected_default = 0
                selected_value = price_cols[idx_product].selectbox(
                    driver_name,
                    options=shock_options,
                    index=shock_options.index(selected_default),
                    key=f"quick::{_scenario_key_suffix(driver_name)}",
                )
                if float(selected_value) != float(current_value):
                    updated_table = _update_scenario_control_value(
                        st.session_state.assumptions["Scenario Controls"],
                        driver_name,
                        float(selected_value),
                    )
                    st.session_state.assumptions["Scenario Controls"] = updated_table
                    _clear_schedule_editor_state("assump::scenario_controls")
                    _reset_cached_results()

        feed_options = list(range(-50, 51))
        feed_selected_default = int(round(feed_default))
        if feed_selected_default not in feed_options:
            feed_selected_default = 0
        feed_cost = st.selectbox(
            "Feed cost change (%)",
            options=feed_options,
            index=feed_options.index(feed_selected_default),
            key="feed_cost_change_dropdown",
        )
        if float(feed_cost) != float(feed_default):
            updated_table = _update_scenario_control_value(
                st.session_state.assumptions["Scenario Controls"],
                "Feed cost change (%)",
                float(feed_cost),
            )
            st.session_state.assumptions["Scenario Controls"] = updated_table
            _clear_schedule_editor_state("assump::scenario_controls")
            _reset_cached_results()

        assumption_tables["Scenario Controls"] = st.session_state.assumptions[
            "Scenario Controls"
        ]
        run_clicked = st.button("Run model", type="primary")
        run_feedback_slot = st.empty()

        st.markdown("#### Production Time Horizon")
        production_table = _ensure_production_horizon_table(
            st.session_state.assumptions.get("Production Horizon")
        )
        st.session_state.assumptions["Production Horizon"] = production_table

        defaults = production_table.iloc[0]
        start_default = int(defaults.get("Start Year", 2024))
        end_default = int(defaults.get("End Year", start_default))

        st.session_state.setdefault("production_start_year", start_default)
        st.session_state.setdefault("production_end_year", end_default)

        year_options = _production_year_options(start_default, end_default)
        if st.session_state.production_start_year not in year_options:
            st.session_state.production_start_year = start_default

        start_col, end_col = st.columns(2)

        start_value = start_col.selectbox(
            "Start year",
            options=year_options,
            key="production_start_year",
        )

        valid_end_options = [year for year in year_options if year >= start_value]
        if not valid_end_options:
            valid_end_options = [start_value]

        if st.session_state.production_end_year not in valid_end_options:
            st.session_state.production_end_year = valid_end_options[0]

        end_value = end_col.selectbox(
            "End year",
            options=valid_end_options,
            key="production_end_year",
        )

        if end_value < start_value:
            st.session_state.production_end_year = start_value
            end_value = start_value

        if start_value != start_default or end_value != end_default:
            updated_table = pd.DataFrame({"Start Year": [start_value], "End Year": [end_value]})
            st.session_state.assumptions["Production Horizon"] = updated_table
            assumption_tables["Production Horizon"] = updated_table
            _sync_production_horizon(start_value, end_value)
            _ensure_default_results_loaded()
            _ensure_active_scenario_selection()
            _maybe_rerun()
        else:
            assumption_tables["Production Horizon"] = production_table

        st.markdown("#### Herd Plan (Heads)")
        herd_plan = _ensure_herd_plan_table(
            st.session_state.assumptions.get("Herd Plan", pd.DataFrame())
        )
        st.session_state.assumptions["Herd Plan"] = herd_plan
        herd_plan = render_herd_plan_editor(
            herd_plan=herd_plan,
            save_table=_save_herd_plan_and_sync_opening_cohorts,
            render_row_editor=_render_schedule_row_editor,
            clear_editor_state=_clear_schedule_editor_state,
            apply_increment_fn=_apply_herd_yearly_increment,
            ensure_fn=_ensure_herd_plan_table,
        )
        st.session_state.assumptions = _sync_opening_herd_cohorts_in_assumptions(
            st.session_state.assumptions
        )
        assumption_tables["Herd Plan"] = st.session_state.assumptions["Herd Plan"]

        st.markdown("### 2. Biological Engine")
        st.caption(
            "These tables drive herd reproduction, lactation, finishing, slaughter, and replacement logic. "
            "They now sit upstream of pricing quantities and herd-linked operating costs. Annual biological rates "
            "such as mortality and culling are converted internally to the active model grain."
        )
        biological_editors = [
            BiologicalEditorDefinition(
                "Biological System Settings",
                "Manage model grain, herd-plan control mode, and biological timing settings.",
                _ensure_biological_system_settings_table,
                "assump::biological_settings",
            ),
            BiologicalEditorDefinition(
                "Breeding & Reproduction Biology",
                "Defines age at first kidding, gestation timing, fertility, kidding rates, mortality, culling, and replacement retention. Age fields are in months; kidding is annualized per doe per year; mortality and cull rates are annual percentages.",
                _ensure_breeding_reproduction_biology_table,
                "assump::breeding_biology",
            ),
            BiologicalEditorDefinition(
                "Lactation Biology",
                "Defines lactation length, dry period, peak month, parity yield factors, and the curve shape used for milk output.",
                _ensure_lactation_biology_table,
                "assump::lactation_biology",
            ),
            BiologicalEditorDefinition(
                "Finishing & Slaughter Biology",
                "Defines months to market weight, slaughter timing, live-sale mix, yield assumptions, and finishing mortality. Age and weight-gain fields are monthly/month-based; mortality is an annual percentage.",
                _ensure_finishing_slaughter_biology_table,
                "assump::finishing_biology",
            ),
            BiologicalEditorDefinition(
                "Opening Herd Cohorts",
                "Provide the opening biological starting herd by cohort, age, sex, purpose, parity, and current lactation state.",
                _ensure_opening_herd_cohorts_table,
                "assump::opening_herd_cohorts",
            ),
            BiologicalEditorDefinition(
                "Cohort Allocation Rules",
                "Controls how surviving kids are retained into replacements versus routed into finishing and sale streams.",
                _ensure_cohort_allocation_rules_table,
                "assump::cohort_allocation_rules",
            ),
            BiologicalEditorDefinition(
                "Biological Cost Drivers",
                "Maps feed, healthcare, and utilities to breeding, replacement, finishing, lactating, or total-herd stage bases.",
                _ensure_biological_cost_drivers_table,
                "assump::biological_cost_drivers",
            ),
        ]
        st.session_state.assumptions = render_biological_assumption_editor(
            definitions=biological_editors,
            assumptions=st.session_state.assumptions,
            render_row_editor=_render_schedule_row_editor,
        )
        st.session_state.assumptions = _sync_biological_start_date_in_assumptions(
            st.session_state.assumptions,
            period_type=st.session_state.get("schedule_period_type", "quarterly"),
        )
        st.session_state.assumptions = _sync_opening_herd_cohorts_in_assumptions(
            st.session_state.assumptions
        )
        for definition in biological_editors:
            assumption_tables[definition.name] = st.session_state.assumptions[definition.name]

        if _is_breeding_to_unit_mode(st.session_state.assumptions):
            st.markdown("### 3. Kid Routing")
            st.caption(
                "Route weaned kids from the Breeding unit into external sale or one downstream operating unit. "
                "This is the explicit transfer layer between the breeding engine and the downstream business."
            )

            routing_table = _ensure_kid_routing_rules_table(
                st.session_state.assumptions.get("Kid Routing Rules")
            )
            st.session_state.assumptions["Kid Routing Rules"] = routing_table

            def _save_kid_routing(updated: pd.DataFrame) -> None:
                st.session_state.assumptions["Kid Routing Rules"] = _ensure_kid_routing_rules_table(updated)
                _reset_cached_results()

            _render_schedule_row_editor(
                "assump::kid_routing_rules",
                routing_table,
                _save_kid_routing,
            )
            assumption_tables["Kid Routing Rules"] = st.session_state.assumptions["Kid Routing Rules"]

            transfer_pricing_table = _ensure_internal_transfer_pricing_table(
                st.session_state.assumptions.get("Internal Transfer Pricing")
            )
            st.session_state.assumptions["Internal Transfer Pricing"] = transfer_pricing_table

            def _save_transfer_pricing(updated: pd.DataFrame) -> None:
                st.session_state.assumptions["Internal Transfer Pricing"] = _ensure_internal_transfer_pricing_table(updated)
                _reset_cached_results()

            st.markdown("#### Internal Transfer Pricing")
            _render_schedule_row_editor(
                "assump::internal_transfer_pricing",
                transfer_pricing_table,
                _save_transfer_pricing,
            )
            assumption_tables["Internal Transfer Pricing"] = st.session_state.assumptions["Internal Transfer Pricing"]

            downstream_intake = _ensure_downstream_intake_rules_table(
                st.session_state.assumptions.get("Downstream Intake Rules")
            )
            st.session_state.assumptions["Downstream Intake Rules"] = downstream_intake

            def _save_downstream_intake(updated: pd.DataFrame) -> None:
                st.session_state.assumptions["Downstream Intake Rules"] = _ensure_downstream_intake_rules_table(updated)
                _reset_cached_results()

            st.markdown("### 4. Downstream Unit Rules")
            _render_schedule_row_editor(
                "assump::downstream_intake_rules",
                downstream_intake,
                _save_downstream_intake,
            )
            assumption_tables["Downstream Intake Rules"] = st.session_state.assumptions["Downstream Intake Rules"]

            elimination_rules = _ensure_transfer_elimination_rules_table(
                st.session_state.assumptions.get("Transfer Elimination Rules")
            )
            st.session_state.assumptions["Transfer Elimination Rules"] = elimination_rules

            def _save_elimination_rules(updated: pd.DataFrame) -> None:
                st.session_state.assumptions["Transfer Elimination Rules"] = _ensure_transfer_elimination_rules_table(updated)
                _reset_cached_results()

            st.markdown("#### Transfer Elimination Rules")
            _render_schedule_row_editor(
                "assump::transfer_elimination_rules",
                elimination_rules,
                _save_elimination_rules,
            )
            assumption_tables["Transfer Elimination Rules"] = st.session_state.assumptions["Transfer Elimination Rules"]

        st.markdown("### 5. Commercial Drivers")
        st.caption(
            "These assumptions define product pricing logic and growth patterns that should align with the revenue "
            "and gross margin structure used in the Input Schedule."
        )
        st.markdown("#### Production Quantity Drivers")
        production_drivers = _ensure_production_driver_table(
            st.session_state.assumptions.get("Production Drivers")
        )
        st.session_state.assumptions["Production Drivers"] = production_drivers
        st.caption(
            "These biological and commercial drivers convert herd size into saleable output. Dairy products use the "
            "lactation stream, while livestock products use the saleable-herd and slaughter stream. Slaughter rates "
            "and driver growth are annual assumptions that the model converts to the active schedule grain."
        )
        production_drivers = _sync_production_driver_table_to_products(
            production_drivers,
            active_products,
        )
        st.session_state.assumptions["Production Drivers"] = production_drivers

        def _save_production_drivers(updated: pd.DataFrame) -> None:
            ensured = _sync_production_driver_table_to_products(updated, active_products)
            if _frames_equal(
                st.session_state.assumptions.get("Production Drivers"),
                ensured,
            ):
                return
            st.session_state.assumptions["Production Drivers"] = ensured
            _reset_cached_results()

        def _save_production_driver_subset(products: Sequence[str], updated: pd.DataFrame) -> None:
            merged = _merge_production_driver_subset(
                st.session_state.assumptions.get("Production Drivers"),
                updated,
                products,
            )
            if _frames_equal(
                st.session_state.assumptions.get("Production Drivers"),
                merged,
            ):
                return
            st.session_state.assumptions["Production Drivers"] = merged
            _reset_cached_results()

        driver_disabled_columns: list[str] = ["Product", "Unit"]
        visible_dairy_products = [
            product for product in _PRODUCTION_DRIVER_DAIRY_PRODUCTS if product in active_products
        ]
        visible_livestock_products = [
            product for product in _PRODUCTION_DRIVER_SLAUGHTER_PRODUCTS if product in active_products
        ]
        dairy_drivers = st.session_state.assumptions["Production Drivers"].loc[
            st.session_state.assumptions["Production Drivers"]["Product"].isin(
                visible_dairy_products
            )
        ]
        slaughter_drivers = st.session_state.assumptions["Production Drivers"].loc[
            st.session_state.assumptions["Production Drivers"]["Product"].isin(
                visible_livestock_products
            )
        ]
        with st.expander("Manage Production Driver Columns", expanded=False):
            st.caption(
                "Manage custom columns and use the advanced full-table editor when you need to edit every active Production Quantity Driver field in one place."
            )
            add_key = "assump::production_drivers::new_column"
            _safe_session_state_setdefault(add_key, "")
            add_cols = st.columns([2.2, 1])
            new_column = add_cols[0].text_input("New column name", key=add_key)
            if add_cols[1].button("Add column", key=f"{add_key}::add", use_container_width=True):
                cleaned = _normalise_production_driver_column_name(new_column)
                if not cleaned:
                    st.warning("Enter a column name before adding it.")
                elif cleaned in st.session_state.assumptions["Production Drivers"].columns:
                    st.warning(f"`{cleaned}` already exists.")
                else:
                    updated = _add_production_driver_column(
                        st.session_state.assumptions["Production Drivers"],
                        cleaned,
                    )
                    st.session_state.assumptions["Production Drivers"] = updated
                    _safe_session_state_set(add_key, "")
                    _reset_cached_results()
                    _maybe_rerun()

            removable_columns = _production_driver_custom_columns(
                st.session_state.assumptions["Production Drivers"]
            )
            remove_cols = st.columns([2.2, 1])
            columns_to_remove = remove_cols[0].multiselect(
                "Custom columns",
                options=removable_columns,
                key="assump::production_drivers::remove_columns",
                disabled=not removable_columns,
            )
            if remove_cols[1].button(
                "Remove columns",
                key="assump::production_drivers::remove_columns::apply",
                use_container_width=True,
                disabled=not columns_to_remove,
            ):
                updated = _remove_production_driver_columns(
                    st.session_state.assumptions["Production Drivers"],
                    columns_to_remove,
                )
                st.session_state.assumptions["Production Drivers"] = updated
                _safe_session_state_set("assump::production_drivers::remove_columns", [])
                _reset_cached_results()
                _maybe_rerun()

            st.markdown("##### Advanced Production Driver Editor")
            st.caption(
                "This editor unlocks all active columns, including Product and Unit. Product choices stay constrained to the active business type."
            )
            production_driver_editor = st.data_editor(
                st.session_state.assumptions["Production Drivers"],
                use_container_width=True,
                hide_index=True,
                key="assump::production_drivers::advanced",
                column_config=_production_driver_column_config(
                    st.session_state.assumptions["Production Drivers"],
                    active_products,
                ),
                disabled=[],
            )
            _save_production_drivers(production_driver_editor)

        st.markdown("##### Guided Editors")
        st.caption(
            "Use these simpler editors for routine updates. Product and Unit stay fixed here to keep the core driver rows stable."
        )
        if visible_dairy_products and visible_livestock_products:
            st.markdown("**Dairy Drivers**")
            dairy_editor = st.data_editor(
                dairy_drivers,
                use_container_width=True,
                hide_index=True,
                key="assump::production_drivers::dairy",
                column_config=_production_driver_column_config(dairy_drivers, visible_dairy_products),
                disabled=driver_disabled_columns,
            )
            _save_production_driver_subset(visible_dairy_products, dairy_editor)

            st.markdown("**Livestock & Slaughter Drivers**")
            slaughter_editor = st.data_editor(
                slaughter_drivers,
                use_container_width=True,
                hide_index=True,
                key="assump::production_drivers::slaughter",
                column_config=_production_driver_column_config(slaughter_drivers, visible_livestock_products),
                disabled=driver_disabled_columns,
            )
            _save_production_driver_subset(
                visible_livestock_products,
                slaughter_editor,
            )
        elif visible_dairy_products:
            st.markdown("**Dairy Drivers**")
            dairy_editor = st.data_editor(
                dairy_drivers,
                use_container_width=True,
                hide_index=True,
                key="assump::production_drivers::dairy",
                column_config=_production_driver_column_config(dairy_drivers, visible_dairy_products),
                disabled=driver_disabled_columns,
            )
            _save_production_driver_subset(visible_dairy_products, dairy_editor)
        elif visible_livestock_products:
            st.markdown("**Livestock & Slaughter Drivers**")
            slaughter_editor = st.data_editor(
                slaughter_drivers,
                use_container_width=True,
                hide_index=True,
                key="assump::production_drivers::slaughter",
                column_config=_production_driver_column_config(slaughter_drivers, visible_livestock_products),
                disabled=driver_disabled_columns,
            )
            _save_production_driver_subset(
                visible_livestock_products,
                slaughter_editor,
            )
        assumption_tables["Production Drivers"] = st.session_state.assumptions[
            "Production Drivers"
        ]

        st.markdown("#### Pricing Assumptions")
        st.session_state.assumptions = _sync_commercial_assumptions_to_core(
            st.session_state.assumptions,
            st.session_state.core_schedule,
        )
        pricing_table = st.session_state.assumptions["Pricing"]
        period_label = (
            "quarter"
            if st.session_state.get("schedule_period_type") == "quarterly"
            else "month"
        )
        st.caption(
            "Activate only the products you want in each period, set the planned quantity for that period, and use "
            "allocation percentages where one production stream is shared across multiple outputs."
        )

        st.session_state.setdefault("pricing_plan_product", active_products[0] if active_products else "Milk")
        st.session_state.setdefault("pricing_plan_active", True)
        st.session_state.setdefault("pricing_plan_allocation_pct", 100.0)
        st.session_state.setdefault("pricing_plan_quantity_mode", "Derived")
        st.session_state.setdefault("pricing_plan_quantity", 0.0)
        st.session_state.setdefault("pricing_plan_growth_pct", 0.0)
        st.session_state.setdefault("pricing_plan_period_start", None)
        st.session_state.setdefault("pricing_plan_period_end", None)

        product_options = sorted(
            {
                str(product).strip()
                for product in pricing_table.get("Product", pd.Series(dtype=str))
                .dropna()
                .tolist()
                if str(product).strip()
            }
        )
        if product_options and st.session_state.get("pricing_plan_product") not in product_options:
            st.session_state.pricing_plan_product = product_options[0]
        period_options = (
            pricing_table.get("Period", pd.Series(dtype=str))
            .dropna()
            .astype(str)
            .drop_duplicates()
            .tolist()
        )
        if period_options:
            if st.session_state.get("pricing_plan_period_start") not in period_options:
                st.session_state.pricing_plan_period_start = period_options[0]
            if st.session_state.get("pricing_plan_period_end") not in period_options:
                st.session_state.pricing_plan_period_end = period_options[-1]

        st.caption(
            "Plan product activation by date range. The available products follow the selected business type and use "
            "the matching dairy or livestock quantity logic."
        )
        plan_product_col, plan_active_col, plan_alloc_col, plan_mode_col, plan_qty_col, plan_growth_col = st.columns(
            [1.1, 0.8, 1.0, 1.1, 1.2, 1.0]
        )
        plan_product_col.selectbox(
            "Plan product",
            options=product_options,
            key="pricing_plan_product",
        )
        plan_active_col.checkbox("Active", key="pricing_plan_active")
        plan_alloc_col.number_input(
            "Allocation (%)",
            min_value=0.0,
            max_value=100.0,
            step=1.0,
            key="pricing_plan_allocation_pct",
        )
        plan_mode_col.selectbox(
            "Quantity mode",
            options=["Derived", "Manual Override"],
            key="pricing_plan_quantity_mode",
        )
        plan_qty_col.number_input(
            f"Manual qty / {period_label}",
            min_value=0.0,
            step=10.0,
            key="pricing_plan_quantity",
        )
        plan_growth_col.number_input(
            "Qty yearly growth (%)",
            min_value=-100.0,
            max_value=300.0,
            step=0.1,
            key="pricing_plan_growth_pct",
        )
        range_start_col, range_end_col, plan_btn_col = st.columns([1.1, 1.1, 1])
        range_start_col.selectbox(
            "From period",
            options=period_options,
            key="pricing_plan_period_start",
        )
        range_end_col.selectbox(
            "To period",
            options=period_options,
            key="pricing_plan_period_end",
        )
        if plan_btn_col.button("Apply plan to range", key="pricing_apply_plan"):
            pricing_table = _apply_pricing_product_plan(
                pricing_table,
                st.session_state.get("pricing_plan_product", "Milk"),
                active=bool(st.session_state.get("pricing_plan_active", True)),
                allocation_pct=float(
                    st.session_state.get("pricing_plan_allocation_pct", 100.0)
                ),
                quantity_mode=str(
                    st.session_state.get("pricing_plan_quantity_mode", "Derived")
                ),
                base_quantity=float(st.session_state.get("pricing_plan_quantity", 0.0)),
                yearly_growth_pct=float(
                    st.session_state.get("pricing_plan_growth_pct", 0.0)
                ),
                period_start=st.session_state.get("pricing_plan_period_start"),
                period_end=st.session_state.get("pricing_plan_period_end"),
            )
            st.session_state.assumptions["Pricing"] = pricing_table
            _reset_cached_results()
        st.session_state.assumptions = render_pricing_manual_editor(
            pricing_table=pricing_table,
            assumptions=st.session_state.assumptions,
            core_schedule=st.session_state.core_schedule,
            sync_assumptions_fn=_sync_commercial_assumptions_to_core,
            refresh_quantities_fn=_derive_pricing_quantities_from_production,
            pricing_context_fn=_pricing_schedule_context,
            pricing_validation_fn=_pricing_validation_messages,
            revenue_by_period_fn=_pricing_revenue_by_period,
            family_summary_fn=_pricing_family_summary,
            quantity_by_period_fn=_pricing_quantity_by_period,
            add_row_fn=_add_pricing_row,
            remove_row_fn=_remove_pricing_row,
            apply_increment_fn=_apply_pricing_yearly_increment,
            render_row_editor=_render_schedule_row_editor,
            clear_editor_state=_clear_schedule_editor_state,
            invalidate_results_fn=_reset_cached_results,
            active_products=active_products,
            period_label=period_label,
        )


        assumption_tables["Pricing"] = st.session_state.assumptions["Pricing"]

        st.markdown("### 6. Operating Drivers")
        st.caption(
            "These assumptions establish the operating-cost baseline that should stay coordinated with the COGS, "
            "variable expenses, and wage schedules on the Input Schedule page."
        )
        st.markdown("#### Operating Cost Assumptions")
        st.caption(
            "Fields `variable_feed_cost_per_herd`, `variable_healthcare_cost_per_herd`, and "
            "`fixed_utility_cost_per_herd` are treated as unit_cost_per_head_per_month values. "
            "Monthly total cost = unit cost × herd heads × months."
        )
        operating_table = _ensure_operating_cost_table(
            st.session_state.assumptions.get("Operating Costs")
        )
        operating_columns = [
            "Year",
            "Field",
            "Category",
            "unit_cost_per_head_per_month",
            "Inflation %",
        ]

        def _save_operating_defaults(template_editor: pd.DataFrame) -> None:
            _set_template(
                "operating_rows",
                _dataframe_to_template(template_editor, operating_columns),
            )

        def _apply_operating_defaults(template_editor: pd.DataFrame) -> pd.DataFrame:
            _save_operating_defaults(template_editor)
            st.session_state["default_operating_editor_seed"] = template_editor
            return _default_operating_cost_table()

        def _restore_operating_defaults() -> pd.DataFrame:
            baseline_template = _template_copy(DEFAULT_OPERATING_COST_ROWS)
            _set_template("operating_rows", baseline_template)
            st.session_state["default_operating_editor_seed"] = _template_to_dataframe(
                baseline_template,
                operating_columns,
            )
            return _default_operating_cost_table()

        def _save_operating_costs(updated: pd.DataFrame) -> None:
            ensured = _ensure_operating_cost_table(updated)
            if not _frames_equal(
                ensured, st.session_state.assumptions.get("Operating Costs")
            ):
                st.session_state.assumptions["Operating Costs"] = ensured
                _reset_cached_results()
            else:
                st.session_state.assumptions["Operating Costs"] = ensured

        st.session_state.assumptions["Operating Costs"] = render_operating_cost_editor(
            operating_table=operating_table,
            save_table=_save_operating_costs,
            render_row_editor=_render_schedule_row_editor,
            clear_editor_state=_clear_schedule_editor_state,
            add_row_fn=_add_operating_cost_row,
            remove_row_fn=_remove_operating_cost_row,
            apply_increment_fn=_apply_operating_cost_increment,
            ensure_fn=_ensure_operating_cost_table,
            get_default_frame=lambda: st.session_state.get(
                "default_operating_editor_seed"
            )
            if isinstance(
                st.session_state.get("default_operating_editor_seed"), pd.DataFrame
            )
            else _template_to_dataframe(
                _get_template("operating_rows", DEFAULT_OPERATING_COST_ROWS),
                operating_columns,
            ),
            save_defaults_fn=_save_operating_defaults,
            apply_defaults_fn=_apply_operating_defaults,
            restore_defaults_fn=_restore_operating_defaults,
        )

        assumption_tables["Operating Costs"] = st.session_state.assumptions[
            "Operating Costs"
        ]

        st.markdown("#### Variable Expense Master Inputs")
        variable_assumptions = _render_assumption_master_table(
            assumption_key="Variable Expenses",
            schedule_name="Variable Expenses Schedule",
            label_column="Item",
            all_label="All items",
            create_label="Create new item",
            new_label_prompt="New variable expense item",
            caption=(
                "Set the baseline monthly amount for each variable expense item here. These rows become the source "
                "of truth for the Input Schedule variable-expense grid."
            ),
            propagation_note=(
                "Annual increase compounds from the base monthly amount. When the schedule grain is quarterly, "
                "the applied schedule multiplies the monthly base by 3 for each quarter."
            ),
            ensure_fn=_ensure_variable_expense_input_table,
            default_fn=_default_variable_expense_input_table,
            propagate_fn=_propagate_variable_expense_inputs_to_schedule,
            add_row_factory=lambda label: {
                "Item": label,
                "Amount per Period": np.nan,
                "Yearly Increase %": 0.0,
            },
            column_config={
                "Item": st.column_config.TextColumn("Item"),
                "Amount per Period": st.column_config.NumberColumn(
                    "Monthly Amount", format="%.2f", step=100.0
                ),
                "Yearly Increase %": st.column_config.NumberColumn(
                    "Annual Increase (%)", format="%.2f", step=0.1
                ),
            },
            editor_key="assump_variable_expense_master",
            editor_identifier="detail::variable_expenses",
            add_choice_key="assump_variable_add_choice",
            add_name_key="assump_variable_new_name",
            remove_choice_key="assump_variable_remove_choice",
            increment_target_key="assump_variable_increment_target",
            increment_pct_key="assump_variable_increment_pct",
        )
        assumption_tables["Variable Expenses"] = variable_assumptions

        st.markdown("#### Direct Wage Master Inputs")
        direct_assumptions = _render_assumption_master_table(
            assumption_key="Direct Wages",
            schedule_name="Direct Wages Schedule",
            label_column="Position",
            all_label="All positions",
            create_label="Create new position",
            new_label_prompt="New direct wage position",
            caption=(
                "Maintain itemised direct labour here by position, head count, and monthly salary per head. "
                "This master table should be set before refining direct labour in the Input Schedule page."
            ),
            propagation_note=(
                "Annual increase applies to `Monthly Salary per Head`, then `Total Salary` is recomputed as "
                "`Head Count × Monthly Salary per Head`. Quarterly schedules convert the monthly salary to quarter totals."
            ),
            ensure_fn=_ensure_direct_wage_input_table,
            default_fn=_default_direct_wage_input_table,
            propagate_fn=_propagate_direct_wage_inputs_to_schedule,
            add_row_factory=lambda label: {
                "Position": label,
                "Head Count": 1.0,
                "Monthly Salary per Head": np.nan,
                "Total Salary": np.nan,
                "Yearly Increase %": 0.0,
            },
            column_config={
                "Position": st.column_config.TextColumn("Position"),
                "Head Count": st.column_config.NumberColumn(
                    "Head Count", format="%.0f", step=1.0
                ),
                "Monthly Salary per Head": st.column_config.NumberColumn(
                    "Monthly Salary per Head", format="%.2f", step=100.0
                ),
                "Total Salary": st.column_config.NumberColumn(
                    "Total Salary", format="%.2f"
                ),
                "Yearly Increase %": st.column_config.NumberColumn(
                "Annual Increase (%)", format="%.2f", step=0.1
                ),
            },
            editor_key="assump_direct_wage_master",
            editor_identifier="detail::direct_wages",
            add_choice_key="assump_direct_add_choice",
            add_name_key="assump_direct_new_name",
            remove_choice_key="assump_direct_remove_choice",
            increment_target_key="assump_direct_increment_target",
            increment_pct_key="assump_direct_increment_pct",
            disabled_columns=["Total Salary"],
        )
        assumption_tables["Direct Wages"] = direct_assumptions

        st.markdown("#### Admin Wage Master Inputs")
        admin_assumptions = _render_assumption_master_table(
            assumption_key="Admin Wages",
            schedule_name="Admin Wages Schedule",
            label_column="Position",
            all_label="All positions",
            create_label="Create new position",
            new_label_prompt="New admin wage position",
            caption=(
                "Set the baseline administrative labour structure here so the Input Schedule inherits a coherent "
                "salary plan across the full production horizon."
            ),
            propagation_note=(
                "Annual increase applies to `Monthly Salary per Head`, then `Total Salary` is recomputed as "
                "`Head Count × Monthly Salary per Head`. Quarterly schedules convert the monthly salary to quarter totals."
            ),
            ensure_fn=_ensure_admin_wage_input_table,
            default_fn=_default_admin_wage_input_table,
            propagate_fn=_propagate_admin_wage_inputs_to_schedule,
            add_row_factory=lambda label: {
                "Position": label,
                "Head Count": 1.0,
                "Monthly Salary per Head": np.nan,
                "Total Salary": np.nan,
                "Yearly Increase %": 0.0,
            },
            column_config={
                "Position": st.column_config.TextColumn("Position"),
                "Head Count": st.column_config.NumberColumn(
                    "Head Count", format="%.0f", step=1.0
                ),
                "Monthly Salary per Head": st.column_config.NumberColumn(
                    "Monthly Salary per Head", format="%.2f", step=100.0
                ),
                "Total Salary": st.column_config.NumberColumn(
                    "Total Salary", format="%.2f"
                ),
                "Yearly Increase %": st.column_config.NumberColumn(
                "Annual Increase (%)", format="%.2f", step=0.1
                ),
            },
            editor_key="assump_admin_wage_master",
            editor_identifier="detail::admin_wages",
            add_choice_key="assump_admin_add_choice",
            add_name_key="assump_admin_new_name",
            remove_choice_key="assump_admin_remove_choice",
            increment_target_key="assump_admin_increment_target",
            increment_pct_key="assump_admin_increment_pct",
            disabled_columns=["Total Salary"],
        )
        assumption_tables["Admin Wages"] = admin_assumptions
    
        st.markdown("### 4. Capital, Funding & Valuation")
        st.caption(
            "These assumptions support the capital structure, capex, ownership, and investor-return logic that is "
            "captured in the capital and asset support schedules on the Input Schedule page."
        )
        st.markdown("#### Capital & Financing Assumptions")
        st.caption(
            "Use annual return and interest percentages here. Terms remain year-based; drawdown and contribution timing are handled in the dated support schedules."
        )
        capital_table = _ensure_capital_financing_table(
            st.session_state.assumptions.get("Capital & Financing")
        )
        st.session_state.assumptions["Capital & Financing"] = capital_table
    
        def _save_capital(updated: pd.DataFrame) -> None:
            ensured = _ensure_capital_financing_table(updated)
            st.session_state.assumptions["Capital & Financing"] = ensured
    
        _render_schedule_row_editor(
            "assump::capital_financing",
            st.session_state.assumptions["Capital & Financing"],
            _save_capital,
        )
        assumption_tables["Capital & Financing"] = st.session_state.assumptions[
            "Capital & Financing"
        ]
    
        st.markdown("#### Valuation Assumptions")
        st.caption(
            "Discount rate, working-capital, liquidity, and covenant assumptions. WACC and terminal growth are annual percentages; IRR and NPV are computed outputs."
        )
        include_valuation = st.checkbox("Include valuation assumptions", value=True)
        valuation_table = _ensure_valuation_inputs_table(
            st.session_state.assumptions.get("Valuation Inputs")
        )
        st.session_state.assumptions["Valuation Inputs"] = valuation_table
    
        def _save_valuation(updated: pd.DataFrame) -> None:
            ensured = _ensure_valuation_inputs_table(updated)
            st.session_state.assumptions["Valuation Inputs"] = ensured
    
        _render_schedule_row_editor(
            "assump::valuation_inputs",
            st.session_state.assumptions["Valuation Inputs"],
            _save_valuation,
        )
    
        if include_valuation:
            valuation_inputs = _valuation_table_to_inputs(
                st.session_state.assumptions["Valuation Inputs"]
            )
        else:
            valuation_inputs = {}
    
        assumption_tables["Valuation Inputs"] = st.session_state.assumptions[
            "Valuation Inputs"
        ]

    if active_main_section == "Financials":
        st.subheader("Financial Statements")
        if st.session_state.results is None:
            st.info("Run the model to generate the financial statements.")
        else:
            _render_stale_results_notice()
            results = st.session_state.results
            scenario_label = results.get("selected_scenario", "Scenario")
            reporting_context = _cached_result_view(
                "reporting_views",
                scenario_label,
                lambda: _reporting_views_for_result(results),
            )
            entity_options = reporting_context.get("entity_options", ["Consolidated"])
            default_entity = reporting_context.get("default_entity", entity_options[0])
            if st.session_state.get("reporting_entity_selector") not in entity_options:
                st.session_state["reporting_entity_selector"] = default_entity
            selected_reporting_entity = st.selectbox(
                "Reporting entity",
                options=entity_options,
                key="reporting_entity_selector",
            )
            base_reporting_schedule = reporting_context.get("base_schedules", {}).get(
                selected_reporting_entity,
                results["base"],
            )
            scenario_reporting_schedule = reporting_context.get("scenario_schedules", {}).get(
                selected_reporting_entity,
                results["scenario"],
            )
            financial_section_labels = [
                "Statement of Financial Performance",
                "Statement of Financial Position",
                "Statement of Cash Flow",
            ]
            active_financial_section = _section_selector(
                "Financial Statement",
                financial_section_labels,
                key="financial_statement_selector",
                help="Only the selected statement is rendered.",
                )
            try:
                financial_views = _cached_result_view(
                    "financial_statements",
                    f"{scenario_label}::{selected_reporting_entity}::{active_financial_section}",
                    lambda: (
                        {
                            "base": results["model"].statement_of_financial_performance(
                                base_reporting_schedule, annual=True
                            ),
                            "scenario": results["model"].statement_of_financial_performance(
                                scenario_reporting_schedule, annual=True
                            ),
                        }
                        if active_financial_section == financial_section_labels[0]
                        else {
                            "base": results["model"].statement_of_financial_position(
                                base_reporting_schedule, annual=True
                            ),
                            "scenario": results["model"].statement_of_financial_position(
                                scenario_reporting_schedule, annual=True
                            ),
                        }
                        if active_financial_section == financial_section_labels[1]
                        else {
                            "base": results["model"].statement_of_cash_flow(
                                base_reporting_schedule, annual=True
                            ),
                            "scenario": results["model"].statement_of_cash_flow(
                                scenario_reporting_schedule, annual=True
                            ),
                        }
                    ),
                )
                if active_financial_section == financial_section_labels[0]:
                    sop_base = financial_views["base"]
                    sop_scenario = financial_views["scenario"]
                    st.dataframe(
                        pd.concat(
                            {"Base": sop_base, scenario_label: sop_scenario}, axis=1
                        )
                        .swaplevel(axis=1)
                        .sort_index(axis=1, level=0)
                    )
                    _render_financial_performance_charts(
                        sop_base, sop_scenario, scenario_label
                    )
                elif active_financial_section == financial_section_labels[1]:
                    sofp_base = financial_views["base"]
                    sofp_scenario = financial_views["scenario"]
                    st.dataframe(
                        pd.concat(
                            {"Base": sofp_base, scenario_label: sofp_scenario}, axis=1
                        )
                        .swaplevel(axis=1)
                        .sort_index(axis=1, level=0)
                    )
                    _render_financial_position_charts(
                        sofp_base, sofp_scenario, scenario_label
                    )
                else:
                    socf_base = financial_views["base"]
                    socf_scenario = financial_views["scenario"]
                    st.dataframe(
                        pd.concat(
                            {"Base": socf_base, scenario_label: socf_scenario}, axis=1
                        )
                        .swaplevel(axis=1)
                        .sort_index(axis=1, level=0)
                    )
                    _render_cash_flow_charts(
                        socf_base, socf_scenario, scenario_label
                    )
            except ValueError as exc:
                st.info(str(exc))

    if run_clicked:
        def _show_run_validation_error(message: str) -> None:
            if run_feedback_slot is not None:
                run_feedback_slot.error(message)
            else:
                st.error(message)

        core_clean = _clean_editor_table(core_editor)
        if core_clean is None:
            _show_run_validation_error(
                "Provide at least one period in the core schedule before running the scenario."
            )
            return

        try:
            core_prepared = _prepare_timeline_table(core_clean)
        except ValueError as exc:
            _show_run_validation_error(f"Core Schedule: {exc}")
            return

        try:
            _prepare_detail_tables_for_schedule(core_clean, detail_tables_for_run)
            schedule_df = _build_schedule_dataframe(
                core_clean,
                detail_tables_for_run,
                assumption_tables,
            )
        except ValueError as exc:
            _show_run_validation_error(str(exc))
            return

        if schedule_df["Revenue"].isna().all():
            _show_run_validation_error("Core Schedule must include revenue values.")
            return
        if schedule_df["COGS"].isna().all():
            _show_run_validation_error("COGS Schedule must include at least one value.")
            return

        production_horizon = assumption_tables.get("Production Horizon")
        horizon_filtered = schedule_df
        if production_horizon is not None and not production_horizon.empty:
            start_year = pd.to_numeric(
                production_horizon.get("Start Year"), errors="coerce"
            ).dropna()
            end_year = pd.to_numeric(
                production_horizon.get("End Year"), errors="coerce"
            ).dropna()
            if not start_year.empty and not end_year.empty:
                start = int(start_year.iloc[0])
                end = int(end_year.iloc[0])
                if start > end:
                    _show_run_validation_error(
                        "Production start year must be before the end year."
                    )
                    return
                mask = (horizon_filtered.index.year >= start) & (
                    horizon_filtered.index.year <= end
                )
                horizon_filtered = horizon_filtered.loc[mask]
                if horizon_filtered.empty:
                    _show_run_validation_error(
                        "No schedule periods fall within the selected production horizon."
                    )
                    return
                schedule_df = horizon_filtered
        combined_supplementary = _sync_asset_schedule_from_capex_in_supplementary(
            dict(supplementary_tables)
        )
        for name, table in assumption_tables.items():
            cleaned = _clean_editor_table(table)
            if cleaned is not None:
                combined_supplementary[f"Assumptions - {name}"] = cleaned

        custom_adjustments = _scenario_controls_value_map(
            assumption_tables.get("Scenario Controls", pd.DataFrame())
        )
        custom_adjustments["Feed cost change (%)"] = float(feed_cost)
        selected_business_type = _selected_business_type(assumption_tables)
        scenario_active_products = _active_products_for_business_type(selected_business_type)
        primary_scenario_product = scenario_active_products[0] if scenario_active_products else "Milk"
        primary_price_driver = _price_change_driver(primary_scenario_product)
        primary_price_adjustment = float(custom_adjustments.get(primary_price_driver, 0.0))

        current_presets = _current_scenario_presets()
        matches_preset = any(
            all(
                np.isclose(
                    float(custom_adjustments.get(driver, 0.0)),
                    float(preset["adjustments"].get(driver, 0.0)),
                )
                for driver in DEFAULT_SCENARIO_ADJUSTMENTS
            )
            for preset in current_presets.values()
        )

        scenario_suite = _build_scenario_suite()

        if not matches_preset:
            suffix = _format_scenario_label(
                primary_scenario_product,
                int(round(primary_price_adjustment)),
                int(round(custom_adjustments["Feed cost change (%)"])),
            )
            custom_label = (
                "Custom Scenario"
                if suffix == "Base Scenario"
                else f"Custom Scenario - {suffix}"
            )
            scenario_suite = _build_scenario_suite(custom_label, custom_adjustments)

        try:
            model, _, scenario_results = _execute_scenario_suite(
                schedule_df,
                valuation_inputs,
                combined_supplementary,
                scenario_suite,
            )
        except ValueError as exc:
            _show_run_validation_error(str(exc))
            return

        if run_feedback_slot is not None:
            run_feedback_slot.success("Scenario suite complete. Opening Dashboard...")
        else:
            st.success("Scenario suite complete")
        st.session_state["model_last_run_at"] = pd.Timestamp.utcnow().strftime(
            "%Y-%m-%d %H:%M UTC"
        )

        assumptions_snapshot = {
            name: table.copy()
            for name, table in assumption_tables.items()
            if isinstance(table, pd.DataFrame)
        }
        detail_snapshot = {
            name: table.copy()
            for name, table in st.session_state.detail_schedules.items()
            if isinstance(table, pd.DataFrame)
        }
        for payload in scenario_results.values():
            payload["assumption_tables"] = assumptions_snapshot
            payload["detail_tables"] = detail_snapshot

        st.session_state.all_scenario_results = scenario_results

        previous_selection = st.session_state.get("selected_scenario_name")
        if previous_selection not in scenario_results:
            previous_selection = "Base Case Scenario"
            if previous_selection not in scenario_results:
                previous_selection = next(iter(scenario_results))

        st.session_state.selected_scenario_name = previous_selection
        st.session_state.results = scenario_results[previous_selection]
        st.session_state.excel_bytes_map = {}
        st.session_state[MODEL_VIEW_CACHE_KEY] = {}
        st.session_state[MODEL_VALIDATION_CACHE_KEY] = {}
        st.session_state[MODEL_LAST_RUN_VERSION_KEY] = int(
            _safe_session_state_get(MODEL_INPUT_VERSION_KEY, 0) or 0
        )
        st.session_state[MODEL_RESULTS_STALE_KEY] = False
        st.session_state["_pending_main_section_selector"] = "Dashboard"
        _maybe_rerun()

    results = st.session_state.results
    results_stale = bool(st.session_state.get(MODEL_RESULTS_STALE_KEY, True))
    _render_excel_download_panel(excel_download_container, results, results_stale)

    if results is None:
        if active_main_section in {"Financials", "Dashboard"}:
            st.info(
                "Complete the assumptions and schedules, then press *Run model* "
                "to generate the scenario outputs."
            )
    elif active_main_section in {"Financials", "Dashboard"}:
        _render_stale_results_notice()
        model = results["model"]
        scenario = results["scenario"]
        selected_scenario = results.get("selected_scenario", "Scenario")
        model.scenario_name = selected_scenario

        valuation_issues = _valuation_diagnostic_messages(model)
        if valuation_issues:
            st.warning("Valuation diagnostics: " + " ".join(f"- {msg}" for msg in valuation_issues))

        valuation_summary = results.get("valuation", {}) or {}
        debt_capacity_annual = results.get("debt_capacity_annual")
        if not isinstance(debt_capacity_annual, pd.DataFrame):
            debt_capacity_annual = pd.DataFrame()
        valuation_metrics = {
            "WACC": valuation_summary.get("discount_rate", model.wacc()),
            "NPV": valuation_summary.get("npv"),
            "IRR": valuation_summary.get("irr"),
            "Terminal Value": valuation_summary.get("terminal_value", model.terminal_value()),
            "Min DSCR": (
                pd.to_numeric(debt_capacity_annual.get("DSCR"), errors="coerce").min()
                if not debt_capacity_annual.empty and "DSCR" in debt_capacity_annual.columns
                else None
            ),
            "Min Cash Headroom": (
                pd.to_numeric(
                    debt_capacity_annual.get("Cash Reserve Headroom"), errors="coerce"
                ).min()
                if not debt_capacity_annual.empty
                and "Cash Reserve Headroom" in debt_capacity_annual.columns
                else None
            ),
        }
        summary_cols = st.columns(len(valuation_metrics))
        for idx, (label, value) in enumerate(valuation_metrics.items()):
            if value is None or pd.isna(value):
                summary_cols[idx].metric(label, "N/A")
            elif label in {"WACC", "IRR"}:
                summary_cols[idx].metric(label, f"{value * 100:.2f}%")
            elif label == "Min DSCR":
                summary_cols[idx].metric(label, f"{value:.2f}x")
            else:
                summary_cols[idx].metric(label, f"{value:,.2f}")

    if active_main_section == "Dashboard":
        st.subheader("Dashboard")
        if results is None:
            st.info("Run the model to populate the dashboard charts.")
            st.markdown("---")
            st.subheader("Supplementary Schedules")
            st.info("Supplementary schedules will appear once the model has been run.")
        else:
            _render_stale_results_notice()
            reporting_context = _cached_result_view(
                "reporting_views",
                results.get("selected_scenario", "Scenario"),
                lambda: _reporting_views_for_result(results),
            )
            entity_options = reporting_context.get("entity_options", ["Consolidated"])
            default_entity = reporting_context.get("default_entity", entity_options[0])
            if st.session_state.get("reporting_entity_selector") not in entity_options:
                st.session_state["reporting_entity_selector"] = default_entity
            selected_reporting_entity = st.selectbox(
                "Reporting entity",
                options=entity_options,
                key="reporting_entity_selector",
            )
            dashboard_outputs = _cached_result_view(
                "dashboard_outputs",
                results.get("selected_scenario", "Scenario"),
                lambda: _dashboard_outputs_for_entity(results, selected_reporting_entity),
                extra=selected_reporting_entity,
            )
            scenario = dashboard_outputs["scenario"]
            valuation_summary = dashboard_outputs["valuation_summary"]
            model_audit = dashboard_outputs["model_audit"]
            working_capital_annual = dashboard_outputs["working_capital_annual"]
            debt_capacity_annual = dashboard_outputs["debt_capacity_annual"]
            ufcf_schedule_annual = dashboard_outputs["ufcf_schedule_annual"]
            kpis = dashboard_outputs["kpis"]
            break_even = dashboard_outputs["break_even"]
            pricing_assumptions = dashboard_outputs["pricing_assumptions"]
            scenario_comparison = _cached_result_view(
                "scenario_viability",
                results.get("selected_scenario", "Scenario"),
                lambda: _reporting_scenario_viability_table(
                    st.session_state.get("all_scenario_results", {}),
                    selected_reporting_entity,
                ),
                extra=selected_reporting_entity,
            )

            st.markdown("#### Investor Viability Snapshot")
            viability_cols = st.columns(4)
            viability_metrics = [
                (
                    "NPV",
                    valuation_summary.get("npv"),
                    "{:,.2f}",
                ),
                (
                    "IRR",
                    valuation_summary.get("irr"),
                    "{:.2%}",
                ),
                (
                    "Payback",
                    valuation_summary.get("payback_years"),
                    "{:.2f} years",
                ),
                (
                    "Covenant Breach Periods",
                    (
                        int(debt_capacity_annual["Covenant Breach"].sum())
                        if not debt_capacity_annual.empty
                        and "Covenant Breach" in debt_capacity_annual.columns
                        else None
                    ),
                    "{:,.0f}",
                ),
            ]
            for idx, (label, value, fmt) in enumerate(viability_metrics):
                if value is None or pd.isna(value):
                    viability_cols[idx].metric(label, "N/A")
                else:
                    viability_cols[idx].metric(label, fmt.format(value))

            if not scenario_comparison.empty:
                st.markdown("#### Scenario Viability Comparison")
                st.dataframe(_format_kpis_for_display(scenario_comparison))

            st.markdown("#### Model Audit")
            _render_model_audit(model_audit)

            if not pricing_assumptions.empty:
                st.markdown("#### Commercial Product View")
                product_view_col1, product_view_col2 = st.columns(2)
                with product_view_col1:
                    product_revenue_summary = dashboard_outputs["product_revenue_summary"]
                    st.markdown("**Revenue by Product**")
                    st.dataframe(product_revenue_summary, use_container_width=True)
                    if not product_revenue_summary.empty:
                        product_chart = product_revenue_summary.set_index("Product")["Total Revenue"]
                        st.bar_chart(product_chart)
                with product_view_col2:
                    product_qty_summary = dashboard_outputs["product_qty_summary"]
                    st.markdown("**Quantity by Period**")
                    st.dataframe(product_qty_summary, use_container_width=True)
                    qty_chart = product_qty_summary.set_index("Period") if "Period" in product_qty_summary.columns else pd.DataFrame()
                    if not qty_chart.empty:
                        st.line_chart(qty_chart)

            st.subheader("KPIs (Annual)")
            st.dataframe(_format_kpis_for_display(kpis))

            col1, col2 = st.columns(2)
            with col1:
                st.markdown("#### Revenue vs NPAT")
                revenue_chart_cols = [
                    col for col in ["Revenue_adj", "NPAT_adj", "Revenue", "NPAT"] if col in scenario.columns
                ]
                if {"Revenue_adj", "NPAT_adj"}.issubset(scenario.columns):
                    st.line_chart(scenario[["Revenue_adj", "NPAT_adj"]])
                elif {"Revenue", "NPAT"}.issubset(scenario.columns):
                    st.line_chart(scenario[["Revenue", "NPAT"]])
                elif revenue_chart_cols:
                    st.line_chart(scenario[revenue_chart_cols])
                else:
                    st.info("Revenue and profitability series are not available for this view.")
                st.markdown("#### Expense Breakdown")
                expense_cols = [
                    col
                    for col in [
                        "COGS_adj",
                        "Variable Expenses",
                        "Fixed Expenses",
                        "Direct Wages",
                        "Admin Wages",
                    ]
                    if col in scenario
                ]
                if expense_cols:
                    st.area_chart(scenario[expense_cols])
                else:
                    st.info("Add expense series to the schedule to view this chart.")

            with col2:
                st.markdown("#### Gross Margin vs EBITDA")
                if {"Gross Margin_adj", "EBITDA_adj"}.issubset(scenario.columns):
                    st.line_chart(scenario[["Gross Margin_adj", "EBITDA_adj"]])
                elif {"Gross Margin", "EBITDA"}.issubset(scenario.columns):
                    st.line_chart(scenario[["Gross Margin", "EBITDA"]])
                else:
                    st.info("Margin series are not available for this view.")
                st.markdown("#### Break-even Revenue")
                if not break_even.empty and "Break-even Revenue" in break_even.columns:
                    st.bar_chart(break_even["Break-even Revenue"])
                else:
                    st.info("Break-even schedule is not available for this view.")

            viability_col1, viability_col2 = st.columns(2)
            with viability_col1:
                st.markdown("#### Working Capital")
                wc_cols = [
                    col
                    for col in [
                        "Accounts Receivable",
                        "Inventory",
                        "Accounts Payable",
                        "Net Working Capital",
                    ]
                    if not working_capital_annual.empty and col in working_capital_annual.columns
                ]
                if wc_cols:
                    st.line_chart(working_capital_annual[wc_cols])
                else:
                    st.info("Working-capital schedule becomes available after valuation inputs are assembled.")

                st.markdown("#### UFCF")
                if not ufcf_schedule_annual.empty and "UFCF" in ufcf_schedule_annual.columns:
                    st.bar_chart(ufcf_schedule_annual["UFCF"])
                else:
                    st.info("UFCF schedule is not available for this scenario.")

            with viability_col2:
                st.markdown("#### Debt Capacity")
                debt_cols = [
                    col
                    for col in ["DSCR", "Interest Coverage", "Cash Reserve Headroom"]
                    if not debt_capacity_annual.empty and col in debt_capacity_annual.columns
                ]
                if debt_cols:
                    st.line_chart(debt_capacity_annual[debt_cols])
                else:
                    st.info("Debt-capacity schedule is not available for this scenario.")

                if not debt_capacity_annual.empty:
                    display_cols = [
                        col
                        for col in [
                            "DSCR",
                            "DSCR Headroom",
                            "Interest Coverage",
                            "Interest Coverage Headroom",
                            "Cash Reserve Headroom",
                            "Covenant Breach",
                        ]
                        if col in debt_capacity_annual.columns
                    ]
                    if display_cols:
                        st.markdown("#### Covenant Headroom")
                        st.dataframe(_format_kpis_for_display(debt_capacity_annual[display_cols]))

            st.download_button(
                "Download Scenario CSV",
                scenario.to_csv().encode("utf-8"),
                file_name=f"{_scenario_key_suffix(selected_reporting_entity)}_scenario_timeseries.csv",
                mime="text/csv",
            )

            st.markdown("---")
            st.subheader("Supplementary Schedules")
            supplementary_render = dashboard_outputs["supplementary"]
            for name in [
                "Loan Facilities",
                "Equity Facilities",
                "Capitalisation Table",
                "Capex Schedule",
                "Asset Schedules",
                "Outputs",
                "Benchmark KPIs",
                "Commercial Revenue by Product",
                "Commercial Quantity by Period",
                "Kid Availability Schedule",
                "Kid Routing Schedule",
                "Internal Transfer Schedule",
                "Breeding External Sales Schedule",
                "Downstream Intake Schedule",
                "Breeding Unit Schedule",
                "Destination Unit Schedule",
                "Internal Transfer Elimination Schedule",
                "Unit Revenue Bridge",
                "Reporting Schedule - Breeding",
                "Reporting Schedule - Consolidated",
                "Reporting Schedule - Meat",
                "Reporting Schedule - Milk-Cheese",
                "Reporting Schedule - Combined",
                "Debt Schedule",
                "Equity Schedule",
                "Working Capital Schedule",
                "Debt Capacity Schedule",
                "UFCF Schedule",
                "Model Audit Summary",
                "Model Audit Findings",
            ]:
                _render_table(name, supplementary_render.get(name))

    if active_main_section == "Advanced Analytics":
        st.subheader("Advanced Analytics")
        st.markdown(
            "Use the framework below to configure inputs, assumptions, model drivers, "
            "and scenarios for each analytics tool."
        )
        _render_analytics_framework(results)
        if results is None:
            st.info("Run the model to view advanced analytics.")
        else:
            _render_stale_results_notice()
            scenario = results["scenario"]
            model = results["model"]
            try:
                adv_annual = _cached_result_view(
                    "advanced_analytics",
                    results.get("selected_scenario", "Scenario"),
                    lambda: model.advanced_analytics(scenario, window=3, annual=True),
                )

                def _render_analytics(
                    block_name: str, payload: Dict[str, object], scenario_label: str
                ) -> None:
                    st.markdown(f"#### {block_name} Advanced Analytics")
                    for key, item in payload.items():
                        title = item.get("title", key.replace("_", " ").title())
                        description = item.get("description", "")
                        tables = item.get("tables", {})
                        with st.expander(title, expanded=False):
                            if description:
                                st.caption(description)
                            if isinstance(tables, dict):
                                for table_name, table in tables.items():
                                    st.markdown(f"**{table_name}**")
                                    if not isinstance(table, pd.DataFrame):
                                        st.info("No data available for this table.")
                                        continue

                                    override = _get_analytics_override(
                                        scenario_label, block_name, key, table_name
                                    )
                                    display_df = override if override is not None else table
                                    edit_flag_key = _analytics_edit_flag_key(
                                        scenario_label, block_name, key, table_name
                                    )
                                    editor_key = _analytics_editor_key(
                                        scenario_label, block_name, key, table_name
                                    )

                                    if not isinstance(display_df, pd.DataFrame):
                                        st.info("No data available for this table.")
                                        continue

                                    editing = st.session_state.get(edit_flag_key, False)

                                    if editing:
                                        editor_df, meta = _prepare_editor_table(display_df)
                                        working_key = f"{editor_key}::working"
                                        working_df = st.session_state.get(working_key)
                                        if working_df is None:
                                            working_df = editor_df.copy(deep=True)
                                            st.session_state[working_key] = working_df

                                        st.dataframe(working_df)
                                        if working_df.empty:
                                            st.info(
                                                "This table has no rows to edit. Update the model inputs to populate it."
                                            )
                                        else:
                                            row_indices = list(range(len(working_df)))
                                            selected_row = st.selectbox(
                                                "Select a row to edit",
                                                row_indices,
                                                format_func=lambda idx: _format_row_label(
                                                    working_df, idx
                                                ),
                                                key=f"{editor_key}_row_selector",
                                            )
                                            dtype_map = working_df.dtypes.to_dict()
                                            row_series = working_df.iloc[selected_row]
                                            with st.form(f"{editor_key}_row_form"):
                                                updated_values: Dict[str, Any] = {}
                                                for column in working_df.columns:
                                                    widget_key = (
                                                        f"{editor_key}::{selected_row}::{column}"
                                                    )
                                                    cell_value = row_series[column]
                                                    updated_values[column] = _render_row_input(
                                                        column,
                                                        cell_value,
                                                        dtype_map[column],
                                                        widget_key,
                                                    )

                                                submitted = st.form_submit_button(
                                                    "Apply Row Changes"
                                                )
                                                if submitted:
                                                    for column, raw_value in (
                                                        updated_values.items()
                                                    ):
                                                        coerced = _coerce_row_value(
                                                            raw_value, dtype_map[column]
                                                        )
                                                        working_df.iat[
                                                            selected_row,
                                                            working_df.columns.get_loc(
                                                                column
                                                            ),
                                                        ] = coerced
                                                    st.session_state[working_key] = (
                                                        working_df
                                                    )
                                                    _maybe_rerun()

                                        action_cols = st.columns(3)
                                        if action_cols[0].button(
                                            "Save Changes",
                                            key=f"save_{editor_key}",
                                        ):
                                            current_df = st.session_state.get(
                                                working_key, editor_df
                                            )
                                            restored = _restore_editor_table(
                                                current_df, meta
                                            )
                                            _set_analytics_override(
                                                scenario_label,
                                                block_name,
                                                key,
                                                table_name,
                                                restored,
                                            )
                                            st.session_state.pop(working_key, None)
                                            st.session_state[edit_flag_key] = False
                                            _maybe_rerun()

                                        if action_cols[1].button(
                                            "Cancel",
                                            key=f"cancel_{editor_key}",
                                        ):
                                            st.session_state.pop(working_key, None)
                                            st.session_state[edit_flag_key] = False
                                            _maybe_rerun()

                                        if action_cols[2].button(
                                            "Restore Original",
                                            key=f"reset_{editor_key}",
                                        ):
                                            _clear_analytics_override(
                                                scenario_label,
                                                block_name,
                                                key,
                                                table_name,
                                            )
                                            st.session_state.pop(working_key, None)
                                            st.session_state[edit_flag_key] = False
                                            _maybe_rerun()
                                    else:
                                        st.dataframe(display_df)
                                        if override is not None:
                                            st.caption("Manual override applied.")
                                        button_cols = st.columns(2)
                                        if button_cols[0].button(
                                            "Edit Table",
                                            key=f"edit_{editor_key}",
                                        ):
                                            st.session_state[edit_flag_key] = True
                                            _maybe_rerun()

                                        if button_cols[1].button(
                                            "Clear Manual Override",
                                            key=f"clear_{editor_key}",
                                            disabled=override is None,
                                        ):
                                            _clear_analytics_override(
                                                scenario_label,
                                                block_name,
                                                key,
                                                table_name,
                                            )
                                            _maybe_rerun()
                            else:
                                st.info("No tables available for this analysis.")

                selected_scenario_name = results.get("selected_scenario", "Scenario")
                _render_analytics("Annual", adv_annual, selected_scenario_name)
            except ValueError as exc:
                st.info(str(exc))

    if active_main_section == "AI Decision Making":
        _render_ai_orchestration_layer(results)

# ---------------------------------------------------------------------------
# Scenario state hooks — called by the parent NumQuants shell.
# ---------------------------------------------------------------------------

_GOAT_STATE_KEYS = [
    "core_schedule",           # pd.DataFrame — the main financial schedule
    "detail_schedules",        # dict[str, pd.DataFrame] — supplementary detail tabs
    "assumptions",             # dict[str, pd.DataFrame] — assumption tables
    "supplementary",           # dict[str, pd.DataFrame] — supplementary tables
    "selected_scenario_name",  # str
    "schedule_period_type",    # str ('monthly' | 'quarterly')
    DEFAULT_INPUT_CONFIG_KEY,  # "default_input_templates"
]


def get_state() -> dict:
    """Snapshot all user-editable schedule and assumption data."""
    import streamlit as _st
    return {k: _st.session_state[k] for k in _GOAT_STATE_KEYS if k in _st.session_state}


def set_state(state: dict) -> None:
    """Restore a previously saved state snapshot.

    Writes to session_state before main() runs. main() checks
    'if key not in st.session_state' before setting defaults, so restored
    values take precedence over defaults.
    """
    import streamlit as _st
    for k, v in state.items():
        if k in _GOAT_STATE_KEYS:
            _st.session_state[k] = v


if __name__ == "__main__":
    main()



