from __future__ import annotations

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from goat_financial_model import InputSchedule
from goat_financial_model.reporting import generate_excel_workbook, prepare_export_bundle


def _build_sample_schedule() -> pd.DataFrame:
    periods = pd.date_range("2026-01-31", periods=6, freq="M")
    revenue = pd.Series([100000, 110000, 120000, 130000, 140000, 150000], index=periods, dtype=float)
    cogs = revenue * 0.45
    gross_margin = revenue - cogs
    variable_expenses = revenue * 0.1
    direct_wages = revenue * 0.06
    fixed_expenses = pd.Series(12000.0, index=periods)
    admin_wages = pd.Series(4500.0, index=periods)
    ebitda = gross_margin - variable_expenses - direct_wages - fixed_expenses - admin_wages
    depreciation = pd.Series(2500.0, index=periods)
    ebit = ebitda - depreciation
    interest = pd.Series(600.0, index=periods)
    npbt = ebit - interest
    tax = npbt * 0.25
    npat = npbt - tax
    cfo = ebitda - 5000.0
    cfi = pd.Series(-8000.0, index=periods)
    cff = pd.Series(2500.0, index=periods)
    net_cash_flow = cfo + cfi + cff
    current_assets = pd.Series(60000.0, index=periods) + net_cash_flow.cumsum()
    non_current_assets = pd.Series(125000.0, index=periods)
    current_liabilities = pd.Series(18000.0, index=periods)
    non_current_liabilities = pd.Series(42000.0, index=periods)
    equity = current_assets + non_current_assets - current_liabilities - non_current_liabilities
    return pd.DataFrame(
        {
            "Revenue": revenue,
            "COGS": cogs,
            "Gross Margin": gross_margin,
            "Variable Expenses": variable_expenses,
            "Direct Wages": direct_wages,
            "Fixed Expenses": fixed_expenses,
            "Admin Wages": admin_wages,
            "EBITDA": ebitda,
            "Depreciation & Amortization": depreciation,
            "EBIT": ebit,
            "Interest Expense": interest,
            "Net Profit Before Tax": npbt,
            "Tax Expense": tax,
            "NPAT": npat,
            "CFO": cfo,
            "CFI": cfi,
            "CFF": cff,
            "Net Cash Flow": net_cash_flow,
            "Current Assets": current_assets,
            "Non-current Assets": non_current_assets,
            "Current Liabilities": current_liabilities,
            "Non-current Liabilities": non_current_liabilities,
            "Equity": equity,
        }
    )


def test_excel_export_keeps_kpi_units_and_adds_chart_pack():
    schedule_df = _build_sample_schedule()
    model = InputSchedule(
        data=schedule_df,
        valuation_inputs={"WACC": 0.12, "Terminal Value": 100000.0},
    ).to_model()
    kpis = model.kpis(schedule_df, annual=True)
    bundle = prepare_export_bundle(
        model,
        scenario_name="Base Case Scenario",
        author_name="Analyst",
        base_df=schedule_df,
        scenario_df=schedule_df,
        kpis_df=kpis,
        break_even_df=model.break_even(schedule_df, annual=True),
        scenario_inputs={"WACC": 0.12},
    )

    workbook_bytes = generate_excel_workbook(bundle)
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert "Overview" in workbook.sheetnames
    assert "Charts" in workbook.sheetnames
    assert "KPIs (Annual)" in workbook.sheetnames
    assert any(name.startswith("Statement of Financial Performa") for name in workbook.sheetnames)
    assert any(name.startswith("Statement of Financial Positi") for name in workbook.sheetnames)
    assert any(name.startswith("Statement of Cash Flows") for name in workbook.sheetnames)
    assert len(workbook["Charts"]._charts) >= 3

    sheet = workbook["KPIs (Annual)"]
    headers = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}
    npv_cell = sheet.cell(row=2, column=headers["NPV"])
    irr_cell = sheet.cell(row=2, column=headers["IRR"])

    assert npv_cell.value == pytest.approx(kpis.iloc[0]["NPV"])
    assert irr_cell.value == pytest.approx(kpis.iloc[0]["IRR"])
    assert irr_cell.number_format == "0.0%"


def test_export_surfaces_statement_build_failures():
    schedule_df = _build_sample_schedule()

    class BrokenStatementModel:
        def __init__(self, model):
            self._model = model

        def __getattr__(self, name):
            return getattr(self._model, name)

        def statement_of_financial_position(self, scenario_df, annual=True):
            raise ValueError("balance sheet mapping failed")

    base_model = InputSchedule(
        data=schedule_df,
        valuation_inputs={"WACC": 0.12, "Terminal Value": 100000.0},
    ).to_model()
    model = BrokenStatementModel(base_model)

    bundle = prepare_export_bundle(
        model,
        scenario_name="Broken Statement Scenario",
        author_name="Analyst",
        base_df=schedule_df,
        scenario_df=schedule_df,
        kpis_df=base_model.kpis(schedule_df, annual=True),
        break_even_df=base_model.break_even(schedule_df, annual=True),
    )

    assert bundle.statement_errors["Statement of Financial Position"] == "balance sheet mapping failed"
    workbook = load_workbook(BytesIO(generate_excel_workbook(bundle)))

    assert "Statement Warnings" in workbook.sheetnames
    warning_sheet = workbook["Statement Warnings"]
    assert warning_sheet["A2"].value == "Statement of Financial Position"
    assert warning_sheet["B2"].value == "balance sheet mapping failed"
